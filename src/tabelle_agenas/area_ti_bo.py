"""
Tabella AGENAS area combinata Anestesia, TI e Blocco Operatorio.
"""

from openpyxl.styles import Font, PatternFill

from src.stili_excel import (
    THIN_BORDER, FILL_A, FILL_B, FILL_HEADER,
    FONT_HEADER, ALIGN_CENTER,
)


# ============================================================
# TABELLA COMBINATA AREA TI + BLOCCO OPERATORIO
# ============================================================

def _scrivi_tabella_agenas_area_ti_bo(
        ws, start_row, df_area, fabb_ti_presidio, fabb_so_presidio,
        mapping_uo_ti, sede_completa):
    """Tabella riepilogativa AREA ANESTESIA, TERAPIA INTENSIVA E
    BLOCCO OPERATORIO.

    Unisce i fabbisogni TI (§ 8.1.1) e BO (§ 8.1.2), raccoglie
    TUTTO il personale afferente alla SC/SSD di Anestesia e Terapia
    Intensiva e calcola le carenze/eccedenze sull'area combinata.

    Restituisce la riga successiva libera.
    """
    import re as _re

    fabb_ti = fabb_ti_presidio or {}
    fabb_so = fabb_so_presidio or {}

    if not fabb_ti and not fabb_so:
        return start_row

    # --- Stili locali ---
    FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
    FONT_NORMAL  = Font(size=10)

    row = start_row + 1
    N_COLS_MERGE = 8

    # Titolo sezione
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="AREA ANESTESIA, TERAPIA INTENSIVA E BLOCCO OPERATORIO"
            ).font = FONT_SECTION
    row += 1

    # PL intensivi e n. sale dal fabbisogno
    pl_ti = 0
    for d in fabb_ti.values():
        pl_ti = d.get('pl', 0)
        break
    n_sale = 0
    for d in fabb_so.values():
        n_sale = d.get('n_sale', 0)
        break

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value=f"Presidio: {sede_completa}  \u2013  "
                  f"PL Intensivi: {pl_ti}  \u2013  "
                  f"Sale Operative: {n_sale}"
            ).font = FONT_NORMAL
    row += 2

    # Intestazioni
    headers = ['Profilo', 'FTE TI', 'FTE BO', 'FTE Area',
               'T.I.', 'T.D.', 'Totale', 'Esito']
    N_COLS = len(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # Profili unificati (merge TI + BO)
    _PROFILI_AREA = [
        {
            'label': 'Dirigenti Medici',
            'key': 'DIRIGENTE_MEDICO',
            'ti_key': 'DIRIGENTE_MEDICO',
            'so_key': 'DIRIGENTE_MEDICO_ANESTESISTA',
        },
        {
            'label': 'Infermieri',
            'key': 'INFERMIERE',
            'ti_key': 'INFERMIERE',
            'so_key': 'INFERMIERE',
        },
        {
            'label': 'Operatori Socio Sanitari',
            'key': 'OPERATORE_SOCIO_SANITARIO',
            'ti_key': None,
            'so_key': 'OPERATORE_SOCIO_SANITARIO',
        },
    ]

    # Mapping profilo personale → profilo unificato dell'area
    _MAPPING_PROFILI_AREA = [
        {'pattern': r'DIRIGENTE MEDICO|DIR\.? MED',
         'profilo_agenas': 'DIRIGENTE_MEDICO'},
        {'pattern': r'INFERMIERE|INFERMIERA',
         'profilo_agenas': 'INFERMIERE'},
        {'pattern': r'OPERATORE SOCIO',
         'profilo_agenas': 'OPERATORE_SOCIO_SANITARIO'},
    ]

    # --- Conta in servizio: TUTTO il personale della UO ---
    in_servizio = {}
    in_servizio_td = {}
    for _, r in df_area.iterrows():
        ssd = str(r.get('_REPARTO', ''))
        profilo = str(r.get('Profilo Professionale', '')).strip().upper()
        quantita_ti = int(r.get('Quantità T.I.', 0))
        quantita_td = int(r.get('Quantità T.D.', 0))

        # Verifica se la UO è dell'area Anestesia e TI
        uo_match = False
        for m_uo in mapping_uo_ti:
            if _re.search(m_uo['pattern'], ssd, _re.IGNORECASE):
                uo_match = True
                break
        if not uo_match:
            continue

        # Mappa il profilo (mapping unificato)
        for m_prof in _MAPPING_PROFILI_AREA:
            if _re.search(m_prof['pattern'], profilo):
                pa = m_prof['profilo_agenas']
                in_servizio[pa] = in_servizio.get(pa, 0) + quantita_ti
                in_servizio_td[pa] = in_servizio_td.get(pa, 0) + quantita_td
                break

    # --- Righe dati ---
    FILL_OK = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                          fill_type='solid')
    FILL_CARENZA = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                               fill_type='solid')
    toggle = 0
    for prof_info in _PROFILI_AREA:
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1

        fte_ti = 0
        if prof_info['ti_key'] and prof_info['ti_key'] in fabb_ti:
            fte_ti = fabb_ti[prof_info['ti_key']]['fte_atteso']
        fte_so = 0
        if prof_info['so_key'] and prof_info['so_key'] in fabb_so:
            fte_so = fabb_so[prof_info['so_key']]['fte_atteso']
        fte_area = fte_ti + fte_so

        servizio_ti = in_servizio.get(prof_info['key'], 0)
        servizio_td = in_servizio_td.get(prof_info['key'], 0)
        servizio = servizio_ti + servizio_td

        if servizio < fte_area:
            delta = fte_area - servizio
            esito = f'CARENZA ({delta})'
            fill_esito = FILL_CARENZA
        elif servizio > fte_area:
            delta = servizio - fte_area
            esito = f'ECCEDENZA (+{delta})'
            fill_esito = FILL_OK
        else:
            esito = 'CONFORME'
            fill_esito = FILL_OK

        vals = [
            prof_info['label'],
            fte_ti, fte_so, fte_area,
            servizio_ti, servizio_td, servizio, esito,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        row += 1

    # Note piè di pagina
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="FTE TI = (PL / Rapporto) \u00d7 (24 \u00d7 365) "
                  "/ Ore_annue \u00d7 1,15  |  "
                  "FTE BO = N.Sale \u00d7 Staff/Sala \u00d7 (24 \u00d7 250) "
                  "/ Ore_annue \u00d7 1,15.  "
                  "Le carenze sono calcolate sull\u2019area complessiva: "
                  "tutto il personale afferente alla SC/SSD di Anestesia "
                  "e Terapia Intensiva (inclusi i CDC condivisi)."
            ).font = Font(italic=True, size=9, color='555555')

    return row + 1
