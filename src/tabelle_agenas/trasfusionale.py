"""
Tabelle AGENAS trasfusionale (standard + fabbisogno UOC speciale).
Estratte da report_fabbisogno.py per modularizzazione.

Funzioni incluse:
  - _scrivi_tabella_agenas_trasfusionale
  - _scrivi_tabella_fabbisogno_uoc_trasfusionale
"""

from openpyxl.styles import Font, Alignment, PatternFill

from src.stili_excel import (
    THIN_BORDER, FILL_A, FILL_B, FILL_HEADER,
    FONT_HEADER, ALIGN_CENTER,
)


# ============================================================
# TABELLA AGENAS TRASFUSIONALE IN CALCE AL FOGLIO
# ============================================================

def _scrivi_tabella_agenas_trasfusionale(
        ws, start_row, df_luogo, fabb_trasf_presidio,
        mapping_uo_trasf, mapping_profili_trasf, sede_completa,
        livello_label, df_completo=None):
    """Aggiunge una tabella riepilogativa AGENAS trasfusionale
    in calce al foglio di dettaglio di un presidio ospedaliero.

    Se *df_completo* è fornito, mostra la distribuzione del personale
    trasfusionale su tutte le sedi dell'UOC con colonne per-sede,
    totale, range AGENAS e valutazione sulla sede principale.

    Restituisce la riga successiva libera.
    """
    import re as _re

    if not fabb_trasf_presidio:
        return start_row

    # --- Stili locali ---
    FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
    FONT_NORMAL  = Font(size=10)
    FONT_NOTE    = Font(italic=True, size=9, color='555555')
    FILL_OK      = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                               fill_type='solid')
    FILL_CARENZA = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                               fill_type='solid')

    # --- Conta personale trasfusionale per sede e profilo ---
    sedi_trasf = {}      # {sede_fisica: {profilo_agenas: count_ti}}
    sedi_trasf_td = {}   # {sede_fisica: {profilo_agenas: count_td}}

    def _conta(df_src):
        for _, r in df_src.iterrows():
            ssd = str(r.get('_REPARTO', ''))
            profilo = str(
                r.get('Profilo Professionale', '')).strip().upper()
            quantita_ti = int(r.get('Quantità T.I.', 0))
            quantita_td = int(r.get('Quantità T.D.', 0))
            sede_r = str(r.get('Sede', ''))
            uo_ok = False
            for m_uo in mapping_uo_trasf:
                if _re.search(m_uo['pattern'], ssd.upper()):
                    uo_ok = True
                    break
            if not uo_ok:
                continue
            for m_prof in mapping_profili_trasf:
                if _re.search(m_prof['pattern'], profilo):
                    pa = m_prof['profilo_agenas']
                    sedi_trasf.setdefault(sede_r, {})
                    sedi_trasf[sede_r][pa] = (
                        sedi_trasf[sede_r].get(pa, 0) + quantita_ti)
                    sedi_trasf_td.setdefault(sede_r, {})
                    sedi_trasf_td[sede_r][pa] = (
                        sedi_trasf_td[sede_r].get(pa, 0) + quantita_td)
                    break

    if df_completo is not None:
        _conta(df_completo)
    else:
        _conta(df_luogo)

    # Ordina: sede principale prima, poi alfabeticamente
    sedi_ordered = sorted(
        sedi_trasf.keys(),
        key=lambda s: (
            0 if sede_completa in s or s in sede_completa else 1, s),
    )

    multi_sede = df_completo is not None and len(sedi_ordered) > 1

    def _abbrev(sede_name):
        """'CAMPOBASSO - P.O. CARDARELLI' -> 'Cardarelli'"""
        parts = sede_name.split(' - ')
        if len(parts) > 1:
            short = parts[-1].replace('P.O. ', '').strip().title()
            short = short.replace('San ', 'S.')
            return short
        return sede_name[:15]

    livello_leggibile = {
        'PRESIDIO_I_LIVELLO':  'Presidio di I livello',
        'PRESIDIO_II_LIVELLO': 'Presidio di II livello',
    }.get(livello_label, livello_label)

    nomi_profili = {
        'DIRIGENTE_SANITARIO':       'Dir. Sanitari (*)',
        'INFERMIERE':                'Infermieri',
        'OPERATORE_SOCIO_SANITARIO': 'Operatori Socio Sanitari',
    }

    row = start_row + 1

    if multi_sede:
        # ===== Layout multi-sede =====
        abbrevs = [_abbrev(s) for s in sedi_ordered]
        headers = (
            ['Profilo'] + abbrevs
            + ['T.I.', 'T.D.', 'Totale UOC', 'Range AGENAS', 'Esito']
        )
        N_COLS = len(headers)

        # Titolo
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS)
        ws.cell(row=row, column=1,
                value="AREA MEDICINA TRASFUSIONALE - Standard AGENAS"
                ).font = FONT_SECTION
        row += 1

        # Sotto-titolo
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS)
        ws.cell(row=row, column=1,
                value=(f"UOC Medicina Trasfusionale  \u2013  "
                       f"{livello_leggibile} ({_abbrev(sede_completa)})")
                ).font = FONT_NORMAL
        row += 1

        # Sedi gestite
        sedi_desc = ', '.join(f"P.O. {a}" for a in abbrevs)
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS)
        ws.cell(row=row, column=1,
                value=f"Sedi gestite: {sedi_desc}"
                ).font = Font(italic=True, size=10)
        row += 2

        # Intestazioni
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.alignment = ALIGN_CENTER
            c.border = THIN_BORDER
        row += 1

        # Righe dati
        toggle = 0
        for prof_key, rng in fabb_trasf_presidio.items():
            fill = FILL_A if toggle % 2 == 0 else FILL_B
            toggle += 1
            v_min, v_max = rng['min'], rng['max']
            range_str = f'{v_min} - {v_max}'

            counts = [sedi_trasf.get(s, {}).get(prof_key, 0)
                      + sedi_trasf_td.get(s, {}).get(prof_key, 0)
                      for s in sedi_ordered]
            tot_ti = sum(sedi_trasf.get(s, {}).get(prof_key, 0)
                         for s in sedi_ordered)
            tot_td = sum(sedi_trasf_td.get(s, {}).get(prof_key, 0)
                         for s in sedi_ordered)
            totale = tot_ti + tot_td
            servizio_princ = counts[0]   # sede principale (TI+TD)

            if servizio_princ < v_min:
                esito = f'CARENZA (min {v_min - servizio_princ})'
                fill_esito = FILL_CARENZA
            elif servizio_princ > v_max:
                esito = f'ECCEDENZA (+{servizio_princ - v_max})'
                fill_esito = FILL_OK
            else:
                esito = 'IN RANGE'
                fill_esito = FILL_OK

            vals = ([nomi_profili.get(prof_key, prof_key)]
                    + counts + [tot_ti, tot_td, totale, range_str, esito])
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = fill
                c.border = THIN_BORDER
            ws.cell(row=row, column=N_COLS).fill = fill_esito
            row += 1

        # Note piè di pagina
        row += 1
        notes = [
            "(*) Dirigenti Sanitari Medici e Non Medici",
            ("Range AGENAS riferito a singolo centro di I livello "
             "(Tab. 15 AGENAS)."),
            (f"Esito valutato sulla sola sede di "
             f"{_abbrev(sede_completa)}."),
            ("Maggiorazione +15% per turnazione, guardie, "
             "ferie, malattie e altre indisponibilità."),
        ]
        for note in notes:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=N_COLS)
            ws.cell(row=row, column=1, value=note).font = FONT_NOTE
            row += 1

    else:
        # ===== Layout singola sede (originale) =====
        N_COLS_MERGE = 6

        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS_MERGE)
        ws.cell(row=row, column=1,
                value="AREA MEDICINA TRASFUSIONALE - Standard AGENAS"
                ).font = FONT_SECTION
        row += 1

        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS_MERGE)
        ws.cell(row=row, column=1,
                value=f"Presidio: {sede_completa}  \u2013  {livello_leggibile}"
                ).font = FONT_NORMAL
        row += 2

        headers = ['Profilo', 'T.I.', 'T.D.', 'Totale', 'Range AGENAS',
                   'Esito']
        N_COLS = len(headers)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.alignment = ALIGN_CENTER
            c.border = THIN_BORDER
        row += 1

        # Trova in_servizio per la sede corrente
        in_servizio_sede = {}
        in_servizio_sede_td = {}
        for s in sedi_ordered:
            if sede_completa in s or s in sede_completa:
                in_servizio_sede = sedi_trasf.get(s, {})
                in_servizio_sede_td = sedi_trasf_td.get(s, {})
                break
        if not in_servizio_sede and sedi_ordered:
            in_servizio_sede = sedi_trasf.get(sedi_ordered[0], {})
            in_servizio_sede_td = sedi_trasf_td.get(sedi_ordered[0], {})

        toggle = 0
        for prof_key, rng in fabb_trasf_presidio.items():
            fill = FILL_A if toggle % 2 == 0 else FILL_B
            toggle += 1
            servizio_ti = in_servizio_sede.get(prof_key, 0)
            servizio_td = in_servizio_sede_td.get(prof_key, 0)
            servizio = servizio_ti + servizio_td
            v_min, v_max = rng['min'], rng['max']
            range_str = f'{v_min} - {v_max}'

            if servizio < v_min:
                esito = f'CARENZA (min {v_min - servizio})'
                fill_esito = FILL_CARENZA
            elif servizio > v_max:
                esito = f'ECCEDENZA (+{servizio - v_max})'
                fill_esito = FILL_OK
            else:
                esito = 'IN RANGE'
                fill_esito = FILL_OK

            vals = [nomi_profili.get(prof_key, prof_key),
                    servizio_ti, servizio_td, servizio, range_str, esito]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = fill
                c.border = THIN_BORDER
            ws.cell(row=row, column=N_COLS).fill = fill_esito
            row += 1

        row += 1
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS_MERGE)
        ws.cell(row=row, column=1,
                value="(*) Dirigenti Sanitari Medici e Non Medici. "
                      "Maggiorazione +15% per turnazione, guardie, "
                      "ferie, malattie e altre indisponibilità."
                ).font = FONT_NOTE

    return row + 1


# ============================================================
# TABELLA FABBISOGNO UOC TRASFUSIONALE (speciale – Primaria)
# ============================================================

def _scrivi_tabella_fabbisogno_uoc_trasfusionale(
        ws, start_row, df_completo, fabb_speciale, sede_completa):
    """Aggiunge la tabella con il fabbisogno UOC richiesto dalla Primaria.

    Mostra per ogni sede e profilo: in servizio, richiesti, delta.
    Restituisce la riga successiva libera.
    """
    import re as _re

    if not fabb_speciale:
        return start_row

    sedi_cfg    = fabb_speciale['sedi']
    mapping_uo  = fabb_speciale['mapping_uo']
    mapping_pr  = fabb_speciale['mapping_profili']
    ordine      = fabb_speciale['ordine_profili']

    # Conta in servizio: {sede: {chiave_profilo: n}}
    in_servizio = {}
    in_servizio_td = {}
    for _, r in df_completo.iterrows():
        ssd     = str(r.get('_REPARTO', ''))
        profilo = str(r.get('Profilo Professionale', '')).strip().upper()
        qti     = int(r.get('Quantità T.I.', 0))
        qtd     = int(r.get('Quantità T.D.', 0))
        sede_r  = str(r.get('Sede', ''))

        uo_ok = False
        for m in mapping_uo:
            if _re.search(m['pattern'], ssd.upper()):
                uo_ok = True
                break
        if not uo_ok:
            continue

        for mp in mapping_pr:
            if _re.search(mp['pattern'], profilo):
                ck = mp['chiave']
                in_servizio.setdefault(sede_r, {})
                in_servizio[sede_r][ck] = (
                    in_servizio[sede_r].get(ck, 0) + qti)
                in_servizio_td.setdefault(sede_r, {})
                in_servizio_td[sede_r][ck] = (
                    in_servizio_td[sede_r].get(ck, 0) + qtd)
                break

    # Labels leggibili
    labels = {mp['chiave']: mp['label'] for mp in mapping_pr}

    # Stili
    FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
    FONT_NORMAL  = Font(size=10)
    FONT_NOTE    = Font(italic=True, size=9, color='555555')
    FILL_OK      = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                               fill_type='solid')
    FILL_CARENZA = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                               fill_type='solid')

    row = start_row + 1

    # --- Intestazione ---
    # Colonne: Profilo | per-sede (T.I. / T.D. / Tot / Rich / Delta) × N
    #        | UOC (T.I. / T.D. / Tot / Rich / Delta)
    n_sedi = len(sedi_cfg)
    SUB = 5   # sotto-colonne per gruppo
    N_COLS = 1 + SUB * n_sedi + SUB

    # Titolo
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1,
            value="FABBISOGNO UOC MEDICINA TRASFUSIONALE \u2013 "
                  "Dotazione segnalata dal Primario"
            ).font = FONT_SECTION
    row += 1

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1,
            value="Fabbisogno espresso per operare in sicurezza "
                  "su 3 sedi ospedaliere (non rientra negli standard "
                  "AGENAS per singolo presidio)"
            ).font = FONT_NOTE
    row += 2

    # Riga intestazioni: merge per gruppo sede
    # Prima riga: nomi sedi (merge SUB colonne ciascuna) + UOC Totale
    col = 2
    for s in sedi_cfg:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + SUB - 1)
        c = ws.cell(row=row, column=col, value=s['abbreviazione'])
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        for ci2 in range(col + 1, col + SUB):
            ws.cell(row=row, column=ci2).border = THIN_BORDER
        col += SUB

    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + SUB - 1)
    c = ws.cell(row=row, column=col, value='TOTALE UOC')
    c.font = FONT_HEADER
    c.fill = FILL_HEADER
    c.alignment = ALIGN_CENTER
    c.border = THIN_BORDER
    for ci2 in range(col + 1, col + SUB):
        ws.cell(row=row, column=ci2).border = THIN_BORDER
    row += 1

    # Seconda riga: Profilo + (T.I. | T.D. | Totale | Rich. | Delta) × (n_sedi + 1)
    sub_headers = ['T.I.', 'T.D.', 'Totale', 'Rich.', 'Delta']
    hdr_vals = ['Profilo']
    for _ in range(n_sedi + 1):
        hdr_vals.extend(sub_headers)

    for ci, h in enumerate(hdr_vals, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # --- Righe dati ---
    toggle = 0
    for prof_key in ordine:
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1

        vals = [labels.get(prof_key, prof_key)]
        tot_ti = 0
        tot_td = 0
        tot_rich = 0

        for s in sedi_cfg:
            # Trova la sede corrispondente nei dati
            serv_ti = 0
            serv_td = 0
            for sede_k, counts in in_servizio.items():
                if s['nome'] in sede_k or sede_k in s['nome']:
                    serv_ti = counts.get(prof_key, 0)
                    serv_td = in_servizio_td.get(sede_k, {}).get(
                        prof_key, 0)
                    break
            serv_tot = serv_ti + serv_td
            rich = s['richiesti'].get(prof_key, 0)
            delta = serv_tot - rich
            tot_ti += serv_ti
            tot_td += serv_td
            tot_rich += rich
            vals.extend([serv_ti, serv_td, serv_tot, rich, delta])

        grand_tot = tot_ti + tot_td
        tot_delta = grand_tot - tot_rich
        vals.extend([tot_ti, tot_td, grand_tot, tot_rich, tot_delta])

        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
            c.alignment = ALIGN_CENTER if ci > 1 else Alignment()

        # Colora le celle Delta (ultima sotto-colonna di ogni gruppo)
        for si in range(n_sedi + 1):
            delta_col = 1 + (si + 1) * SUB   # ultima sotto-col del gruppo
            delta_val = ws.cell(row=row, column=delta_col).value
            if delta_val is not None and delta_val < 0:
                ws.cell(row=row, column=delta_col).fill = FILL_CARENZA
            elif delta_val is not None and delta_val >= 0:
                ws.cell(row=row, column=delta_col).fill = FILL_OK

        row += 1

    # --- Riga TOTALE ---
    row_tot = row
    ws.cell(row=row_tot, column=1, value='TOTALE').font = Font(bold=True, size=10)
    ws.cell(row=row_tot, column=1).border = THIN_BORDER

    for ci in range(2, N_COLS + 1):
        total = 0
        for r_data in range(row_tot - len(ordine), row_tot):
            v = ws.cell(row=r_data, column=ci).value
            if isinstance(v, (int, float)):
                total += int(v)
        c = ws.cell(row=row_tot, column=ci, value=total)
        c.font = Font(bold=True, size=10)
        c.border = THIN_BORDER
        c.alignment = ALIGN_CENTER

        # Colora Delta totale
        is_delta_col = (ci - 1) % SUB == 0 and ci > 1
        if is_delta_col:
            if total < 0:
                c.fill = FILL_CARENZA
            else:
                c.fill = FILL_OK

    row = row_tot + 2

    # --- Note sedi ---
    for s in sedi_cfg:
        if s.get('nota'):
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=N_COLS)
            ws.cell(row=row, column=1,
                    value=f"{s['abbreviazione']}: {s['nota']}"
                    ).font = FONT_NOTE
            row += 1

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1,
            value="Fonte: dotazione segnalata dal Primario UOC Medicina "
                  "Trasfusionale per operare in sicurezza."
            ).font = FONT_NOTE

    return row + 1
