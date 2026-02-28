"""
Tabella AGENAS radiologia in calce al foglio presidio.
"""

from openpyxl.styles import Font, PatternFill

from src.stili_excel import (
    THIN_BORDER, FILL_A, FILL_B, FILL_HEADER,
    FONT_HEADER, ALIGN_CENTER,
)


# ============================================================
# TABELLA AGENAS RADIOLOGIA IN CALCE AL FOGLIO
# ============================================================

def _scrivi_tabella_agenas_radiologia(
        ws, start_row, df_luogo, fabb_radio_presidio,
        mapping_uo_radio, mapping_profili_radio, sede_completa,
        livello_label, df_area=None):
    """Aggiunge una tabella riepilogativa AGENAS radiologia
    in calce al foglio di dettaglio di un presidio ospedaliero.

    Se df_area è fornito, contabilizza anche il personale
    radiologico presente in sedi territoriali della stessa area.

    Restituisce la riga successiva libera.
    """
    import re as _re

    if not fabb_radio_presidio:
        return start_row

    # --- Stili locali ---
    FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
    FONT_NORMAL  = Font(size=10)

    row = start_row + 1
    N_COLS_MERGE = 6  # per merge titolo/sottotitolo

    # Titolo sezione
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    c = ws.cell(row=row, column=1,
                value="AREA SERVIZI DI RADIOLOGIA - Standard AGENAS")
    c.font = FONT_SECTION
    row += 1

    # Nota presidio + livello
    livello_leggibile = {
        'OSPEDALE_DI_BASE':     'Ospedale di Base',
        'PRESIDIO_I_LIVELLO':   'Presidio di I livello',
        'PRESIDIO_II_LIVELLO':  'Presidio di II livello',
    }.get(livello_label, livello_label)

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value=f"Presidio: {sede_completa}  –  {livello_leggibile}"
            ).font = FONT_NORMAL
    row += 2

    # Intestazioni
    headers = ['Profilo', 'T.I.', 'T.D.', 'Totale', 'Range AGENAS',
               'Esito']
    N_COLS = len(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # Nomi leggibili dei profili AGENAS radiologia
    nomi_profili = {
        'DIRIGENTE_MEDICO_RADIOLOGIA': 'Dir. Medici Radiologia (*)',
        'TECNICO_RADIOLOGIA':          'Tecnici di Radiologia',
        'INFERMIERE':                  'Infermieri',
        'OPERATORE_SOCIO_SANITARIO':   'Operatori Socio Sanitari',
    }

    # Calcola gli in servizio per ciascun profilo AGENAS radiologia
    in_servizio = {}
    in_servizio_td = {}

    def _conta_radio(df_src, solo_territoriali=False):
        """Conta personale radiologico nel DataFrame fornito."""
        for _, r in df_src.iterrows():
            ssd = str(r.get('_REPARTO', ''))
            sede_r = str(r.get('Sede', ''))
            profilo = str(r.get('Profilo Professionale', '')
                          ).strip().upper()
            quantita_ti = int(r.get('Quantità T.I.', 0))
            quantita_td = int(r.get('Quantità T.D.', 0))

            # Se richiesto solo territoriali, salta i P.O.
            if solo_territoriali and 'P.O.' in sede_r:
                continue

            # Verifica se la UO è dell'area radiologica
            uo_radio = False
            for m_uo in mapping_uo_radio:
                if _re.search(m_uo['pattern'], ssd.upper()):
                    uo_radio = True
                    break

            # Per sedi territoriali: verifica anche CDC
            if not uo_radio and 'P.O.' not in sede_r:
                cdc_val = str(r.get('Centro di Costo', '')).upper()
                if _re.search(
                        r'CENTRO RADIOLOGICO|RADIOLOGIA', cdc_val):
                    uo_radio = True

            # Per sedi territoriali: TS RADIOLOGIA conta sempre
            if not uo_radio and 'P.O.' not in sede_r:
                for m_prof in mapping_profili_radio:
                    if (m_prof['profilo_agenas'] == 'TECNICO_RADIOLOGIA'
                            and _re.search(m_prof['pattern'], profilo)):
                        uo_radio = True
                        break

            if not uo_radio:
                continue

            # Mappa il profilo
            for m_prof in mapping_profili_radio:
                if _re.search(m_prof['pattern'], profilo):
                    pa = m_prof['profilo_agenas']
                    in_servizio[pa] = in_servizio.get(pa, 0) + quantita_ti
                    in_servizio_td[pa] = in_servizio_td.get(pa, 0) + quantita_td
                    break

    # Conta personale P.O. (da df_luogo)
    _conta_radio(df_luogo)

    # Conta anche personale da sedi territoriali dell'area
    if df_area is not None:
        _conta_radio(df_area, solo_territoriali=True)

    # Righe dati
    FILL_OK = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                          fill_type='solid')
    FILL_CARENZA = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                               fill_type='solid')
    toggle = 0
    for prof_key, rng in fabb_radio_presidio.items():
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1
        servizio_ti = in_servizio.get(prof_key, 0)
        servizio_td = in_servizio_td.get(prof_key, 0)
        servizio = servizio_ti + servizio_td
        v_min = rng['min']
        v_max = rng['max']
        range_str = f'{v_min} - {v_max}'

        if servizio < v_min:
            esito = f'CARENZA (min {v_min - servizio})'
            fill_esito = FILL_CARENZA
        elif servizio > v_max:
            esito = f'ECCEDENZA (+{servizio - v_max})'
            fill_esito = FILL_OK
        else:
            esito = 'IN RANGE'
            fill_esito = FILL_OK

        vals = [
            nomi_profili.get(prof_key, prof_key),
            servizio_ti, servizio_td, servizio, range_str, esito,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        # Colora la cella Esito
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        row += 1

    # Note piè di pagina
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="(*) comprese guardie e apicalità. "
                  "Maggiorazione +15% per turnazione, guardie, "
                  "ferie, malattie e altre indisponibilità."
            ).font = Font(
                italic=True, size=9, color='555555')

    if df_area is not None:
        row += 1
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS_MERGE)
        ws.cell(row=row, column=1,
                value="I dati includono il personale delle sedi "
                      "territoriali dell'area."
                ).font = Font(italic=True, size=9, color='555555')

    return row + 1
