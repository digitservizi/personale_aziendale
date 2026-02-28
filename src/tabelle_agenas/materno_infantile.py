"""
Tabella AGENAS materno-infantile in calce al foglio presidio.
"""

from openpyxl.styles import Font, PatternFill

from src.stili_excel import (
    THIN_BORDER, FILL_A, FILL_B, FILL_HEADER,
    FONT_HEADER, ALIGN_CENTER,
)


# ============================================================
# TABELLA AGENAS MATERNO-INFANTILE IN CALCE AL FOGLIO
# ============================================================

def _scrivi_tabella_agenas_materno_infantile(
        ws, start_row, df_area, fabb_presidio,
        mapping_uo, mapping_profili, sede_completa):
    """Aggiunge una tabella riepilogativa AGENAS materno-infantile
    in calce al foglio di dettaglio di un presidio ospedaliero.

    Restituisce la riga successiva libera.
    """
    import re as _re

    if not fabb_presidio:
        return start_row

    # --- Stili locali ---
    FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
    FONT_NORMAL  = Font(size=10)

    row = start_row + 1
    N_COLS_MERGE = 6  # per merge titolo/sottotitolo

    # Titolo sezione
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    c = ws.cell(row=row, column=1,
                value="AREA MATERNO INFANTILE - Standard AGENAS")
    c.font = FONT_SECTION
    row += 1

    # Nota presidio
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value=f"Presidio: {sede_completa}"
            ).font = FONT_NORMAL
    row += 2

    # Intestazioni
    headers = ['Profilo', 'T.I.', 'T.D.', 'Totale', 'Range AGENAS',
               'Esito']
    N_COLS = len(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # Nomi leggibili dei profili AGENAS materno-infantile
    nomi_profili = {
        'DIRIGENTE_MEDICO_PEDIATRIA':  'Dir. Medici Pediatria/Neonat.',
        'DIRIGENTE_MEDICO_OSTETRICIA': 'Dir. Medici Ostetricia/Ginec.',
        'OSTETRICA':                   'Ostetriche',
        'INFERMIERE':                  'Infermieri',
        'OPERATORE_SOCIO_SANITARIO':   'Operatori Socio Sanitari',
    }

    # Calcola gli in servizio per ciascun profilo AGENAS
    in_servizio = {}
    in_servizio_td = {}
    for _, r in df_area.iterrows():
        ssd = str(r.get('_REPARTO', ''))
        profilo = str(r.get('Profilo Professionale', '')
                      ).strip().upper()
        quantita_ti = int(r.get('Quantità T.I.', 0))
        quantita_td = int(r.get('Quantità T.D.', 0))

        # Verifica se la UO è dell'area materno-infantile
        uo_match = False
        profilo_uo = None
        for m_uo in mapping_uo:
            if _re.search(m_uo['pattern'], ssd.upper()):
                uo_match = True
                profilo_uo = m_uo.get('profilo_agenas')
                break
        if not uo_match:
            continue

        # Mappa il profilo
        for m_prof in mapping_profili:
            if _re.search(m_prof['pattern'], profilo):
                pa = m_prof['profilo_agenas']
                # I dirigenti medici prendono il profilo dalla UO
                if pa in ('DIRIGENTE_MEDICO_PEDIATRIA',
                          'DIRIGENTE_MEDICO_OSTETRICIA'):
                    if profilo_uo:
                        pa = profilo_uo
                    else:
                        continue
                in_servizio[pa] = in_servizio.get(pa, 0) + quantita_ti
                in_servizio_td[pa] = in_servizio_td.get(pa, 0) + quantita_td
                break

    # Righe dati
    FILL_OK = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                          fill_type='solid')
    FILL_CARENZA = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                               fill_type='solid')
    toggle = 0
    for prof_key, rng in fabb_presidio.items():
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1
        servizio_ti = in_servizio.get(prof_key, 0)
        servizio_td = in_servizio_td.get(prof_key, 0)
        servizio = servizio_ti + servizio_td
        v_min = rng['min']
        v_max = rng['max']
        range_str = f'{v_min} - {v_max}'

        if servizio < v_min:
            esito = f'CARENZA (min {v_min - servizio})'
            fill_esito = FILL_CARENZA
        elif servizio > v_max:
            esito = f'ECCEDENZA (+{servizio - v_max})'
            fill_esito = FILL_OK
        else:
            esito = 'IN RANGE'
            fill_esito = FILL_OK

        vals = [
            nomi_profili.get(prof_key, prof_key),
            servizio_ti, servizio_td, servizio, range_str, esito,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        # Colora la cella Esito
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        row += 1

    # Note piè di pagina
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="(*) I medici comprendono guardie e apicalità. "
                  "Maggiorazione +15% per turnazione, guardie, "
                  "ferie, malattie e altre indisponibilità."
            ).font = Font(
                italic=True, size=9, color='555555')

    return row + 1
