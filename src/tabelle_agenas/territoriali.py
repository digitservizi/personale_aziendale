"""
Tabella AGENAS territoriale generica (basata su popolazione).
Estratta da report_fabbisogno.py per modularizzazione.
"""

from openpyxl.styles import Font, PatternFill

from src.stili_excel import (
    THIN_BORDER, FILL_A, FILL_B, FILL_HEADER,
    FONT_HEADER, ALIGN_CENTER,
)


# ============================================================
# TABELLA AGENAS TERRITORIALE GENERICA (basata su popolazione)
# ============================================================

def _scrivi_tabella_agenas_territoriale(
        ws, start_row, df_area, indicatori, fabb_area,
        citta):
    """Aggiunge una tabella AGENAS per standard territoriali (tassi per popolazione).

    Conta il personale in servizio (T.I.) nelle UO che matchano i pattern,
    confronta con il fabbisogno atteso calcolato sulla popolazione.

    Supporta due modalità:
      - Tasso singolo (SM, NPIA, Carcere): valuta >= minimo
      - Range min/regime (SerD): valuta come intervallo

    Restituisce la riga successiva libera.
    """
    import re as _re

    if not fabb_area:
        return start_row

    titolo = indicatori.get('titolo', 'Standard AGENAS')
    riferimento = indicatori.get('riferimento', '')
    uo_patterns = indicatori.get('unita_operative', [])
    profili_attesi = fabb_area.get('profili', [])
    pop_rif = fabb_area.get('popolazione_rif', 0)
    base = fabb_area.get('base', 0)
    fascia = indicatori.get('fascia_popolazione', '')

    # Etichetta per la popolazione
    if fascia == 'detenuti':
        pop_label = f"Detenuti: {pop_rif}"
    elif fascia == 'gte_18':
        pop_label = f"Pop. ≥ 18 anni: {pop_rif:,}".replace(',', '.')
    elif fascia == 'range_15_64':
        pop_label = f"Pop. 15-64 anni: {pop_rif:,}".replace(',', '.')
    elif fascia == 'range_1_17':
        pop_label = f"Pop. 1-17 anni: {pop_rif:,}".replace(',', '.')
    else:
        pop_label = f"Pop.: {pop_rif:,}".replace(',', '.')

    # --- Stili locali ---
    FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
    FONT_NORMAL  = Font(size=10)
    FONT_NOTE    = Font(size=9, italic=True)

    # Determina se è un range (SerD) o tasso singolo
    has_regime = any(p['atteso_regime'] is not None for p in profili_attesi)

    row = start_row + 1

    # -- Determina quante colonne serve il merge --
    if has_regime:
        headers = ['Profilo', 'T.I.', 'T.D.', 'Totale', 'Atteso min.',
                   'Atteso a regime', 'Esito']
    else:
        headers = ['Profilo', 'T.I.', 'T.D.', 'Totale', 'Standard min.',
                   'Esito']
    N_COLS = len(headers)

    # Titolo sezione
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1,
            value=f"{titolo} - Standard AGENAS").font = FONT_SECTION
    row += 1

    # Sottotitolo: area e popolazione
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1,
            value=f"Area: {citta}  –  {pop_label}").font = FONT_NORMAL
    row += 2

    # Intestazioni
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # --- Conta in servizio per ogni profilo ---
    # Filtra righe che matchano le UO
    FILL_OK = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                          fill_type='solid')
    FILL_CARENZA = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                               fill_type='solid')
    FILL_REGIME = PatternFill(start_color='B4C6E7', end_color='B4C6E7',
                              fill_type='solid')

    totale_servizio = 0
    totale_servizio_td = 0
    totale_atteso_min = 0
    totale_atteso_regime = 0
    toggle = 0

    for prof in profili_attesi:
        # Conta personale in servizio che matcha le qualifiche
        # e che si trova nelle UO corrispondenti
        servizio = 0
        servizio_td_val = 0
        for _, r in df_area.iterrows():
            reparto = str(r.get('_REPARTO', ''))
            cdc = str(r.get('Centro di Costo', ''))
            qualifica = str(r.get('Profilo Professionale', '')).strip()
            quantita_ti = int(r.get('Quantità T.I.', 0))
            quantita_td = int(r.get('Quantità T.D.', 0))

            # Verifica se il reparto o il centro di costo matcha una UO
            uo_match = False
            if uo_patterns:
                for pat in uo_patterns:
                    if (_re.search(pat, reparto, _re.IGNORECASE)
                            or _re.search(pat, cdc, _re.IGNORECASE)):
                        uo_match = True
                        break
            else:
                uo_match = True  # nessun filtro UO, conta tutto

            if not uo_match:
                continue

            # Verifica se la qualifica è tra quelle del profilo
            for q in prof['qualifiche']:
                if _re.search(q, qualifica, _re.IGNORECASE):
                    servizio += quantita_ti
                    servizio_td_val += quantita_td
                    break

        atteso_min = prof['atteso_min']
        atteso_regime = prof['atteso_regime']

        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1

        totale_prof = servizio + servizio_td_val

        if has_regime and atteso_regime is not None:
            # Modalità range
            amin = int(round(atteso_min))
            areg = int(round(atteso_regime))
            if totale_prof < amin:
                esito = f'CARENZA ({amin - totale_prof})'
                fill_esito = FILL_CARENZA
            elif totale_prof >= areg:
                esito = 'A REGIME'
                fill_esito = FILL_REGIME
            else:
                esito = 'IN RANGE'
                fill_esito = FILL_OK

            vals = [prof['nome'], servizio, servizio_td_val, totale_prof,
                    amin, areg, esito]
        else:
            # Modalità singola (solo minimo)
            amin = int(round(atteso_min))
            if totale_prof >= amin:
                esito = 'OK'
                fill_esito = FILL_OK
            else:
                esito = f'CARENZA ({amin - totale_prof})'
                fill_esito = FILL_CARENZA

            vals = [prof['nome'], servizio, servizio_td_val, totale_prof,
                    amin, esito]

        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        # Colora la cella esito
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        row += 1

        totale_servizio += servizio
        totale_servizio_td += servizio_td_val
        totale_atteso_min += atteso_min
        totale_atteso_regime += (atteso_regime if atteso_regime else 0)

    # Riga TOTALE
    t_amin = int(round(totale_atteso_min))
    grand_tot = totale_servizio + totale_servizio_td
    if has_regime:
        t_areg = int(round(totale_atteso_regime))
        if grand_tot < t_amin:
            esito_tot = f'CARENZA ({t_amin - grand_tot})'
            fill_tot = FILL_CARENZA
        elif grand_tot >= t_areg:
            esito_tot = 'A REGIME'
            fill_tot = FILL_REGIME
        else:
            esito_tot = 'IN RANGE'
            fill_tot = FILL_OK
        vals_tot = ['TOTALE', totale_servizio, totale_servizio_td,
                    grand_tot, t_amin, t_areg, esito_tot]
    else:
        if grand_tot >= t_amin:
            esito_tot = 'OK'
            fill_tot = FILL_OK
        else:
            esito_tot = f'CARENZA ({t_amin - grand_tot})'
            fill_tot = FILL_CARENZA
        vals_tot = ['TOTALE', totale_servizio, totale_servizio_td,
                    grand_tot, t_amin, esito_tot]

    for ci, v in enumerate(vals_tot, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = Font(bold=True, size=10)
        c.fill = FILL_HEADER
        c.border = THIN_BORDER
    ws.cell(row=row, column=N_COLS).fill = fill_tot
    row += 1

    # Note
    row += 1
    if riferimento:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS)
        ws.cell(row=row, column=1, value=riferimento).font = FONT_NOTE
        row += 1

    if fascia == 'detenuti':
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS)
        ws.cell(row=row, column=1,
                value=(f"Tasso calcolato ogni {base} detenuti. "
                       "Standard SM + SerD penitenziario.")).font = FONT_NOTE
        row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS)
        ws.cell(row=row, column=1,
                value=(f"Tasso per {base:,} abitanti ({fascia})."
                       ).replace(',', '.')).font = FONT_NOTE
        row += 1

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1,
            value="Maggiorazione +15% per turnazione, guardie, "
                  "ferie, malattie e altre indisponibilità."
            ).font = FONT_NOTE
    row += 1

    return row
