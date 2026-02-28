"""
Tabella AGENAS terapia intensiva (§ 8.1.1) in calce al foglio presidio.
"""

from openpyxl.styles import Font, PatternFill

from src.stili_excel import (
    THIN_BORDER, FILL_A, FILL_B, FILL_HEADER,
    FONT_HEADER, ALIGN_CENTER,
)


# ============================================================
# TABELLA AGENAS TERAPIA INTENSIVA (§ 8.1.1) IN CALCE AL FOGLIO
# ============================================================

def _scrivi_tabella_agenas_terapia_intensiva(
        ws, start_row, df_area, fabb_ti_presidio,
        mapping_uo_ti, mapping_profili_ti,
        sede_completa):
    """Aggiunge una tabella riepilogativa AGENAS per l'area
    Terapia Intensiva (§ 8.1.1) nel foglio RIEPILOGO.

    Il fabbisogno è calcolato con la formula:
      FTE = (PL / rapporto_letti) × (ore_turno × 365) / ore_annue_eff

    Colonne: Profilo | Rapporto | PL | FTE Atteso | T.I. | T.D. |
             Totale | Esito

    Restituisce la riga successiva libera.
    """
    import re as _re

    if not fabb_ti_presidio:
        return start_row

    # --- Stili locali ---
    FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
    FONT_NORMAL  = Font(size=10)

    row = start_row + 1
    N_COLS_MERGE = 8

    # Titolo sezione
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="AREA TERAPIA INTENSIVA - Standard AGENAS (§ 8.1.1)"
            ).font = FONT_SECTION
    row += 1

    # Nota presidio
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value=f"Presidio: {sede_completa}"
            ).font = FONT_NORMAL
    row += 2

    # Intestazioni
    headers = ['Profilo', 'Rapporto', 'PL Intensivi',
               'FTE Atteso', 'T.I.', 'T.D.', 'Totale', 'Esito']
    N_COLS = len(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # Nomi leggibili dei profili
    nomi_profili = {
        'DIRIGENTE_MEDICO': 'Dirigenti Medici',
        'INFERMIERE':       'Infermieri',
    }

    # --- Conta in servizio per ciascun profilo ---
    # Filtra solo il personale il cui Centro di Costo contiene
    # "TERAPIA INTENSIVA" (esclude Blocco Operatorio e CDC condivisi
    # come "Anestesia e T.I. – Camere Operatorie").
    _RE_CDC_TI = _re.compile(r'TERAPIA INTENSIVA', _re.IGNORECASE)
    in_servizio = {}
    in_servizio_td = {}
    for _, r in df_area.iterrows():
        cdc = str(r.get('Centro di Costo', ''))
        profilo = str(r.get('Profilo Professionale', '')).strip().upper()
        quantita_ti = int(r.get('Quantità T.I.', 0))
        quantita_td = int(r.get('Quantità T.D.', 0))

        if not _RE_CDC_TI.search(cdc):
            continue

        # Mappa il profilo
        for m_prof in mapping_profili_ti:
            if _re.search(m_prof['pattern'], profilo):
                pa = m_prof['profilo_agenas']
                in_servizio[pa] = in_servizio.get(pa, 0) + quantita_ti
                in_servizio_td[pa] = in_servizio_td.get(pa, 0) + quantita_td
                break

    # --- Righe dati ---
    FILL_OK = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                          fill_type='solid')
    FILL_CARENZA = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                               fill_type='solid')
    toggle = 0
    for prof_key, dati in fabb_ti_presidio.items():
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1

        fte_atteso = dati['fte_atteso']
        rapporto = dati['rapporto']
        pl = dati['pl']

        servizio_ti = in_servizio.get(prof_key, 0)
        servizio_td = in_servizio_td.get(prof_key, 0)
        servizio = servizio_ti + servizio_td

        if servizio < fte_atteso:
            delta = fte_atteso - servizio
            esito = f'CARENZA ({delta})'
            fill_esito = FILL_CARENZA
        elif servizio > fte_atteso:
            delta = servizio - fte_atteso
            esito = f'ECCEDENZA (+{delta})'
            fill_esito = FILL_OK
        else:
            esito = 'CONFORME'
            fill_esito = FILL_OK

        rapporto_str = f'1 ogni {rapporto} letti'

        vals = [
            nomi_profili.get(prof_key, prof_key),
            rapporto_str, pl, fte_atteso,
            servizio_ti, servizio_td, servizio, esito,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        row += 1

    # Note piè di pagina
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="Formula: FTE = (PL / Rapporto) × (24 × 365) "
                  "/ Ore annue effettive × 1,15.  "
                  "Fonte: AGENAS § 8.1.1 – Area Intensiva.  "
                  "Maggiorazione +15% per turnazione, guardie, "
                  "ferie, malattie e altre indisponibilità."
            ).font = Font(italic=True, size=9, color='555555')

    return row + 1
