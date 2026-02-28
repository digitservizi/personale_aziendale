"""
Tabella AGENAS sale operatorie (§ 8.1.2) in calce al foglio presidio.
"""

from openpyxl.styles import Font, PatternFill

from src.stili_excel import (
    THIN_BORDER, FILL_A, FILL_B, FILL_HEADER,
    FONT_HEADER, ALIGN_CENTER,
)


# ============================================================
# TABELLA AGENAS SALE OPERATORIE IN CALCE AL FOGLIO
# ============================================================

def _scrivi_tabella_agenas_sale_operatorie(
        ws, start_row, df_luogo, fabb_so_presidio,
        mapping_uo_so, mapping_profili_so, sede_completa):
    """Aggiunge una tabella riepilogativa AGENAS sale operatorie (§ 8.1.2)
    in calce al foglio di dettaglio di un presidio ospedaliero.

    Restituisce la riga successiva libera.
    """
    import re as _re

    if not fabb_so_presidio:
        return start_row

    # --- Stili locali ---
    FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
    FONT_NORMAL  = Font(size=10)

    row = start_row + 1
    N_COLS_MERGE = 8  # colonne: Profilo|Staff/Sala|N.Sale|FTE|TI|TD|Tot|Esito

    # Titolo sezione
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    c = ws.cell(row=row, column=1,
                value="AREA SALE OPERATORIE - Standard AGENAS (§ 8.1.2)")
    c.font = FONT_SECTION
    row += 1

    # Recupero n_sale dal primo profilo (uguale per tutti)
    first_prof = next(iter(fabb_so_presidio.values()))
    n_sale = first_prof['n_sale']
    ore_cop = first_prof['ore_copertura']
    gg_anno = first_prof['giorni_anno']

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value=f"Presidio: {sede_completa}  –  "
                  f"N. sale operative: {n_sale}  –  "
                  f"Copertura: {ore_cop}h × {gg_anno} gg/anno"
            ).font = FONT_NORMAL
    row += 2

    # Intestazioni
    headers = ['Profilo', 'Staff/Sala', 'N. Sale', 'FTE Atteso',
               'T.I.', 'T.D.', 'Totale', 'Esito']
    N_COLS = len(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # Nomi leggibili
    nomi_profili = {
        'DIRIGENTE_MEDICO_ANESTESISTA': 'Medici Anestesisti (*)',
        'INFERMIERE':                   'Infermieri',
        'OPERATORE_SOCIO_SANITARIO':    'Operatori Socio Sanitari',
    }

    # Calcola gli in servizio – filtra per UO di Blocco Operatorio
    # e per UO di Anestesia/Rianimazione (anestesisti)
    in_servizio = {}
    in_servizio_td = {}
    for _, r in df_luogo.iterrows():
        cdc = str(r.get('Centro di Costo', ''))
        reparto = str(r.get('_REPARTO', ''))
        profilo = str(r.get('Profilo Professionale', '')).strip().upper()
        quantita_ti = int(r.get('Quantità T.I.', 0))
        quantita_td = int(r.get('Quantità T.D.', 0))

        # Verifica se la UO/CDC è dell'area sale operatorie
        uo_match = False
        for m_uo in mapping_uo_so:
            pat = m_uo['pattern']
            if (_re.search(pat, cdc, _re.IGNORECASE)
                    or _re.search(pat, reparto, _re.IGNORECASE)):
                uo_match = True
                break
        if not uo_match:
            continue

        # Escludi il personale il cui CDC è "Terapia Intensiva":
        # è già conteggiato nella tabella TI (§ 8.1.1) e non va
        # doppio-contato nella tabella Sale Operatorie.
        if _re.search(r'TERAPIA INTENSIVA', cdc, _re.IGNORECASE):
            continue

        for m_prof in mapping_profili_so:
            if _re.search(m_prof['pattern'], profilo):
                pa = m_prof['profilo_agenas']
                in_servizio[pa] = in_servizio.get(pa, 0) + quantita_ti
                in_servizio_td[pa] = in_servizio_td.get(pa, 0) + quantita_td
                break

    # Righe dati
    FILL_OK = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                          fill_type='solid')
    FILL_CARENZA = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                               fill_type='solid')
    toggle = 0
    for prof_key, dati in fabb_so_presidio.items():
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1

        fte_atteso = dati['fte_atteso']
        pps = dati['personale_per_sala']

        servizio_ti = in_servizio.get(prof_key, 0)
        servizio_td = in_servizio_td.get(prof_key, 0)
        servizio = servizio_ti + servizio_td

        if servizio < fte_atteso:
            delta = fte_atteso - servizio
            esito = f'CARENZA ({delta})'
            fill_esito = FILL_CARENZA
        elif servizio > fte_atteso:
            delta = servizio - fte_atteso
            esito = f'ECCEDENZA (+{delta})'
            fill_esito = FILL_OK
        else:
            esito = 'CONFORME'
            fill_esito = FILL_OK

        vals = [
            nomi_profili.get(prof_key, prof_key),
            pps, n_sale, fte_atteso,
            servizio_ti, servizio_td, servizio, esito,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        row += 1

    # Note piè di pagina
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="(*) Solo anestesisti; i medici specialisti (chirurghi, "
                  "ortopedici, ecc.) sono già conteggiati nelle tabelle "
                  "per livello di presidio."
            ).font = Font(italic=True, size=9, color='555555')
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="Formula: FTE = N.Sale × Staff/Sala × "
                  f"({ore_cop} × {gg_anno}) / Ore annue eff. × 1,15.  "
                  "Fonte: AGENAS § 8.1.2 – Sale Operatorie.  "
                  "Maggiorazione +15% per turnazione, guardie, "
                  "ferie, malattie e altre indisponibilità."
            ).font = Font(italic=True, size=9, color='555555')

    return row + 1
