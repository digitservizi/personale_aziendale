"""
Stili Excel condivisi e funzioni di formattazione dei fogli.
Tutte le costanti grafiche e le utility di scrittura celle risiedono qui
per garantire uniformità tra i diversi report.
"""

from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side,
)
from openpyxl.utils import get_column_letter

# ============================================================
# COSTANTI DI STILE
# ============================================================
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

FILL_A = PatternFill(start_color='DCE6F1', end_color='DCE6F1',
                     fill_type='solid')          # azzurro chiaro
FILL_B = PatternFill(start_color='FFFFFF', end_color='FFFFFF',
                     fill_type='solid')          # bianco

FILL_HEADER = PatternFill(start_color='4472C4', end_color='4472C4',
                          fill_type='solid')     # blu intestazione

FILL_AGENAS = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                          fill_type='solid')     # verde chiaro (rif. AGENAS)

FILL_TOTALE = PatternFill(start_color='D9E2F3', end_color='D9E2F3',
                          fill_type='solid')     # azzurro medio (riga totale)

FONT_HEADER = Font(bold=True, color='FFFFFF', size=11)
FONT_TITLE  = Font(bold=True, size=13)
FONT_TOTALE = Font(bold=True, size=11)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')


# ============================================================
# FUNZIONI DI SCRITTURA
# ============================================================

def scrivi_titolo(ws, titolo, n_cols):
    """Scrive una riga titolo unificata (riga 1) su un foglio Excel."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1)
    cell.value = titolo
    cell.font = FONT_TITLE
    cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 22


def scrivi_intestazioni(ws, colonne, riga=2):
    """Scrive le intestazioni colonna con stile header (bianco su blu)."""
    for col_idx, col_name in enumerate(colonne, 1):
        c = ws.cell(row=riga, column=col_idx, value=col_name)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER


def scrivi_riga_dati(ws, riga_idx, valori, fill):
    """Scrive una riga di dati con fill e bordi."""
    for col_idx, val in enumerate(valori, 1):
        c = ws.cell(row=riga_idx, column=col_idx, value=val)
        c.fill = fill
        c.border = THIN_BORDER


def scrivi_riga_totale(ws, riga_idx, valori):
    """Scrive una riga TOTALE in grassetto con sfondo dedicato."""
    for col_idx, val in enumerate(valori, 1):
        c = ws.cell(row=riga_idx, column=col_idx, value=val)
        c.fill = FILL_TOTALE
        c.font = FONT_TOTALE
        c.border = THIN_BORDER


def auto_larghezza_colonne(ws, colonne=None, riga_inizio=1,
                          larghezza_min=0):
    """Imposta la larghezza delle colonne in base al contenuto.

    *colonne* può essere:
      - una lista di nomi colonna (ne determina anche il numero),
      - un intero (numero di colonne, senza nomi di intestazione),
      - None  → usa tutte le colonne già presenti nel foglio.

    Le celle che fanno parte di unioni multi-colonna vengono ignorate
    per evitare di gonfiare la larghezza di una singola colonna.
    """
    if colonne is None:
        n_cols = ws.max_column or 1
        col_names = [None] * n_cols
    elif isinstance(colonne, int):
        n_cols = colonne
        col_names = [None] * n_cols
    else:
        n_cols = len(colonne)
        col_names = list(colonne)

    merged = list(ws.merged_cells.ranges)

    for i in range(n_cols):
        col_idx = i + 1
        max_len = len(str(col_names[i])) if col_names[i] else 0
        for row_obj in ws.iter_rows(min_row=riga_inizio,
                                    min_col=col_idx, max_col=col_idx):
            cell = row_obj[0]
            if cell.value is None:
                continue
            # Salta celle appartenenti a unioni multi-colonna
            in_merge = False
            for mr in merged:
                if (mr.min_col <= col_idx <= mr.max_col
                        and mr.min_row <= cell.row <= mr.max_row
                        and mr.min_col != mr.max_col):
                    in_merge = True
                    break
            if in_merge:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            max_len + 3, larghezza_min)


def fill_alternato(idx):
    """Restituisce FILL_A o FILL_B in base alla parità dell'indice."""
    return FILL_A if idx % 2 == 0 else FILL_B
