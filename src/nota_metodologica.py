"""
Foglio «NOTA METODOLOGICA» del report fabbisogno.
Estratto da report_fabbisogno.py per modularizzazione.

Funzioni incluse:
  - _scrivi_foglio_metodologia
"""

from collections import OrderedDict

from openpyxl.styles import Font, Alignment, PatternFill

from src.stili_excel import (
    THIN_BORDER, FILL_A, FILL_B, FILL_HEADER,
    FONT_HEADER, FONT_TITLE, ALIGN_CENTER,
    auto_larghezza_colonne,
)
from src.config import (
    LIVELLO_PRESIDIO, PARTI_PER_PRESIDIO,
    POPOLAZIONE_AREA, DETENUTI_PER_ISTITUTO,
)


# ============================================================
# FOGLIO NOTA METODOLOGICA
# ============================================================

def _scrivi_foglio_metodologia(wb, indicators, posti_letto_citta, anno_analisi):
    """Inserisce come primo foglio la nota metodologica completa:
    indicatori, posti letto rilevati, formula e arrotondamenti."""
    ws = wb.create_sheet(title='NOTA METODOLOGICA', index=0)

    # --- Stili locali ---
    FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
    FONT_SUBSECT = Font(bold=True, size=11)
    FONT_NORMAL  = Font(size=10)
    FONT_BOLD_SM = Font(bold=True, size=10)
    FONT_FORMULA = Font(bold=True, size=11, name='Consolas')
    FILL_FORMULA = PatternFill(start_color='FFF2CC', end_color='FFF2CC',
                               fill_type='solid')
    N_COLS = 9
    row = 1

    # ============ TITOLO ============
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    c = ws.cell(row=row, column=1)
    c.value = (f"NOTA METODOLOGICA \u2013 Criteri di Calcolo del "
               f"Fabbisogno di Personale ({anno_analisi})")
    c.font = FONT_TITLE
    c.alignment = ALIGN_CENTER
    ws.row_dimensions[row].height = 30
    row += 2

    # ============================================================
    # 1. INDICATORI PER FIGURA PROFESSIONALE
    # ============================================================
    ws.cell(row=row, column=1,
            value="1. INDICATORI PER FIGURA PROFESSIONALE").font = FONT_SECTION
    row += 1
    d1 = ws.cell(row=row, column=1)
    d1.value = (
        "I parametri seguenti, differenziati per figura professionale "
        "e intensit\u00e0 assistenziale del reparto, alimentano la "
        "formula di calcolo del fabbisogno teorico."
    )
    d1.font = FONT_NORMAL
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    row += 2

    PROFILE_LABELS = {
        'DIRIGENTE_MEDICO': 'Dirigente Medico',
        'INFERMIERE': 'Infermiere',
        'OPERATORE_SOCIO_SANITARIO': 'Operatore Socio Sanitario',
    }
    INTENSITY_LABELS = {
        'Intensiva': 'Intensiva', 'Alta': 'Alta',
        'MedioAlta': 'Medio/Alta', 'Media': 'Media',
        'MedioBassa': 'Medio/Bassa', 'Bassa': 'Bassa',
        'DH_DS': 'DH / Day Surgery',
    }
    INTENSITY_ORDER = [
        'Intensiva', 'Alta', 'MedioAlta', 'Media',
        'MedioBassa', 'Bassa', 'DH_DS',
    ]
    IND_HEADERS = [
        'Intensit\u00e0', 'Tasso Occupazione', 'Coeff. Complessit\u00e0',
        'Ore per Turno', 'Ore Annue Effettive',
    ]

    for prof_key, prof_label in PROFILE_LABELS.items():
        if prof_key not in indicators:
            continue
        ws.cell(row=row, column=1, value=prof_label).font = FONT_SUBSECT
        row += 1
        for ci, cn in enumerate(IND_HEADERS, 1):
            c = ws.cell(row=row, column=ci, value=cn)
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.alignment = ALIGN_CENTER
            c.border = THIN_BORDER
        row += 1

        prof_data = indicators[prof_key]
        toggle = 0
        for int_key in INTENSITY_ORDER:
            if int_key not in prof_data:
                continue
            d = prof_data[int_key]
            fill = FILL_A if toggle % 2 == 0 else FILL_B
            vals = [
                INTENSITY_LABELS.get(int_key, int_key),
                f"{d['TassoOccupazione']:.0f}%",
                d['CoefficienteComplessita'],
                d['OreEffettuateTurni'],
                d['OreAnnueLavoroEffettivo'],
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = fill
                c.border = THIN_BORDER
                if ci > 1:
                    c.alignment = ALIGN_CENTER
            toggle += 1
            row += 1
        row += 1  # gap tra profili

    row += 1

    # ============================================================
    # 2. POSTI LETTO RILEVATI
    # ============================================================
    ws.cell(row=row, column=1,
            value="2. POSTI LETTO RILEVATI").font = FONT_SECTION
    row += 1
    d2 = ws.cell(row=row, column=1)
    d2.value = (
        "Posti letto per sede e unit\u00e0 operativa con il livello di "
        "intensit\u00e0 assistenziale assegnato. Solo le unit\u00e0 con "
        "posti letto attivi generano un fabbisogno calcolato; le altre "
        'sono indicate come \"Servizio privo di posti letto\".'
    )
    d2.font = FONT_NORMAL
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    row += 2

    PL_HEADERS = [
        'Reparto (SSD)', 'Ordinari', 'DH', 'UTIC',
        'DS', 'Intensit\u00e0',
    ]

    # Raggruppa per sede
    sedi_pl = OrderedDict()
    for (sede, ssd), pl in sorted(posti_letto_citta.items(),
                                  key=lambda x: (x[0][0], x[0][1])):
        sedi_pl.setdefault(sede, []).append((ssd, pl))

    for sede_nome, reparti in sedi_pl.items():
        # Sotto-titolo sede
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=len(PL_HEADERS))
        c = ws.cell(row=row, column=1, value=sede_nome)
        c.font = FONT_SUBSECT
        ws.row_dimensions[row].height = 18
        row += 1

        # Intestazioni tabellina
        for ci, cn in enumerate(PL_HEADERS, 1):
            c = ws.cell(row=row, column=ci, value=cn)
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.alignment = ALIGN_CENTER
            c.border = THIN_BORDER
        row += 1

        # Righe dati
        for toggle_idx, (ssd, pl) in enumerate(reparti):
            fill = FILL_A if toggle_idx % 2 == 0 else FILL_B
            vals = [
                ssd,
                int(pl['ordinari']), int(pl['dh']),
                int(pl.get('utic', 0)),
                int(pl.get('ds', 0)),
                pl.get('intensita', ''),
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = fill
                c.border = THIN_BORDER
            row += 1

        row += 1  # gap tra sedi

    row += 1

    # ============================================================
    # 3. FORMULA DI CALCOLO
    # ============================================================
    ws.cell(row=row, column=1,
            value="3. FORMULA DI CALCOLO DEL FABBISOGNO").font = FONT_SECTION
    row += 1
    d3 = ws.cell(row=row, column=1)
    d3.value = (
        "Il fabbisogno teorico per ciascun reparto ospedaliero \u00e8 "
        "calcolato con la seguente formula:"
    )
    d3.font = FONT_NORMAL
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    row += 2

    # --- box formula ---
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    fc = ws.cell(row=row, column=1)
    fc.value = ("Fabbisogno = (PL_omog \u00d7 Tasso_occ \u00d7 "
                "Coeff_compl \u00d7 Ore_turni \u00d7 365) / Ore_annue_eff")
    fc.font = FONT_FORMULA
    fc.alignment = ALIGN_CENTER
    for ci in range(1, N_COLS + 1):
        ws.cell(row=row, column=ci).fill = FILL_FORMULA
        ws.cell(row=row, column=ci).border = THIN_BORDER
    row += 2

    ws.cell(row=row, column=1, value="dove:").font = FONT_BOLD_SM
    row += 1

    DEFS = [
        ("PL_omog",
         "Posti Letto Omogeneizzati = Ordinari \u00d7 1  +  DH \u00d7 0,5  "
         "+  UTIC \u00d7 1  +  Breast Unit Ord. \u00d7 1  +  "
         "Breast Unit DH \u00d7 0,5"),
        ("Tasso_occ",
         "Tasso di occupazione (da tabella indicatori, espresso in %)"),
        ("Coeff_compl",
         "Coefficiente di complessit\u00e0 assistenziale "
         "(da tabella indicatori)"),
        ("Ore_turni",
         "Ore effettuate per turno (da tabella indicatori)"),
        ("365",
         "Giorni di assistenza annui"),
        ("Ore_annue_eff",
         "Ore annue di lavoro effettivo per operatore "
         "(da tabella indicatori)"),
    ]
    for sigla, desc in DEFS:
        ws.cell(row=row, column=1, value=f"  \u2022 {sigla}").font = FONT_BOLD_SM
        ws.merge_cells(start_row=row, start_column=3,
                       end_row=row, end_column=N_COLS)
        ws.cell(row=row, column=3, value=desc).font = FONT_NORMAL
        row += 1

    row += 1
    odc = ws.cell(row=row, column=1)
    odc.value = (
        "Per gli Ospedali di Comunit\u00e0 (OdC) il fabbisogno \u00e8 "
        "definito in misura fissa dal DM 77/2022 e non deriva dalla "
        "formula indicata."
    )
    odc.font = FONT_NORMAL
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    row += 2

    # ============================================================
    # 4. REGOLE DI ARROTONDAMENTO
    # ============================================================
    ws.cell(row=row, column=1,
            value="4. REGOLE DI ARROTONDAMENTO").font = FONT_SECTION
    row += 1
    d4 = ws.cell(row=row, column=1)
    d4.value = (
        "Il fabbisogno calcolato \u00e8 un valore decimale. Poich\u00e9 "
        "il personale \u00e8 indivisibile, si applicano le seguenti "
        "regole di arrotondamento:"
    )
    d4.font = FONT_NORMAL
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    row += 2

    # Header tabella regole
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=2)
    for ci in (1, 2):
        c = ws.cell(row=row, column=ci,
                    value='Caso' if ci == 1 else None)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    ws.merge_cells(start_row=row, start_column=3,
                   end_row=row, end_column=N_COLS)
    for ci in range(3, N_COLS + 1):
        c = ws.cell(row=row, column=ci,
                    value='Regola applicata' if ci == 3 else None)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    RULES = [
        ('Fabbisogno < 1\n(profilo assente)',
         'Se il profilo professionale non ha personale in servizio '
         'nel reparto, il fabbisogno minimo \u00e8 fissato a 1 unit\u00e0.',
         FILL_A),
        ('Fabbisogno < 1\n(profilo presente)',
         'Arrotondamento standard: se \u2265 0,5 \u2192 1;  '
         'se < 0,5 \u2192 0.',
         FILL_A),
        ('Fabbisogno \u2265 1',
         "Arrotondamento all'intero pi\u00f9 vicino: parte decimale "
         '\u2265 0,5 \u2192 eccesso;  < 0,5 \u2192 difetto.',
         FILL_B),
    ]
    for caso, regola, fill in RULES:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=2)
        for ci in (1, 2):
            c = ws.cell(row=row, column=ci,
                        value=caso if ci == 1 else None)
            c.font = FONT_BOLD_SM
            c.fill = fill
            c.border = THIN_BORDER
            c.alignment = Alignment(wrap_text=True, vertical='center')
        ws.merge_cells(start_row=row, start_column=3,
                       end_row=row, end_column=N_COLS)
        for ci in range(3, N_COLS + 1):
            c = ws.cell(row=row, column=ci,
                        value=regola if ci == 3 else None)
            c.font = FONT_NORMAL
            c.fill = fill
            c.border = THIN_BORDER
            c.alignment = Alignment(wrap_text=True, vertical='center')
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1
    nota_fin = ws.cell(row=row, column=1)
    nota_fin.value = (
        "I servizi privi di posti letto (ambulatori, servizi "
        "territoriali, ecc.) non generano un fabbisogno calcolato "
        'e sono indicati come "Servizio privo di posti letto".'
    )
    nota_fin.font = FONT_NORMAL
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    row += 2

    # ============================================================
    # 5. STANDARD AGENAS – AREE OSPEDALIERE
    # ============================================================
    ws.cell(row=row, column=1,
            value="5. STANDARD AGENAS \u2013 AREE OSPEDALIERE"
            ).font = FONT_SECTION
    row += 1

    d5 = ws.cell(row=row, column=1)
    d5.value = (
        "Le tabelle AGENAS definiscono dotazioni minime e massime "
        "di personale (in FTE) per area funzionale e per livello di "
        "presidio ospedaliero, secondo la classificazione del "
        "DM 70/2015. I dati sono letti dai file XML di configurazione "
        "nella cartella configurazione/. Il confronto \u00e8 tra il "
        "Totale in servizio (T.I. + T.D.) e il range previsto.\n\n"
        "A tutti gli standard AGENAS \u00e8 applicata una "
        "maggiorazione organica del +15% per copertura della "
        "turnazione, guardie interdivisionali, visite specialistiche, "
        "ferie, malattie e altre indisponibilit\u00e0 di personale."
    )
    d5.font = FONT_NORMAL
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    row += 2

    # --- Helper per scrivere una tabella AGENAS ospedaliera ---
    def _scrivi_tabella_standard(ws, row, titolo, fonte, descrizione,
                                 livelli, profili_label, note_extra=None,
                                 uo_patterns=None, mapping_prof=None):
        """Scrive una sotto-sezione con tabella standard AGENAS.
        livelli: lista di tuple (nome_livello, {profilo: (min, max), ...})
        """
        ws.cell(row=row, column=1, value=titolo).font = FONT_SUBSECT
        row += 1

        # Fonte
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS)
        ws.cell(row=row, column=1,
                value=f"Fonte: {fonte}").font = Font(
                    italic=True, size=9, color='555555')
        row += 1

        # Descrizione
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS)
        ws.cell(row=row, column=1, value=descrizione).font = FONT_NORMAL
        row += 1

        if note_extra:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=N_COLS)
            ws.cell(row=row, column=1, value=note_extra).font = Font(
                italic=True, size=9)
            row += 1

        row += 1

        # Tabella: header
        n_prof = len(profili_label)
        # Col 1: Livello, poi 2 colonne per ogni profilo (min, max)
        # Header riga 1: Livello | profilo1 (merge 2) | profilo2 (merge 2) ...
        c = ws.cell(row=row, column=1, value='Livello Presidio')
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        for pi, plabel in enumerate(profili_label):
            col_start = 2 + pi * 2
            col_end = col_start + 1
            ws.merge_cells(start_row=row, start_column=col_start,
                           end_row=row, end_column=col_end)
            for ci2 in range(col_start, col_end + 1):
                c2 = ws.cell(row=row, column=ci2,
                             value=plabel if ci2 == col_start else None)
                c2.font = FONT_HEADER
                c2.fill = FILL_HEADER
                c2.alignment = ALIGN_CENTER
                c2.border = THIN_BORDER
        row += 1

        # Header riga 2: vuoto | min | max | min | max ...
        ws.cell(row=row, column=1, value='').font = FONT_HEADER
        ws.cell(row=row, column=1).fill = FILL_HEADER
        ws.cell(row=row, column=1).border = THIN_BORDER
        for pi in range(n_prof):
            for off, lab in enumerate(('Min', 'Max')):
                c2 = ws.cell(row=row, column=2 + pi * 2 + off, value=lab)
                c2.font = FONT_HEADER
                c2.fill = FILL_HEADER
                c2.alignment = ALIGN_CENTER
                c2.border = THIN_BORDER
        row += 1

        # Righe dati
        LIVELLO_LABELS = {
            'OSPEDALE_DI_BASE': 'Ospedale di Base',
            'PRESIDIO_I_LIVELLO': 'I Livello',
            'PRESIDIO_II_LIVELLO': 'II Livello',
        }
        for ti, (liv_nome, dati) in enumerate(livelli):
            fill = FILL_A if ti % 2 == 0 else FILL_B
            c = ws.cell(row=row, column=1,
                        value=LIVELLO_LABELS.get(liv_nome, liv_nome))
            c.font = FONT_BOLD_SM
            c.fill = fill
            c.border = THIN_BORDER
            for pi, plabel in enumerate(profili_label):
                prof_key = list(dati.keys())[pi] if pi < len(dati) else None
                if prof_key:
                    mn, mx = dati[prof_key]
                else:
                    mn, mx = '-', '-'
                for off, val in enumerate((mn, mx)):
                    c2 = ws.cell(row=row, column=2 + pi * 2 + off,
                                 value=val)
                    c2.font = FONT_NORMAL
                    c2.fill = fill
                    c2.alignment = ALIGN_CENTER
                    c2.border = THIN_BORDER
            row += 1

        row += 1

        # UO matching
        if uo_patterns:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=N_COLS)
            ws.cell(row=row, column=1,
                    value=f"UO incluse (pattern regex): {uo_patterns}"
                    ).font = Font(size=9, color='444444')
            row += 1

        # Mapping profili
        if mapping_prof:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=N_COLS)
            ws.cell(row=row, column=1,
                    value=f"Mapping profili personale \u2192 AGENAS: "
                          f"{mapping_prof}"
                    ).font = Font(size=9, color='444444')
            row += 1

        row += 1
        return row

    # --- 5a. Area Materno Infantile (Tab. 11) ---
    MAT_LIVELLI = [
        ('500\u20131.500 parti', {
            'Dir. Med. Pediatria': (6, 13),
            'Dir. Med. Ostetricia': (6, 13),
            'Ostetriche': (12, 24),
            'Infermieri': (12, 36),
            'OSS': (6, 12),
        }),
        ('1.500\u20132.000 parti', {
            'Dir. Med. Pediatria': (13, 18),
            'Dir. Med. Ostetricia': (13, 18),
            'Ostetriche': (24, 33),
            'Infermieri': (18, 46),
            'OSS': (12, 18),
        }),
        ('>2.000 parti', {
            'Dir. Med. Pediatria': (15, 30),
            'Dir. Med. Ostetricia': (15, 30),
            'Ostetriche': (33, 60),
            'Infermieri': (40, 80),
            'OSS': (18, 30),
        }),
    ]
    row = _scrivi_tabella_standard(
        ws, row,
        titolo="5a. Area Materno Infantile (Tab. 11)",
        fonte="Tabella 11 AGENAS \u2013 indicatori_agenas_materno_infantile.xml",
        descrizione=(
            "Parametro di ingresso: numero di parti annui per presidio. "
            "La fascia di volume determina il range min\u2013max per ciascun "
            "profilo professionale. Per la selezione del personale, i "
            "dirigenti medici sono classificati per UO di assegnazione "
            "(Pediatria/Neonatologia vs Ostetricia/Ginecologia)."
        ),
        livelli=MAT_LIVELLI,
        profili_label=['Dir.Med.Ped.', 'Dir.Med.Ost.',
                       'Ostetriche', 'Infermieri', 'OSS'],
        uo_patterns='PEDIATRIA|NEONATOLOGIA|OSTETRICIA|GINECOLOGIA',
        mapping_prof=(
            'Dir.Med. \u2192 specializzazione per UO; '
            'OSTETRICA \u2192 Ostetriche; INFERMIERE; OSS'
        ),
    )

    # --- 5b. Area Radiologia (Tab. 13) ---
    RAD_LIVELLI = [
        ('OSPEDALE_DI_BASE', {
            'Dir. Medici': (2, 7), 'Tecnici Rad.': (4, 18),
            'Infermieri': (0, 3), 'OSS': (1, 1),
        }),
        ('PRESIDIO_I_LIVELLO', {
            'Dir. Medici': (4, 18), 'Tecnici Rad.': (10, 36),
            'Infermieri': (3, 12), 'OSS': (1, 2),
        }),
        ('PRESIDIO_II_LIVELLO', {
            'Dir. Medici': (16, 55), 'Tecnici Rad.': (36, 110),
            'Infermieri': (10, 40), 'OSS': (2, 5),
        }),
    ]
    row = _scrivi_tabella_standard(
        ws, row,
        titolo="5b. Area Servizi di Radiologia (Tab. 13)",
        fonte="Tabella 13 AGENAS \u2013 indicatori_agenas_radiologia.xml",
        descrizione=(
            "Classificazione per livello di presidio (DM 70/2015). "
            "Include: Radiologia, Radioterapia, Medicina Nucleare, "
            "Neuroradiologia. Il conteggio include anche il personale "
            "radiologico assegnato alle sedi territoriali della stessa area."
        ),
        livelli=RAD_LIVELLI,
        profili_label=['Dir.Medici (*)', 'Tecnici Rad.',
                       'Infermieri', 'OSS'],
        uo_patterns=(
            'RADIOLOGIA|RADIODIAGNOSTICA|RADIOTERAPIA|'
            'MEDICINA NUCLEARE|NEURORADIOLOGIA|DIAGNOSTICA PER IMMAGINI'
        ),
        mapping_prof=(
            'DIR.MEDICO \u2192 Dir.Medici; TS RADIOLOGIA \u2192 '
            'Tecnici Rad.; INFERMIERE; OSS'
        ),
    )

    # --- 5c. Area Laboratorio (Tab. 14) ---
    LAB_LIVELLI = [
        ('OSPEDALE_DI_BASE', {
            'Dir. Sanitari': (0, 6), 'Infermieri': (1, 2),
            'OSS': (1, 1),
        }),
        ('PRESIDIO_I_LIVELLO', {
            'Dir. Sanitari': (4, 20), 'Infermieri': (2, 4),
            'OSS': (1, 2),
        }),
        ('PRESIDIO_II_LIVELLO', {
            'Dir. Sanitari': (18, 38), 'Infermieri': (4, 9),
            'OSS': (2, 3),
        }),
    ]
    row = _scrivi_tabella_standard(
        ws, row,
        titolo="5c. Area Servizi di Laboratorio (Tab. 14)",
        fonte="Tabella 14 AGENAS \u2013 indicatori_agenas_laboratorio.xml",
        descrizione=(
            "Classificazione per livello di presidio. Include: "
            "Laboratorio Analisi, Microbiologia e Virologia, Genetica, "
            "Tossicologia. Escluse le UO con tabelle AGENAS dedicate "
            "(Anatomia Patologica, Trasfusionale)."
        ),
        livelli=LAB_LIVELLI,
        profili_label=['Dir.Sanitari (*)', 'Infermieri', 'OSS'],
        uo_patterns='LABORATORIO ANALISI|MICROBIOLOG|VIROLOG|TOSSICOLOG',
        mapping_prof=(
            'DIR.MEDICO + DIR.BIOLOGO + DIR.CHIMICO + DIR.FARMACISTA '
            '\u2192 Dir.Sanitari; INFERMIERE; OSS. '
            'Escluse UO: TRASFUSIONALE|SIMT|ANATOMIA PATOLOGIC'
        ),
        note_extra=(
            "(*) Dirigenti Sanitari = Dirigenti Medici + Biologi + "
            "Chimici + Farmacisti (comprese guardie e apicalit\u00e0)."
        ),
    )

    # --- 5d. Area Trasfusionale (Tab. 15) ---
    TRASF_LIVELLI = [
        ('PRESIDIO_I_LIVELLO', {
            'Dir. Sanitari': (1, 6), 'Infermieri': (3, 4),
            'OSS': (1, 1),
        }),
        ('PRESIDIO_II_LIVELLO', {
            'Dir. Sanitari': (3, 16), 'Infermieri': (5, 6),
            'OSS': (1, 1),
        }),
    ]
    row = _scrivi_tabella_standard(
        ws, row,
        titolo="5d. Area Medicina Trasfusionale (Tab. 15)",
        fonte="Tabella 15 AGENAS \u2013 indicatori_agenas_trasfusionale.xml",
        descrizione=(
            "Solo presidi di I e II livello (Ospedale di Base non "
            "previsto). Include: Centro Trasfusionale, SIMT."
        ),
        livelli=TRASF_LIVELLI,
        profili_label=['Dir.Sanitari (*)', 'Infermieri', 'OSS'],
        uo_patterns='TRASFUSIONALE|SIMT',
        mapping_prof=(
            'DIR.MEDICO + DIR.BIOLOGO + DIR.CHIMICO '
            '\u2192 Dir.Sanitari; INFERMIERE; OSS'
        ),
    )

    # --- Sotto-sezione: dotazione segnalata dal Primario ---
    ws.cell(row=row, column=1,
            value="    Dotazione segnalata dal Primario (UOC Trasfusionale)"
            ).font = FONT_SUBSECT
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1, value=(
        "Fonte: indicatori_trasfusionale_speciale.xml \u2013 "
        "Dr.ssa Matilde Caruso, Primario UOC Medicina Trasfusionale. "
        "Dotazione minima richiesta per operare in sicurezza su 3 "
        "sedi con specializzazioni distribuite. "
        "Questa tabella integra (e non sostituisce) lo standard "
        "AGENAS Tab. 15."
    )).font = FONT_NORMAL
    row += 2

    # Tabella dotazione speciale
    TRASF_SPEC_PROF = ['Medici', 'Biologi', 'Infermieri',
                       'TSLB', 'OSS', 'Amm.vi']
    TRASF_SPEC_SEDI = [
        ('Cardarelli (CB)', [4, 5, 3, 9, 1, 1],
         'Sede principale, II livello diagnostica'),
        ('Veneziale (IS)', [4, 3, 3, 7, 1, 1],
         'Officina Trasfusionale Unica Regionale'),
        ('S.Timoteo (TE)', [5, 2, 7, 6, 1, 1],
         'Polo clinico (ematologia, TAO, coagulazione)'),
    ]
    # Header
    c = ws.cell(row=row, column=1, value='Sede')
    c.font = FONT_HEADER; c.fill = FILL_HEADER
    c.alignment = ALIGN_CENTER; c.border = THIN_BORDER
    for pi, pl in enumerate(TRASF_SPEC_PROF):
        c2 = ws.cell(row=row, column=2 + pi, value=pl)
        c2.font = FONT_HEADER; c2.fill = FILL_HEADER
        c2.alignment = ALIGN_CENTER; c2.border = THIN_BORDER
    c3 = ws.cell(row=row, column=2 + len(TRASF_SPEC_PROF), value='Note')
    c3.font = FONT_HEADER; c3.fill = FILL_HEADER
    c3.alignment = ALIGN_CENTER; c3.border = THIN_BORDER
    row += 1
    for ti, (sede, vals, nota) in enumerate(TRASF_SPEC_SEDI):
        fill = FILL_A if ti % 2 == 0 else FILL_B
        c = ws.cell(row=row, column=1, value=sede)
        c.font = FONT_BOLD_SM; c.fill = fill; c.border = THIN_BORDER
        for pi, v in enumerate(vals):
            c2 = ws.cell(row=row, column=2 + pi, value=v)
            c2.font = FONT_NORMAL; c2.fill = fill
            c2.alignment = ALIGN_CENTER; c2.border = THIN_BORDER
        c3 = ws.cell(row=row, column=2 + len(TRASF_SPEC_PROF), value=nota)
        c3.font = Font(size=9, italic=True); c3.fill = fill
        c3.border = THIN_BORDER
        row += 1
    # Riga totale
    fill = FILL_B
    c = ws.cell(row=row, column=1, value='TOTALE 3 Sedi')
    c.font = FONT_BOLD_SM; c.fill = fill; c.border = THIN_BORDER
    totali_spec = [sum(s[1][pi] for s in TRASF_SPEC_SEDI)
                   for pi in range(len(TRASF_SPEC_PROF))]
    for pi, v in enumerate(totali_spec):
        c2 = ws.cell(row=row, column=2 + pi, value=v)
        c2.font = FONT_BOLD_SM; c2.fill = fill
        c2.alignment = ALIGN_CENTER; c2.border = THIN_BORDER
    ws.cell(row=row, column=2 + len(TRASF_SPEC_PROF),
            value='').fill = fill
    ws.cell(row=row, column=2 + len(TRASF_SPEC_PROF)).border = THIN_BORDER
    row += 2

    # --- 5e. Anatomia Patologica (Tab. 16) ---
    AP_LIVELLI = [
        ('OSPEDALE_DI_BASE', {
            'Dir. Sanitari': (0, 3), 'Infermieri': (1, 1),
            'OSS': (1, 1),
        }),
        ('PRESIDIO_I_LIVELLO', {
            'Dir. Sanitari': (0, 8), 'Infermieri': (1, 1),
            'OSS': (1, 1),
        }),
        ('PRESIDIO_II_LIVELLO', {
            'Dir. Sanitari': (3, 14), 'Infermieri': (1, 2),
            'OSS': (1, 2),
        }),
    ]
    row = _scrivi_tabella_standard(
        ws, row,
        titolo="5e. Area Anatomia Patologica (Tab. 16)",
        fonte="Tabella 16 AGENAS \u2013 indicatori_agenas_anatomia_patologica.xml",
        descrizione=(
            "Classificazione per livello di presidio. "
            "Dirigenti Sanitari Medici e Non Medici, comprese guardie "
            "e apicalit\u00e0."
        ),
        livelli=AP_LIVELLI,
        profili_label=['Dir.Sanitari (*)', 'Infermieri', 'OSS'],
        uo_patterns='ANATOMIA PATOLOGIC',
    )

    # --- 5f. Tecnici di Laboratorio (Tab. 17) ---
    TL_LIVELLI = [
        ('OSPEDALE_DI_BASE', {'TS Lab. Biomedico': (6, 20)}),
        ('PRESIDIO_I_LIVELLO', {'TS Lab. Biomedico': (10, 50)}),
        ('PRESIDIO_II_LIVELLO', {'TS Lab. Biomedico': (45, 130)}),
    ]
    row = _scrivi_tabella_standard(
        ws, row,
        titolo="5f. Tecnici di Laboratorio (Tab. 17, \u00a77.1.9)",
        fonte="Tabella 17 AGENAS \u2013 indicatori_agenas_tecnici_laboratorio.xml",
        descrizione=(
            "Classificazione per livello di presidio. Il conteggio \u00e8 "
            "per ruolo (TS Laboratorio Biomedico) su tutto il presidio, "
            "indipendentemente dalla unit\u00e0 operativa di assegnazione "
            "(cross-UO). Vengono contati tutti i TSLB di tutte le UO del "
            "presidio."
        ),
        livelli=TL_LIVELLI,
        profili_label=['TS Lab. Biomedico'],
        note_extra=(
            "Nessun filtro su UO: il conteggio \u00e8 trasversale a "
            "Laboratorio, Anatomia Patologica, Trasfusionale, etc."
        ),
        mapping_prof='TS LABORATORIO BIOMEDICO \u2192 TS Lab. Biomedico',
    )

    # --- 5g. Medicina Legale (Tab. 18) ---
    ML_LIVELLI = [
        ('PRESIDIO_I_LIVELLO', {
            'Dir. Medici': (1, 2), 'Infermieri': (1, 1),
            'OSS': (1, 1),
        }),
        ('PRESIDIO_II_LIVELLO', {
            'Dir. Medici': (1, 3), 'Infermieri': (1, 2),
            'OSS': (2, 3),
        }),
    ]
    row = _scrivi_tabella_standard(
        ws, row,
        titolo="5g. Medicina Legale (Tab. 18, \u00a77.1.10)",
        fonte="Tabella 18 AGENAS \u2013 indicatori_agenas_medicina_legale.xml",
        descrizione=(
            "Solo presidi di I e II livello (non previsto per "
            "Ospedale di Base)."
        ),
        livelli=ML_LIVELLI,
        profili_label=['Dir. Medici', 'Infermieri', 'OSS'],
        uo_patterns='MEDICINA LEGALE',
    )

    # --- 5h. Emergenza-Urgenza (Tab. 20) ---
    EU_LIVELLI = [
        ('OSPEDALE_DI_BASE', {
            'Dir. Medici': (6, 14), 'Infermieri': (12, 18),
            'OSS': (3, 5),
        }),
        ('PRESIDIO_I_LIVELLO', {
            'Dir. Medici': (12, 24), 'Infermieri': (16, 48),
            'OSS': (6, 10),
        }),
        ('PRESIDIO_II_LIVELLO', {
            'Dir. Medici': (24, 40), 'Infermieri': (48, 70),
            'OSS': (9, 12),
        }),
    ]
    row = _scrivi_tabella_standard(
        ws, row,
        titolo="5h. Area Emergenza-Urgenza (Tab. 20)",
        fonte="Tabella 20 AGENAS \u2013 indicatori_agenas_emergenza_urgenza.xml",
        descrizione=(
            "Classificazione: PS (Osp. di Base), DEA I (I livello), "
            "DEA II (II livello). Valori minimi in FTE riferiti ad "
            "apertura proporzionata sulle 24 ore (DM 70/2015). "
            "Include le UO di Pronto Soccorso e MCAU; non include "
            "il 118 (emergenza territoriale)."
        ),
        livelli=EU_LIVELLI,
        profili_label=['Dir. Medici', 'Infermieri', 'OSS'],
        uo_patterns='PRONTO SOCCORSO|MCAU|MEDICINA.*CHIRURGIA.*URGENZA',
    )

    # --- 5i. Terapia Intensiva (§ 8.1.1) ---
    ws.cell(row=row, column=1,
            value="5i. Area Terapia Intensiva (§ 8.1.1)"
            ).font = Font(bold=True, size=11, color='1F4E79')
    row += 1
    ws.cell(row=row, column=1).value = (
        "Fonte: § 8.1.1 AGENAS – indicatori_agenas_terapia_intensiva.xml"
    )
    ws.cell(row=row, column=1).font = Font(italic=True, size=9,
                                           color='444444')
    row += 1
    ws.cell(row=row, column=1).value = (
        "Lo standard è espresso come rapporto letti/operatore per turno "
        "(non come range min-max). Il fabbisogno FTE è calcolato con la "
        "formula:\n"
        "   FTE = (PL_intensivi / Rapporto_letti) × (24 × 365) "
        "/ Ore_annue_eff × 1,15\n"
        "dove PL_intensivi sono i posti letto ordinari delle UO di "
        "Anestesia e Terapia Intensiva / Rianimazione. "
        "La copertura è H24 (24 ore/giorno). "
        "Lo standard OSS non è definito per la forte variabilità "
        "tra strutture. La maggiorazione +15% copre turnazione, guardie, "
        "ferie, malattie e altre indisponibilità."
    )
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True,
                                                     vertical='top')
    row += 2

    # Tabella standard rapporti
    ti_headers = ['Profilo', 'Rapporto per turno']
    for ci, h in enumerate(ti_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1
    ti_data = [
        ('Medico Anestesista-Rianimatore', '1 ogni 8 letti'),
        ('Infermiere', '1 ogni 2 letti'),
    ]
    for i, (prof, rapp) in enumerate(ti_data):
        fill = FILL_A if i % 2 == 0 else FILL_B
        for ci, v in enumerate([prof, rapp], 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        row += 1
    row += 1

    # UO intercettate
    ws.cell(row=row, column=1).value = (
        "UO intercettate (regex): "
        "ANESTESIA.*TERAPIA INTENSIVA|RIANIMAZIONE"
    )
    ws.cell(row=row, column=1).font = Font(italic=True, size=9,
                                           color='444444')
    row += 2

    # --- 5l. Sale Operatorie (§ 8.1.2) ---
    ws.cell(row=row, column=1,
            value="5l. Area Sale Operatorie (§ 8.1.2)"
            ).font = Font(bold=True, size=11, color='1F4E79')
    row += 1
    ws.cell(row=row, column=1).value = (
        "Fonte: § 8.1.2 AGENAS – indicatori_agenas_sale_operatorie.xml"
    )
    ws.cell(row=row, column=1).font = Font(italic=True, size=9,
                                           color='444444')
    row += 1
    ws.cell(row=row, column=1).value = (
        "Per ogni seduta operatoria (durata minima 6 ore) è prevista la "
        "presenza minima di 1 medico anestesista, 3 infermieri e "
        "1 operatore sociosanitario (§ 8.1.2 AGENAS). "
        "Il numero di medici specialisti (chirurghi, ortopedici, ecc.) "
        "è già conteggiato nei valori minimi e nei pesi delle tabelle "
        "per livello di presidio.\n"
        "Il fabbisogno FTE è calcolato con la formula:\n"
        "   FTE = N.Sale × Staff/Sala × (Ore_copertura × Giorni_anno) "
        "/ Ore_annue_eff × 1,15\n"
        "La maggiorazione +15% copre turnazione, guardie, "
        "ferie, malattie e altre indisponibilità."
    )
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True,
                                                     vertical='top')
    row += 2

    # Tabella standard per sala
    so_headers = ['Profilo', 'Personale per sala']
    for ci, h in enumerate(so_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1
    so_data = [
        ('Medico Anestesista', '1'),
        ('Infermieri', '3'),
        ('Operatore Socio Sanitario', '1'),
    ]
    for i, (prof, n) in enumerate(so_data):
        fill = FILL_A if i % 2 == 0 else FILL_B
        for ci, v in enumerate([prof, n], 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        row += 1
    row += 1

    # Sale per presidio
    ws.cell(row=row, column=1).value = (
        "N. sale operatorie configurate: vedi config.py."
    )
    ws.cell(row=row, column=1).font = Font(italic=True, size=9,
                                           color='444444')
    row += 2

    # --- 5m. Area Combinata TI + Blocco Operatorio ---
    ws.cell(row=row, column=1,
            value="5m. Area Anestesia, Terapia Intensiva e "
                  "Blocco Operatorio (combinata)"
            ).font = Font(bold=True, size=11, color='1F4E79')
    row += 1
    ws.cell(row=row, column=1).value = (
        "Terapia Intensiva e Blocco Operatorio sono aree funzionalmente "
        "inscindibili: il blocco operatorio \u00e8 di continuit\u00e0 "
        "alla terapia intensiva e il personale afferente \u00e8 spesso "
        "condiviso tra le due aree.\n\n"
        "Per ciascun presidio vengono prodotte tre tabelle:\n"
        "  1) Terapia Intensiva (stand-alone): FTE calcolato sui posti "
        "letto intensivi, personale assegnato al CDC "
        "\u00abTerapia Intensiva\u00bb;\n"
        "  2) Blocco Operatorio (stand-alone): FTE calcolato sulle "
        "sale operatorie, personale assegnato al CDC "
        "\u00abBlocco Operatorio\u00bb;\n"
        "  3) Area Combinata: somma dei fabbisogni TI + BO, contro "
        "TUTTO il personale afferente alla SC/SSD di Anestesia e "
        "Terapia Intensiva (compresi i CDC condivisi quali "
        "\u00abAnestesia e Camere Operatorie\u00bb, "
        "\u00abCosti Comuni\u00bb, \u00abTerapia del Dolore\u00bb).\n\n"
        "Le carenze/eccedenze nella tabella combinata sono calcolate "
        "sull\u2019area complessiva, permettendo di valutare se il "
        "totale del personale \u00e8 adeguato all\u2019insieme delle "
        "esigenze di TI e Blocco Operatorio."
    )
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True,
                                                     vertical='top')
    row += 2

    # ============================================================
    # 6. STANDARD AGENAS – SERVIZI TERRITORIALI
    # ============================================================
    ws.cell(row=row, column=1,
            value="6. STANDARD AGENAS \u2013 SERVIZI TERRITORIALI"
            ).font = FONT_SECTION
    row += 1

    d6 = ws.cell(row=row, column=1)
    d6.value = (
        "Gli standard AGENAS per i servizi territoriali sono espressi "
        "come tassi di operatori per popolazione di riferimento. "
        "Il fabbisogno atteso \u00e8 calcolato con la formula "
        "seguente, dove la base e la fascia di popolazione variano "
        "per ogni area. Il confronto \u00e8 tra il Totale in servizio "
        "(T.I. + T.D.) e il valore atteso."
    )
    d6.font = FONT_NORMAL
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    row += 2

    # Formula generica
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    fc2 = ws.cell(row=row, column=1)
    fc2.value = ("Atteso = Popolazione_fascia \u00d7 Tasso / "
                 "Base_popolazione")
    fc2.font = FONT_FORMULA
    fc2.alignment = ALIGN_CENTER
    for ci in range(1, N_COLS + 1):
        ws.cell(row=row, column=ci).fill = FILL_FORMULA
        ws.cell(row=row, column=ci).border = THIN_BORDER
    row += 2

    # --- Helper per scrivere tabella territoriale ---
    def _scrivi_tabella_territoriale(ws, row, titolo, fonte, descrizione,
                                     base_pop, fascia_label,
                                     profili_tassi, uo_patterns,
                                     is_range=False):
        """profili_tassi: lista di tuple (nome, tasso) o
           (nome, tasso_min, tasso_regime) se is_range=True.
        """
        ws.cell(row=row, column=1, value=titolo).font = FONT_SUBSECT
        row += 1
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS)
        ws.cell(row=row, column=1,
                value=f"Fonte: {fonte}").font = Font(
                    italic=True, size=9, color='555555')
        row += 1
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS)
        ws.cell(row=row, column=1, value=descrizione).font = FONT_NORMAL
        row += 2

        # Parametri
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS)
        ws.cell(row=row, column=1,
                value=(f"Base popolazione: {base_pop:,} \u2013 "
                       f"Fascia: {fascia_label} \u2013 "
                       f"UO: {uo_patterns}")
                ).font = Font(size=9, color='444444')
        row += 1

        # Tabella profili tassi
        if is_range:
            hdrs = ['Profilo', 'Tasso Min', 'Tasso Regime',
                    'Qualifiche incluse']
        else:
            hdrs = ['Profilo', 'Tasso', 'Qualifiche incluse']

        for ci, cn in enumerate(hdrs, 1):
            c = ws.cell(row=row, column=ci, value=cn)
            c.font = FONT_HEADER; c.fill = FILL_HEADER
            c.alignment = ALIGN_CENTER; c.border = THIN_BORDER
        row += 1

        for ti, pt in enumerate(profili_tassi):
            fill = FILL_A if ti % 2 == 0 else FILL_B
            if is_range:
                nome, t_min, t_max, quals = pt
                vals = [nome, t_min, t_max, quals]
            else:
                nome, tasso, quals = pt
                vals = [nome, tasso, quals]
            for ci, v in enumerate(vals, 1):
                c2 = ws.cell(row=row, column=ci, value=v)
                c2.font = FONT_NORMAL; c2.fill = fill
                c2.border = THIN_BORDER
                if ci > 1 and ci < len(vals) + 1 - (1 if not is_range else 1):
                    c2.alignment = ALIGN_CENTER
            row += 1

        row += 1
        return row

    # --- 6a. Salute Mentale ---
    SM_PROFILI = [
        ('Medici Psichiatri', 1.0,
         'DIRIGENTE MEDICO'),
        ('Psicologi Psicoterapeuti', 0.5,
         'DIRIGENTE PSICOLOGO'),
        ('Prof. Sanitarie + Ass. Sociali', 5.0,
         'INFERMIERE, TRP, ASSISTENTE SOCIALE, '
         'EDUCATORE PROF., OSS'),
        ('Altro Personale', 0.2,
         'COLL. AMM.VO, ASS. AMM.VO, '
         'OPERATORE TECNICO'),
    ]
    row = _scrivi_tabella_territoriale(
        ws, row,
        titolo="6a. Area Salute Mentale Adulti",
        fonte=("DPR 1/11/1999 + AGENAS servizi territoriali \u2013 "
               "indicatori_agenas_salute_mentale.xml"),
        descrizione=(
            "Tasso su 10.000 abitanti \u2265 18 anni. "
            "Totale minimo: \u2265 6,7 operatori / 10.000 ab. "
            "Il totale dei profili (1,0 + 0,5 + 5,0 + 0,2 = 6,7) "
            "rappresenta la soglia minima complessiva."
        ),
        base_pop=10_000,
        fascia_label='residenti \u2265 18 anni',
        profili_tassi=SM_PROFILI,
        uo_patterns='CSM|SPDC|SALUTE MENTALE',
    )

    # --- 6b. Dipendenze (SerD) ---
    SERD_PROFILI = [
        ('Medico', 3.0, 4.0, 'DIRIGENTE MEDICO'),
        ('Psicologo', 3.0, 3.5, 'DIRIGENTE PSICOLOGO'),
        ('Infermiere', 4.0, 6.0, 'INFERMIERE'),
        ('Educatore Prof. / TeRP', 2.5, 3.5,
         'EDUCATORE PROF., TRP'),
        ('Assistente Sociale', 2.0, 3.0, 'ASSISTENTE SOCIALE'),
        ('Amministrativo', 0.5, 1.0,
         'COLL. AMM.VO, ASS. AMM.VO, COLL. TECNICO'),
    ]
    row = _scrivi_tabella_territoriale(
        ws, row,
        titolo="6b. Area Dipendenze Patologiche (SerD)",
        fonte=("Tabella 1 AGENAS \u2013 "
               "indicatori_agenas_dipendenze.xml"),
        descrizione=(
            "Tasso su 100.000 residenti 15\u201364 anni. "
            "Ogni profilo ha due valori: standard minimo e standard "
            "a regime. L\u2019esito \u00e8 a 3 stati: CARENZA (sotto il "
            "min), IN RANGE (tra min e regime), A REGIME (\u2265 regime)."
        ),
        base_pop=100_000,
        fascia_label='residenti 15\u201364 anni',
        profili_tassi=SERD_PROFILI,
        uo_patterns='DIPENDENZ',
        is_range=True,
    )

    # --- 6c. NPIA ---
    NPIA_PROFILI = [
        ('Dirigenza Sanitaria', 6.0,
         'DIR. MEDICO NPI + DIR. PSICOLOGO'),
        ('Prof. San. + Ass. Sociali', 10.0,
         'INFERMIERE, TNPEE, EDUCATORE, TRP, '
         'FISIOTERAPISTA, LOGOPEDISTA, ASS. SOC.'),
        ('Altro Personale', 0.2,
         'COLL. AMM.VO, ASS. AMM.VO, OP. TECNICO'),
    ]
    row = _scrivi_tabella_territoriale(
        ws, row,
        titolo="6c. Area NPIA \u2013 Neuropsichiatria Infanzia e Adolescenza",
        fonte=("Documento AGENAS servizi territoriali \u2013 "
               "indicatori_agenas_npia.xml"),
        descrizione=(
            "Tasso su 10.000 abitanti 1\u201317 anni. "
            "Il profilo \u00abDirigenza Sanitaria\u00bb aggrega medici NPI "
            "e psicologi psicoterapeuti."
        ),
        base_pop=10_000,
        fascia_label='residenti 1\u201317 anni',
        profili_tassi=NPIA_PROFILI,
        uo_patterns='NEUROPSICH',
    )

    # --- 6d. Salute in Carcere ---
    CARC_PROFILI = [
        ('Medici SM+SerD', 2.0,
         'DIRIGENTE MEDICO (1 SM + 1 SerD)'),
        ('Psicologi SM+SerD', 2.0,
         'DIRIGENTE PSICOLOGO (1 SM + 1 SerD)'),
        ('Infermieri', 1.0,
         'INFERMIERE (SM: 1 prof.san. / 350 det.)'),
        ('Assistenti Sociali', 1.0,
         'ASSISTENTE SOCIALE (SerD: 1 / 350 det.)'),
    ]
    row = _scrivi_tabella_territoriale(
        ws, row,
        titolo="6d. Area Salute in Carcere",
        fonte=("Standard minimi AGENAS \u2013 "
               "indicatori_agenas_carcere.xml"),
        descrizione=(
            "Tasso ogni 350 detenuti (dato per istituto penitenziario). "
            "Standard combinato Salute Mentale + SerD penitenziario."
        ),
        base_pop=350,
        fascia_label='detenuti per istituto',
        profili_tassi=CARC_PROFILI,
        uo_patterns='DETENUTI|CARCER|PENITEN',
    )

    # ============================================================
    # 7. DATI DI INGRESSO UTILIZZATI
    # ============================================================
    ws.cell(row=row, column=1,
            value="7. DATI DI INGRESSO UTILIZZATI"
            ).font = FONT_SECTION
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1, value=(
        "I dati seguenti sono configurati in src/config.py e "
        "determinano la classificazione dei presidi e il calcolo "
        "degli attesi territoriali."
    )).font = FONT_NORMAL
    row += 2

    # --- 7a. Livello presidio ---
    ws.cell(row=row, column=1,
            value="7a. Livello dei presidi ospedalieri (DM 70/2015)"
            ).font = FONT_SUBSECT
    row += 1
    PRES_LABELS = {
        'OSPEDALE_DI_BASE': 'Ospedale di Base',
        'PRESIDIO_I_LIVELLO': 'I Livello',
        'PRESIDIO_II_LIVELLO': 'II Livello',
    }
    hdrs_pres = ['Presidio', 'Livello']
    for ci, cn in enumerate(hdrs_pres, 1):
        c = ws.cell(row=row, column=ci, value=cn)
        c.font = FONT_HEADER; c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER; c.border = THIN_BORDER
    row += 1
    for ti, (pres, liv) in enumerate(LIVELLO_PRESIDIO.items()):
        fill = FILL_A if ti % 2 == 0 else FILL_B
        c = ws.cell(row=row, column=1, value=pres)
        c.font = FONT_NORMAL; c.fill = fill; c.border = THIN_BORDER
        c2 = ws.cell(row=row, column=2,
                     value=PRES_LABELS.get(liv, liv))
        c2.font = FONT_NORMAL; c2.fill = fill
        c2.alignment = ALIGN_CENTER; c2.border = THIN_BORDER
        row += 1
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1, value=(
        "Il livello determina quale riga delle tabelle AGENAS "
        "ospedaliere (Tab. 13\u201320) viene applicata al presidio."
    )).font = Font(size=9, color='444444')
    row += 2

    # --- 7b. Parti per presidio ---
    ws.cell(row=row, column=1,
            value="7b. Parti annui per presidio (Tab. 11)"
            ).font = FONT_SUBSECT
    row += 1
    hdrs_parti = ['Presidio', 'N. Parti annui', 'Fascia AGENAS']
    for ci, cn in enumerate(hdrs_parti, 1):
        c = ws.cell(row=row, column=ci, value=cn)
        c.font = FONT_HEADER; c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER; c.border = THIN_BORDER
    row += 1
    for ti, (pres, npar) in enumerate(PARTI_PER_PRESIDIO.items()):
        fill = FILL_A if ti % 2 == 0 else FILL_B
        if npar < 500:
            fascia = '< 500 (sotto soglia)'
        elif npar <= 1500:
            fascia = '500\u20131.500'
        elif npar <= 2000:
            fascia = '1.500\u20132.000'
        else:
            fascia = '> 2.000'
        c = ws.cell(row=row, column=1, value=pres)
        c.font = FONT_NORMAL; c.fill = fill; c.border = THIN_BORDER
        c2 = ws.cell(row=row, column=2, value=npar)
        c2.font = FONT_NORMAL; c2.fill = fill
        c2.alignment = ALIGN_CENTER; c2.border = THIN_BORDER
        c3 = ws.cell(row=row, column=3, value=fascia)
        c3.font = FONT_NORMAL; c3.fill = fill
        c3.alignment = ALIGN_CENTER; c3.border = THIN_BORDER
        row += 1
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1, value=(
        "Il numero di parti determina la fascia della Tabella 11 "
        "AGENAS (Area Materno Infantile)."
    )).font = Font(size=9, color='444444')
    row += 2

    # --- 7c. Popolazione per area distrettuale ---
    ws.cell(row=row, column=1,
            value="7c. Popolazione per area distrettuale (ISTAT)"
            ).font = FONT_SUBSECT
    row += 1
    POP_HDRS = ['Area', 'Totale', '\u2265 18 anni',
                '15\u201364 anni', '1\u201317 anni']
    for ci, cn in enumerate(POP_HDRS, 1):
        c = ws.cell(row=row, column=ci, value=cn)
        c.font = FONT_HEADER; c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER; c.border = THIN_BORDER
    row += 1
    for ti, (area, dati) in enumerate(POPOLAZIONE_AREA.items()):
        fill = FILL_A if ti % 2 == 0 else FILL_B
        totale = dati.get('totale', 0)
        # gestione totale come float (es. 116.489 = 116489)
        if isinstance(totale, float) and totale < 1000:
            totale = int(totale * 1000)
        vals = [
            area, int(totale),
            int(dati.get('gte_18', 0)),
            int(dati.get('range_15_64', 0)),
            int(dati.get('range_1_17', 0)),
        ]
        for ci, v in enumerate(vals, 1):
            c2 = ws.cell(row=row, column=ci, value=v)
            c2.font = FONT_NORMAL; c2.fill = fill; c2.border = THIN_BORDER
            if ci > 1:
                c2.alignment = ALIGN_CENTER
                c2.number_format = '#,##0'
        row += 1
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1, value=(
        "Le fasce di popolazione sono utilizzate come parametro nella "
        "formula: Atteso = Popolazione_fascia \u00d7 Tasso / Base. "
        "Ogni area territoriale usa la fascia appropriata."
    )).font = Font(size=9, color='444444')
    row += 2

    # --- 7d. Detenuti per istituto ---
    ws.cell(row=row, column=1,
            value="7d. Detenuti per istituto penitenziario"
            ).font = FONT_SUBSECT
    row += 1
    hdrs_det = ['Area', 'N. Detenuti', 'Atteso (base 350)']
    for ci, cn in enumerate(hdrs_det, 1):
        c = ws.cell(row=row, column=ci, value=cn)
        c.font = FONT_HEADER; c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER; c.border = THIN_BORDER
    row += 1
    for ti, (area, ndet) in enumerate(DETENUTI_PER_ISTITUTO.items()):
        fill = FILL_A if ti % 2 == 0 else FILL_B
        c = ws.cell(row=row, column=1, value=area)
        c.font = FONT_NORMAL; c.fill = fill; c.border = THIN_BORDER
        c2 = ws.cell(row=row, column=2, value=ndet)
        c2.font = FONT_NORMAL; c2.fill = fill
        c2.alignment = ALIGN_CENTER; c2.border = THIN_BORDER
        c3 = ws.cell(row=row, column=3,
                     value=f"{ndet / 350:.2f} unit\u00e0 base")
        c3.font = FONT_NORMAL; c3.fill = fill
        c3.alignment = ALIGN_CENTER; c3.border = THIN_BORDER
        row += 1
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1, value=(
        "Es. Campobasso: 184 detenuti, tasso Medici = 2,0/350 \u2192 "
        "atteso = 184 \u00d7 2,0 / 350 = 1,05 \u2248 1 medico."
    )).font = Font(size=9, color='444444')
    row += 2

    # ============================================================
    # 8. METODO DI CONFRONTO CON GLI STANDARD AGENAS
    # ============================================================
    ws.cell(row=row, column=1,
            value="8. METODO DI CONFRONTO"
            ).font = FONT_SECTION
    row += 1

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1, value=(
        "Per ogni tabella AGENAS, il personale in servizio \u00e8 "
        "conteggiato come somma di T.I. (tempo indeterminato) e T.D. "
        "(tempo determinato). Le tabelle mostrano le tre componenti "
        "T.I., T.D. e Totale. L\u2019esito \u00e8 calcolato "
        "confrontando il Totale con il range standard:"
    )).font = FONT_NORMAL
    row += 2

    ws.cell(row=row, column=1,
            value="Procedimento:").font = FONT_BOLD_SM
    row += 1
    STEPS = [
        "1. Identificare il livello del presidio (o la fascia di attivit\u00e0) "
        "dalla configurazione (config.py).",
        "2. Selezionare la riga corrispondente nella tabella AGENAS "
        "(XML di configurazione).",
        "3. Per ogni profilo, estrarre i valori Min e Max (ospedalieri) "
        "o il Tasso (territoriali).",
        "4. Contare il personale T.I. e T.D. per UO e profilo dal DB "
        "del personale (matching regex su UO e qualifica).",
        "5. Calcolare Totale = T.I. + T.D. e confrontare con lo standard.",
        "6. Determinare l\u2019esito secondo i criteri della tabella seguente.",
    ]
    for step in STEPS:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS)
        ws.cell(row=row, column=1, value=step).font = FONT_NORMAL
        row += 1
    row += 1

    METODO_ESITI = [
        ('Area ospedaliera\n(range min\u2013max)',
         '\u2022 CARENZA: Totale < min \u2192 indicata la differenza\n'
         '\u2022 IN RANGE: min \u2264 Totale \u2264 max\n'
         '\u2022 ECCEDENZA: Totale > max \u2192 indicato il surplus',
         FILL_A),
        ('Area territoriale\n(tasso singolo)',
         '\u2022 OK: Totale \u2265 valore atteso\n'
         '\u2022 CARENZA: Totale < valore atteso '
         '\u2192 indicata la differenza',
         FILL_B),
        ('Area territoriale\n(range min/regime)',
         '\u2022 CARENZA: Totale < atteso minimo\n'
         '\u2022 IN RANGE: atteso min \u2264 Totale < atteso a regime\n'
         '\u2022 A REGIME: Totale \u2265 atteso a regime',
         FILL_A),
    ]
    # Header
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=3)
    for ci in range(1, 4):
        c = ws.cell(row=row, column=ci,
                    value='Tipologia' if ci == 1 else None)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    ws.merge_cells(start_row=row, start_column=4,
                   end_row=row, end_column=N_COLS)
    for ci in range(4, N_COLS + 1):
        c = ws.cell(row=row, column=ci,
                    value='Criteri di valutazione' if ci == 4 else None)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    for tipo, criteri, fill in METODO_ESITI:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=3)
        for ci in range(1, 4):
            c = ws.cell(row=row, column=ci,
                        value=tipo if ci == 1 else None)
            c.font = FONT_BOLD_SM
            c.fill = fill
            c.border = THIN_BORDER
            c.alignment = Alignment(wrap_text=True, vertical='center')
        ws.merge_cells(start_row=row, start_column=4,
                       end_row=row, end_column=N_COLS)
        for ci in range(4, N_COLS + 1):
            c = ws.cell(row=row, column=ci,
                        value=criteri if ci == 4 else None)
            c.font = FONT_NORMAL
            c.fill = fill
            c.border = THIN_BORDER
            c.alignment = Alignment(wrap_text=True, vertical='center')
        ws.row_dimensions[row].height = 50
        row += 1

    row += 2
    # ============================================================
    # 9. FONTI DATI E RIFERIMENTI NORMATIVI
    # ============================================================
    ws.cell(row=row, column=1,
            value="9. FONTI DATI E RIFERIMENTI NORMATIVI"
            ).font = FONT_SECTION
    row += 1
    FONTI = [
        "\u2022 DM 70/2015 \u2013 Regolamento standard qualitativi, "
        "strutturali, tecnologici e quantitativi delle strutture "
        "ospedaliere (classificazione presidi: Base, I, II livello).",
        "\u2022 DM 77/2022 \u2013 Standard per l\u2019assistenza "
        "territoriale (Ospedali di Comunit\u00e0, Case di Comunit\u00e0).",
        "\u2022 DPR 1 Novembre 1999 \u2013 Progetto Obiettivo Salute "
        "Mentale (standard operatori salute mentale adulti: 6,7/10.000).",
        "\u2022 AGENAS \u2013 Standard di personale per le strutture "
        "ospedaliere e territoriali del SSN (Tabelle 11, 13, 14, "
        "15, 16, 17, 18, 20 e standard territoriali SerD, NPIA, Carcere).",
        "\u2022 Dati ISTAT \u2013 Popolazione residente per fasce "
        "d\u2019et\u00e0 e genere, per area distrettuale "
        "(fonte: config.py / POPOLAZIONE_AREA).",
        "\u2022 Dati aziendali \u2013 Parti annui per presidio "
        "(SDO), posti letto attivi, detenuti per istituto "
        "penitenziario, dotazione organica comunicata dai Primari.",
        "\u2022 File XML di configurazione \u2013 Tutti gli standard "
        "e i mapping sono definiti nei file XML nella cartella "
        "configurazione/ e possono essere aggiornati senza "
        "modificare il codice sorgente.",
    ]
    for testo in FONTI:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS)
        ws.cell(row=row, column=1, value=testo).font = FONT_NORMAL
        row += 1

    auto_larghezza_colonne(ws, colonne=N_COLS, larghezza_min=10)

    # Forza la colonna A a una larghezza leggibile per le tabelle
    # indicatori, evitando che i testi descrittivi (merge multi-colonna)
    # la allarghino a dismisura.
    ws.column_dimensions['A'].width = 40
