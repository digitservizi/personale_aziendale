"""
Report Ospedali di Comunità – confronto personale vs fabbisogno DM 77/2022.
Include foglio RIEPILOGO e fogli dettaglio per sede ODC.
"""

import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from src.stili_excel import (
    THIN_BORDER, FILL_A, FILL_B,
    scrivi_titolo, scrivi_intestazioni, scrivi_riga_dati,
    auto_larghezza_colonne,
)
from src.caricamento_dati import (
    carica_dataframe, normalizza_colonne_personale,
    normalizza_colonne_pensionamenti,
)
from src.caricamento_xml import carica_fabbisogno_odc_dm77


# ============================================================
# HELPER INTERNO
# ============================================================

def _scrivi_foglio_odc(ws, titolo, colonne, righe_df):
    """Scrive un foglio ODC con titolo, intestazioni e righe colorate."""
    scrivi_titolo(ws, titolo, len(colonne))
    scrivi_intestazioni(ws, colonne)

    prev_struttura = None
    fill_idx = 0
    for r_idx, (_, riga) in enumerate(righe_df.iterrows(), start=3):
        struttura = riga.get('Struttura', '')
        if struttura != prev_struttura:
            if prev_struttura is not None:
                fill_idx += 1
            prev_struttura = struttura
        fill = FILL_A if fill_idx % 2 == 0 else FILL_B
        valori = [riga[c] for c in colonne]
        scrivi_riga_dati(ws, r_idx, valori, fill)

    auto_larghezza_colonne(ws, colonne)


# ============================================================
# REPORT ODC
# ============================================================

def genera_report_odc(personale_file, pensionamenti_file, lista_odc,
                      indicatori_odc_file, output_file, anno_analisi):
    """
    Genera il report XLSX degli Ospedali di Comunità con confronto
    DM 77/2022.
    """
    print(f"\n{'=' * 70}")
    print("REPORT OSPEDALI DI COMUNITÀ - CONFRONTO DM 77/2022")
    print(f"{'=' * 70}\n")

    # Carica personale
    personale_df = carica_dataframe(personale_file)
    personale_df = normalizza_colonne_personale(personale_df)
    personale_df = personale_df[
        personale_df['DESC_NATURA'].str.upper() == "TEMPO INDETERMINATO"
    ]

    # Pensionamenti
    pensionamenti_df = carica_dataframe(pensionamenti_file)
    pensionamenti_df = normalizza_colonne_pensionamenti(pensionamenti_df)

    if 'DT_CESSAZIONE' in personale_df.columns:
        personale_df = personale_df.rename(
            columns={'DT_CESSAZIONE': 'DT_CESSAZIONE_PERS'}
        )
    personale_df = pd.merge(
        personale_df,
        pensionamenti_df[['MATR.', 'DT_CESSAZIONE']],
        on='MATR.',
        how='left',
    )

    anni_pensionamento = [anno_analisi + 1, anno_analisi + 2, anno_analisi + 3]

    # Fabbisogno DM 77
    fabbisogno_dm77 = {}
    mappa_profili_dm77 = {}
    if indicatori_odc_file and os.path.exists(indicatori_odc_file):
        fabbisogno_dm77, mappa_profili_dm77 = carica_fabbisogno_odc_dm77(
            indicatori_odc_file
        )

    # Mapping inverso: categoria DM77 → set di PROFILO_RAGGRUPPATO (upper)
    # Usato per determinare a quale categoria DM77 appartiene un profilo
    # senza sovrascrivere PROFILO_RAGGRUPPATO (che resta da profili_atto_aziendale.xml)
    cat_dm77_per_profilo = {}   # PROFILO_UPPER → CATEGORIA_DM77_UPPER
    for prof_upper, cat_upper in mappa_profili_dm77.items():
        cat_dm77_per_profilo[prof_upper] = cat_upper

    righe_report = []

    for odc in lista_odc:
        nome_odc = odc['nome']
        print(f"  === {nome_odc} ===")

        for struttura in odc['strutture']:
            nome_struttura = struttura['nome']
            pattern_cdc = struttura['pattern_cdc']
            posti_letto_ord = struttura['posti_letto']['ordinari']
            note = struttura.get('note', '')

            is_udi = (
                'U.D.I' in nome_struttura.upper()
                or 'OSP.*COMUNITA' in pattern_cdc.upper()
            )

            mask = personale_df['DESC_TIPO_CDC'].str.contains(
                pattern_cdc, case=False, na=False, regex=True
            )
            personale_struttura = personale_df[mask]

            if len(personale_struttura) == 0 and posti_letto_ord == 0:
                continue

            # NON sovrascrivere PROFILO_RAGGRUPPATO: restano i nomi
            # dall'unico punto di verità (profili_atto_aziendale.xml).
            # Il mapping DM77 è usato solo per il confronto fabbisogno.

            profili = personale_struttura.groupby(
                'PROFILO_RAGGRUPPATO'
            ).agg(
                IN_SERVIZIO=('PROFILO_RAGGRUPPATO', 'size')
            ).reset_index()

            if len(profili) == 0:
                riga = {
                    'ODC': nome_odc,
                    'STRUTTURA': nome_struttura,
                    'POSTI_LETTO': posti_letto_ord,
                    'TIPO': 'DM 77' if is_udi else 'Struttura',
                    'PROFILO': '-',
                    'CATEGORIA_DM77': '-',
                    'IN_SERVIZIO': 0,
                    'FABBISOGNO_DM77': '-',
                    'DELTA': '-',
                    'NOTE': note,
                }
                for anno in anni_pensionamento:
                    riga[f'PENSIONAMENTI_{anno}'] = 0
                    riga[f'PROIEZIONE_{anno}'] = 0
                righe_report.append(riga)
                print(
                    f"    {nome_struttura:45s}  PL: {posti_letto_ord:3d}  "
                    f"Personale: 0"
                )
                continue

            # Per le strutture U.D.I. calcolo fabbisogno per categoria DM77
            # aggregando i profili che vi ricadono
            totale_struttura = 0
            totale_fabb = 0
            categorie_presenti = {}  # cat_upper → totale in servizio
            fabb_gia_contati = set()

            if is_udi and fabbisogno_dm77:
                # Raggruppa i profili presenti per categoria DM77
                for _, pr in profili.iterrows():
                    cat = cat_dm77_per_profilo.get(
                        pr['PROFILO_RAGGRUPPATO'].strip().upper()
                    )
                    if cat:
                        categorie_presenti[cat] = (
                            categorie_presenti.get(cat, 0)
                            + pr['IN_SERVIZIO']
                        )

            for _, pr in profili.iterrows():
                profilo = pr['PROFILO_RAGGRUPPATO']
                in_servizio = pr['IN_SERVIZIO']
                totale_struttura += in_servizio

                cat_dm77 = cat_dm77_per_profilo.get(
                    profilo.strip().upper()
                )

                if is_udi and fabbisogno_dm77 and cat_dm77:
                    fabb = fabbisogno_dm77.get(cat_dm77, 0)
                    tot_cat = categorie_presenti.get(cat_dm77, 0)
                    delta = tot_cat - fabb
                    fabb_str = fabb
                    # Somma fabbisogno solo una volta per categoria
                    if cat_dm77 not in fabb_gia_contati:
                        totale_fabb += fabb
                        fabb_gia_contati.add(cat_dm77)
                else:
                    cat_dm77 = '-'
                    fabb_str = '-'
                    delta = '-'

                pers_profilo = personale_struttura[
                    personale_struttura['PROFILO_RAGGRUPPATO'] == profilo
                ]
                pens = {}
                for anno in anni_pensionamento:
                    pens[anno] = pers_profilo['DT_CESSAZIONE'].apply(
                        lambda x, a=anno: (
                            1 if pd.to_datetime(x, errors='coerce').year == a
                            else 0
                        )
                    ).sum()

                riga = {
                    'ODC': nome_odc,
                    'STRUTTURA': nome_struttura,
                    'POSTI_LETTO': posti_letto_ord,
                    'TIPO': 'DM 77' if is_udi else 'Struttura',
                    'PROFILO': profilo,
                    'CATEGORIA_DM77': cat_dm77 if cat_dm77 else '-',
                    'IN_SERVIZIO': in_servizio,
                    'FABBISOGNO_DM77': fabb_str,
                    'DELTA': delta,
                    'NOTE': note,
                }
                for anno in anni_pensionamento:
                    riga[f'PENSIONAMENTI_{anno}'] = pens[anno]
                pens_cum = 0
                for anno in anni_pensionamento:
                    pens_cum += pens[anno]
                    riga[f'PROIEZIONE_{anno}'] = in_servizio - pens_cum
                righe_report.append(riga)

            # Categorie DM 77 completamente mancanti
            if is_udi and fabbisogno_dm77:
                categorie_coperte = set()
                for _, pr in profili.iterrows():
                    cat = cat_dm77_per_profilo.get(
                        pr['PROFILO_RAGGRUPPATO'].strip().upper()
                    )
                    if cat:
                        categorie_coperte.add(cat)

                for profilo_dm, fabb_dm in fabbisogno_dm77.items():
                    if profilo_dm not in categorie_coperte:
                        riga = {
                            'ODC': nome_odc,
                            'STRUTTURA': nome_struttura,
                            'POSTI_LETTO': posti_letto_ord,
                            'TIPO': 'DM 77',
                            'PROFILO': f'[MANCANTE] {profilo_dm}',
                            'CATEGORIA_DM77': profilo_dm,
                            'IN_SERVIZIO': 0,
                            'FABBISOGNO_DM77': fabb_dm,
                            'DELTA': -fabb_dm,
                            'NOTE': 'Categoria DM 77 assente',
                        }
                        for anno in anni_pensionamento:
                            riga[f'PENSIONAMENTI_{anno}'] = 0
                            riga[f'PROIEZIONE_{anno}'] = 0
                        righe_report.append(riga)
                        totale_fabb += fabb_dm

            stato = ''
            if is_udi and fabbisogno_dm77:
                diff = totale_struttura - totale_fabb
                stato = f'  DM77: {totale_fabb}  Delta: {diff:+d}'
            print(
                f"    {nome_struttura:45s}  PL: {posti_letto_ord:3d}  "
                f"Personale: {totale_struttura:3d}{stato}"
            )

        print()

    # ------------------------------------------------------------------
    # Genera XLSX
    # ------------------------------------------------------------------
    report_df = pd.DataFrame(righe_report)

    rename_cols = {
        'ODC': 'ODC',
        'STRUTTURA': 'Struttura',
        'POSTI_LETTO': 'Posti Letto',
        'TIPO': 'Tipo',
        'PROFILO': 'Profilo Professionale',
        'CATEGORIA_DM77': 'Categoria DM 77',
        'IN_SERVIZIO': 'In Servizio',
        'FABBISOGNO_DM77': 'Fabbisogno DM 77',
        'DELTA': 'Delta',
        'NOTE': 'Note',
    }
    for anno in anni_pensionamento:
        rename_cols[f'PENSIONAMENTI_{anno}'] = f'Pens./Ces. {anno}'
        rename_cols[f'PROIEZIONE_{anno}'] = f'Proiez. {anno}'
    report_df = report_df.rename(columns=rename_cols)

    wb = Workbook()
    wb.remove(wb.active)
    colonne_dati = [c for c in report_df.columns
                    if c not in ('ODC', 'Categoria DM 77')]

    # ====== FOGLIO RIEPILOGO ======
    # Aggrega per sede + profilo professionale (nomi da profili_atto_aziendale.xml)
    # Il fabbisogno DM77 è mostrato per ciascun profilo che rientra in una
    # categoria DM77; il delta è a livello di categoria (aggregato).
    ws_riep = wb.create_sheet(title='RIEPILOGO')
    col_pens_odc = [f'Pens./Ces. {a}' for a in anni_pensionamento]
    col_riep = (
        ['Sede', 'Profilo Professionale', 'In Servizio',
         'Fabbisogno DM 77', 'Delta']
        + col_pens_odc + ['Proiez. tot.']
    )

    scrivi_titolo(ws_riep, 'Riepilogo Ospedali di Comunità', len(col_riep))
    scrivi_intestazioni(ws_riep, col_riep)

    riep_row = 3
    fill_idx_riep = 0
    prev_odc_riep = None

    for nome_odc, df_odc_riep in report_df.groupby('ODC', sort=False):
        if nome_odc != prev_odc_riep:
            if prev_odc_riep is not None:
                fill_idx_riep += 1
            prev_odc_riep = nome_odc
        fill_riep = FILL_A if fill_idx_riep % 2 == 0 else FILL_B

        # Solo le strutture U.D.I. hanno fabbisogno DM77;
        # le altre strutture non hanno Categoria DM 77
        for profilo, df_prof in df_odc_riep.groupby(
            'Profilo Professionale', sort=True
        ):
            in_serv = int(df_prof['In Servizio'].sum())

            # Determina la categoria DM77 del profilo (se esiste)
            cat_vals = df_prof['Categoria DM 77'].dropna().unique()
            cat_vals = [c for c in cat_vals if c != '-']
            cat_dm77 = cat_vals[0] if cat_vals else None

            if cat_dm77:
                # Fabbisogno della categoria DM77
                fabb = fabbisogno_dm77.get(cat_dm77, 0)
                # Totale in servizio dell'intera categoria nella sede
                mask_cat = df_odc_riep['Categoria DM 77'] == cat_dm77
                tot_cat_serv = int(df_odc_riep.loc[
                    mask_cat, 'In Servizio'
                ].sum())
                delta = tot_cat_serv - fabb
            else:
                fabb = '-'
                delta = '-'

            pens_vals = [int(df_prof[cp].sum()) for cp in col_pens_odc]
            proiezione = in_serv - sum(pens_vals)

            valori = (
                [nome_odc, profilo, in_serv, fabb, delta]
                + pens_vals + [proiezione]
            )
            scrivi_riga_dati(ws_riep, riep_row, valori, fill_riep)
            riep_row += 1

    auto_larghezza_colonne(ws_riep, col_riep)

    # ====== FOGLI DETTAGLIO ======
    for nome_odc, df_sede in report_df.groupby('ODC', sort=False):
        df_sede = df_sede.drop(columns=['ODC', 'Categoria DM 77'])
        sheet_name = str(nome_odc)[:31]
        ws = wb.create_sheet(title=sheet_name)
        _scrivi_foglio_odc(
            ws, f'Ospedali di Comunità - {nome_odc}',
            list(df_sede.columns), df_sede,
        )

    wb.save(output_file)

    print(f"  Report salvato in: {output_file}")
    print(f"{'=' * 70}\n")

    return report_df
