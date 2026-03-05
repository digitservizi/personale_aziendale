"""
Dettaglio nominativo del personale per area AGENAS.

Genera 'dettaglio_agenas_{anno}.xlsx' con:
  - Un foglio RIEPILOGO con il conteggio per area
  - Un foglio per ciascuna area AGENAS con l'elenco nominativo
    del personale afferente (match su UO + profilo)

Le aree sono:
  Materno Infantile, Radiologia, Anatomia Patologica,
  Laboratorio, Tecnici Laboratorio, Medicina Legale,
  Trasfusionale, Emergenza-Urgenza, Terapia Intensiva,
  Sale Operatorie, Salute Mentale, Dipendenze, NPIA,
  Salute in Carcere
"""

import os
import re
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.caricamento_dati import carica_dataframe, normalizza_colonne_personale
from src.stili_excel import auto_larghezza_colonne

# ─────────────────────────────────────────────────────────────
# STILI (stessi di dettaglio_atto_aziendale.py)
# ─────────────────────────────────────────────────────────────
_FILL_HEADER = PatternFill('solid', fgColor='1F4E79')
_FILL_SUBHDR = PatternFill('solid', fgColor='2E75B6')
_FILL_A      = PatternFill('solid', fgColor='DCE6F1')
_FILL_B      = PatternFill('solid', fgColor='FFFFFF')
_FILL_TOTALE = PatternFill('solid', fgColor='BDD7EE')

_FONT_HEADER = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
_FONT_TITOLO = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
_FONT_BODY   = Font(name='Calibri', size=10)
_FONT_TOTALE = Font(name='Calibri', bold=True, size=10)

_THIN   = Side(style='thin', color='B8CCE4')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_C  = Alignment(horizontal='center', vertical='center', wrap_text=False)
_ALIGN_L  = Alignment(horizontal='left',   vertical='center', wrap_text=False)
_ALIGN_CW = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Colonne del foglio nominativo
_COLS = [
    'Matricola', 'Cognome', 'Nome', 'Data Nascita',
    'Forma Contrattuale', 'Profilo AGENAS',
    'Sede Fisica', 'Centro di Costo', 'SSD', 'Data Assunzione',
    'Data Cessazione',
]
_N_COLS = len(_COLS)


# ─────────────────────────────────────────────────────────────
# HELPER: FORMA CONTRATTUALE
# ─────────────────────────────────────────────────────────────
def _forma_contrattuale(desc_natura: str) -> str:
    n = str(desc_natura).strip().upper()
    if n.startswith('COMANDATO IN USCITA'):
        return 'T.I. – Comando in uscita'
    if n in ('PENITENZIARIO INDETERMINATO',):
        return 'T.I.'
    if n.startswith('TEMPO INDETERMINATO ASP'):
        return 'T.I.'
    if n == 'TEMPO INDETERMINATO':
        return 'T.I.'
    if '15_OCTIES' in n:
        return 'T.D. – Art. 15 Octies'
    if n == 'TEMPO DETERMINATO ART. 15 SEPTIES DLGS 502/92':
        return 'T.D. – Art. 15 Septies'
    if n == 'TEMPO DETERMINATO':
        return 'T.D.'
    if n == 'T.D. SPECIALIZZANDI':
        return 'Specializzando'
    if n == 'UNIVERSITARI H19':
        return 'Universitario'
    return str(desc_natura).strip()


def _ordine_natura(desc_natura: str) -> int:
    n = str(desc_natura).strip().upper()
    if n == 'TEMPO INDETERMINATO':          return 1
    if n == 'PENITENZIARIO INDETERMINATO':  return 2
    if n.startswith('COMANDATO IN USCITA'): return 3
    if n.startswith('TEMPO INDETERMINATO ASP'): return 4
    if n == 'TEMPO DETERMINATO':            return 5
    if n == 'TEMPO DETERMINATO ART. 15 SEPTIES DLGS 502/92': return 6
    if '15_OCTIES' in n:                   return 7
    if n == 'T.D. SPECIALIZZANDI':         return 8
    if n == 'UNIVERSITARI H19':            return 9
    return 99


# ─────────────────────────────────────────────────────────────
# HELPER: MATCH PERSONALE SU AREA AGENAS
# ─────────────────────────────────────────────────────────────
def _filtra_per_area(df: pd.DataFrame,
                     mapping_uo: list,
                     mapping_profili: list,
                     esclusioni: list = None,
                     match_cdc: bool = False,
                     solo_profilo: bool = False) -> pd.DataFrame:
    """Filtra il DataFrame raw del personale per una data area AGENAS.

    Parametri
    ---------
    df               : DataFrame normalizzato (da normalizza_colonne_personale)
    mapping_uo       : lista di {pattern} su DESC_SC_SSD_SS
    mapping_profili  : lista di {pattern, profilo_agenas} su PROFILO_RAGGRUPPATO
    esclusioni       : lista di pattern da escludere su DESC_SC_SSD_SS
    match_cdc        : se True, controlla anche DESC_TIPO_CDC per il match UO
    solo_profilo     : se True, ignora il filtro UO (tecnici lab, ecc.)
    """
    ssd_col  = df['DESC_SC_SSD_SS'].astype(str).str.upper()
    cdc_col  = df['DESC_TIPO_CDC'].astype(str).str.upper()
    prof_col = df['PROFILO_RAGGRUPPATO'].astype(str).str.upper()

    # Maschera UO
    if solo_profilo or not mapping_uo:
        mask_uo = pd.Series(True, index=df.index)
    else:
        mask_uo = pd.Series(False, index=df.index)
        for m in mapping_uo:
            pat = m['pattern']
            hits = ssd_col.str.contains(pat, na=False, regex=True)
            if match_cdc:
                hits |= cdc_col.str.contains(pat, na=False, regex=True)
            mask_uo |= hits

    # Esclusioni
    if esclusioni:
        for ep in esclusioni:
            mask_uo &= ~ssd_col.str.contains(ep, na=False, regex=True)

    # Maschera profilo
    mask_prof = pd.Series(False, index=df.index)
    profilo_agenas_map = {}   # idx → etichetta profilo agenas
    for mp in mapping_profili:
        pat  = mp['pattern']
        lbl  = mp.get('profilo_agenas', mp['pattern'])
        hits = prof_col.str.contains(pat, na=False, regex=True)
        for idx in hits[hits].index:
            if idx not in profilo_agenas_map:
                profilo_agenas_map[idx] = lbl
        mask_prof |= hits

    subset = df[mask_uo & mask_prof].copy()
    # Aggiunge colonna leggibile
    subset['_PROFILO_AGENAS'] = subset.index.map(
        lambda i: profilo_agenas_map.get(i, '')
    )
    return subset


def _filtra_territoriale(df: pd.DataFrame,
                         uo_patterns: list,
                         profili: list) -> pd.DataFrame:
    """Filtra per aree territoriali (salute mentale, dipendenze, npia, carcere).

    uo_patterns : lista di stringhe pattern su DESC_SC_SSD_SS o DESC_TIPO_CDC
    profili     : lista di {nome, qualifiche: [pattern]}
    """
    ssd_col  = df['DESC_SC_SSD_SS'].astype(str).str.upper()
    cdc_col  = df['DESC_TIPO_CDC'].astype(str).str.upper()
    qual_col = df['PROFILO_RAGGRUPPATO'].astype(str).str.upper()

    # Maschera UO
    if uo_patterns:
        mask_uo = pd.Series(False, index=df.index)
        for pat in uo_patterns:
            mask_uo |= ssd_col.str.contains(pat, na=False, regex=True)
            mask_uo |= cdc_col.str.contains(pat, na=False, regex=True)
    else:
        mask_uo = pd.Series(True, index=df.index)

    # Maschera profilo
    mask_prof = pd.Series(False, index=df.index)
    profilo_agenas_map = {}
    for prof in profili:
        nome = prof['nome']
        for q in prof.get('qualifiche', []):
            hits = qual_col.str.contains(q, na=False, regex=True)
            for idx in hits[hits].index:
                if idx not in profilo_agenas_map:
                    profilo_agenas_map[idx] = nome
            mask_prof |= hits

    subset = df[mask_uo & mask_prof].copy()
    subset['_PROFILO_AGENAS'] = subset.index.map(
        lambda i: profilo_agenas_map.get(i, '')
    )
    return subset


# ─────────────────────────────────────────────────────────────
# SCRITTURA FOGLIO NOMINATIVO
# ─────────────────────────────────────────────────────────────
def _scrivi_foglio_nominativo(wb: Workbook, sheet_name: str,
                               titolo: str, df_pers: pd.DataFrame) -> None:
    ws = wb.create_sheet(title=sheet_name[:31])

    # Riga 1: Titolo
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=_N_COLS)
    tc = ws.cell(row=1, column=1, value=titolo)
    tc.font      = _FONT_TITOLO
    tc.fill      = _FILL_SUBHDR
    tc.alignment = _ALIGN_CW
    tc.border    = _BORDER
    ws.row_dimensions[1].height = 22

    # Riga 2: Intestazioni
    for ci, col in enumerate(_COLS, 1):
        c = ws.cell(row=2, column=ci, value=col)
        c.font      = _FONT_HEADER
        c.fill      = _FILL_HEADER
        c.alignment = _ALIGN_C
        c.border    = _BORDER
    ws.row_dimensions[2].height = 18

    # Dati
    color_toggle = 0
    prev_grp     = None
    row          = 3

    for _, r in df_pers.iterrows():
        grp = r.get('DESC_SEDE_FISICA', '')
        if grp != prev_grp:
            color_toggle = 1 - color_toggle
            prev_grp = grp
        fill = _FILL_A if color_toggle else _FILL_B

        dt_nasc = r.get('DT_NASCITA', '')
        dt_ass  = r.get('PRIMA_DATA_ASSUNZIONE', '')
        dt_cess = r.get('DT_CESSAZIONE', '')
        try:
            dt_nasc = pd.to_datetime(dt_nasc).strftime('%d/%m/%Y')
        except Exception:
            dt_nasc = str(dt_nasc) if pd.notna(dt_nasc) else ''
        try:
            dt_ass = pd.to_datetime(dt_ass).strftime('%d/%m/%Y')
        except Exception:
            dt_ass = str(dt_ass) if pd.notna(dt_ass) else ''
        try:
            dt_cess = pd.to_datetime(dt_cess).strftime('%d/%m/%Y')
        except Exception:
            dt_cess = str(dt_cess) if pd.notna(dt_cess) else ''

        valori = [
            r.get('MATR.', ''),
            r.get('PF_COGNOME', ''),
            r.get('PF_NOME', ''),
            dt_nasc,
            _forma_contrattuale(r.get('DESC_NATURA', '')),
            r.get('_PROFILO_AGENAS', ''),
            r.get('DESC_SEDE_FISICA', ''),
            r.get('DESC_TIPO_CDC', ''),
            r.get('DESC_SC_SSD', ''),
            dt_ass,
            dt_cess,
        ]
        for ci, val in enumerate(valori, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font      = _FONT_BODY
            c.fill      = fill
            c.border    = _BORDER
            c.alignment = _ALIGN_C if ci in (1, 4, 5, 10, 11) else _ALIGN_L
        row += 1

    # Riga totale
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=_N_COLS)
    tc2 = ws.cell(row=row, column=1,
                  value=f'Totale dipendenti: {len(df_pers)}')
    tc2.font      = _FONT_TOTALE
    tc2.fill      = _FILL_TOTALE
    tc2.alignment = _ALIGN_C
    tc2.border    = _BORDER

    # Larghezze
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 26
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 28
    ws.column_dimensions['H'].width = 42
    ws.column_dimensions['I'].width = 36
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 16
    ws.freeze_panes = 'A3'


# ─────────────────────────────────────────────────────────────
# FOGLIO RIEPILOGO
# ─────────────────────────────────────────────────────────────
def _scrivi_foglio_riepilogo(wb: Workbook, titolo_wb: str,
                              righe: list, anno_analisi: int) -> None:
    ws = wb.create_sheet(title='RIEPILOGO', index=0)
    COLS = ['Area AGENAS', 'T.I.', 'T.D.', 'Totale']
    n_cols = len(COLS)

    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1, value=titolo_wb)
    tc.font      = _FONT_TITOLO
    tc.fill      = _FILL_SUBHDR
    tc.alignment = _ALIGN_CW
    tc.border    = _BORDER
    ws.row_dimensions[1].height = 22

    for ci, col in enumerate(COLS, 1):
        c = ws.cell(row=2, column=ci, value=col)
        c.font      = _FONT_HEADER
        c.fill      = _FILL_HEADER
        c.alignment = _ALIGN_C
        c.border    = _BORDER
    ws.row_dimensions[2].height = 18

    row = 3
    for idx, r in enumerate(righe):
        fill = _FILL_A if idx % 2 == 0 else _FILL_B
        nat = r.get('natura', '')
        n_ti  = r.get('n_ti', 0)
        n_td  = r.get('n_td', 0)
        n_tot = r.get('n_tot', 0)
        for ci, val in enumerate([r['nome'], n_ti, n_td, n_tot], 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font      = _FONT_BODY
            c.fill      = fill
            c.border    = _BORDER
            c.alignment = _ALIGN_C if ci > 1 else _ALIGN_L
        row += 1

    # Totale generale
    tot_ti  = sum(r.get('n_ti', 0)  for r in righe)
    tot_td  = sum(r.get('n_td', 0)  for r in righe)
    tot_tot = sum(r.get('n_tot', 0) for r in righe)
    for ci, val in enumerate(['TOTALE', tot_ti, tot_td, tot_tot], 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font      = _FONT_TOTALE
        c.fill      = _FILL_TOTALE
        c.border    = _BORDER
        c.alignment = _ALIGN_C if ci > 1 else _ALIGN_L

    ws.column_dimensions['A'].width = 42
    for col_letter in ['B', 'C', 'D']:
        ws.column_dimensions[col_letter].width = 14
    ws.freeze_panes = 'A3'


# ─────────────────────────────────────────────────────────────
# FUNZIONE PRINCIPALE
# ─────────────────────────────────────────────────────────────
def genera_dettaglio_agenas(
        personale_file: str,
        anno_analisi: int,
        output_dir: str,
        pensionamenti_file: str = None,
        # indicatori ospedalieri (opzionali)
        indicatori_agenas=None,           # materno infantile
        indicatori_radiologia=None,
        indicatori_anatomia_pat=None,
        indicatori_laboratorio=None,
        indicatori_tecnici_lab=None,
        indicatori_med_legale=None,
        indicatori_trasfusionale=None,
        indicatori_emergenza=None,
        indicatori_terapia_intensiva=None,
        indicatori_sale_operatorie=None,
        # indicatori territoriali
        indicatori_salute_mentale=None,
        indicatori_dipendenze=None,
        indicatori_npia=None,
        indicatori_carcere=None,
) -> str:
    """Genera 'dettaglio_agenas_{anno}.xlsx' con il personale per area AGENAS."""

    personale_df = carica_dataframe(personale_file)
    personale_df = normalizza_colonne_personale(personale_df)

    # DT_CESSAZIONE è già nel file personale; assicura la colonna esista
    if 'DT_CESSAZIONE' not in personale_df.columns:
        personale_df['DT_CESSAZIONE'] = ''

    # ── Escludi dipendenti già cessati ────────────────────────
    _oggi = date.today()
    _dt_cess = pd.to_datetime(personale_df['DT_CESSAZIONE'], errors='coerce')
    _mask_cessati = _dt_cess.notna() & (_dt_cess.dt.date <= _oggi)
    _n_cessati = int(_mask_cessati.sum())
    if _n_cessati:
        print(f"  [Dettaglio AGENAS] Esclusi {_n_cessati} dipendenti già cessati alla data {_oggi}")
        personale_df = personale_df[~_mask_cessati].copy()

    nat_upper = personale_df['DESC_NATURA'].str.upper()
    # Allineato alla logica di report_fabbisogno.py:
    # T.D. = contratti temporanei + specializzandi + universitari
    # T.I. = tutto il resto
    mask_td = (
        nat_upper.isin([
            'TEMPO DETERMINATO',
            'TEMPO DETERMINATO ART. 15 SEPTIES DLGS 502/92',
            'UNIVERSITARI H19',
            'T.D. SPECIALIZZANDI',
        ]) | nat_upper.str.contains('15_OCTIES', na=False)
    )
    mask_ti = ~mask_td

    wb = Workbook()
    wb.remove(wb.active)

    righe_riep = []

    # ── Helper interno ─────────────────────────────────────────────
    def _aggiungi_area(nome_area, df_area, sheet_name):
        """Ordina, scrive il foglio nominativo e raccoglie i dati per il riepilogo."""
        if df_area.empty:
            return

        df_area = df_area.copy()
        df_area['_ORD'] = df_area['DESC_NATURA'].apply(_ordine_natura)
        df_area = df_area.sort_values(
            ['DESC_SEDE_FISICA', 'DESC_TIPO_CDC', '_PROFILO_AGENAS',
             '_ORD', 'PF_COGNOME', 'PF_NOME']
        ).drop(columns=['_ORD'])

        n_ti  = int(mask_ti[df_area.index].sum())
        n_td  = int(mask_td[df_area.index].sum())
        n_tot = len(df_area)

        titolo = (
            f'{nome_area} – {anno_analisi} '
            f'(T.I.: {n_ti} | T.D.: {n_td} | Tot.: {n_tot})'
        )
        _scrivi_foglio_nominativo(wb, sheet_name, titolo, df_area)
        righe_riep.append({
            'nome':  nome_area,
            'n_ti':  n_ti,
            'n_td':  n_td,
            'n_tot': n_tot,
        })

    # ── Aree ospedaliere ───────────────────────────────────────────

    if indicatori_agenas:
        df_a = _filtra_per_area(
            personale_df,
            indicatori_agenas.get('mapping_uo', []),
            indicatori_agenas.get('mapping_profili', []),
        )
        _aggiungi_area('Materno Infantile', df_a, 'MATERNO INFANTILE')

    if indicatori_radiologia:
        df_a = _filtra_per_area(
            personale_df,
            indicatori_radiologia.get('mapping_uo', []),
            indicatori_radiologia.get('mapping_profili', []),
            match_cdc=True,
        )
        _aggiungi_area('Servizi di Radiologia', df_a, 'RADIOLOGIA')

    if indicatori_anatomia_pat:
        df_a = _filtra_per_area(
            personale_df,
            indicatori_anatomia_pat.get('mapping_uo', []),
            indicatori_anatomia_pat.get('mapping_profili', []),
        )
        _aggiungi_area('Anatomia Patologica', df_a, 'ANATOMIA PATOLOGICA')

    if indicatori_laboratorio:
        df_a = _filtra_per_area(
            personale_df,
            indicatori_laboratorio.get('mapping_uo', []),
            indicatori_laboratorio.get('mapping_profili', []),
            esclusioni=indicatori_laboratorio.get('esclusioni', []),
        )
        _aggiungi_area('Servizi di Laboratorio', df_a, 'LABORATORIO')

    if indicatori_tecnici_lab:
        # Tecnici: no filtro UO, solo profilo
        df_a = _filtra_per_area(
            personale_df,
            indicatori_tecnici_lab.get('mapping_uo', []),
            indicatori_tecnici_lab.get('mapping_profili', []),
            solo_profilo=True,
        )
        _aggiungi_area('Tecnici di Laboratorio', df_a, 'TECNICI LABORATORIO')

    if indicatori_med_legale:
        df_a = _filtra_per_area(
            personale_df,
            indicatori_med_legale.get('mapping_uo', []),
            indicatori_med_legale.get('mapping_profili', []),
        )
        _aggiungi_area('Medicina Legale', df_a, 'MEDICINA LEGALE')

    if indicatori_trasfusionale:
        df_a = _filtra_per_area(
            personale_df,
            indicatori_trasfusionale.get('mapping_uo', []),
            indicatori_trasfusionale.get('mapping_profili', []),
        )
        _aggiungi_area('Medicina Trasfusionale', df_a, 'TRASFUSIONALE')

    if indicatori_emergenza:
        df_a = _filtra_per_area(
            personale_df,
            indicatori_emergenza.get('mapping_uo', []),
            indicatori_emergenza.get('mapping_profili', []),
            match_cdc=True,
        )
        _aggiungi_area('Emergenza-Urgenza', df_a, 'EMERGENZA-URGENZA')

    if indicatori_terapia_intensiva:
        # TI: filtra su CDC che contiene "TERAPIA INTENSIVA"
        mask_ti_cdc = (
            personale_df['DESC_TIPO_CDC'].astype(str)
            .str.upper().str.contains('TERAPIA INTENSIVA', na=False)
        )
        df_a = _filtra_per_area(
            personale_df[mask_ti_cdc],
            [],   # nessun filtro UO aggiuntivo: già filtrato per CDC
            indicatori_terapia_intensiva.get('mapping_profili', []),
            solo_profilo=True,
        )
        _aggiungi_area('Terapia Intensiva', df_a, 'TERAPIA INTENSIVA')

    if indicatori_sale_operatorie:
        df_a = _filtra_per_area(
            personale_df,
            indicatori_sale_operatorie.get('mapping_uo', []),
            indicatori_sale_operatorie.get('mapping_profili', []),
            match_cdc=True,
        )
        # Escludi eventuali record TI già conteggiati
        if not df_a.empty:
            mask_no_ti = ~(
                df_a['DESC_TIPO_CDC'].astype(str)
                .str.upper().str.contains('TERAPIA INTENSIVA', na=False)
            )
            df_a = df_a[mask_no_ti]
        _aggiungi_area('Sale Operatorie', df_a, 'SALE OPERATORIE')

    # ── Aree territoriali ──────────────────────────────────────────

    for indicatori, nome_area, sheet_name in [
        (indicatori_salute_mentale, 'Salute Mentale Adulti',  'SALUTE MENTALE'),
        (indicatori_dipendenze,     'Dipendenze Patologiche', 'DIPENDENZE'),
        (indicatori_npia,           'NPIA',                   'NPIA'),
        (indicatori_carcere,        'Salute in Carcere',      'CARCERE'),
    ]:
        if not indicatori:
            continue
        uo_patterns = indicatori.get('unita_operative', [])
        profili     = indicatori.get('profili', [])
        df_a = _filtra_territoriale(personale_df, uo_patterns, profili)
        _aggiungi_area(nome_area, df_a, sheet_name)

    # ── Riepilogo ──────────────────────────────────────────────────
    if righe_riep:
        _scrivi_foglio_riepilogo(
            wb,
            f'Dettaglio Nominativo per Area AGENAS – {anno_analisi}',
            righe_riep,
            anno_analisi,
        )

    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f'dettaglio_agenas_{anno_analisi}.xlsx')
    wb.save(output_file)
    print(f"  Dettaglio AGENAS salvato in: {output_file}")
    return output_file
