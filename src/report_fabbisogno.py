"""
Elaborazione principale: raggruppamento personale, calcolo fabbisogno,
generazione dei report Excel per città con RIEPILOGO e fogli dettaglio.
"""

import math
import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from src.config import (
    FILE_OUTPUT,
    PARTI_PER_PRESIDIO, LIVELLO_PRESIDIO,
    SALE_OPERATORIE_PER_PRESIDIO,
)
from src.stili_excel import (
    THIN_BORDER, FILL_A, FILL_B, FILL_HEADER, FILL_AGENAS,
    FONT_HEADER, ALIGN_CENTER,
    scrivi_titolo, scrivi_intestazioni, scrivi_riga_dati,
    auto_larghezza_colonne,
)
from src.caricamento_dati import (
    carica_dataframe, normalizza_colonne_personale,
    normalizza_colonne_pensionamenti, pulisci_prefisso_cdc,
)
from src.caricamento_xml import carica_fabbisogno_odc_dm77
from src.posti_letto import leggi_posti_letto_csv
from src.calcolo_fabbisogno import (
    read_indicators, calculate_fabbisogno, match_profilo,
    scrivi_controprova_xlsx,
    calcola_fabbisogno_agenas_materno_infantile,
    calcola_fabbisogno_agenas_radiologia,
    calcola_fabbisogno_agenas_trasfusionale,
    calcola_fabbisogno_agenas_anatomia_patologica,
    calcola_fabbisogno_agenas_laboratorio,
    calcola_fabbisogno_agenas_tecnici_laboratorio,
    calcola_fabbisogno_agenas_medicina_legale,
    calcola_fabbisogno_agenas_emergenza_urgenza,
    calcola_fabbisogno_agenas_terapia_intensiva,
    calcola_fabbisogno_agenas_sale_operatorie,
)


# ============================================================
# HELPER INTERNI
# ============================================================

def _scrivi_tabella_sanitario(ws, start_row, merged_df_subset):
    """Scrive la tabella Rapporto Sanitario / Non Sanitario su *ws*
    a partire da *start_row*.  Restituisce la riga successiva libera."""
    df_tmp = merged_df_subset.copy()
    df_tmp['_AREA'] = df_tmp['DESC_RUOLO'].apply(
        lambda r: 'Sanitario'
        if str(r).upper().startswith('RUOLO SANIT')
           or str(r).upper().startswith('RUOLO SOCIO')
        else 'Non Sanitario'
    )
    df_tmp['_COMP'] = df_tmp['DESC_TIPO_DIPENDENTE'].apply(
        lambda t: 'Dirigenza' if str(t).startswith('Dirig') else 'Comparto'
    )

    ct = pd.crosstab(
        df_tmp['_AREA'], df_tmp['_COMP'],
        margins=True, margins_name='Totale',
    )
    for c_name in ('Dirigenza', 'Comparto', 'Totale'):
        if c_name not in ct.columns:
            ct[c_name] = 0
    ct = ct[['Dirigenza', 'Comparto', 'Totale']]

    totale_gen = ct.loc['Totale', 'Totale'] if 'Totale' in ct.index else 1
    if totale_gen == 0:
        totale_gen = 1

    row = start_row
    # Titoletto sezione
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=5)
    t_cell = ws.cell(row=row, column=1,
                     value='RAPPORTO PERSONALE SANITARIO / NON SANITARIO')
    t_cell.font = Font(bold=True, size=12)
    t_cell.alignment = ALIGN_CENTER
    row += 1

    # Intestazione
    rap_headers = ['', 'Dirigenza', 'Comparto', 'Totale', '%']
    for ci, cn in enumerate(rap_headers, 1):
        c = ws.cell(row=row, column=ci, value=cn)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # Righe dati
    for idx_area, area_label in enumerate(
            ['Sanitario', 'Non Sanitario', 'Totale']):
        fill_r = FILL_A if idx_area % 2 == 0 else FILL_B
        dir_val  = int(ct.loc[area_label, 'Dirigenza']) if area_label in ct.index else 0
        comp_val = int(ct.loc[area_label, 'Comparto'])  if area_label in ct.index else 0
        tot_val  = int(ct.loc[area_label, 'Totale'])    if area_label in ct.index else 0
        pct = f'{tot_val / totale_gen * 100:.1f}%'

        font_r = Font(bold=True, size=10) if area_label == 'Totale' else None
        vals_r = [area_label, dir_val, comp_val, tot_val, pct]
        for ci, v in enumerate(vals_r, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill_r
            c.border = THIN_BORDER
            if ci > 1:
                c.alignment = ALIGN_CENTER
            if font_r:
                c.font = font_r
        row += 1

    return row

def _scrivi_foglio_con_titolo(wb, sheet_name, df, titolo):
    """Scrive un DataFrame in un foglio Excel con titolo, intestazioni
    e righe colorate a bande per reparto (_REPARTO).
    Le celle della colonna 'Fabbisogno AGENAS' con valore non vuoto
    vengono evidenziate in verde (FILL_AGENAS)."""
    ws = wb.create_sheet(title=sheet_name[:31])

    col_reparto_name = '_REPARTO'
    has_reparto = col_reparto_name in df.columns
    if has_reparto:
        col_reparto_idx = list(df.columns).index(col_reparto_name)
        visible_cols = [c for c in df.columns if c != col_reparto_name]
    else:
        col_reparto_idx = None
        visible_cols = list(df.columns)

    n_vis = len(visible_cols)

    # Indice (1-based) della colonna AGENAS tra le visibili
    agenas_col_name = 'Fabbisogno AGENAS'
    agenas_out_idx = (visible_cols.index(agenas_col_name) + 1
                      if agenas_col_name in visible_cols else None)

    # Indice (1-based) della colonna Fabbisogno Teorico tra le visibili
    fabb_teor_col_name = 'Fabbisogno Teorico'
    fabb_teor_out_idx = (visible_cols.index(fabb_teor_col_name) + 1
                         if fabb_teor_col_name in visible_cols else None)

    # Titolo
    scrivi_titolo(ws, titolo, n_vis)

    # Intestazioni
    scrivi_intestazioni(ws, visible_cols)

    # Dati con colore alternato per reparto
    color_toggle = 0
    prev_reparto = None
    for r_idx, row_data in enumerate(df.itertuples(index=False), 3):
        current_reparto = row_data[col_reparto_idx] if has_reparto else None
        if current_reparto != prev_reparto:
            color_toggle = 1 - color_toggle
            prev_reparto = current_reparto
        fill = FILL_A if color_toggle else FILL_B

        # Determina se la riga ha un riferimento AGENAS
        agenas_val = None
        if agenas_out_idx is not None:
            # Trova il valore AGENAS nella tupla (considerando _REPARTO skip)
            vis_idx = 0
            for ci, v in enumerate(row_data):
                if has_reparto and ci == col_reparto_idx:
                    continue
                vis_idx += 1
                if vis_idx == agenas_out_idx:
                    agenas_val = v
                    break

        c_out = 1
        for c_idx, val in enumerate(row_data):
            if has_reparto and c_idx == col_reparto_idx:
                continue
            c = ws.cell(row=r_idx, column=c_out, value=val)
            # Cella AGENAS → sfondo verde se la riga ha riferimento
            # Cella Fabbisogno Teorico → sfondo verde se valore valido
            #   (escluso solo "Servizio privo di posti letto")
            verde = False
            if agenas_out_idx and c_out == agenas_out_idx and agenas_val:
                verde = True
            elif (fabb_teor_out_idx and c_out == fabb_teor_out_idx
                  and val is not None and val != ''
                  and str(val) != 'Servizio privo di posti letto'):
                verde = True
            if verde:
                c.fill = FILL_AGENAS
            else:
                c.fill = fill
            c.border = THIN_BORDER
            c_out += 1

    auto_larghezza_colonne(ws, visible_cols)
    return ws


# ============================================================
# IMPORT DEI MODULI ESTRATTI
# ============================================================
from src.tabelle_agenas import (
    _scrivi_tabella_agenas_territoriale,
    _scrivi_tabella_agenas_materno_infantile,
    _scrivi_tabella_agenas_radiologia,
    _scrivi_tabella_agenas_emergenza_urgenza,
    _scrivi_tabella_agenas_terapia_intensiva,
    _scrivi_tabella_agenas_sale_operatorie,
    _scrivi_tabella_agenas_area_ti_bo,
    _scrivi_tabella_agenas_anatomia_patologica,
    _scrivi_tabella_agenas_laboratorio,
    _scrivi_tabella_agenas_tecnici_laboratorio,
    _scrivi_tabella_agenas_medicina_legale,
    _scrivi_tabella_agenas_trasfusionale,
    _scrivi_tabella_fabbisogno_uoc_trasfusionale,
)
from src.riepilogo_agenas_aziendale import scrivi_foglio_riepilogo_agenas
from src.riepilogo_fabbisogno_teorico import scrivi_foglio_riepilogo_fabbisogno_teorico
from src.riepilogo_veterinari import scrivi_foglio_veterinari
from src.nota_metodologica import _scrivi_foglio_metodologia


# ============================================================
# ELABORAZIONE PRINCIPALE
# ============================================================

def process_data(personale_file, pensionamenti_file, posti_letto_csv,
                 indicators_file, debug_file, anno_analisi,
                 indicatori_odc_file=None,
                 indicatori_agenas=None, parti_per_presidio=None,
                 indicatori_radiologia=None, livello_presidio=None,
                 indicatori_trasfusionale=None,
                 fabb_trasf_speciale=None,
                 indicatori_anatomia_pat=None,
                 indicatori_laboratorio=None,
                 indicatori_tecnici_lab=None,
                 indicatori_med_legale=None,
                 indicatori_emergenza=None,
                 indicatori_salute_mentale=None,
                 fabb_salute_mentale=None,
                 indicatori_dipendenze=None,
                 fabb_dipendenze=None,
                 indicatori_npia=None,
                 fabb_npia=None,
                 indicatori_carcere=None,
                 fabb_carcere=None,
                 indicatori_terapia_intensiva=None,
                 indicatori_sale_operatorie=None,
                 lista_odc=None):
    """Elaborazione principale dei dati del personale."""

    _data_esecuzione = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    calcoli_log = []
    avvisi_log  = []
    odc_log     = []
    arrot_log   = []

    # Caricamento fabbisogno DM 77 per OdC
    fabbisogno_dm77 = {}
    mappa_profili_dm77 = {}
    if indicatori_odc_file and os.path.exists(indicatori_odc_file):
        fabbisogno_dm77, mappa_profili_dm77 = carica_fabbisogno_odc_dm77(
            indicatori_odc_file
        )
        print(f"Fabbisogno DM 77 OdC caricato: {fabbisogno_dm77}")
    else:
        print("ATTENZIONE: File indicatori DM 77 OdC non trovato. "
              "Le OdC non avranno fabbisogno calcolato.")

    # Caricamento dati
    print(f"Caricamento personale da: {personale_file}")
    personale_raw_df = carica_dataframe(personale_file)   # copia grezza
    personale_df = normalizza_colonne_personale(personale_raw_df.copy())

    print(f"Caricamento pensionamenti da: {pensionamenti_file}")
    pensionamenti_df = carica_dataframe(pensionamenti_file)
    pensionamenti_df = normalizza_colonne_pensionamenti(pensionamenti_df)

    # Riempimento NaN nelle colonne di raggruppamento con segnaposto
    # (evita che pandas.groupby scarti silenziosamente le righe)

    # DESC_SEDE_FISICA: etichetta differenziata in base al CDC
    def _sede_placeholder(row):
        if pd.notna(row['DESC_SEDE_FISICA']):
            return row['DESC_SEDE_FISICA']
        cdc = str(row.get('DESC_TIPO_CDC', '')).upper()
        if 'LUNGHE ASSENZE' in cdc:
            return 'LUNGHE ASSENZE - LUNGHE ASSENZE'
        if 'ATTESA DI ASSEGNAZIONE' in cdc:
            return 'IN ATTESA DI ASSEGNAZIONE - IN ATTESA DI ASSEGNAZIONE'
        return '(Non assegnata) - (Non assegnata)'

    personale_df['DESC_SEDE_FISICA'] = personale_df.apply(
        _sede_placeholder, axis=1
    )

    _FILL_NA = {
        'DESC_SC_SSD_SS':   '(Reparto non assegnato)',
        'DESC_TIPO_CDC':    '(CDC non assegnato)',
    }
    personale_df = personale_df.fillna(_FILL_NA)

    # Conteggio T.D. (tutti i contratti non-TI) prima di filtrare
    personale_df['_IS_TD'] = (
        personale_df['DESC_NATURA'].str.upper() != 'TEMPO INDETERMINATO'
    ).astype(int)
    td_counts = personale_df[personale_df['_IS_TD'] == 1].groupby(
        ['DESC_SEDE_FISICA', 'DESC_SC_SSD_SS', 'DESC_TIPO_CDC',
         'PROFILO_RAGGRUPPATO']
    ).size().reset_index(name='QUANTITA_TD')

    # Copia completa per analisi composizione (T.I. + T.D.)
    personale_all_df = personale_df.copy()
    personale_all_df.drop(columns=['_IS_TD'], inplace=True)

    personale_df = personale_df[
        personale_df['DESC_NATURA'].str.upper() == "TEMPO INDETERMINATO"
    ].copy()
    personale_df.drop(columns=['_IS_TD'], inplace=True)

    # Colonna assunti nell'anno
    col_assunti = f'DI CUI ASSUNTI NEL {anno_analisi}'
    if 'PRIMA_DATA_ASSUNZIONE' in personale_df.columns:
        personale_df[col_assunti] = personale_df['PRIMA_DATA_ASSUNZIONE'].apply(
            lambda x: 1 if pd.to_datetime(
                x, dayfirst=True, errors='coerce'
            ).year == anno_analisi else 0
        )
    else:
        print("ATTENZIONE: Colonna data assunzione non trovata.")
        personale_df[col_assunti] = 0

    # Colonne ALTRO (presenti solo nei vecchi CSV)
    if 'ALTRO' in personale_df.columns:
        personale_df['Maternità'] = personale_df['ALTRO'].apply(
            lambda x: 1 if x == 'MAT' else 0
        )
        personale_df['Aspettativa'] = personale_df['ALTRO'].apply(
            lambda x: 1 if x == 'ASP' else 0
        )
        personale_df['104'] = personale_df['ALTRO'].apply(
            lambda x: 1 if x == '104' else 0
        )
    else:
        personale_df['Maternità'] = 0
        personale_df['Aspettativa'] = 0
        personale_df['104'] = 0

    personale_df['DESC_DISCIPLINE'] = (
        personale_df['DESC_DISCIPLINE'].str.upper().fillna('')
    )

    # Merge con pensionamenti
    if 'DT_CESSAZIONE' in personale_df.columns:
        personale_df = personale_df.rename(
            columns={'DT_CESSAZIONE': 'DT_CESSAZIONE_PERS'}
        )
    merged_df = pd.merge(
        personale_df,
        pensionamenti_df[['MATR.', 'DT_CESSAZIONE']],
        on='MATR.',
        how='left',
    )

    anni_pensionamento = [anno_analisi + 1, anno_analisi + 2, anno_analisi + 3]
    for anno in anni_pensionamento:
        merged_df[f'PENSIONAMENTI_{anno}'] = merged_df['DT_CESSAZIONE'].apply(
            lambda x, a=anno: (
                1 if pd.to_datetime(x, errors='coerce').year == a else 0
            )
        )

    # Posti letto
    print(f"Caricamento posti letto da: {posti_letto_csv}")
    posti_letto = leggi_posti_letto_csv(posti_letto_csv)

    # Indicatori
    print(f"Caricamento indicatori da: {indicators_file}")
    indicators = read_indicators(indicators_file)

    # Raggruppamento
    col_pensionamenti = [f'PENSIONAMENTI_{a}' for a in anni_pensionamento]
    agg_dict = {
        'QUANTITA': ('PROFILO_RAGGRUPPATO', 'size'),
        f'ASSUNTI_{anno_analisi}': (col_assunti, 'sum'),
    }
    for col in col_pensionamenti:
        agg_dict[col] = (col, 'sum')

    grouped = merged_df.groupby(
        ['DESC_SEDE_FISICA', 'DESC_SC_SSD_SS', 'DESC_TIPO_CDC',
         'PROFILO_RAGGRUPPATO']
    ).agg(**agg_dict).reset_index()

    # Merge conteggio T.D. (outer per includere combo T.D.-only)
    grouped = grouped.merge(
        td_counts,
        on=['DESC_SEDE_FISICA', 'DESC_SC_SSD_SS', 'DESC_TIPO_CDC',
            'PROFILO_RAGGRUPPATO'],
        how='outer',
    )
    # Riempi NaN per righe T.D.-only o T.I.-only
    grouped['QUANTITA_TD'] = grouped['QUANTITA_TD'].fillna(0).astype(int)
    grouped['QUANTITA'] = grouped['QUANTITA'].fillna(0).astype(int)
    grouped[f'ASSUNTI_{anno_analisi}'] = (
        grouped[f'ASSUNTI_{anno_analisi}'].fillna(0).astype(int)
    )
    for _cp in col_pensionamenti:
        grouped[_cp] = grouped[_cp].fillna(0).astype(int)

    # CDC ripulito
    grouped['CENTRO_DI_COSTO'] = grouped['DESC_TIPO_CDC'].apply(
        pulisci_prefisso_cdc
    )

    # ------------------------------------------------------------------
    # CALCOLO FABBISOGNO
    # ------------------------------------------------------------------
    mapping_intensita = {
        'Intensiva': 'Intensiva',
        'Alta': 'Alta',
        'MedioAlta': 'MedioAlta',
        'Medio/Alta': 'MedioAlta',
        'Media': 'Media',
        'MedioBassa': 'MedioBassa',
        'Medio/Bassa': 'MedioBassa',
        'Bassa': 'Bassa',
        'DH_DS': 'DH_DS',
    }

    fabbisogno_teorico_list = []
    intensita_list = []

    for _, row in grouped.iterrows():
        ssd = row['DESC_SC_SSD_SS']
        sede_row = row['DESC_SEDE_FISICA']
        pl_key = (sede_row, ssd)

        if pl_key not in posti_letto:
            fabbisogno_teorico_list.append('Servizio privo di posti letto')
            intensita_list.append('-')
            continue

        pl = posti_letto[pl_key]
        sede_fisica = pl.get('DESC_SEDE_FISICA', '')
        ordinari = int(pl['ordinari'])
        dh = int(pl['dh'])
        utic = int(pl.get('utic', 0))
        intensita = pl['intensita']

        # ---- OdC: fabbisogno fisso DM 77 ----
        if str(sede_fisica).startswith('OdC '):
            intensita_list.append('DM 77')
            if fabbisogno_dm77:
                profilo_upper = row['PROFILO_RAGGRUPPATO'].strip().upper()
                cat_dm77 = mappa_profili_dm77.get(profilo_upper, profilo_upper)
                fabbisogno_odc = fabbisogno_dm77.get(cat_dm77, 0)
                fabbisogno_teorico_list.append(
                    fabbisogno_odc if fabbisogno_odc > 0 else 'non previsto'
                )
                odc_log.append({
                    'sede':    sede_fisica,
                    'reparto': ssd,
                    'profilo': row['PROFILO_RAGGRUPPATO'],
                    'fabb':    fabbisogno_odc,
                })
            else:
                fabbisogno_teorico_list.append('Servizio privo di posti letto')
            continue

        # ---- Sede NON ospedaliera: l'intensità non è applicabile ----
        if 'P.O.' not in sede_row:
            fabbisogno_teorico_list.append('Servizio privo di posti letto')
            intensita_list.append('-')
            continue

        # Intensità non configurata
        if not intensita:
            fabbisogno_teorico_list.append('Servizio privo di posti letto')
            intensita_list.append('-')
            continue

        # Posti letto tutti a 0
        if (ordinari == 0 and dh == 0 and utic == 0):
            fabbisogno_teorico_list.append('Servizio privo di posti letto')
            intensita_list.append(intensita)
            continue

        intensita_list.append(intensita)
        intensita_xml = mapping_intensita.get(intensita, intensita)

        profilo_key = match_profilo(
            row['PROFILO_RAGGRUPPATO'], indicators, avvisi_log
        )
        if profilo_key is not None:
            fabb = calculate_fabbisogno(
                ordinari=ordinari, dh=dh, utic=utic,
                intensity=intensita_xml,
                indicators=indicators[profilo_key],
                reparto=ssd, profilo=profilo_key,
                sede=sede_row, calcoli_log=calcoli_log,
            )
            fabbisogno_teorico_list.append(fabb)
        else:
            fabbisogno_teorico_list.append(
                'Figura professionale non censita'
            )

    grouped['FABBISOGNO_TEORICO'] = fabbisogno_teorico_list
    grouped['INTENSITA'] = intensita_list

    # ------------------------------------------------------------------
    # FABBISOGNO AGENAS
    # ------------------------------------------------------------------
    fabb_agenas_per_presidio = {}
    if indicatori_agenas and parti_per_presidio:
        fabb_agenas_per_presidio = calcola_fabbisogno_agenas_materno_infantile(
            indicatori_agenas, parti_per_presidio,
        )
        print(f"Fabbisogno AGENAS materno-infantile calcolato per "
              f"{len(fabb_agenas_per_presidio)} presidi")

    # FABBISOGNO AGENAS – AREA RADIOLOGIA
    fabb_radio_per_presidio = {}
    if indicatori_radiologia and livello_presidio:
        fabb_radio_per_presidio = calcola_fabbisogno_agenas_radiologia(
            indicatori_radiologia, livello_presidio,
        )
        print(f"Fabbisogno AGENAS radiologia calcolato per "
              f"{len(fabb_radio_per_presidio)} presidi")

    # FABBISOGNO AGENAS – AREA TRASFUSIONALE
    fabb_trasf_per_presidio = {}
    if indicatori_trasfusionale and livello_presidio:
        fabb_trasf_per_presidio = calcola_fabbisogno_agenas_trasfusionale(
            indicatori_trasfusionale, livello_presidio,
        )
        print(f"Fabbisogno AGENAS trasfusionale calcolato per "
              f"{len(fabb_trasf_per_presidio)} presidi")

    # FABBISOGNO AGENAS – AREA ANATOMIA PATOLOGICA
    fabb_anapato_per_presidio = {}
    if indicatori_anatomia_pat and livello_presidio:
        fabb_anapato_per_presidio = calcola_fabbisogno_agenas_anatomia_patologica(
            indicatori_anatomia_pat, livello_presidio,
        )
        print(f"Fabbisogno AGENAS anatomia patologica calcolato per "
              f"{len(fabb_anapato_per_presidio)} presidi")

    # FABBISOGNO AGENAS – AREA SERVIZI DI LABORATORIO
    fabb_lab_per_presidio = {}
    if indicatori_laboratorio and livello_presidio:
        fabb_lab_per_presidio = calcola_fabbisogno_agenas_laboratorio(
            indicatori_laboratorio, livello_presidio,
        )
        print(f"Fabbisogno AGENAS laboratorio calcolato per "
              f"{len(fabb_lab_per_presidio)} presidi")

    # FABBISOGNO AGENAS – TECNICI DI LABORATORIO
    fabb_teclab_per_presidio = {}
    if indicatori_tecnici_lab and livello_presidio:
        fabb_teclab_per_presidio = calcola_fabbisogno_agenas_tecnici_laboratorio(
            indicatori_tecnici_lab, livello_presidio,
        )
        print(f"Fabbisogno AGENAS tecnici laboratorio calcolato per "
              f"{len(fabb_teclab_per_presidio)} presidi")

    # FABBISOGNO AGENAS – MEDICINA LEGALE
    fabb_medleg_per_presidio = {}
    if indicatori_med_legale and livello_presidio:
        fabb_medleg_per_presidio = calcola_fabbisogno_agenas_medicina_legale(
            indicatori_med_legale, livello_presidio,
        )
        print(f"Fabbisogno AGENAS medicina legale calcolato per "
              f"{len(fabb_medleg_per_presidio)} presidi")

    # FABBISOGNO AGENAS – EMERGENZA-URGENZA
    fabb_emergenza_per_presidio = {}
    if indicatori_emergenza and livello_presidio:
        fabb_emergenza_per_presidio = calcola_fabbisogno_agenas_emergenza_urgenza(
            indicatori_emergenza, livello_presidio,
        )
        print(f"Fabbisogno AGENAS emergenza-urgenza calcolato per "
              f"{len(fabb_emergenza_per_presidio)} presidi")

    # FABBISOGNO AGENAS – TERAPIA INTENSIVA (§ 8.1.1)
    fabb_ti_per_presidio = {}
    if indicatori_terapia_intensiva:
        fabb_ti_per_presidio = calcola_fabbisogno_agenas_terapia_intensiva(
            indicatori_terapia_intensiva, posti_letto, indicators,
        )
        print(f"Fabbisogno AGENAS terapia intensiva calcolato per "
              f"{len(fabb_ti_per_presidio)} presidi")

    # FABBISOGNO AGENAS – SALE OPERATORIE (§ 8.1.2)
    fabb_so_per_presidio = {}
    if indicatori_sale_operatorie:
        fabb_so_per_presidio = calcola_fabbisogno_agenas_sale_operatorie(
            indicatori_sale_operatorie, indicators,
            SALE_OPERATORIE_PER_PRESIDIO,
        )
        print(f"Fabbisogno AGENAS sale operatorie calcolato per "
              f"{len(fabb_so_per_presidio)} presidi")

    # Mapping UO → profilo AGENAS materno-infantile
    mapping_uo_agenas = (
        indicatori_agenas.get('mapping_uo', []) if indicatori_agenas else []
    )
    mapping_profili_agenas = (
        indicatori_agenas.get('mapping_profili', []) if indicatori_agenas else []
    )

    # Mapping UO → profilo AGENAS radiologia
    mapping_uo_radio = (
        indicatori_radiologia.get('mapping_uo', []) if indicatori_radiologia else []
    )
    mapping_profili_radio = (
        indicatori_radiologia.get('mapping_profili', []) if indicatori_radiologia else []
    )

    # Mapping UO → profilo AGENAS trasfusionale
    mapping_uo_trasf = (
        indicatori_trasfusionale.get('mapping_uo', []) if indicatori_trasfusionale else []
    )
    mapping_profili_trasf = (
        indicatori_trasfusionale.get('mapping_profili', []) if indicatori_trasfusionale else []
    )

    # Mapping UO → profilo AGENAS anatomia patologica
    mapping_uo_anapato = (
        indicatori_anatomia_pat.get('mapping_uo', []) if indicatori_anatomia_pat else []
    )
    mapping_profili_anapato = (
        indicatori_anatomia_pat.get('mapping_profili', []) if indicatori_anatomia_pat else []
    )

    # Mapping UO → profilo AGENAS laboratorio
    mapping_uo_lab = (
        indicatori_laboratorio.get('mapping_uo', []) if indicatori_laboratorio else []
    )
    mapping_profili_lab = (
        indicatori_laboratorio.get('mapping_profili', []) if indicatori_laboratorio else []
    )
    esclusioni_lab = (
        indicatori_laboratorio.get('esclusioni', []) if indicatori_laboratorio else []
    )

    # Mapping profilo AGENAS tecnici laboratorio (nessun filtro UO)
    mapping_profili_teclab = (
        indicatori_tecnici_lab.get('mapping_profili', []) if indicatori_tecnici_lab else []
    )

    # Mapping UO → profilo AGENAS medicina legale
    mapping_uo_medleg = (
        indicatori_med_legale.get('mapping_uo', []) if indicatori_med_legale else []
    )
    mapping_profili_medleg = (
        indicatori_med_legale.get('mapping_profili', []) if indicatori_med_legale else []
    )

    # Mapping UO → profilo AGENAS emergenza-urgenza
    mapping_uo_emergenza = (
        indicatori_emergenza.get('mapping_uo', []) if indicatori_emergenza else []
    )
    mapping_profili_emergenza = (
        indicatori_emergenza.get('mapping_profili', []) if indicatori_emergenza else []
    )

    # Mapping UO → profilo AGENAS terapia intensiva
    mapping_uo_ti = (
        indicatori_terapia_intensiva.get('mapping_uo', [])
        if indicatori_terapia_intensiva else []
    )
    mapping_profili_ti = (
        indicatori_terapia_intensiva.get('mapping_profili', [])
        if indicatori_terapia_intensiva else []
    )

    # Mapping UO → profilo AGENAS sale operatorie
    mapping_uo_so = (
        indicatori_sale_operatorie.get('mapping_uo', [])
        if indicatori_sale_operatorie else []
    )
    mapping_profili_so = (
        indicatori_sale_operatorie.get('mapping_profili', [])
        if indicatori_sale_operatorie else []
    )

    import re as _re_agenas

    # Costruisci lista pattern CDC per strutture OdC (DM 77)
    odc_cdc_patterns = []
    if lista_odc:
        for odc in lista_odc:
            for struttura in odc.get('strutture', []):
                pat = struttura.get('pattern_cdc', '')
                if pat:
                    odc_cdc_patterns.append(pat)

    fabbisogno_agenas_list = []
    for _, row in grouped.iterrows():
        sede_row = row['DESC_SEDE_FISICA']
        ssd = row['DESC_SC_SSD_SS']
        profilo = row['PROFILO_RAGGRUPPATO'].strip().upper()

        # === OSPEDALI DI COMUNITÀ (DM 77) ===
        if odc_cdc_patterns:
            desc_cdc = str(row.get('DESC_TIPO_CDC', ''))
            is_odc = False
            for pat_odc in odc_cdc_patterns:
                if _re_agenas.search(pat_odc, desc_cdc, _re_agenas.IGNORECASE):
                    is_odc = True
                    break
            if is_odc:
                fabbisogno_agenas_list.append(
                    "Vedi Report OdC - DM 77")
                continue

        # === AREA MATERNO-INFANTILE ===
        presidio_match_mi = None
        for presidio_nome in fabb_agenas_per_presidio:
            if presidio_nome in sede_row:
                presidio_match_mi = presidio_nome
                break

        if presidio_match_mi and fabb_agenas_per_presidio.get(presidio_match_mi):
            uo_materno = False
            for m_uo in mapping_uo_agenas:
                if _re_agenas.search(m_uo['pattern'], ssd.upper()):
                    uo_materno = True
                    break
            if uo_materno:
                fabbisogno_agenas_list.append(
                    "Vedi Tab. 11 nel Riepilogo")
                continue

        # === AREA RADIOLOGIA ===
        presidio_match_r = None
        for presidio_nome in fabb_radio_per_presidio:
            if presidio_nome in sede_row:
                presidio_match_r = presidio_nome
                break

        # Fallback per sedi territoriali: cerca il P.O. della stessa area
        if not presidio_match_r and 'P.O.' not in sede_row:
            area_sede = sede_row.split(' - ')[0].strip()
            for presidio_nome in fabb_radio_per_presidio:
                if presidio_nome.startswith(area_sede + ' - '):
                    presidio_match_r = presidio_nome
                    break

        if presidio_match_r and fabb_radio_per_presidio.get(presidio_match_r):
            # Verifica se la UO è dell'area radiologica
            uo_radio = False
            for m_uo in mapping_uo_radio:
                if _re_agenas.search(m_uo['pattern'], ssd.upper()):
                    uo_radio = True
                    break

            # Sedi territoriali: verifica anche CDC (Centro Radiologico)
            if not uo_radio and 'P.O.' not in sede_row:
                cdc_val = str(row.get('DESC_TIPO_CDC', '')).upper()
                if _re_agenas.search(
                        r'CENTRO RADIOLOGICO|RADIOLOGIA', cdc_val):
                    uo_radio = True

            # Sedi territoriali: TS RADIOLOGIA conta sempre
            if not uo_radio and 'P.O.' not in sede_row:
                for m_prof in mapping_profili_radio:
                    if (m_prof['profilo_agenas'] == 'TECNICO_RADIOLOGIA'
                            and _re_agenas.search(m_prof['pattern'], profilo)):
                        uo_radio = True
                        break

            if uo_radio:
                fabbisogno_agenas_list.append(
                    "Vedi Tab. 13 nel Riepilogo")
                continue

        # === AREA TRASFUSIONALE ===
        presidio_match_t = None
        for presidio_nome in fabb_trasf_per_presidio:
            if presidio_nome in sede_row:
                presidio_match_t = presidio_nome
                break

        if presidio_match_t and fabb_trasf_per_presidio.get(presidio_match_t):
            uo_trasf = False
            for m_uo in mapping_uo_trasf:
                if _re_agenas.search(m_uo['pattern'], ssd.upper()):
                    uo_trasf = True
                    break
            if uo_trasf:
                fabbisogno_agenas_list.append(
                    "Vedi Tab. 15 nel Riepilogo")
                continue

        # === AREA ANATOMIA PATOLOGICA ===
        presidio_match_ap = None
        for presidio_nome in fabb_anapato_per_presidio:
            if presidio_nome in sede_row:
                presidio_match_ap = presidio_nome
                break

        if presidio_match_ap and fabb_anapato_per_presidio.get(presidio_match_ap):
            uo_anapato = False
            for m_uo in mapping_uo_anapato:
                if _re_agenas.search(m_uo['pattern'], ssd.upper()):
                    uo_anapato = True
                    break
            if uo_anapato:
                fabbisogno_agenas_list.append(
                    "Vedi Tab. 16 nel Riepilogo")
                continue

        # === AREA SERVIZI DI LABORATORIO ===
        presidio_match_lb = None
        for presidio_nome in fabb_lab_per_presidio:
            if presidio_nome in sede_row:
                presidio_match_lb = presidio_nome
                break

        if presidio_match_lb and fabb_lab_per_presidio.get(presidio_match_lb):
            uo_lab = False
            for m_uo in mapping_uo_lab:
                if _re_agenas.search(m_uo['pattern'], ssd.upper()):
                    uo_lab = True
                    break

            # Escludi UO che appartengono ad altre aree AGENAS
            if uo_lab and esclusioni_lab:
                for excl_pattern in esclusioni_lab:
                    if _re_agenas.search(excl_pattern, ssd.upper()):
                        uo_lab = False
                        break

            if uo_lab:
                fabbisogno_agenas_list.append(
                    "Vedi Tab. 14 nel Riepilogo")
                continue

        # === TECNICI DI LABORATORIO (conteggio per ruolo, no filtro UO) ===
        presidio_match_tl = None
        for presidio_nome in fabb_teclab_per_presidio:
            if presidio_nome in sede_row:
                presidio_match_tl = presidio_nome
                break

        if presidio_match_tl and fabb_teclab_per_presidio.get(presidio_match_tl):
            fabb_presidio_tl = fabb_teclab_per_presidio[presidio_match_tl]
            profilo_agenas_tl = None
            for m_prof in mapping_profili_teclab:
                if _re_agenas.search(m_prof['pattern'], profilo):
                    profilo_agenas_tl = m_prof['profilo_agenas']
                    break
            if profilo_agenas_tl and profilo_agenas_tl in fabb_presidio_tl:
                fabbisogno_agenas_list.append(
                    "Vedi Tab. 17 nel Riepilogo")
                continue

        # === MEDICINA LEGALE ===
        presidio_match_ml = None
        for presidio_nome in fabb_medleg_per_presidio:
            if presidio_nome in sede_row:
                presidio_match_ml = presidio_nome
                break

        if presidio_match_ml and fabb_medleg_per_presidio.get(presidio_match_ml):
            uo_medleg = False
            for m_uo in mapping_uo_medleg:
                if _re_agenas.search(m_uo['pattern'], ssd.upper()):
                    uo_medleg = True
                    break
            if uo_medleg:
                fabbisogno_agenas_list.append(
                    "Vedi Tab. 18 nel Riepilogo")
                continue

        # === EMERGENZA-URGENZA ===
        presidio_match_eu = None
        for presidio_nome in fabb_emergenza_per_presidio:
            if presidio_nome in sede_row:
                presidio_match_eu = presidio_nome
                break

        if presidio_match_eu and fabb_emergenza_per_presidio.get(presidio_match_eu):
            uo_emergenza = False
            for m_uo in mapping_uo_emergenza:
                if _re_agenas.search(m_uo['pattern'], ssd.upper()):
                    uo_emergenza = True
                    break
            if uo_emergenza:
                fabbisogno_agenas_list.append(
                    "Vedi Tab. 20 nel Riepilogo")
                continue

        # === TERAPIA INTENSIVA (§ 8.1.1) ===
        presidio_match_ti = None
        for presidio_nome in fabb_ti_per_presidio:
            if presidio_nome in sede_row:
                presidio_match_ti = presidio_nome
                break

        if presidio_match_ti and fabb_ti_per_presidio.get(presidio_match_ti):
            uo_ti = False
            for m_uo in mapping_uo_ti:
                if _re_agenas.search(m_uo['pattern'], ssd.upper()):
                    uo_ti = True
                    break
            if uo_ti:
                fabbisogno_agenas_list.append(
                    "Vedi § 8.1.1 nel Riepilogo")
                continue

        # === SALE OPERATORIE (§ 8.1.2) ===
        presidio_match_so = None
        for presidio_nome in fabb_so_per_presidio:
            if presidio_nome in sede_row:
                presidio_match_so = presidio_nome
                break

        if presidio_match_so and fabb_so_per_presidio.get(presidio_match_so):
            uo_so = False
            for m_uo in mapping_uo_so:
                if _re_agenas.search(m_uo['pattern'], ssd.upper()):
                    uo_so = True
                    break
            if uo_so:
                fabbisogno_agenas_list.append(
                    "Vedi § 8.1.2 nel Riepilogo")
                continue

        # --- Helper: match territoriale su SSD, CDC e sede ---
        cdc_row = str(row.get('CENTRO_DI_COSTO', '')).upper()

        def _match_territoriale(patterns):
            for pat in patterns:
                if (_re_agenas.search(pat, ssd.upper())
                        or _re_agenas.search(pat, cdc_row)
                        or _re_agenas.search(pat, sede_row.upper())):
                    return True
            return False

        # === AREA SALUTE MENTALE ADULTI (§ 6a) ===
        if indicatori_salute_mentale:
            uo_sm_patterns = indicatori_salute_mentale.get(
                'unita_operative', [])
            if _match_territoriale(uo_sm_patterns):
                fabbisogno_agenas_list.append(
                    "Vedi § 6a nel Riepilogo")
                continue

        # === AREA DIPENDENZE PATOLOGICHE - SerD (§ 6b) ===
        if indicatori_dipendenze:
            uo_dip_patterns = indicatori_dipendenze.get(
                'unita_operative', [])
            if _match_territoriale(uo_dip_patterns):
                fabbisogno_agenas_list.append(
                    "Vedi § 6b nel Riepilogo")
                continue

        # === AREA NPIA (§ 6c) ===
        if indicatori_npia:
            uo_npia_patterns = indicatori_npia.get(
                'unita_operative', [])
            if _match_territoriale(uo_npia_patterns):
                fabbisogno_agenas_list.append(
                    "Vedi § 6c nel Riepilogo")
                continue

        # === AREA SALUTE IN CARCERE (§ 6d) ===
        if indicatori_carcere:
            uo_carc_patterns = indicatori_carcere.get(
                'unita_operative', [])
            if _match_territoriale(uo_carc_patterns):
                fabbisogno_agenas_list.append(
                    "Vedi § 6d nel Riepilogo")
                continue

        fabbisogno_agenas_list.append('')

    grouped['FABBISOGNO_AGENAS'] = fabbisogno_agenas_list

    # ------------------------------------------------------------------
    # ARROTONDAMENTO
    # ------------------------------------------------------------------
    def arrotonda_fabbisogno(row):
        val_raw = row['FABBISOGNO_TEORICO']
        try:
            val = float(val_raw)
        except (ValueError, TypeError):
            return val_raw
        quantita = row['QUANTITA']
        sede = row['DESC_SEDE_FISICA']
        cdc = row['CENTRO_DI_COSTO']
        profilo = row['PROFILO_RAGGRUPPATO']
        if val < 1:
            risultato = (
                1 if quantita == 0
                else (1 if val >= 0.5 else 0)
            )
            motivo = (
                "CASO 1 (minimo 1 - profilo assente)" if quantita == 0
                else f"CASO 1 (arrotondamento < 1, quantità={quantita})"
            )
        else:
            risultato = (
                math.floor(val)
                if (val - math.floor(val)) < 0.5
                else math.ceil(val)
            )
            motivo = "CASO 2 (arrotondamento standard)"
        if risultato != val:
            arrot_log.append({
                'sede':    sede,
                'reparto': row.get('DESC_SC_SSD_SS', cdc),
                'profilo': profilo,
                'f_raw':   val,
                'f_fin':   risultato,
                'motivo':  motivo,
            })
        return risultato

    grouped['FABBISOGNO_TEORICO'] = grouped.apply(arrotonda_fabbisogno, axis=1)

    # ------------------------------------------------------------------
    # CONTROPROVA
    # ------------------------------------------------------------------
    scrivi_controprova_xlsx(
        debug_file, anno_analisi, _data_esecuzione,
        calcoli_log, avvisi_log, odc_log, arrot_log,
    )

    # ------------------------------------------------------------------
    # ORDINAMENTO E COLONNE
    # ------------------------------------------------------------------
    grouped['QUANTITA_TOT'] = grouped['QUANTITA'] + grouped['QUANTITA_TD']
    columns_order = [
        'DESC_SEDE_FISICA', 'DESC_SC_SSD_SS', 'CENTRO_DI_COSTO', 'INTENSITA',
        'PROFILO_RAGGRUPPATO', 'QUANTITA', 'QUANTITA_TD', 'QUANTITA_TOT',
        f'ASSUNTI_{anno_analisi}',
    ] + col_pensionamenti + ['FABBISOGNO_TEORICO', 'FABBISOGNO_AGENAS']
    grouped = grouped[columns_order]

    _NO_BEDS = 'Servizio privo di posti letto'

    # Segnaposto per CDC / reparto non assegnati: vanno sempre in fondo
    _PLACEHOLDERS_CDC = {'(CDC non assegnato)'}
    _PLACEHOLDERS_REP = {'(Reparto non assegnato)'}

    def _chiave_ordinamento(row):
        """Restituisce una tupla (sort_cdc, sort_pl) per l'ordinamento.
        sort_cdc = 1 se il CDC o il reparto è un segnaposto → in fondo.
        sort_pl  = 1 se il servizio è privo di posti letto → in fondo
                   (tra quelli con CDC reale).
        """
        cdc = str(row['CENTRO_DI_COSTO']).strip()
        rep = str(row['DESC_SC_SSD_SS']).strip()

        # Personale non assegnato a un vero CDC/reparto → sempre in fondo
        sort_cdc = 1 if (cdc in _PLACEHOLDERS_CDC
                         or rep in _PLACEHOLDERS_REP) else 0

        # Servizio privo di posti letto → dopo quelli con PL reali
        sort_pl = 0
        if row['FABBISOGNO_TEORICO'] == _NO_BEDS:
            sort_pl = 1
        elif isinstance(row['INTENSITA'], str) and row['INTENSITA'].strip() == '-':
            sort_pl = 1

        return sort_cdc, sort_pl

    grouped[['_SORT_CDC', '_SORT_PL']] = pd.DataFrame(
        grouped.apply(_chiave_ordinamento, axis=1).tolist(),
        index=grouped.index,
    )
    grouped = grouped.sort_values(
        ['DESC_SEDE_FISICA', '_SORT_CDC', '_SORT_PL', 'DESC_SC_SSD_SS',
         'CENTRO_DI_COSTO', 'PROFILO_RAGGRUPPATO'],
        ascending=True,
    ).drop(columns=['_SORT_CDC', '_SORT_PL']).reset_index(drop=True)

    grouped = grouped.rename(columns={'DESC_SC_SSD_SS': '_REPARTO'})

    # ------------------------------------------------------------------
    # RINOMINA COLONNE LEGGIBILI
    # ------------------------------------------------------------------
    rename_cols = {
        'DESC_SEDE_FISICA': 'Sede',
        'CENTRO_DI_COSTO': 'Centro di Costo',
        'INTENSITA': 'Intensità',
        'PROFILO_RAGGRUPPATO': 'Profilo Professionale',
        'QUANTITA': 'Quantità T.I.',
        'QUANTITA_TD': 'Quantità T.D.',
        'QUANTITA_TOT': 'Totale',
        f'ASSUNTI_{anno_analisi}': f'Assunti {anno_analisi}',
        'FABBISOGNO_TEORICO': 'Fabbisogno Teorico',
        'FABBISOGNO_AGENAS': 'Fabbisogno AGENAS',
    }
    for anno in anni_pensionamento:
        rename_cols[f'PENSIONAMENTI_{anno}'] = f'Pensionamenti {anno}'
    grouped = grouped.rename(columns=rename_cols)

    # ------------------------------------------------------------------
    # SPLIT PER CITTÀ E GENERAZIONE EXCEL
    # ------------------------------------------------------------------
    sep = ' - '
    ha_pattern = grouped['Sede'].str.contains(
        sep, regex=False, na=False
    ).all()

    if ha_pattern:
        grouped['_CITTA'] = (
            grouped['Sede'].str.split(sep, n=1).str[0].str.strip()
        )
        grouped['_LUOGO'] = (
            grouped['Sede'].str.split(sep, n=1).str[1].str.strip()
        )

        # Aree speciali: non generano file separati per città
        _AREE_SPECIALI = {'(Non assegnata)', 'LUNGHE ASSENZE',
                          'IN ATTESA DI ASSEGNAZIONE'}

        file_outputs = []

        for citta, df_citta in grouped.groupby('_CITTA', sort=True):
            # Le aree speciali non generano file separati
            # (i dati sono nel riepilogo aziendale + personale_non_assegnato)
            if citta in _AREE_SPECIALI:
                continue

            nome_file = f"elaborati/analisi_personale_{citta}_{anno_analisi}.xlsx"
            file_outputs.append(nome_file)
            titolo_area = f"PERSONALE AZIENDALE AREA - {citta}"

            wb = Workbook()
            wb.remove(wb.active)

            # ====== FOGLIO NOTA METODOLOGICA (primo foglio) ======
            sedi_citta = set(df_citta['Sede'].unique())
            pl_citta = {
                k: v for k, v in posti_letto.items()
                if k[0] in sedi_citta
                and (int(v['ordinari']) + int(v['dh'])
                     + int(v.get('utic', 0))
                     + int(v.get('ds', 0))) > 0
            }
            _scrivi_foglio_metodologia(
                wb, indicators, pl_citta, anno_analisi,
            )

            # ====== FOGLIO RIEPILOGO ======
            col_pens = [f'Pensionamenti {a}' for a in anni_pensionamento]
            col_quant_ti = 'Quantità T.I.'
            col_quant_td = 'Quantità T.D.'
            col_quant_tot = 'Totale'

            ws_riep = wb.create_sheet(title='RIEPILOGO')
            col_riep = (
                ['Sede', 'Profilo Professionale',
                 col_quant_ti, col_quant_td, col_quant_tot]
                + col_pens + ['Proiezione']
            )
            n_col_riep = len(col_riep)

            scrivi_titolo(ws_riep, f"RIEPILOGO PERSONALE - {citta}",
                          n_col_riep)
            scrivi_intestazioni(ws_riep, col_riep)

            riep_row = 3
            color_toggle_riep = 0
            prev_luogo_riep = None

            for luogo, df_luogo_riep in df_citta.groupby('_LUOGO', sort=True):
                agg_riep = df_luogo_riep.groupby(
                    'Profilo Professionale'
                ).agg(
                    **{
                        col_quant_ti: (col_quant_ti, 'sum'),
                        col_quant_td: (col_quant_td, 'sum'),
                        **{cp: (cp, 'sum') for cp in col_pens},
                    }
                ).reset_index().sort_values('Profilo Professionale')

                agg_riep[col_quant_tot] = (
                    agg_riep[col_quant_ti] + agg_riep[col_quant_td]
                )
                agg_riep['Proiezione'] = (
                    agg_riep[col_quant_tot] - agg_riep[col_pens].sum(axis=1)
                )

                if luogo != prev_luogo_riep:
                    color_toggle_riep = 1 - color_toggle_riep
                    prev_luogo_riep = luogo
                fill_riep = FILL_A if color_toggle_riep else FILL_B

                for _, r_riep in agg_riep.iterrows():
                    valori = (
                        [f'{citta} - {luogo}',
                         r_riep['Profilo Professionale'],
                         int(r_riep[col_quant_ti]),
                         int(r_riep[col_quant_td]),
                         int(r_riep[col_quant_tot])]
                        + [int(r_riep[cp]) for cp in col_pens]
                        + [int(r_riep['Proiezione'])]
                    )
                    scrivi_riga_dati(ws_riep, riep_row, valori, fill_riep)
                    riep_row += 1

            # --- Tabella rapporto Sanitario / Non Sanitario ---
            sedi_citta_set = set(df_citta['Sede'].unique())
            df_raw_citta = personale_all_df[
                personale_all_df['DESC_SEDE_FISICA'].isin(sedi_citta_set)
            ]
            riep_row = _scrivi_tabella_sanitario(
                ws_riep, riep_row + 2, df_raw_citta
            )

            # ====== FOGLI DETTAGLIO ======
            for luogo, df_luogo in df_citta.groupby('_LUOGO', sort=True):
                df_luogo_exp = df_luogo.drop(columns=['_CITTA', '_LUOGO'])
                ws_det = _scrivi_foglio_con_titolo(
                    wb, luogo, df_luogo_exp, titolo_area
                )

            # ====== TABELLE AGENAS NEL RIEPILOGO ======
            # Identifica il P.O. di questa città per ottenere livello
            presidio_citta = None
            for pn in (livello_presidio or {}):
                if citta.upper() in pn.upper():
                    presidio_citta = pn
                    break

            if presidio_citta:
                livello_label = (
                    livello_presidio.get(presidio_citta, '')
                    if livello_presidio else ''
                )

                # --- Materno-infantile ---
                if fabb_agenas_per_presidio.get(presidio_citta):
                    riep_row = _scrivi_tabella_agenas_materno_infantile(
                        ws_riep, riep_row + 1, df_citta,
                        fabb_agenas_per_presidio[presidio_citta],
                        mapping_uo_agenas, mapping_profili_agenas,
                        presidio_citta,
                    )

                # --- Radiologia ---
                if fabb_radio_per_presidio.get(presidio_citta):
                    riep_row = _scrivi_tabella_agenas_radiologia(
                        ws_riep, riep_row + 1, df_citta,
                        fabb_radio_per_presidio[presidio_citta],
                        mapping_uo_radio, mapping_profili_radio,
                        presidio_citta, livello_label,
                        df_area=df_citta,
                    )

                # --- Anatomia patologica ---
                if fabb_anapato_per_presidio.get(presidio_citta):
                    riep_row = _scrivi_tabella_agenas_anatomia_patologica(
                        ws_riep, riep_row + 1, df_citta,
                        fabb_anapato_per_presidio[presidio_citta],
                        mapping_uo_anapato, mapping_profili_anapato,
                        presidio_citta, livello_label,
                    )

                # --- Servizi di laboratorio ---
                if fabb_lab_per_presidio.get(presidio_citta):
                    riep_row = _scrivi_tabella_agenas_laboratorio(
                        ws_riep, riep_row + 1, df_citta,
                        fabb_lab_per_presidio[presidio_citta],
                        mapping_uo_lab, mapping_profili_lab,
                        esclusioni_lab,
                        presidio_citta, livello_label,
                    )

                # --- Tecnici di laboratorio ---
                if fabb_teclab_per_presidio.get(presidio_citta):
                    riep_row = _scrivi_tabella_agenas_tecnici_laboratorio(
                        ws_riep, riep_row + 1, df_citta,
                        fabb_teclab_per_presidio[presidio_citta],
                        mapping_profili_teclab,
                        presidio_citta, livello_label,
                    )

                # --- Medicina legale ---
                if fabb_medleg_per_presidio.get(presidio_citta):
                    riep_row = _scrivi_tabella_agenas_medicina_legale(
                        ws_riep, riep_row + 1, df_citta,
                        fabb_medleg_per_presidio[presidio_citta],
                        mapping_uo_medleg, mapping_profili_medleg,
                        presidio_citta, livello_label,
                    )

                # --- Trasfusionale ---
                if fabb_trasf_per_presidio.get(presidio_citta):
                    riep_row = _scrivi_tabella_agenas_trasfusionale(
                        ws_riep, riep_row + 1, df_citta,
                        fabb_trasf_per_presidio[presidio_citta],
                        mapping_uo_trasf, mapping_profili_trasf,
                        presidio_citta, livello_label,
                        df_completo=grouped,
                    )

                # --- Fabbisogno UOC Trasfusionale (speciale Primaria) ---
                if (fabb_trasf_speciale
                        and fabb_trasf_speciale.get('sedi')):
                    sede_princ = fabb_trasf_speciale['sedi'][0]['nome']
                    if (citta.upper() in sede_princ.upper()
                            or sede_princ.upper() in presidio_citta.upper()):
                        riep_row = _scrivi_tabella_fabbisogno_uoc_trasfusionale(
                            ws_riep, riep_row + 1, grouped,
                            fabb_trasf_speciale, presidio_citta,
                        )

                # --- Emergenza-Urgenza ---
                if fabb_emergenza_per_presidio.get(presidio_citta):
                    riep_row = _scrivi_tabella_agenas_emergenza_urgenza(
                        ws_riep, riep_row + 1, df_citta,
                        fabb_emergenza_per_presidio[presidio_citta],
                        mapping_uo_emergenza, mapping_profili_emergenza,
                        presidio_citta, livello_label,
                    )

                # --- Terapia Intensiva (§ 8.1.1) ---
                if fabb_ti_per_presidio.get(presidio_citta):
                    riep_row = _scrivi_tabella_agenas_terapia_intensiva(
                        ws_riep, riep_row + 1, df_citta,
                        fabb_ti_per_presidio[presidio_citta],
                        mapping_uo_ti, mapping_profili_ti,
                        presidio_citta,
                    )

                # --- Sale Operatorie (§ 8.1.2) ---
                if fabb_so_per_presidio.get(presidio_citta):
                    riep_row = _scrivi_tabella_agenas_sale_operatorie(
                        ws_riep, riep_row + 1, df_citta,
                        fabb_so_per_presidio[presidio_citta],
                        mapping_uo_so, mapping_profili_so,
                        presidio_citta,
                    )

                # --- Area TI + Blocco Operatorio (combinata) ---
                if (fabb_ti_per_presidio.get(presidio_citta)
                        or fabb_so_per_presidio.get(presidio_citta)):
                    riep_row = _scrivi_tabella_agenas_area_ti_bo(
                        ws_riep, riep_row + 1, df_citta,
                        fabb_ti_per_presidio.get(presidio_citta, {}),
                        fabb_so_per_presidio.get(presidio_citta, {}),
                        mapping_uo_ti, presidio_citta,
                    )

            # ====== TABELLE AGENAS TERRITORIALI (basate su popolazione) ======
            citta_upper = citta.upper()

            # --- Salute Mentale Adulti ---
            if (fabb_salute_mentale
                    and fabb_salute_mentale.get(citta_upper)):
                riep_row = _scrivi_tabella_agenas_territoriale(
                    ws_riep, riep_row + 1, df_citta,
                    indicatori_salute_mentale,
                    fabb_salute_mentale[citta_upper],
                    citta_upper,
                )

            # --- Dipendenze Patologiche (SerD) ---
            if (fabb_dipendenze
                    and fabb_dipendenze.get(citta_upper)):
                riep_row = _scrivi_tabella_agenas_territoriale(
                    ws_riep, riep_row + 1, df_citta,
                    indicatori_dipendenze,
                    fabb_dipendenze[citta_upper],
                    citta_upper,
                )

            # --- NPIA ---
            if fabb_npia and fabb_npia.get(citta_upper):
                riep_row = _scrivi_tabella_agenas_territoriale(
                    ws_riep, riep_row + 1, df_citta,
                    indicatori_npia,
                    fabb_npia[citta_upper],
                    citta_upper,
                )

            # --- Salute in Carcere ---
            if fabb_carcere and fabb_carcere.get(citta_upper):
                riep_row = _scrivi_tabella_agenas_territoriale(
                    ws_riep, riep_row + 1, df_citta,
                    indicatori_carcere,
                    fabb_carcere[citta_upper],
                    citta_upper,
                )

            auto_larghezza_colonne(ws_riep, col_riep)
            wb.save(nome_file)

        # ==============================================================
        # RIEPILOGO AZIENDALE (file separato, dati di tutte le aree)
        # ==============================================================
        col_pens = [f'Pensionamenti {a}' for a in anni_pensionamento]
        col_quant_ti  = 'Quantità T.I.'
        col_quant_td  = 'Quantità T.D.'
        col_quant_tot = 'Totale'

        nome_file_az = f"elaborati/riepilogo_aziendale_{anno_analisi}.xlsx"
        file_outputs.append(nome_file_az)
        wb_az = Workbook()
        wb_az.remove(wb_az.active)

        # --- Foglio 1: RIEPILOGO PER AREA ---
        ws_area = wb_az.create_sheet(title='RIEPILOGO PER AREA')
        col_area = (
            ['Area', 'Profilo Professionale',
             col_quant_ti, col_quant_td, col_quant_tot]
            + col_pens + ['Proiezione']
        )
        scrivi_titolo(ws_area,
                      f"RIEPILOGO AZIENDALE PER AREA - {anno_analisi}",
                      len(col_area))
        scrivi_intestazioni(ws_area, col_area)

        row_area = 3
        color_t_area = 0
        prev_citta_az = None

        for citta_az, df_caz in grouped.groupby('_CITTA', sort=True):
            if citta_az in _AREE_SPECIALI:
                continue  # scritte dopo in tabella separata
            agg_az = df_caz.groupby('Profilo Professionale').agg(
                **{
                    col_quant_ti:  (col_quant_ti, 'sum'),
                    col_quant_td:  (col_quant_td, 'sum'),
                    **{cp: (cp, 'sum') for cp in col_pens},
                }
            ).reset_index().sort_values('Profilo Professionale')
            agg_az[col_quant_tot] = agg_az[col_quant_ti] + agg_az[col_quant_td]
            agg_az['Proiezione'] = (
                agg_az[col_quant_tot] - agg_az[col_pens].sum(axis=1)
            )

            if citta_az != prev_citta_az:
                color_t_area = 1 - color_t_area
                prev_citta_az = citta_az
            fill_az = FILL_A if color_t_area else FILL_B

            for _, r_az in agg_az.iterrows():
                valori = (
                    [citta_az, r_az['Profilo Professionale'],
                     int(r_az[col_quant_ti]),
                     int(r_az[col_quant_td]),
                     int(r_az[col_quant_tot])]
                    + [int(r_az[cp]) for cp in col_pens]
                    + [int(r_az['Proiezione'])]
                )
                scrivi_riga_dati(ws_area, row_area, valori, fill_az)
                row_area += 1

        # --- Tabella separata: personale non assegnato / speciale ---
        df_speciali = grouped[grouped['_CITTA'].isin(_AREE_SPECIALI)]
        if not df_speciali.empty:
            row_area += 1  # riga vuota di separazione

            # Titoletto sezione
            ws_area.merge_cells(
                start_row=row_area, start_column=1,
                end_row=row_area, end_column=len(col_area),
            )
            t_spec = ws_area.cell(
                row=row_area, column=1,
                value='PERSONALE NON ASSEGNATO A SEDE OPERATIVA',
            )
            t_spec.font = Font(bold=True, size=12)
            t_spec.alignment = ALIGN_CENTER
            row_area += 1

            # Intestazioni
            scrivi_intestazioni(ws_area, col_area, riga=row_area)
            row_area += 1

            color_t_spec = 0
            prev_spec = None
            for citta_sp, df_sp in df_speciali.groupby('_CITTA', sort=True):
                agg_sp = df_sp.groupby('Profilo Professionale').agg(
                    **{
                        col_quant_ti:  (col_quant_ti, 'sum'),
                        col_quant_td:  (col_quant_td, 'sum'),
                        **{cp: (cp, 'sum') for cp in col_pens},
                    }
                ).reset_index().sort_values('Profilo Professionale')
                agg_sp[col_quant_tot] = (
                    agg_sp[col_quant_ti] + agg_sp[col_quant_td]
                )
                agg_sp['Proiezione'] = (
                    agg_sp[col_quant_tot] - agg_sp[col_pens].sum(axis=1)
                )

                if citta_sp != prev_spec:
                    color_t_spec = 1 - color_t_spec
                    prev_spec = citta_sp
                fill_sp = FILL_A if color_t_spec else FILL_B

                for _, r_sp in agg_sp.iterrows():
                    valori = (
                        [citta_sp, r_sp['Profilo Professionale'],
                         int(r_sp[col_quant_ti]),
                         int(r_sp[col_quant_td]),
                         int(r_sp[col_quant_tot])]
                        + [int(r_sp[cp]) for cp in col_pens]
                        + [int(r_sp['Proiezione'])]
                    )
                    scrivi_riga_dati(ws_area, row_area, valori, fill_sp)
                    row_area += 1

        auto_larghezza_colonne(ws_area, col_area)

        # --- Foglio 2: RIEPILOGO PER PROFILO ---
        ws_prof = wb_az.create_sheet(title='RIEPILOGO PER PROFILO')
        col_prof = (
            ['Profilo Professionale',
             col_quant_ti, col_quant_td, col_quant_tot]
            + col_pens + ['Proiezione']
        )
        scrivi_titolo(ws_prof,
                      f"RIEPILOGO AZIENDALE PER PROFILO - {anno_analisi}",
                      len(col_prof))
        scrivi_intestazioni(ws_prof, col_prof)

        agg_prof = grouped.groupby('Profilo Professionale').agg(
            **{
                col_quant_ti:  (col_quant_ti, 'sum'),
                col_quant_td:  (col_quant_td, 'sum'),
                **{cp: (cp, 'sum') for cp in col_pens},
            }
        ).reset_index().sort_values('Profilo Professionale')
        agg_prof[col_quant_tot] = agg_prof[col_quant_ti] + agg_prof[col_quant_td]
        agg_prof['Proiezione'] = (
            agg_prof[col_quant_tot] - agg_prof[col_pens].sum(axis=1)
        )

        row_prof = 3
        for idx_p, (_, r_p) in enumerate(agg_prof.iterrows()):
            fill_p = FILL_A if idx_p % 2 == 0 else FILL_B
            valori = (
                [r_p['Profilo Professionale'],
                 int(r_p[col_quant_ti]),
                 int(r_p[col_quant_td]),
                 int(r_p[col_quant_tot])]
                + [int(r_p[cp]) for cp in col_pens]
                + [int(r_p['Proiezione'])]
            )
            scrivi_riga_dati(ws_prof, row_prof, valori, fill_p)
            row_prof += 1

        # Riga TOTALE aziendale
        fill_tot = FILL_HEADER
        tot_vals = (
            ['TOTALE AZIENDALE',
             int(agg_prof[col_quant_ti].sum()),
             int(agg_prof[col_quant_td].sum()),
             int(agg_prof[col_quant_tot].sum())]
            + [int(agg_prof[cp].sum()) for cp in col_pens]
            + [int(agg_prof['Proiezione'].sum())]
        )
        for ci, v in enumerate(tot_vals, 1):
            c = ws_prof.cell(row=row_prof, column=ci, value=v)
            c.font = FONT_HEADER
            c.fill = fill_tot
            c.border = THIN_BORDER
            if ci > 1:
                c.alignment = ALIGN_CENTER
        row_prof += 1

        # Tabella Sanitario / Non Sanitario aziendale
        row_prof = _scrivi_tabella_sanitario(
            ws_prof, row_prof + 2, personale_all_df
        )

        auto_larghezza_colonne(ws_prof, col_prof)

        # --- Foglio 3: COMPOSIZIONE PER AREA ---
        ws_comp = wb_az.create_sheet(title='COMPOSIZIONE PER AREA')

        # Classificazione su tutto il personale (TI + TD)
        sep_az = ' - '
        pa = personale_all_df.copy()
        pa['_AREA_SAN'] = pa['DESC_RUOLO'].apply(
            lambda r: 'Sanitario'
            if str(r).upper().startswith('RUOLO SANIT')
               or str(r).upper().startswith('RUOLO SOCIO')
            else 'Non Sanitario'
        )
        pa['_LIVELLO'] = pa['DESC_TIPO_DIPENDENTE'].apply(
            lambda t: 'Dirigenza' if str(t).startswith('Dirig') else 'Comparto'
        )
        pa['_NATURA'] = pa['DESC_NATURA'].str.upper().apply(
            lambda n: 'T.I.' if n == 'TEMPO INDETERMINATO' else 'T.D.'
        )
        if pa['DESC_SEDE_FISICA'].str.contains(sep_az, regex=False).all():
            pa['_CITTA'] = (
                pa['DESC_SEDE_FISICA'].str.split(sep_az, n=1).str[0].str.strip()
            )
        else:
            pa['_CITTA'] = pa['DESC_SEDE_FISICA']

        # Pivot: Area x (Area_San, Livello, Natura)
        comp_cols = [
            'Sanitario\nDirigenza\nT.I.', 'Sanitario\nDirigenza\nT.D.',
            'Sanitario\nDirigenza\nTotale',
            'Sanitario\nComparto\nT.I.', 'Sanitario\nComparto\nT.D.',
            'Sanitario\nComparto\nTotale',
            'Sanitario\nTotale', 'Sanitario\n%',
            'Non Sanitario\nDirigenza\nT.I.', 'Non Sanitario\nDirigenza\nT.D.',
            'Non Sanitario\nDirigenza\nTotale',
            'Non Sanitario\nComparto\nT.I.', 'Non Sanitario\nComparto\nT.D.',
            'Non Sanitario\nComparto\nTotale',
            'Non Sanitario\nTotale', 'Non Sanitario\n%',
            'Totale'
        ]
        all_columns_comp = ['Area'] + comp_cols

        scrivi_titolo(ws_comp,
                      f"COMPOSIZIONE PERSONALE PER AREA - {anno_analisi}",
                      len(all_columns_comp))

        # Intestazioni (riga 2) con wrap_text
        for ci, cn in enumerate(all_columns_comp, 1):
            c = ws_comp.cell(row=2, column=ci, value=cn)
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.alignment = Alignment(horizontal='center', vertical='center',
                                    wrap_text=True)
            c.border = THIN_BORDER
        ws_comp.row_dimensions[2].height = 50

        def _conta(df_sub, area_san, livello, natura):
            mask = (df_sub['_AREA_SAN'] == area_san) & (df_sub['_LIVELLO'] == livello)
            if natura != 'Totale':
                mask = mask & (df_sub['_NATURA'] == natura)
            return int(mask.sum())

        row_comp = 3
        rows_data_comp = []
        rows_data_spec = []

        for citta_c in sorted(pa['_CITTA'].unique()):
            df_c = pa[pa['_CITTA'] == citta_c]
            tot_c = len(df_c)
            s_dir_ti  = _conta(df_c, 'Sanitario', 'Dirigenza', 'T.I.')
            s_dir_td  = _conta(df_c, 'Sanitario', 'Dirigenza', 'T.D.')
            s_dir_tot = s_dir_ti + s_dir_td
            s_com_ti  = _conta(df_c, 'Sanitario', 'Comparto', 'T.I.')
            s_com_td  = _conta(df_c, 'Sanitario', 'Comparto', 'T.D.')
            s_com_tot = s_com_ti + s_com_td
            s_tot     = s_dir_tot + s_com_tot
            s_pct     = f'{s_tot / tot_c * 100:.1f}%' if tot_c else '0.0%'

            n_dir_ti  = _conta(df_c, 'Non Sanitario', 'Dirigenza', 'T.I.')
            n_dir_td  = _conta(df_c, 'Non Sanitario', 'Dirigenza', 'T.D.')
            n_dir_tot = n_dir_ti + n_dir_td
            n_com_ti  = _conta(df_c, 'Non Sanitario', 'Comparto', 'T.I.')
            n_com_td  = _conta(df_c, 'Non Sanitario', 'Comparto', 'T.D.')
            n_com_tot = n_com_ti + n_com_td
            n_tot     = n_dir_tot + n_com_tot
            n_pct     = f'{n_tot / tot_c * 100:.1f}%' if tot_c else '0.0%'

            vals_c = [citta_c,
                      s_dir_ti, s_dir_td, s_dir_tot,
                      s_com_ti, s_com_td, s_com_tot,
                      s_tot, s_pct,
                      n_dir_ti, n_dir_td, n_dir_tot,
                      n_com_ti, n_com_td, n_com_tot,
                      n_tot, n_pct,
                      tot_c]
            if citta_c in _AREE_SPECIALI:
                rows_data_spec.append(vals_c)
            else:
                rows_data_comp.append(vals_c)

        # Riga totale aziendale
        tot_all = len(pa)
        def _conta_all(area_san, livello, natura):
            return _conta(pa, area_san, livello, natura)

        sa_dir_ti  = _conta_all('Sanitario', 'Dirigenza', 'T.I.')
        sa_dir_td  = _conta_all('Sanitario', 'Dirigenza', 'T.D.')
        sa_dir_tot = sa_dir_ti + sa_dir_td
        sa_com_ti  = _conta_all('Sanitario', 'Comparto', 'T.I.')
        sa_com_td  = _conta_all('Sanitario', 'Comparto', 'T.D.')
        sa_com_tot = sa_com_ti + sa_com_td
        sa_tot     = sa_dir_tot + sa_com_tot
        sa_pct     = f'{sa_tot / tot_all * 100:.1f}%' if tot_all else '0.0%'

        na_dir_ti  = _conta_all('Non Sanitario', 'Dirigenza', 'T.I.')
        na_dir_td  = _conta_all('Non Sanitario', 'Dirigenza', 'T.D.')
        na_dir_tot = na_dir_ti + na_dir_td
        na_com_ti  = _conta_all('Non Sanitario', 'Comparto', 'T.I.')
        na_com_td  = _conta_all('Non Sanitario', 'Comparto', 'T.D.')
        na_com_tot = na_com_ti + na_com_td
        na_tot     = na_dir_tot + na_com_tot
        na_pct     = f'{na_tot / tot_all * 100:.1f}%' if tot_all else '0.0%'

        totale_row = ['TOTALE AZIENDALE',
                      sa_dir_ti, sa_dir_td, sa_dir_tot,
                      sa_com_ti, sa_com_td, sa_com_tot,
                      sa_tot, sa_pct,
                      na_dir_ti, na_dir_td, na_dir_tot,
                      na_com_ti, na_com_td, na_com_tot,
                      na_tot, na_pct,
                      tot_all]

        # Scrittura righe area
        for idx_c, vals_c in enumerate(rows_data_comp):
            fill_c = FILL_A if idx_c % 2 == 0 else FILL_B
            for ci, v in enumerate(vals_c, 1):
                c = ws_comp.cell(row=row_comp, column=ci, value=v)
                c.fill = fill_c
                c.border = THIN_BORDER
                if ci > 1:
                    c.alignment = ALIGN_CENTER
            row_comp += 1

        # Scrittura riga totale
        for ci, v in enumerate(totale_row, 1):
            c = ws_comp.cell(row=row_comp, column=ci, value=v)
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.border = THIN_BORDER
            if ci > 1:
                c.alignment = ALIGN_CENTER
        row_comp += 1

        # --- Tabella separata: personale non assegnato ---
        if rows_data_spec:
            row_comp += 1  # riga vuota

            ws_comp.merge_cells(
                start_row=row_comp, start_column=1,
                end_row=row_comp, end_column=len(all_columns_comp),
            )
            t_sp2 = ws_comp.cell(
                row=row_comp, column=1,
                value='PERSONALE NON ASSEGNATO A SEDE OPERATIVA',
            )
            t_sp2.font = Font(bold=True, size=12)
            t_sp2.alignment = ALIGN_CENTER
            row_comp += 1

            # Intestazioni
            for ci, cn in enumerate(all_columns_comp, 1):
                c = ws_comp.cell(row=row_comp, column=ci, value=cn)
                c.font = FONT_HEADER
                c.fill = FILL_HEADER
                c.alignment = Alignment(horizontal='center',
                                        vertical='center', wrap_text=True)
                c.border = THIN_BORDER
            ws_comp.row_dimensions[row_comp].height = 50
            row_comp += 1

            for idx_s, vals_s in enumerate(rows_data_spec):
                fill_s = FILL_A if idx_s % 2 == 0 else FILL_B
                for ci, v in enumerate(vals_s, 1):
                    c = ws_comp.cell(row=row_comp, column=ci, value=v)
                    c.fill = fill_s
                    c.border = THIN_BORDER
                    if ci > 1:
                        c.alignment = ALIGN_CENTER
                row_comp += 1

        auto_larghezza_colonne(ws_comp, all_columns_comp, larghezza_min=10)

        # --- Foglio 4: FABBISOGNO AGENAS (riepilogo aziendale) ---
        scrivi_foglio_riepilogo_agenas(
            wb_az, grouped, livello_presidio,
            # Ospedaliere
            fabb_agenas_per_presidio,
            mapping_uo_agenas, mapping_profili_agenas,
            fabb_radio_per_presidio,
            mapping_uo_radio, mapping_profili_radio,
            fabb_anapato_per_presidio,
            mapping_uo_anapato, mapping_profili_anapato,
            fabb_lab_per_presidio,
            mapping_uo_lab, mapping_profili_lab, esclusioni_lab,
            fabb_teclab_per_presidio, mapping_profili_teclab,
            fabb_medleg_per_presidio,
            mapping_uo_medleg, mapping_profili_medleg,
            fabb_trasf_per_presidio,
            mapping_uo_trasf, mapping_profili_trasf,
            fabb_trasf_speciale,
            fabb_emergenza_per_presidio,
            mapping_uo_emergenza, mapping_profili_emergenza,
            fabb_ti_per_presidio,
            mapping_uo_ti, mapping_profili_ti,
            fabb_so_per_presidio,
            mapping_uo_so, mapping_profili_so,
            # Territoriali
            indicatori_salute_mentale, fabb_salute_mentale,
            indicatori_dipendenze, fabb_dipendenze,
            indicatori_npia, fabb_npia,
            indicatori_carcere, fabb_carcere,
        )

        # --- Foglio 5: FABBISOGNO TEORICO (riepilogo aziendale) ---
        scrivi_foglio_riepilogo_fabbisogno_teorico(
            wb_az, grouped, livello_presidio,
        )

        # --- Foglio 6: VETERINARI (riepilogo aziendale) ---
        scrivi_foglio_veterinari(wb_az, grouped)

        wb_az.save(nome_file_az)

        # ==============================================================
        # FILE DETTAGLIO PERSONALE NON ASSEGNATO (dati sorgente grezzi)
        # ==============================================================
        _col_sede = 'DESC_SEDE_FISICA'
        if _col_sede in personale_raw_df.columns:
            df_na_raw = personale_raw_df[
                personale_raw_df[_col_sede].isna()
            ].copy()
            if not df_na_raw.empty:
                # Aggiungo colonna con la classificazione assegnata
                def _classifica(row):
                    cdc = str(row.get('DESC_TIPO_CDC', '')).upper()
                    if 'LUNGHE ASSENZE' in cdc:
                        return 'LUNGHE ASSENZE'
                    if 'ATTESA DI ASSEGNAZIONE' in cdc:
                        return 'IN ATTESA DI ASSEGNAZIONE'
                    return '(Non assegnata)'
                df_na_raw.insert(0, 'Classificazione', df_na_raw.apply(
                    _classifica, axis=1))

                nome_file_na = (
                    f"elaborati/personale_non_assegnato_{anno_analisi}.xlsx"
                )
                with pd.ExcelWriter(nome_file_na, engine='openpyxl') as writer:
                    df_na_raw.to_excel(writer, index=False,
                                       sheet_name='Non assegnati')
                    ws_na = writer.sheets['Non assegnati']
                    auto_larghezza_colonne(ws_na, len(df_na_raw.columns))

                file_outputs.append(nome_file_na)

        print(f"\n{'=' * 70}")
        print(f"ELABORAZIONE COMPLETATA - Anno di analisi: {anno_analisi}")
        print(f"{'=' * 70}")
        for f in file_outputs:
            print(f"Output salvato in: {f}")
        print(f"Controprova salvata in: {debug_file}")
        print(f"Righe elaborate: {len(grouped)}")
        print(f"{'=' * 70}\n")

    else:
        # Fallback: unico file, unico foglio
        grouped.to_excel(FILE_OUTPUT, index=False)

        print(f"\n{'=' * 70}")
        print(f"ELABORAZIONE COMPLETATA - Anno di analisi: {anno_analisi}")
        print(f"{'=' * 70}")
        print(f"Output salvato in: {FILE_OUTPUT}")
        print(f"Controprova salvata in: {debug_file}")
        print(f"Righe elaborate: {len(grouped)}")
        print(f"{'=' * 70}\n")
