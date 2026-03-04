"""
Esportazione in PDF (A3 landscape) dei report Excel aziendali.

Converte i file:
  - riepilogo_aziendale_YYYY.xlsx
  - odc_dm77_YYYY.xlsx

in un unico PDF multi-pagina, leggibile e stampabile,
senza spezzare le tabelle a metà pagina.
"""

import os
import re
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A3
from reportlab.lib.units import mm, cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    PageBreak, KeepTogether,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# ============================================================
# Costanti pagina A3 landscape
# ============================================================
PAGE_SIZE = landscape(A3)       # 420 x 297 mm
PAGE_W, PAGE_H = PAGE_SIZE
MARGIN = 15 * mm

# Colori presi da stili_excel.py
CLR_HEADER_BG = colors.HexColor('#4472C4')
CLR_HEADER_FG = colors.white
CLR_ROW_A     = colors.HexColor('#DCE6F1')
CLR_ROW_B     = colors.white
CLR_TOTALE_BG = colors.HexColor('#D9E2F3')
CLR_OK        = colors.HexColor('#C6EFCE')
CLR_CARENZA   = colors.HexColor('#FFC7CE')
CLR_REGIME    = colors.HexColor('#B4C6E7')
CLR_SECTION   = colors.HexColor('#1F4E79')
CLR_DIVIDER   = colors.HexColor('#002060')


# ============================================================
# Stili paragrafo
# ============================================================
_STYLES = getSampleStyleSheet()

STYLE_TITLE = ParagraphStyle(
    'PDFTitle', parent=_STYLES['Heading1'],
    fontSize=16, leading=20, alignment=TA_CENTER,
    textColor=CLR_SECTION, spaceAfter=4 * mm,
)
STYLE_SECTION = ParagraphStyle(
    'PDFSection', parent=_STYLES['Heading2'],
    fontSize=13, leading=16, alignment=TA_LEFT,
    textColor=CLR_SECTION, spaceBefore=3 * mm, spaceAfter=2 * mm,
)
STYLE_SUBSECTION = ParagraphStyle(
    'PDFSubsection', parent=_STYLES['Heading3'],
    fontSize=11, leading=14, alignment=TA_LEFT,
    textColor=colors.HexColor('#2E75B6'), spaceBefore=2 * mm,
    spaceAfter=1 * mm,
)
STYLE_NOTE = ParagraphStyle(
    'PDFNote', parent=_STYLES['Normal'],
    fontSize=7.5, leading=9, alignment=TA_LEFT,
    textColor=colors.HexColor('#333333'), spaceAfter=1 * mm,
)
STYLE_DIVIDER = ParagraphStyle(
    'PDFDivider', parent=_STYLES['Heading1'],
    fontSize=15, leading=20, alignment=TA_CENTER,
    textColor=colors.white, backColor=CLR_DIVIDER,
    spaceBefore=0, spaceAfter=4 * mm,
    borderPadding=(3 * mm, 0, 3 * mm, 0),
)
STYLE_SHEET_TITLE = ParagraphStyle(
    'PDFSheetTitle', parent=_STYLES['Heading1'],
    fontSize=14, leading=18, alignment=TA_CENTER,
    textColor=CLR_SECTION, spaceBefore=2 * mm, spaceAfter=3 * mm,
)
STYLE_FOOTER = ParagraphStyle(
    'PDFFooter', parent=_STYLES['Normal'],
    fontSize=7, leading=9, alignment=TA_CENTER,
    textColor=colors.grey,
)


# ============================================================
# Utility: lettura celle Excel
# ============================================================

def _cell_text(cell):
    """Restituisce il testo di una cella, gestendo None e numeri."""
    v = cell.value
    if v is None:
        return ''
    return str(v)


def _cell_is_bold(cell):
    return cell.font and cell.font.bold


def _cell_font_size(cell):
    if cell.font and cell.font.size:
        return cell.font.size
    return 10


def _cell_fill_hex(cell):
    """Restituisce il colore di sfondo come stringa hex (senza #), o None."""
    fill = cell.fill
    if fill and fill.start_color and fill.start_color.rgb:
        rgb = str(fill.start_color.rgb)
        if rgb not in ('00000000', 'FFFFFFFF', '0', ''):
            # openpyxl restituisce AARRGGBB, togliamo alpha
            if len(rgb) == 8:
                return rgb[2:]
            return rgb
    return None


def _is_section_title(cell):
    """True se la cella è un titolo di sezione (bold, size >= 12)."""
    return _cell_is_bold(cell) and _cell_font_size(cell) >= 12


def _is_header_row(cell):
    """True se la riga è un header di tabella (sfondo blu header).

    Le righe TOTALE con lo stesso sfondo NON sono header.
    """
    h = _cell_fill_hex(cell)
    if not (h and h.upper() == '4472C4'):
        return False
    # Escludi righe il cui testo inizia con "TOTALE"
    txt = _cell_text(cell).strip().upper()
    if txt.startswith('TOTALE'):
        return False
    return True


def _is_totale_row(cell):
    """True se la riga è una riga TOTALE (grassetto + sfondo totale)."""
    v = _cell_text(cell).strip()
    return v == 'TOTALE' and _cell_is_bold(cell)


def _is_divider_row(ws, row_idx):
    """True se la riga è un divisore di area (sfondo scuro pieno)."""
    cell = ws.cell(row=row_idx, column=1)
    h = _cell_fill_hex(cell)
    if not h:
        return False
    # I divisori hanno sfondo scuro tipo 002060 e altezza 30
    return h.upper() in ('002060', '1F4E79') and _cell_is_bold(cell)


def _is_group_header_row(ws, row_idx, max_col):
    """True se la riga è una riga di gruppo (celle unite con fill 4472C4)
    ma la colonna 1 è vuota – tipico header multi-livello con ospedali."""
    cell1 = ws.cell(row=row_idx, column=1)
    if _cell_text(cell1).strip():
        return False   # la col 1 non è vuota → non è una riga di gruppo
    # Cerca almeno una cella con fill 4472C4 nelle colonne successive
    for c in range(2, max_col + 1):
        h = _cell_fill_hex(ws.cell(row=row_idx, column=c))
        if h and h.upper() == '4472C4':
            return True
    return False


def _row_is_empty(ws, row_idx, max_col):
    """True se tutte le celle della riga sono vuote."""
    for c in range(1, max_col + 1):
        if ws.cell(row=row_idx, column=c).value is not None:
            return False
    return True


# ============================================================
# Parsing di un foglio Excel in blocchi logici
# ============================================================

def _parse_sheet_blocks(ws):
    """Analizza un foglio e lo spezza in blocchi logici.

    Ogni blocco è un dict:
      - 'type': 'title' | 'section' | 'divider' | 'table' | 'note' | 'empty'
      - 'rows': lista di righe (ciascuna = lista di valori stringa)
      - 'raw_rows': range (start, end) righe originali (1-based)
      - 'styles': info stile per rendering
    """
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    # Determina le colonne realmente usate
    real_max_col = 1
    for r in range(1, max_row + 1):
        for c in range(max_col, 0, -1):
            if ws.cell(row=r, column=c).value is not None:
                real_max_col = max(real_max_col, c)
                break
    max_col = real_max_col

    # Mappa merged cells
    merged_map = {}
    for mr in ws.merged_cells.ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if r == mr.min_row and c == mr.min_col:
                    continue
                merged_map[(r, c)] = (mr.min_row, mr.min_col)

    def get_val(r, c):
        if (r, c) in merged_map:
            mr, mc = merged_map[(r, c)]
            return _cell_text(ws.cell(row=mr, column=mc))
        return _cell_text(ws.cell(row=r, column=c))

    blocks = []
    row = 1
    while row <= max_row:
        cell1 = ws.cell(row=row, column=1)
        text1 = _cell_text(cell1).strip()

        # Riga vuota (ma non riga di gruppo con header spostato sulle colonne successive)
        if _row_is_empty(ws, row, max_col):
            row += 1
            continue

        # Divisore di area (sfondo scuro pieno)
        if _is_divider_row(ws, row):
            blocks.append({
                'type': 'divider',
                'text': text1,
                'raw_rows': (row, row),
            })
            row += 1
            continue

        # Titolo di sezione (bold, size >= 12)
        if _is_section_title(cell1) and text1:
            blocks.append({
                'type': 'section',
                'text': text1,
                'raw_rows': (row, row),
            })
            row += 1
            continue

        # Riga di gruppo (header multi-livello: C1 vuota, ospedali su C2+)
        # Deve essere seguita immediatamente da una riga header normale
        is_grp = _is_group_header_row(ws, row, max_col)

        # Header di tabella (sfondo blu) – o riga di gruppo seguita da header
        if _is_header_row(cell1) or is_grp:
            # Inizio di una tabella
            table_start = row
            group_header = None
            group_spans = {}

            if is_grp:
                # Questa è la riga di gruppo; la riga successiva è l'header vero
                group_row_idx = row
                # Costruiamo la mappa span dalla riga di gruppo
                for mr in ws.merged_cells.ranges:
                    if mr.min_row == group_row_idx and mr.min_col >= 2:
                        lbl = _cell_text(
                            ws.cell(row=mr.min_row, column=mr.min_col)).strip()
                        group_spans[mr.min_col - 1] = (lbl, mr.max_col - 1)
                # Raccogliamo anche le celle non unite con fill 4472C4
                for c in range(2, max_col + 1):
                    if (c - 1) not in {s for s, _ in
                                       [(mc, v) for mc, v in group_spans.items()]}:
                        # controlla se la cella ha fill 4472C4 e non è nel merged_map
                        if (group_row_idx, c) not in merged_map:
                            h = _cell_fill_hex(ws.cell(row=group_row_idx, column=c))
                            if h and h.upper() == '4472C4':
                                lbl = _cell_text(
                                    ws.cell(row=group_row_idx, column=c)).strip()
                                if lbl:
                                    group_spans[c - 1] = (lbl, c - 1)
                # Salva i valori della riga di gruppo come group_header raw
                group_header_raw = [get_val(row, c) for c in range(1, max_col + 1)]
                row += 1  # passa all'header vero
                header = [get_val(row, c) for c in range(1, max_col + 1)]
                group_header = group_header_raw
                row += 1
            else:
                header = [get_val(row, c) for c in range(1, max_col + 1)]
                row += 1
                # Controllo: la prossima riga è un'altra riga header? (doppio header)
                if row <= max_row and _is_header_row(ws.cell(row=row, column=1)):
                    group_header = header
                    group_spans = {}
                    gr = row - 1
                    for mr in ws.merged_cells.ranges:
                        if mr.min_row == gr and mr.min_col >= 1:
                            lbl = _cell_text(
                                ws.cell(row=mr.min_row, column=mr.min_col)).strip()
                            group_spans[mr.min_col - 1] = (lbl, mr.max_col - 1)
                    header = [get_val(row, c) for c in range(1, max_col + 1)]
                    row += 1
            data_rows = []
            row_styles = []
            while row <= max_row:
                if _row_is_empty(ws, row, max_col):
                    break
                if _is_section_title(ws.cell(row=row, column=1)):
                    break
                if _is_divider_row(ws, row):
                    break
                if _is_header_row(ws.cell(row=row, column=1)):
                    break
                if _is_group_header_row(ws, row, max_col):
                    break
                vals = [get_val(row, c) for c in range(1, max_col + 1)]
                # Stile riga
                fill_h = _cell_fill_hex(ws.cell(row=row, column=1))
                txt_val = _cell_text(ws.cell(row=row, column=1)).strip().upper()
                is_tot = (txt_val.startswith('TOTALE')
                          or txt_val.startswith('SUBTOTALE'))
                row_styles.append({
                    'fill': fill_h,
                    'bold': _cell_is_bold(ws.cell(row=row, column=1)),
                    'is_totale': is_tot,
                })
                data_rows.append(vals)
                row += 1

            # Rimuovi colonne vuote in coda
            used_cols = max_col
            while used_cols > 1:
                all_empty = (header[used_cols - 1] == '' and
                             all(r[used_cols - 1] == '' for r in data_rows))
                if not all_empty:
                    break
                used_cols -= 1

            header = header[:used_cols]
            if group_header:
                group_header = group_header[:used_cols]
            data_rows = [r[:used_cols] for r in data_rows]

            blocks.append({
                'type': 'table',
                'header': header,
                'group_header': group_header,   # None se non c'è riga di gruppo
                'group_spans': group_spans if group_header else {},
                'data': data_rows,
                'row_styles': row_styles,
                'n_cols': used_cols,
                'raw_rows': (table_start, row - 1),
            })
            continue

        # Testo generico (note, sottotitoli, ecc.)
        if text1:
            style = 'subsection' if _cell_is_bold(cell1) else 'note'
            # Verifica se è un sotto-header di tabella
            if _cell_font_size(cell1) >= 11 and _cell_is_bold(cell1):
                style = 'subsection'
            blocks.append({
                'type': style,
                'text': text1,
                'raw_rows': (row, row),
            })
        row += 1

    return blocks, max_col


# ============================================================
# Rendering blocchi → flowable ReportLab
# ============================================================

def _esito_color(text):
    """Colore sfondo per celle esito."""
    t = str(text).upper()
    if 'CARENZA' in t:
        return CLR_CARENZA
    if 'ECCEDENZA' in t or t == 'OK' or t == 'A REGIME':
        return CLR_OK
    if 'IN RANGE' in t:
        return CLR_OK
    return None


def _build_table(block, avail_width):
    """Costruisce un oggetto Table da un blocco di tipo 'table'."""
    header = block['header']
    data = block['data']
    n_cols = block['n_cols']
    row_styles_info = block['row_styles']
    group_header = block.get('group_header')   # riga di gruppo (es. ospedali)
    group_spans = block.get('group_spans', {}) # {col_start_0based: (label, col_end_0based)}

    if not header:
        return None

    # ---- Classificazione colonne: testuale vs numerica ----
    # Per ogni colonna calcola la max lunghezza dei DATI (escluso header)
    col_data_lens = []
    for ci in range(n_cols):
        max_data = 0
        for dr in data:
            if ci < len(dr):
                max_data = max(max_data, len(str(dr[ci])))
        col_data_lens.append(max(max_data, 2))

    col_hdr_lens = [len(str(header[ci])) for ci in range(n_cols)]

    # Una colonna è "numerica/corta" se i dati sono <= 8 caratteri
    is_numeric = [dl <= 8 for dl in col_data_lens]

    # ---- Calcolo larghezze colonne ----
    # Strategia: colonne numeriche ricevono larghezza basata sui DATI,
    # colonne testuali ricevono larghezza basata sul contenuto reale.
    # Gli header lunghi vanno a capo (Paragraph wrapping).
    col_weights = []
    for ci in range(n_cols):
        if is_numeric[ci]:
            # Colonne numeriche: larghezza proporzionale ai dati
            # con un minimo ragionevole per l'header (che andrà a capo)
            col_weights.append(max(col_data_lens[ci], 4))
        else:
            # Colonne testuali: usa il massimo tra header e dati
            col_weights.append(max(col_data_lens[ci], col_hdr_lens[ci]))

    total_w = sum(col_weights)
    if total_w == 0:
        total_w = n_cols

    # Larghezza minima colonna
    min_col_w = max(8 * mm, avail_width / (n_cols * 4))

    col_widths = []
    for cw in col_weights:
        w = (cw / total_w) * avail_width
        w = max(w, min_col_w)
        col_widths.append(w)

    # Riscala per occupare tutta la larghezza
    s = sum(col_widths)
    if s != 0:
        scale = avail_width / s
        col_widths = [w * scale for w in col_widths]

    # ---- Font size adattativo ----
    if n_cols <= 5:
        font_size = 10
    elif n_cols <= 7:
        font_size = 9.5
    elif n_cols <= 9:
        font_size = 9
    elif n_cols <= 12:
        font_size = 8.5
    elif n_cols <= 15:
        font_size = 8
    elif n_cols <= 18:
        font_size = 7.5
    else:
        font_size = 7

    header_size = font_size

    # ---- Stili Paragraph per header (word-wrap) ----
    hdr_para_style = ParagraphStyle(
        'TblHdr', parent=_STYLES['Normal'],
        fontSize=header_size, leading=header_size + 2,
        alignment=TA_CENTER, textColor=CLR_HEADER_FG,
        fontName='Helvetica-Bold',
    )
    data_para_style = ParagraphStyle(
        'TblData', parent=_STYLES['Normal'],
        fontSize=font_size, leading=font_size + 2,
        alignment=TA_CENTER, fontName='Helvetica',
    )
    data_left_style = ParagraphStyle(
        'TblDataL', parent=data_para_style,
        alignment=TA_LEFT,
    )
    # Stile per righe TOTALE con sfondo blu (testo bianco)
    data_totale_blue_style = ParagraphStyle(
        'TblTotB', parent=data_para_style,
        textColor=CLR_HEADER_FG, fontName='Helvetica-Bold',
    )
    data_totale_blue_left = ParagraphStyle(
        'TblTotBL', parent=data_totale_blue_style,
        alignment=TA_LEFT,
    )
    data_bold_style = ParagraphStyle(
        'TblBold', parent=data_para_style,
        fontName='Helvetica-Bold',
    )
    data_bold_left = ParagraphStyle(
        'TblBoldL', parent=data_bold_style,
        alignment=TA_LEFT,
    )

    # ---- Costruisci dati tabella con Paragraph ----
    # Riga di gruppo (ospedali) se presente
    n_hdr_rows = 1
    group_row = None
    if group_header:
        n_hdr_rows = 2
        group_row = []
        for ci in range(n_cols):
            lbl = str(group_header[ci]).strip() if ci < len(group_header) else ''
            group_row.append(Paragraph(lbl, hdr_para_style))

    header_row = []
    for ci in range(n_cols):
        header_row.append(Paragraph(str(header[ci]), hdr_para_style))

    if group_row:
        table_data = [group_row, header_row]
    else:
        table_data = [header_row]
    for ri, dr in enumerate(data):
        rs = row_styles_info[ri] if ri < len(row_styles_info) else {}
        is_tot = rs.get('is_totale', False)
        fill_hex = (rs.get('fill') or '').upper()
        is_blue_tot = is_tot and fill_hex == '4472C4'

        row_cells = []
        for ci in range(n_cols):
            val = str(dr[ci]) if ci < len(dr) else ''
            if ci == 0:
                # Prima colonna: allineata a sinistra
                if is_blue_tot:
                    sty = data_totale_blue_left
                elif is_tot:
                    sty = data_bold_left
                else:
                    sty = data_left_style
            else:
                if is_blue_tot:
                    sty = data_totale_blue_style
                elif is_tot:
                    sty = data_bold_style
                else:
                    sty = data_para_style
            row_cells.append(Paragraph(val, sty))
        table_data.append(row_cells)

    tbl = Table(table_data, colWidths=col_widths, repeatRows=n_hdr_rows)

    # ---- Stile tabella ----
    style_cmds = [
        # Header (tutte le righe di intestazione)
        ('BACKGROUND', (0, 0), (-1, n_hdr_rows - 1), CLR_HEADER_BG),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]

    # SPAN per celle unite della riga di gruppo
    if group_row and group_spans:
        for col_start, (lbl, col_end) in group_spans.items():
            if col_end > col_start and col_start < n_cols:
                style_cmds.append(
                    ('SPAN', (col_start, 0), (min(col_end, n_cols - 1), 0)))
        # Le celle non coperte da span nella riga 0 restano singole
        style_cmds.append(('ALIGN', (0, 0), (-1, 0), 'CENTER'))

    # Sfondo righe alternato + riga TOTALE
    for ri, rs in enumerate(row_styles_info):
        data_row = ri + n_hdr_rows  # offset per le righe header
        if rs.get('is_totale'):
            fill_hex = (rs.get('fill') or '').upper()
            if fill_hex == '4472C4':
                style_cmds.append(
                    ('BACKGROUND', (0, data_row), (-1, data_row),
                     CLR_HEADER_BG))
            else:
                style_cmds.append(
                    ('BACKGROUND', (0, data_row), (-1, data_row),
                     CLR_TOTALE_BG))
        else:
            fill = CLR_ROW_A if ri % 2 == 0 else CLR_ROW_B
            style_cmds.append(
                ('BACKGROUND', (0, data_row), (-1, data_row), fill))

        # Colora cella esito (ultima colonna con testo)
        if ri < len(data):
            row_data = data[ri]
            for ci in range(len(row_data) - 1, -1, -1):
                txt = str(row_data[ci])
                ec = _esito_color(txt)
                if ec:
                    style_cmds.append(
                        ('BACKGROUND', (ci, data_row),
                         (ci, data_row), ec))
                    break

    tbl.setStyle(TableStyle(style_cmds))
    return tbl


def _render_blocks(blocks, avail_width):
    """Converte blocchi in flowable ReportLab, raggruppando con KeepTogether.

    Strategia: ogni sezione (titolo + sottotitolo + tabella + note)
    viene raggruppata in KeepTogether per evitare di spezzarla.
    """
    flowables = []
    group = []  # accumula blocchi da tenere insieme

    def flush_group():
        nonlocal group
        if group:
            flowables.append(KeepTogether(group))
            group = []

    seen_divider = False
    consumed = set()   # indici blocchi già consumati dal peek-ahead
    for bi, blk in enumerate(blocks):
        if bi in consumed:
            continue
        btype = blk['type']

        if btype == 'divider':
            flush_group()
            # Nuova pagina per ogni area (tranne la primissima)
            if seen_divider:
                flowables.append(PageBreak())
            seen_divider = True
            flowables.append(Spacer(1, 3 * mm))
            # Divider come paragrafo con sfondo
            flowables.append(Paragraph(blk['text'], STYLE_DIVIDER))
            flowables.append(Spacer(1, 2 * mm))
            continue

        if btype == 'section':
            # Se c'è un gruppo accumulato, flusho prima
            # MA: se il prossimo blocco è una tabella, tieni insieme
            # sezione+tabella. Altrimenti flusha.
            # Peek ahead: il prossimo blocco non-note è tabella?
            next_is_table = False
            for nbi in range(bi + 1, min(bi + 5, len(blocks))):
                nt = blocks[nbi]['type']
                if nt == 'table':
                    next_is_table = True
                    break
                if nt in ('divider', 'section'):
                    break

            if not next_is_table:
                flush_group()

            group.append(Paragraph(blk['text'], STYLE_SECTION))
            continue

        if btype == 'subsection':
            group.append(Paragraph(blk['text'], STYLE_SUBSECTION))
            continue

        if btype == 'note':
            txt = blk['text'].replace('\n', '<br/>')
            group.append(Paragraph(txt, STYLE_NOTE))
            continue

        if btype == 'table':
            tbl = _build_table(blk, avail_width)
            if tbl:
                group.append(Spacer(1, 1 * mm))
                group.append(tbl)
                group.append(Spacer(1, 1.5 * mm))
            # Flush dopo la tabella + eventuali note che seguono
            # Peek: se seguono note, includile nel gruppo
            peek = bi + 1
            while peek < len(blocks) and blocks[peek]['type'] == 'note':
                txt = blocks[peek]['text'].replace('\n', '<br/>')
                group.append(Paragraph(txt, STYLE_NOTE))
                consumed.add(peek)   # marca come già consumato
                peek += 1
            flush_group()
            continue

    flush_group()
    return flowables


# ============================================================
# Numerazione pagine
# ============================================================

def _footer(canvas, doc):
    canvas.saveState()
    canvas.setFont('Helvetica', 7)
    canvas.setFillColor(colors.grey)
    page_text = f"Pag. {doc.page}"
    canvas.drawCentredString(PAGE_W / 2, 8 * mm, page_text)
    # Data generazione
    now = datetime.now().strftime('%d/%m/%Y %H:%M')
    canvas.drawRightString(PAGE_W - MARGIN, 8 * mm,
                           f"Generato il {now}")
    canvas.drawString(MARGIN, 8 * mm, "ASREM – Fabbisogno Personale")
    canvas.restoreState()


# ============================================================
# Classe principale
# ============================================================

class ExportPDF:
    """Genera un PDF A3 landscape dai report Excel aziendali.

    Uso:
        exp = ExportPDF(anno=2026, cartella_elaborati='elaborati')
        exp.genera('elaborati/report_aziendale_2026.pdf')
    """

    def __init__(self, anno, cartella_elaborati='elaborati'):
        self.anno = anno
        self.cartella = cartella_elaborati
        self._xlsx_riepilogo = os.path.join(
            cartella_elaborati, f'riepilogo_aziendale_{anno}.xlsx')
        self._xlsx_odc = os.path.join(
            cartella_elaborati, f'odc_dm77_{anno}.xlsx')

    def genera(self, output_path=None):
        """Genera il PDF combinando riepilogo aziendale + OdC DM77."""
        if output_path is None:
            output_path = os.path.join(
                self.cartella, f'report_aziendale_{self.anno}.pdf')

        doc = SimpleDocTemplate(
            output_path,
            pagesize=PAGE_SIZE,
            leftMargin=MARGIN,
            rightMargin=MARGIN,
            topMargin=MARGIN,
            bottomMargin=MARGIN + 5 * mm,
            title=f'Fabbisogno Personale ASREM {self.anno}',
            author='ASREM – Servizio Programmazione',
        )

        avail_width = PAGE_W - 2 * MARGIN
        story = []

        # ----- COPERTINA -----
        story.append(Spacer(1, 60 * mm))
        story.append(Paragraph(
            f"ASREM – Azienda Sanitaria Regionale del Molise",
            ParagraphStyle('Cover1', parent=_STYLES['Heading1'],
                           fontSize=22, leading=28, alignment=TA_CENTER,
                           textColor=CLR_SECTION)))
        story.append(Spacer(1, 15 * mm))
        story.append(Paragraph(
            f"Fabbisogno di Personale – Anno {self.anno}",
            ParagraphStyle('Cover2', parent=_STYLES['Heading1'],
                           fontSize=18, leading=24, alignment=TA_CENTER,
                           textColor=CLR_DIVIDER)))
        story.append(Spacer(1, 10 * mm))
        story.append(Paragraph(
            "Riepilogo Aziendale e Ospedali di Comunità (DM 77)",
            ParagraphStyle('Cover3', parent=_STYLES['Heading2'],
                           fontSize=14, leading=18, alignment=TA_CENTER,
                           textColor=colors.HexColor('#2E75B6'))))
        story.append(Spacer(1, 30 * mm))
        now = datetime.now().strftime('%d/%m/%Y')
        story.append(Paragraph(
            f"Documento generato il {now}",
            ParagraphStyle('CoverDate', parent=_STYLES['Normal'],
                           fontSize=10, alignment=TA_CENTER,
                           textColor=colors.grey)))
        story.append(PageBreak())

        # ----- RIEPILOGO AZIENDALE -----
        if os.path.exists(self._xlsx_riepilogo):
            story.extend(
                self._processa_workbook(self._xlsx_riepilogo,
                                        "RIEPILOGO AZIENDALE",
                                        avail_width))

        # ----- ODC DM77 -----
        if os.path.exists(self._xlsx_odc):
            story.append(PageBreak())
            story.extend(
                self._processa_workbook(self._xlsx_odc,
                                        "OSPEDALI DI COMUNITÀ (DM 77)",
                                        avail_width))

        # Build
        doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
        print(f"PDF generato: {output_path}")
        return output_path

    def _processa_workbook(self, xlsx_path, titolo_sezione, avail_width):
        """Processa un workbook Excel e restituisce lista di flowables."""
        wb = openpyxl.load_workbook(xlsx_path)
        flowables = []

        # Titolo sezione principale
        flowables.append(Paragraph(titolo_sezione, ParagraphStyle(
            'WBTitle', parent=_STYLES['Heading1'],
            fontSize=18, leading=22, alignment=TA_CENTER,
            textColor=CLR_DIVIDER, spaceAfter=5 * mm)))
        flowables.append(Spacer(1, 3 * mm))

        for si, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            if ws.max_row is None or ws.max_row < 2:
                continue

            # Nuova pagina per ogni foglio (tranne il primo)
            if si > 0:
                flowables.append(PageBreak())

            # Titolo foglio
            flowables.append(Paragraph(
                sheet_name, STYLE_SHEET_TITLE))
            flowables.append(Spacer(1, 2 * mm))

            # Parse e render
            blocks, max_col = _parse_sheet_blocks(ws)
            rendered = _render_blocks(blocks, avail_width)
            flowables.extend(rendered)

        wb.close()
        return flowables
