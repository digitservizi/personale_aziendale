"""
Lettura degli indicatori XML e calcolo del fabbisogno teorico di personale.
"""

import math
import xml.etree.ElementTree as ET
from collections import defaultdict

from src.config import MAGGIORAZIONE_TURNAZIONE


def read_indicators(file_path):
    """Legge il file XML degli indicatori e restituisce un dizionario
    strutturato per categoria → intensità → parametri."""
    tree = ET.parse(file_path)
    root = tree.getroot()
    indicators = {}

    for category in root:
        category_name = category.tag
        indicators[category_name] = {}
        for intensity in category:
            name = intensity.tag
            indicators[category_name][name] = {}
            for element in intensity:
                if 'TassoOccupazione' in element.tag:
                    indicators[category_name][name][element.tag] = float(
                        element.text.strip('%')
                    )
                else:
                    indicators[category_name][name][element.tag] = float(
                        element.text
                    )
    return indicators


def calculate_fabbisogno(ordinari, dh, utic,
                         intensity, indicators,
                         reparto, profilo, sede='', calcoli_log=None):
    """Calcola il fabbisogno teorico di personale per un reparto.

    I posti UTIC vengono sempre calcolati con intensità 'Intensiva'
    (sono posti di terapia intensiva coronarica) e sommati al fabbisogno
    della parte ordinaria+DH calcolata con l'intensità del reparto.

    Se calcoli_log è una lista, vi aggiunge un dict con tutti i parametri
    del calcolo (usato per la controprova).
    """
    # ── Parte ordinaria + DH ─────────────────────────────────────
    pl_ord = ordinari * 1 + dh * 0.5

    if intensity in indicators:
        indicator = indicators[intensity]
    else:
        indicator = indicators.get('Bassa', list(indicators.values())[0])

    tasso_occupazione = indicator['TassoOccupazione'] / 100.0
    coefficiente_complessita = indicator['CoefficienteComplessita']
    ore_effettuate_turni = indicator['OreEffettuateTurni']
    ore_annue_lavoro_effettivo = indicator['OreAnnueLavoroEffettivo']

    fabb_ord = (
        pl_ord * tasso_occupazione
        * coefficiente_complessita * ore_effettuate_turni * 365
    ) / ore_annue_lavoro_effettivo

    # ── Parte UTIC (intensità Intensiva) ─────────────────────────
    fabb_utic = 0.0
    if utic > 0 and intensity != 'Intensiva':
        # I posti UTIC si calcolano sempre con coefficiente Intensiva
        ind_int = indicators.get('Intensiva', indicator)
        t_occ_int   = ind_int['TassoOccupazione'] / 100.0
        coeff_int   = ind_int['CoefficienteComplessita']
        h_tur_int   = ind_int['OreEffettuateTurni']
        h_ann_int   = ind_int['OreAnnueLavoroEffettivo']
        fabb_utic = (
            utic * t_occ_int * coeff_int * h_tur_int * 365
        ) / h_ann_int
    elif utic > 0:
        # Se il reparto è già Intensivo, UTIC rientra nel calcolo ordinario
        fabb_utic = 0.0
        pl_ord += utic  # ri-aggiungi al calcolo ordinario
        fabb_ord = (
            pl_ord * tasso_occupazione
            * coefficiente_complessita * ore_effettuate_turni * 365
        ) / ore_annue_lavoro_effettivo

    fabbisogno_teorico = fabb_ord + fabb_utic
    posti_letto_omogeneizzati = pl_ord + (utic if intensity != 'Intensiva' else 0)

    if calcoli_log is not None:
        log_entry = {
            'sede':    sede,
            'reparto': reparto,
            'profilo': profilo,
            'pl_omg':  round(posti_letto_omogeneizzati, 2),
            'intensita': intensity,
            't_occ':   tasso_occupazione,
            'coeff':   coefficiente_complessita,
            'h_tur':   ore_effettuate_turni,
            'h_ann':   ore_annue_lavoro_effettivo,
            'f_raw':   round(fabbisogno_teorico, 4),
        }
        if utic > 0 and intensity != 'Intensiva':
            log_entry['utic'] = utic
            log_entry['f_ord'] = round(fabb_ord, 4)
            log_entry['f_utic'] = round(fabb_utic, 4)
        calcoli_log.append(log_entry)

    return round(fabbisogno_teorico, 2)


def match_profilo(profilo, indicators, avvisi_log=None):
    """Fa il match del profilo professionale con le intestazioni XML.

    Se avvisi_log è una lista, vi aggiunge un messaggio di avviso quando
    il profilo non è censito negli indicatori.
    """
    profilo_key = profilo.replace(" ", "_").upper()
    if profilo_key in indicators:
        return profilo_key

    if avvisi_log is not None:
        avvisi_log.append(
            f"Profilo non censito negli indicatori XML: '{profilo}' "
            f"→ figura professionale non calcolata"
        )
    return None


# ============================================================
# CONTROPROVA XLSX
# ============================================================

def scrivi_controprova_xlsx(output_file, anno_analisi, data_esecuzione,
                            calcoli_log, avvisi_log, odc_log, arrot_log):
    """Genera la controprova calcoli come file XLSX con 4 fogli.

    Fogli:
      1. Calcoli        – dettaglio per sede/reparto/profilo
      2. OdC DM77       – fabbisogni fissi da normativa
      3. Arrotondamenti – casi in cui il valore grezzo è stato arrotondato
      4. Avvisi         – figure professionali non censite negli indicatori
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Stili ────────────────────────────────────────────────────────
    THIN = Side(style='thin', color='BBBBBB')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def _font_hdr():  return Font(bold=True, color='FFFFFF', size=10)
    def _font_norm(): return Font(size=10)
    def _font_bold(): return Font(bold=True, size=10)
    def _font_warn(): return Font(bold=True, color='C00000', size=10)

    def _fill(hex_color): return PatternFill('solid', fgColor=hex_color)

    FILL_HDR    = _fill('1F4E79')   # blu scuro – header
    FILL_SEDE_A = _fill('D6E4F0')   # azzurro chiaro – sede A
    FILL_SEDE_B = _fill('EBF5FB')   # azzurro chiarissimo – sede B
    FILL_WARN   = _fill('FFF2CC')   # giallo – avvisi
    FILL_ARROT  = _fill('FCE4D6')   # arancio chiaro – arrotondamenti
    FILL_ODC    = _fill('E2EFDA')   # verde chiaro – OdC

    ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=False)
    ALIGN_L = Alignment(horizontal='left',   vertical='center', wrap_text=False)

    def _scrivi_header(ws, cols, fill=FILL_HDR):
        ws.append(cols)
        for cell in ws[ws.max_row]:
            cell.font      = _font_hdr()
            cell.fill      = fill
            cell.alignment = ALIGN_C
            cell.border    = BORDER

    def _scrivi_riga(ws, valori, fill, align=ALIGN_L):
        ws.append(valori)
        for cell in ws[ws.max_row]:
            cell.fill      = fill
            cell.font      = _font_norm()
            cell.alignment = align
            cell.border    = BORDER

    def _autofit(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    def _freeze(ws, cell='B2'):
        ws.freeze_panes = cell

    # ── Workbook ─────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    # ── Riga titolo helper ───────────────────────────────────────────
    def _titolo_foglio(ws, testo):
        ws.append([testo, '', f'Anno: {anno_analisi}', '', f'Generata: {data_esecuzione}'])
        for cell in ws[1]:
            cell.font      = _font_bold()
            cell.fill      = _fill('2E75B6')
            cell.font      = Font(bold=True, color='FFFFFF', size=11)
            cell.alignment = ALIGN_L
        ws.row_dimensions[1].height = 18

    # ================================================================
    # FOGLIO 1 – CALCOLI
    # ================================================================
    ws1 = wb.create_sheet('Calcoli')
    _titolo_foglio(ws1, 'CONTROPROVA CALCOLI FABBISOGNO')
    _scrivi_header(ws1, [
        'Sede', 'Reparto', 'Profilo',
        'PL Omog.', 'Intensità',
        'T. Occ.', 'Coeff.', 'H. Turno', 'H. Annue',
        'F. Grezzo', 'F. Finale',
        'PL UTIC', 'F. Ord.', 'F. UTIC',
    ])

    sedi = sorted(set(e['sede'] for e in calcoli_log))
    fill_map = {s: (FILL_SEDE_A if i % 2 == 0 else FILL_SEDE_B)
                for i, s in enumerate(sedi)}

    # Mappa arrotondamento finale (sede, reparto, profilo) → f_fin
    arrot_map = {
        (a['sede'], a['reparto'], a['profilo']): a['f_fin']
        for a in arrot_log
    }

    for e in calcoli_log:
        fill = fill_map.get(e['sede'], FILL_SEDE_A)
        f_fin = arrot_map.get((e['sede'], e['reparto'], e['profilo']),
                              round(e['f_raw']))
        _scrivi_riga(ws1, [
            e['sede'], e['reparto'], e['profilo'],
            e['pl_omg'], e['intensita'],
            e['t_occ'], e['coeff'], e['h_tur'], e['h_ann'],
            e['f_raw'], f_fin,
            e.get('utic', ''), e.get('f_ord', ''), e.get('f_utic', ''),
        ], fill)
    _autofit(ws1)
    _freeze(ws1, 'C3')

    # ================================================================
    # FOGLIO 2 – OdC DM77
    # ================================================================
    ws2 = wb.create_sheet('OdC DM77')
    _titolo_foglio(ws2, 'OSPEDALI DI COMUNITÀ – Fabbisogni DM 77/2022')
    _scrivi_header(ws2, ['Sede OdC', 'Reparto', 'Profilo', 'Fabbisogno DM77'])

    prev_sede_odc = None
    fill_odc_toggle = True
    for o in odc_log:
        if o['sede'] != prev_sede_odc:
            fill_odc_toggle = not fill_odc_toggle
            prev_sede_odc = o['sede']
        _scrivi_riga(ws2, [
            o['sede'], o['reparto'], o['profilo'], o['fabb'],
        ], FILL_ODC if fill_odc_toggle else _fill('F0FFF0'))
    _autofit(ws2)
    _freeze(ws2, 'B3')

    # ================================================================
    # FOGLIO 3 – ARROTONDAMENTI
    # ================================================================
    ws3 = wb.create_sheet('Arrotondamenti')
    _titolo_foglio(ws3, 'ARROTONDAMENTI APPLICATI')
    _scrivi_header(ws3, ['Sede', 'Reparto', 'Profilo', 'F. Grezzo', 'F. Finale', 'Motivo'])

    for a in arrot_log:
        _scrivi_riga(ws3, [
            a['sede'], a['reparto'], a['profilo'],
            a['f_raw'], a['f_fin'], a['motivo'],
        ], FILL_ARROT)
    _autofit(ws3)
    _freeze(ws3, 'B3')

    # ================================================================
    # FOGLIO 4 – AVVISI
    # ================================================================
    ws4 = wb.create_sheet('Avvisi')
    _titolo_foglio(ws4, 'FIGURE PROFESSIONALI NON CENSITE NEGLI INDICATORI')
    _scrivi_header(ws4, ['Avviso'])

    seen = set()
    for av in avvisi_log:
        if av not in seen:
            ws4.append([av])
            cell = ws4[ws4.max_row][0]
            cell.font      = _font_warn()
            cell.fill      = FILL_WARN
            cell.alignment = ALIGN_L
            cell.border    = BORDER
            seen.add(av)
    _autofit(ws4)

    wb.save(output_file)


# ============================================================
# FABBISOGNO AGENAS – AREA MATERNO INFANTILE
# ============================================================

def calcola_fabbisogno_agenas_materno_infantile(
        indicatori_agenas, parti_per_presidio):
    """Calcola il fabbisogno AGENAS per l'area materno-infantile.

    Per ogni presidio ospedaliero determina la fascia in base al
    numero di parti e restituisce il range FTE (min-max) per profilo.

    Se il numero di parti è inferiore al minimo della fascia più bassa
    (500), i valori vengono ridotti proporzionalmente:
      fabbisogno = valore_fascia × (parti / soglia_minima)

    Restituisce un dict:
      {nome_presidio: {profilo_agenas: {'min': x, 'max': y}, ...}, ...}
    """
    fasce = indicatori_agenas['fasce']
    if not fasce:
        return {}

    # La fascia con il minimo più basso (prima fascia)
    fascia_base = min(fasce, key=lambda f: f['min'])
    soglia_minima = fascia_base['min']  # tipicamente 500

    risultato = {}
    for presidio, parti in parti_per_presidio.items():
        if parti <= 0:
            risultato[presidio] = {}
            continue

        # Cerca la fascia corrispondente
        fascia_trovata = None
        for fascia in fasce:
            if fascia['min'] <= parti < fascia['max']:
                fascia_trovata = fascia
                break

        if fascia_trovata is None and parti >= fasce[-1]['min']:
            # Oltre il massimo: usa la fascia più alta
            fascia_trovata = fasce[-1]

        if fascia_trovata is not None:
            # Fascia trovata: usa min e max
            risultato[presidio] = {
                profilo: {'min': vals['min'], 'max': vals['max']}
                for profilo, vals in fascia_trovata['profili'].items()
            }
        else:
            # Sotto la soglia minima: proporzione lineare con arrotondamento
            fattore = parti / soglia_minima
            fabb_presidio = {}
            for profilo, vals in fascia_base['profili'].items():
                def _arrot(v):
                    fraz = v - math.floor(v)
                    return math.ceil(v) if fraz >= 0.5 else math.floor(v)
                fabb_presidio[profilo] = {
                    'min': _arrot(vals['min'] * fattore),
                    'max': _arrot(vals['max'] * fattore),
                }
            risultato[presidio] = fabb_presidio

    return risultato


# ============================================================
# FABBISOGNO AGENAS – AREA RADIOLOGIA
# ============================================================

def calcola_fabbisogno_agenas_radiologia(
        indicatori_radiologia, livello_presidio):
    """Calcola il fabbisogno AGENAS per l'area dei servizi di radiologia."""
    return _calcola_fabbisogno_agenas_per_livello(
        indicatori_radiologia, livello_presidio, 'radiologia')


def calcola_fabbisogno_agenas_trasfusionale(
        indicatori_trasfusionale, livello_presidio):
    """Calcola il fabbisogno AGENAS per l'area medicina trasfusionale.

    N.B. Solo i presidi il cui livello è presente nella tabella AGENAS
    (I e II livello) verranno valutati. Gli ospedali di base sono esclusi.
    """
    return _calcola_fabbisogno_agenas_per_livello(
        indicatori_trasfusionale, livello_presidio, 'trasfusionale')


def calcola_fabbisogno_agenas_anatomia_patologica(
        indicatori_anatomia_pat, livello_presidio):
    """Calcola il fabbisogno AGENAS per l'area anatomia patologica (Tab. 16).

    Tutti e tre i livelli (ospedale di base, I e II livello)
    sono previsti nella tabella AGENAS.
    """
    return _calcola_fabbisogno_agenas_per_livello(
        indicatori_anatomia_pat, livello_presidio, 'anatomia patologica')


def calcola_fabbisogno_agenas_laboratorio(
        indicatori_laboratorio, livello_presidio):
    """Calcola il fabbisogno AGENAS per l'area servizi di laboratorio (Tab. 14).

    Tutti e tre i livelli (ospedale di base, I e II livello)
    sono previsti nella tabella AGENAS.
    """
    return _calcola_fabbisogno_agenas_per_livello(
        indicatori_laboratorio, livello_presidio, 'laboratorio')


def calcola_fabbisogno_agenas_tecnici_laboratorio(
        indicatori_tecnici_lab, livello_presidio):
    """Calcola il fabbisogno AGENAS per i tecnici di laboratorio (Tab. 17).

    Tutti e tre i livelli (ospedale di base, I e II livello)
    sono previsti nella tabella AGENAS.
    """
    return _calcola_fabbisogno_agenas_per_livello(
        indicatori_tecnici_lab, livello_presidio, 'tecnici laboratorio')


def calcola_fabbisogno_agenas_medicina_legale(
        indicatori_med_legale, livello_presidio):
    """Calcola il fabbisogno AGENAS per la medicina legale (Tab. 18).

    Solo presidi di I e II livello (non Ospedale di Base).
    """
    return _calcola_fabbisogno_agenas_per_livello(
        indicatori_med_legale, livello_presidio, 'medicina legale')


def calcola_fabbisogno_agenas_emergenza_urgenza(
        indicatori_emergenza, livello_presidio):
    """Calcola il fabbisogno AGENAS per l'area emergenza-urgenza (Tab. 20).

    Tutti e tre i livelli (ospedale di base / PS, DEA I, DEA II)
    sono previsti nella tabella AGENAS.
    """
    return _calcola_fabbisogno_agenas_per_livello(
        indicatori_emergenza, livello_presidio, 'emergenza-urgenza')


def _calcola_fabbisogno_agenas_per_livello(
        indicatori, livello_presidio, nome_area):
    """Calcolo generico fabbisogno AGENAS basato su livello presidio.

    Per ogni presidio ospedaliero usa il livello assegnato
    per determinare il range FTE (min-max) per profilo.
    I presidi il cui livello non è definito nella tabella vengono omessi.

    Restituisce un dict:
      {nome_presidio: {profilo_agenas: {'min': x, 'max': y}, ...}, ...}
    """
    livelli = indicatori['livelli']
    if not livelli:
        return {}

    fattore = 1 + MAGGIORAZIONE_TURNAZIONE

    risultato = {}
    for presidio, livello in livello_presidio.items():
        if livello not in livelli:
            # Livello non previsto nella tabella: presidio omesso
            continue

        profili_livello = livelli[livello]
        risultato[presidio] = {
            profilo: {
                'min': math.floor(vals['min'] * fattore + 0.5),
                'max': math.floor(vals['max'] * fattore + 0.5),
            }
            for profilo, vals in profili_livello.items()
        }

    return risultato


# ============================================================
# FABBISOGNO AGENAS TERRITORIALE (basato su popolazione)
# ============================================================

def calcola_fabbisogno_agenas_territoriale(indicatori, popolazione_area,
                                           detenuti_per_istituto=None):
    """Calcola il fabbisogno AGENAS per standard territoriale basato su popolazione.

    Per ogni area distrettuale:
      atteso = tasso × (popolazione_fascia / base_popolazione)

    Per servizi carcerari (fascia = 'detenuti'):
      atteso = tasso × (n_detenuti / base_popolazione)

    Parametri:
      indicatori:            dict da carica_indicatori_agenas_territoriale()
      popolazione_area:      dict {CITTA: {totale, gte_18, range_15_64, range_1_17}}
      detenuti_per_istituto: dict {CITTA: n_detenuti} (solo per carcere)

    Restituisce un dict per area:
      {CITTA: {'popolazione_rif': int,
               'base': int,
               'profili': [{nome, atteso_min, atteso_regime, qualifiche, nota}]}}
    """
    base = indicatori['base_popolazione']
    fascia = indicatori['fascia_popolazione']

    risultato = {}
    for citta, dati_pop in popolazione_area.items():
        # Determina la popolazione di riferimento
        if fascia == 'detenuti':
            if not detenuti_per_istituto:
                continue
            pop_rif = detenuti_per_istituto.get(citta, 0)
            if pop_rif == 0:
                continue
        else:
            pop_rif = dati_pop.get(fascia, 0)
            if pop_rif == 0:
                continue

        moltiplicatore = pop_rif / base
        fattore = 1 + MAGGIORAZIONE_TURNAZIONE
        profili_calc = []
        for prof in indicatori['profili']:
            atteso_min = math.floor(prof['tasso_min'] * moltiplicatore * fattore + 0.5)
            atteso_regime = (math.floor(prof['tasso_regime'] * moltiplicatore * fattore + 0.5)
                            if prof['tasso_regime'] is not None
                            else None)
            profili_calc.append({
                'nome': prof['nome'],
                'atteso_min': atteso_min,
                'atteso_regime': atteso_regime,
                'qualifiche': prof['qualifiche'],
                'nota': prof.get('nota', ''),
            })

        risultato[citta] = {
            'popolazione_rif': pop_rif,
            'base': base,
            'profili': profili_calc,
        }

    nome = indicatori.get('titolo', 'territoriale')
    print(f"Fabbisogno AGENAS {nome} calcolato per {len(risultato)} aree")
    return risultato


# ============================================================
# FABBISOGNO AGENAS – TERAPIA INTENSIVA (§ 8.1.1)
# ============================================================

def calcola_fabbisogno_agenas_terapia_intensiva(
        indicatori_ti, posti_letto, indicators):
    """Calcola il fabbisogno AGENAS per l'Area Terapia Intensiva (§ 8.1.1).

    La formula è:
      FTE = (PL_intensivi / rapporto_letti) × (ore_turno × 365) / ore_annue_eff

    dove PL_intensivi = somma dei posti letto ordinari delle UO che matchano
    i pattern dell'area terapia intensiva.

    Parametri:
      indicatori_ti: dict da carica_indicatori_agenas_terapia_intensiva()
      posti_letto:   dict {(sede, ssd): {ordinari, intensita, ...}}
      indicators:    dict da read_indicators() – per ore_turno / ore_annue

    Restituisce un dict per presidio:
      {presidio: {profilo: {'fte_atteso': float, 'rapporto': int, 'pl': int}}}
    """
    import re

    mapping_uo = indicatori_ti.get('mapping_uo', [])
    standard = indicatori_ti.get('standard', [])

    # Mapping categoria indicatori_medici per ciascun profilo AGENAS
    _CATEGORIA_INDICATORI = {
        'DIRIGENTE_MEDICO': 'DIRIGENTE_MEDICO',
        'INFERMIERE': 'INFERMIERE',
    }

    # Raggruppa PL intensivi per presidio
    pl_per_presidio = defaultdict(int)
    for (sede, ssd), pl in posti_letto.items():
        for m_uo in mapping_uo:
            if re.search(m_uo['pattern'], ssd, re.IGNORECASE):
                pl_per_presidio[sede] += int(pl.get('ordinari', 0))
                break

    risultato = {}
    for presidio, pl_tot in pl_per_presidio.items():
        if pl_tot <= 0:
            continue
        fabb = {}
        for std in standard:
            prof = std['profilo']
            rapporto = std['rapporto_letti']

            # Recupera ore_turno e ore_annue dall'XML indicatori_medici
            cat = _CATEGORIA_INDICATORI.get(prof)
            if cat and cat in indicators and 'Intensiva' in indicators[cat]:
                ind = indicators[cat]['Intensiva']
                ore_turno = ind.get('OreEffettuateTurni', 8.0)
                ore_annue = ind.get('OreAnnueLavoroEffettivo', 1485.0)
            else:
                ore_turno = 8.0
                ore_annue = 1485.0

            # FTE = (PL / rapporto) × (24 × 365) / ore_annue × (1 + magg)
            # Il rapporto indica quanti letti per operatore IN TURNO;
            # la TI opera H24, quindi si moltiplica per 24 ore/giorno
            # (non per la durata del singolo turno).
            # La maggiorazione copre guardie, ferie, malattie, ecc.
            operatori_per_turno = pl_tot / rapporto
            fte_base = operatori_per_turno * (24 * 365) / ore_annue
            fte = fte_base * (1 + MAGGIORAZIONE_TURNAZIONE)

            fabb[prof] = {
                'fte_atteso': math.floor(fte + 0.5),
                'rapporto': rapporto,
                'pl': pl_tot,
                'ore_turno': ore_turno,
                'ore_annue': ore_annue,
            }
        risultato[presidio] = fabb

    return risultato


# ============================================================
# FABBISOGNO AGENAS SALE OPERATORIE (§ 8.1.2)
# ============================================================

def calcola_fabbisogno_agenas_sale_operatorie(indicatori_so, indicators,
                                               sale_per_presidio):
    """Calcola il fabbisogno AGENAS per le Sale Operatorie (§ 8.1.2).

    Per ogni seduta operatoria è prevista la presenza minima di:
      1 medico anestesista, 3 infermieri, 1 OSS.

    Formula:
      FTE = n_sale × personale_per_sala × (ore_copertura × giorni_anno)
            / ore_annue_eff × (1 + maggiorazione)

    Parametri:
      indicatori_so:     dict da carica_indicatori_agenas_sale_operatorie()
      indicators:        dict indicatori da indicatori_medici.xml (per ore_annue)
      sale_per_presidio:  dict {nome_presidio: n_sale} da config.py

    Restituisce un dict per presidio:
      {presidio: {profilo: {'fte_atteso': int, 'personale_per_sala': int,
                            'n_sale': int, 'ore_copertura': int,
                            'giorni_anno': int}}}
    """
    import re

    standard = indicatori_so.get('standard', [])
    parametri = indicatori_so.get('parametri', {})

    ore_copertura = parametri.get('ore_copertura', 24)
    giorni_anno = parametri.get('giorni_anno', 250)

    # Mapping categoria indicatori_medici per ciascun profilo AGENAS
    _CATEGORIA_INDICATORI = {
        'DIRIGENTE_MEDICO_ANESTESISTA': 'DIRIGENTE_MEDICO',
        'INFERMIERE':                   'INFERMIERE',
        'OPERATORE_SOCIO_SANITARIO':    'OSS',
    }

    # Usa direttamente il dict sale_per_presidio da config
    risultato = {}
    for presidio, n_sale in sale_per_presidio.items():
        if n_sale <= 0:
            continue
        fabb = {}
        for std in standard:
            prof = std['profilo']
            pps = std['personale_per_sala']

            # Recupera ore_annue dall'XML indicatori_medici
            cat = _CATEGORIA_INDICATORI.get(prof)
            if cat and cat in indicators and 'Intensiva' in indicators[cat]:
                ore_annue = indicators[cat]['Intensiva'].get(
                    'OreAnnueLavoroEffettivo', 1485.0)
            else:
                ore_annue = 1485.0

            # FTE = n_sale × staff × (ore_copertura × giorni) / ore_annue × (1+magg)
            fte_base = n_sale * pps * (ore_copertura * giorni_anno) / ore_annue
            fte = fte_base * (1 + MAGGIORAZIONE_TURNAZIONE)

            fabb[prof] = {
                'fte_atteso': math.floor(fte + 0.5),
                'personale_per_sala': pps,
                'n_sale': n_sale,
                'ore_copertura': ore_copertura,
                'giorni_anno': giorni_anno,
                'ore_annue': ore_annue,
            }
        risultato[presidio] = fabb

    return risultato
