"""
Report medici – confronto dotazione da Atto Aziendale vs personale in servizio.
"""

from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from src.stili_excel import (
    FILL_A, FILL_B,
    scrivi_titolo, scrivi_intestazioni, scrivi_riga_dati,
    scrivi_riga_totale, auto_larghezza_colonne,
    ALIGN_CENTER,
)
from src.caricamento_dati import (
    carica_dataframe, normalizza_colonne_personale,
    normalizza_colonne_pensionamenti,
)
from src.caricamento_xml import carica_medici_atto_aziendale


# ─────────────────────────────────────────────────────────────
# HELPER: scrive un DataFrame in un foglio con formattazione
# ─────────────────────────────────────────────────────────────

def _scrivi_foglio_atto(wb, sheet_name, df, titolo,
                        col_gruppo='Disciplina'):
    """Scrive *df* in un nuovo foglio del workbook con stile standard."""
    ws = wb.create_sheet(title=sheet_name[:31])
    cols = list(df.columns)
    scrivi_titolo(ws, titolo, len(cols))
    scrivi_intestazioni(ws, cols)
    color_toggle = 0
    prev_grp = None
    last_row = 2
    for r_idx, row_data in enumerate(df.itertuples(index=False), 3):
        grp_idx = cols.index(col_gruppo) if col_gruppo in cols else 0
        current_grp = row_data[grp_idx]
        if current_grp != prev_grp:
            color_toggle = 1 - color_toggle
            prev_grp = current_grp
        scrivi_riga_dati(
            ws, r_idx, list(row_data),
            FILL_A if color_toggle else FILL_B,
        )
        last_row = r_idx

    # Riga totale
    if not df.empty:
        tot_row = last_row + 1
        totali = []
        for c in cols:
            if c == col_gruppo:
                totali.append('TOTALE')
            elif df[c].dtype in ('int64', 'float64'):
                totali.append(int(df[c].sum()))
            else:
                try:
                    s = pd.to_numeric(df[c], errors='coerce').sum()
                    totali.append(int(s) if pd.notna(s) and s != 0 else '')
                except Exception:
                    totali.append('')
        scrivi_riga_totale(ws, tot_row, totali)

    auto_larghezza_colonne(ws, cols)


# ─────────────────────────────────────────────────────────────
# CALCOLO DATI (logica pura, nessuna scrittura Excel)
# ─────────────────────────────────────────────────────────────

def _calcola_report_atto_medici(personale_file, pensionamenti_file,
                                 mapper_atto_aziendale, anno_analisi):
    """Calcola i dati del report medici e ritorna (df_atto, df_fuori)."""

    print(f"\n{'=' * 70}")
    print("REPORT MEDICI - ATTO AZIENDALE vs PERSONALE IN SERVIZIO")
    print(f"{'=' * 70}\n")

    # Carica personale
    personale_df = carica_dataframe(personale_file)
    personale_df = normalizza_colonne_personale(personale_df)

    # Esclusione personale già cessato alla data odierna
    _oggi = date.today()
    if 'DT_CESSAZIONE' in personale_df.columns:
        _dt_cess = pd.to_datetime(
            personale_df['DT_CESSAZIONE'], errors='coerce'
        )
        _mask_cessati = _dt_cess.notna() & (_dt_cess.dt.date <= _oggi)
        _n_cessati = int(_mask_cessati.sum())
        if _n_cessati:
            print(f"  Esclusi {_n_cessati} dipendenti già cessati "
                  f"(DT_CESSAZIONE <= {_oggi:%d/%m/%Y})")
            personale_df = personale_df[~_mask_cessati].copy()

    # Tutti i dirigenti medici + veterinari (qualsiasi natura)
    medici_all = personale_df[
        personale_df['PROFILO_RAGGRUPPATO'].str.upper().isin(
            ['DIRIGENTE MEDICO', 'DIRIGENTE VETERINARIO']
        )
    ].copy()
    medici_all['DISC_UPPER'] = (
        medici_all['DESC_DISCIPLINE'].str.upper().str.strip()
    )
    # Medici senza disciplina specificata → catch-all "Dirigente medico"
    medici_all['DISC_UPPER'] = medici_all['DISC_UPPER'].fillna('DIRIGENTE MEDICO')
    natura_upper = medici_all['DESC_NATURA'].str.upper()

    # Classificazione per tipologia contrattuale
    mask_ti_puro = natura_upper == 'TEMPO INDETERMINATO'
    mask_cmd = natura_upper.str.startswith('COMANDATO IN USCITA')
    mask_penit = natura_upper == 'PENITENZIARIO INDETERMINATO'
    mask_asp = natura_upper.str.startswith('TEMPO INDETERMINATO ASP')
    mask_ti = mask_ti_puro | mask_cmd | mask_penit | mask_asp
    mask_td_puro = natura_upper.isin([
        'TEMPO DETERMINATO',
        'TEMPO DETERMINATO ART. 15 SEPTIES DLGS 502/92',
    ])
    mask_octies = natura_upper.str.contains('15_OCTIES', na=False)
    mask_td = mask_td_puro | mask_octies
    mask_univ = natura_upper.isin(['UNIVERSITARI H19', 'T.D. SPECIALIZZANDI'])

    medici_ti = medici_all[mask_ti]

    # Pensionamenti (su tutto il personale)
    pensionamenti_df = carica_dataframe(pensionamenti_file)
    pensionamenti_df = normalizza_colonne_pensionamenti(pensionamenti_df)

    medici_all_pens = medici_all.copy()
    if 'DT_CESSAZIONE' in medici_all_pens.columns:
        medici_all_pens = medici_all_pens.rename(
            columns={'DT_CESSAZIONE': 'DT_CESSAZIONE_PERS'}
        )
    medici_all_pens = pd.merge(
        medici_all_pens,
        pensionamenti_df[['MATR.', 'DT_CESSAZIONE']],
        on='MATR.',
        how='left',
    )

    # Data uscita unificata: pensionamento OPPURE fine contratto TD
    _dt_pens_m = pd.to_datetime(
        medici_all_pens['DT_CESSAZIONE'], errors='coerce'
    )
    _dt_pers_m = pd.to_datetime(
        medici_all_pens.get('DT_CESSAZIONE_PERS'), errors='coerce'
    ) if 'DT_CESSAZIONE_PERS' in medici_all_pens.columns else pd.Series(
        pd.NaT, index=medici_all_pens.index
    )
    medici_all_pens['_DT_USCITA'] = _dt_pens_m.fillna(_dt_pers_m)

    anni_pensionamento = [anno_analisi + 1, anno_analisi + 2, anno_analisi + 3]

    # Mapper atto aziendale
    mapper = carica_medici_atto_aziendale(mapper_atto_aziendale)
    print(f"Discipline da atto aziendale: {len(mapper)}")
    print(f"Dirigenti nel DB: {len(medici_all)}  "
          f"(TI: {mask_ti.sum()} di cui Cmd: {mask_cmd.sum()}, "
          f"TD: {mask_td.sum()} di cui 15Oct: {mask_octies.sum()}, "
          f"Univ/Spec: {mask_univ.sum()})\n")

    # Costruisci il report
    righe_report = []
    discipline_db_mappate = set()

    for disc in mapper:
        nome_atto = disc['nome_atto']
        dotazione = disc['dotazione']
        voci_db = disc['discipline_db']

        m_all = medici_all[medici_all['DISC_UPPER'].isin(voci_db)]
        nat = m_all['DESC_NATURA'].str.upper()
        n_cmd = int(nat.str.startswith('COMANDATO IN USCITA').sum())
        n_td_puro = int(nat.isin([
            'TEMPO DETERMINATO',
            'TEMPO DETERMINATO ART. 15 SEPTIES DLGS 502/92',
        ]).sum())
        n_octies = int(nat.str.contains('15_OCTIES', na=False).sum())
        n_td = n_td_puro + n_octies
        n_univ = int(nat.isin([
            'UNIVERSITARI H19', 'T.D. SPECIALIZZANDI',
        ]).sum())
        n_tot = len(m_all)               # headcount effettivo
        n_ti = n_tot - n_td - n_univ     # residuale → TI
        delta = n_tot - dotazione

        # Pensionamenti e cessazioni
        m_ti_disc = medici_all_pens[medici_all_pens['DISC_UPPER'].isin(voci_db)]
        dt_cess = m_ti_disc['_DT_USCITA']
        pens = {}
        for anno in anni_pensionamento:
            pens[anno] = int((dt_cess.dt.year == anno).sum())

        discipline_db_mappate.update(voci_db)

        riga = {
            'DISCIPLINA_ATTO_AZIENDALE': nome_atto,
            'DOTAZIONE_ATTO': dotazione,
            'TI': n_ti,
            'CMD_TI': n_cmd,
            'TD': n_td,
            'OCTIES_TD': n_octies,
            'UNIV_SPEC': n_univ,
            'TOTALE': n_tot,
            'DELTA': delta,
        }
        for anno in anni_pensionamento:
            riga[f'PENSIONAMENTI_{anno}'] = pens[anno]

        pens_cumulati = 0
        for anno in anni_pensionamento:
            pens_cumulati += pens[anno]
            riga[f'PROIEZIONE_{anno}'] = n_tot - pens_cumulati

        righe_report.append(riga)

        stato = "OK" if delta >= 0 else f"CARENZA {abs(delta)}"
        cmd_str = f"(cmd:{n_cmd})" if n_cmd else ""
        oct_str = f"(oct:{n_octies})" if n_octies else ""
        print(
            f"  {nome_atto:45s}  Atto: {dotazione:3d}  "
            f"TI: {n_ti:3d}{cmd_str:>7s}  TD: {n_td:2d}{oct_str:>7s}  "
            f"U/S: {n_univ:2d}  Tot: {n_tot:3d}  Delta: {delta:+4d}  [{stato}]"
        )

    # Discipline non mappate
    tutte_disc_db = set(medici_all['DISC_UPPER'].dropna().unique())
    non_mappate = tutte_disc_db - discipline_db_mappate
    medici_non_mappati = medici_all[medici_all['DISC_UPPER'].isin(non_mappate)]

    if non_mappate:
        print(
            f"\n  --- Discipline fuori dall'atto aziendale "
            f"({len(medici_non_mappati)} medici) ---"
        )
        for disc_nm in sorted(non_mappate):
            m_nm = medici_all[medici_all['DISC_UPPER'] == disc_nm]
            nat_nm = m_nm['DESC_NATURA'].str.upper()
            n_cmd = int(nat_nm.str.startswith('COMANDATO IN USCITA').sum())
            n_td_puro = int(nat_nm.isin([
                'TEMPO DETERMINATO',
                'TEMPO DETERMINATO ART. 15 SEPTIES DLGS 502/92',
            ]).sum())
            n_octies = int(nat_nm.str.contains('15_OCTIES', na=False).sum())
            n_td = n_td_puro + n_octies
            n_univ = int(nat_nm.isin([
                'UNIVERSITARI H19', 'T.D. SPECIALIZZANDI',
            ]).sum())
            n_tot = len(m_nm)
            n_ti = n_tot - n_td - n_univ

            m_ti_nm = medici_all_pens[medici_all_pens['DISC_UPPER'] == disc_nm]
            dt_cess_nm = m_ti_nm['_DT_USCITA']
            pens_nm = {}
            for anno in anni_pensionamento:
                pens_nm[anno] = int((dt_cess_nm.dt.year == anno).sum())

            riga = {
                'DISCIPLINA_ATTO_AZIENDALE': f'[FUORI ATTO] {disc_nm}',
                'DOTAZIONE_ATTO': '-',
                'TI': n_ti,
                'CMD_TI': n_cmd,
                'TD': n_td,
                'OCTIES_TD': n_octies,
                'UNIV_SPEC': n_univ,
                'TOTALE': n_tot,
                'DELTA': '-',
            }
            for anno in anni_pensionamento:
                riga[f'PENSIONAMENTI_{anno}'] = pens_nm[anno]
            pens_cum = 0
            for anno in anni_pensionamento:
                pens_cum += pens_nm[anno]
                riga[f'PROIEZIONE_{anno}'] = n_tot - pens_cum
            righe_report.append(riga)

            cmd_str = f"(cmd:{n_cmd})" if n_cmd else ""
            oct_str = f"(oct:{n_octies})" if n_octies else ""
            print(f"  {disc_nm:45s}  TI: {n_ti:3d}{cmd_str:>7s}  TD: {n_td:2d}{oct_str:>7s}  U/S: {n_univ:2d}")

    # Totali
    tot_dot = sum(r['DOTAZIONE_ATTO'] for r in righe_report if isinstance(r['DOTAZIONE_ATTO'], int))
    tot_ti = sum(r['TI'] for r in righe_report)
    tot_cmd = sum(r['CMD_TI'] for r in righe_report)
    tot_td = sum(r['TD'] for r in righe_report)
    tot_oct = sum(r['OCTIES_TD'] for r in righe_report)
    tot_us = sum(r['UNIV_SPEC'] for r in righe_report)
    tot_all = tot_ti + tot_td + tot_us
    cmd_str = f"(cmd:{tot_cmd})" if tot_cmd else ""
    oct_str = f"(oct:{tot_oct})" if tot_oct else ""
    print(
        f"\n  {'TOTALE':45s}  Atto: {tot_dot:3d}  "
        f"TI: {tot_ti:3d}{cmd_str:>7s}  TD: {tot_td:2d}{oct_str:>7s}  "
        f"U/S: {tot_us:2d}  Tot: {tot_all:3d}  Delta: {tot_all - tot_dot:+4d}"
    )
    print(f"  Medici in discipline fuori atto: {len(medici_non_mappati)}")
    print(f"  Totale dirigenti nel DB: {len(medici_all)}")

    # Rinomina colonne
    def rinomina_colonne(df):
        rename = {
            'DISCIPLINA_ATTO_AZIENDALE': 'Disciplina',
            'DOTAZIONE_ATTO': 'Dotazione Atto',
            'TI': 'Tempo Indeterminato',
            'CMD_TI': 'di cui Cmd. Uscita',
            'TD': 'Tempo Determinato',
            'OCTIES_TD': 'di cui 15 Octies',
            'UNIV_SPEC': 'Università e Specializzandi',
            'TOTALE': 'Totale',
            'DELTA': 'Delta',
        }
        for anno in anni_pensionamento:
            rename[f'PENSIONAMENTI_{anno}'] = f'Pens. e cessazioni {anno}'
            rename[f'PROIEZIONE_{anno}'] = f'Proiezione {anno}'
        return df.rename(columns=rename)

    df_atto = pd.DataFrame(
        [r for r in righe_report
         if not str(r.get('DISCIPLINA_ATTO_AZIENDALE', '')).startswith(
             '[FUORI ATTO]'
         )]
    )
    df_fuori = pd.DataFrame(
        [r for r in righe_report
         if str(r.get('DISCIPLINA_ATTO_AZIENDALE', '')).startswith(
             '[FUORI ATTO]'
         )]
    )
    if not df_fuori.empty:
        df_fuori = df_fuori.copy()
        df_fuori['DISCIPLINA_ATTO_AZIENDALE'] = (
            df_fuori['DISCIPLINA_ATTO_AZIENDALE']
            .str.replace(r'^\[FUORI ATTO\] ', '', regex=True)
        )

    df_atto = rinomina_colonne(df_atto)
    df_fuori = rinomina_colonne(df_fuori).drop(
        columns=['Dotazione Atto', 'Delta'], errors='ignore'
    )

    return df_atto, df_fuori


# ─────────────────────────────────────────────────────────────
# FUNZIONE PUBBLICA – report standalone (file XLSX separato)
# ─────────────────────────────────────────────────────────────

def genera_report_atto_aziendale(personale_file, pensionamenti_file,
                                  mapper_atto_aziendale, output_file,
                                  anno_analisi):
    """
    Genera un report XLSX standalone della dirigenza medica confrontando
    la dotazione da atto aziendale con il personale effettivo.
    """
    df_atto, df_fuori = _calcola_report_atto_medici(
        personale_file, pensionamenti_file, mapper_atto_aziendale,
        anno_analisi,
    )

    wb = Workbook()
    wb.remove(wb.active)
    _scrivi_foglio_atto(
        wb, 'Requisiti Atto Aziendale', df_atto,
        'Personale Medico - Requisiti da Atto Aziendale',
    )
    if not df_fuori.empty:
        _scrivi_foglio_atto(
            wb, 'Fuori Atto Aziendale', df_fuori,
            'Personale Medico - Discipline Fuori Atto Aziendale',
        )
    wb.save(output_file)

    report_df = pd.concat([df_atto, df_fuori], ignore_index=True)
    print(f"\n  Report salvato in: {output_file}")
    print(f"{'=' * 70}\n")
    return report_df


# ─────────────────────────────────────────────────────────────
# FUNZIONE PUBBLICA – foglio nel riepilogo aziendale
# ─────────────────────────────────────────────────────────────

def scrivi_foglio_riepilogo_atto_medici(wb, personale_file,
                                         pensionamenti_file,
                                         mapper_atto_aziendale,
                                         anno_analisi):
    """Aggiunge il foglio 'FABBISOGNO ATTO MEDICI' al workbook
    del riepilogo aziendale. Combina requisiti e discipline fuori
    atto in un unico foglio."""
    df_atto, df_fuori = _calcola_report_atto_medici(
        personale_file, pensionamenti_file, mapper_atto_aziendale,
        anno_analisi,
    )

    ws = wb.create_sheet(title='FABBISOGNO ATTO MEDICI')

    # ── Sezione 1: Requisiti da Atto Aziendale ────────────────
    cols_atto = list(df_atto.columns)
    scrivi_titolo(
        ws,
        f'Personale Medico – Atto Aziendale vs Personale ({anno_analisi})',
        len(cols_atto),
    )
    scrivi_intestazioni(ws, cols_atto)

    row = 3
    color_toggle = 0
    prev_grp = None
    for row_data in df_atto.itertuples(index=False):
        current_grp = row_data[0]
        if current_grp != prev_grp:
            color_toggle = 1 - color_toggle
            prev_grp = current_grp
        scrivi_riga_dati(
            ws, row, list(row_data),
            FILL_A if color_toggle else FILL_B,
        )
        row += 1

    # Riga totale atto
    if not df_atto.empty:
        totali = []
        for c in cols_atto:
            if c == 'Disciplina':
                totali.append('TOTALE ATTO')
            elif df_atto[c].dtype in ('int64', 'float64'):
                totali.append(int(df_atto[c].sum()))
            else:
                try:
                    s = pd.to_numeric(df_atto[c], errors='coerce').sum()
                    totali.append(int(s) if pd.notna(s) and s != 0 else '')
                except Exception:
                    totali.append('')
        scrivi_riga_totale(ws, row, totali)
        row += 1

    # ── Sezione 2: Discipline fuori Atto ──────────────────────
    if not df_fuori.empty:
        row += 1  # riga vuota di separazione

        cols_fuori = list(df_fuori.columns)
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=len(cols_fuori),
        )
        cell = ws.cell(
            row=row, column=1,
            value='DISCIPLINE FUORI ATTO AZIENDALE',
        )
        cell.font = Font(bold=True, size=12)
        cell.alignment = ALIGN_CENTER
        row += 1

        scrivi_intestazioni(ws, cols_fuori, riga=row)
        row += 1

        color_toggle = 0
        prev_grp = None
        for row_data in df_fuori.itertuples(index=False):
            current_grp = row_data[0]
            if current_grp != prev_grp:
                color_toggle = 1 - color_toggle
                prev_grp = current_grp
            scrivi_riga_dati(
                ws, row, list(row_data),
                FILL_A if color_toggle else FILL_B,
            )
            row += 1

        # Riga totale fuori atto
        totali_f = []
        for c in cols_fuori:
            if c == 'Disciplina':
                totali_f.append('TOTALE FUORI ATTO')
            elif df_fuori[c].dtype in ('int64', 'float64'):
                totali_f.append(int(df_fuori[c].sum()))
            else:
                try:
                    s = pd.to_numeric(df_fuori[c], errors='coerce').sum()
                    totali_f.append(
                        int(s) if pd.notna(s) and s != 0 else ''
                    )
                except Exception:
                    totali_f.append('')
        scrivi_riga_totale(ws, row, totali_f)

    auto_larghezza_colonne(ws, cols_atto)
