"""
Foglio riepilogo Fabbisogno Teorico aziendale.

Genera un foglio 'FABBISOGNO TEORICO' nel workbook del riepilogo aziendale
che consolida, per ogni presidio ospedaliero, la situazione reparto per
reparto del personale in servizio rispetto al fabbisogno teorico calcolato
in base ai posti letto e all'intensità di cura.

Per ciascun presidio viene scritta una tabella di dettaglio con una riga
per ogni combinazione reparto × profilo professionale (solo quelli con
fabbisogno numerico validato), seguita da un subtotale per profilo.

In calce al foglio, una tabella TOTALE AZIENDALE aggrega i dati di
tutti i presidi.
"""

from openpyxl.styles import Font, PatternFill

from src.stili_excel import (
    THIN_BORDER, FILL_A, FILL_B, FILL_HEADER,
    FONT_HEADER, ALIGN_CENTER,
    auto_larghezza_colonne,
)


# ============================================================
# STILI
# ============================================================

_FILL_DIVIDER = PatternFill(
    start_color='1F4E79', end_color='1F4E79', fill_type='solid')
_FONT_DIVIDER = Font(bold=True, size=14, color='FFFFFF')

_FILL_SUBTOTAL = PatternFill(
    start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
_FONT_SUBTOTAL = Font(bold=True, size=10)

FILL_CARENZA = PatternFill(
    start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
FILL_OK = PatternFill(
    start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
FONT_NOTE = Font(italic=True, size=9, color='555555')

# Colonne tabella dettaglio (reparto × profilo)
_HEADERS_DETTAGLIO = [
    'Reparto', 'Intensità', 'Profilo Professionale',
    'T.I.', 'T.D.', 'Totale', 'Fabb. Teorico', 'Delta',
]
_N_COLS_DET = len(_HEADERS_DETTAGLIO)  # 8

# Colonne tabella riepilogo per profilo
_HEADERS_RIEPILOGO = [
    'Profilo Professionale', 'T.I.', 'T.D.',
    'Totale', 'Fabb. Teorico', 'Delta',
]
_N_COLS_RIEP = len(_HEADERS_RIEPILOGO)  # 6


# ============================================================
# HELPER
# ============================================================

def _is_numeric(val):
    """True se il valore è convertibile a int."""
    try:
        int(val)
        return True
    except (ValueError, TypeError):
        return False


def _fill_delta(delta):
    """FILL_CARENZA se delta < 0, FILL_OK altrimenti."""
    return FILL_CARENZA if delta < 0 else FILL_OK


def _scrivi_divisore(ws, row, titolo):
    """Barra colorata larga _N_COLS_DET colonne come separatore."""
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=_N_COLS_DET)
    c = ws.cell(row=row, column=1, value=titolo)
    c.font = _FONT_DIVIDER
    c.fill = _FILL_DIVIDER
    c.alignment = ALIGN_CENTER
    c.border = THIN_BORDER
    ws.row_dimensions[row].height = 30
    return row + 2


# ============================================================
# TABELLA DETTAGLIO REPARTO × PROFILO
# ============================================================

def _scrivi_tabella_dettaglio(ws, start_row, df_presidio):
    """Scrive intestazioni + righe dettaglio (reparto × profilo).

    Alterna colore per blocco reparto.

    Returns
    -------
    (next_row, totali_per_profilo)
        totali_per_profilo: {profilo: {'ti':, 'td':, 'tot':, 'fabb':}}
    """
    row = start_row

    # Intestazioni
    for ci, h in enumerate(_HEADERS_DETTAGLIO, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 30
    row += 1

    # ------------------------------------------------------------------
    # Aggrega per (reparto, profilo) — uno stesso reparto può avere più
    # CDC, tutti con lo stesso fabbisogno teorico; le quantità vanno
    # sommate, ma il fabbisogno va preso una sola volta.
    # ------------------------------------------------------------------
    agg = {}  # {(reparto, profilo): {ti, td, tot, fabb, intensita}}
    for _, r in df_presidio.iterrows():
        reparto = r['_REPARTO']
        profilo = r['Profilo Professionale']
        k = (reparto, profilo)
        if k not in agg:
            agg[k] = {
                'ti': 0, 'td': 0, 'tot': 0,
                'fabb': int(r['Fabbisogno Teorico']),
                'intensita': r.get('Intensità', ''),
            }
        agg[k]['ti'] += int(r.get('Quantità T.I.', 0))
        agg[k]['td'] += int(r.get('Quantità T.D.', 0))
        agg[k]['tot'] += int(r.get('Totale', 0))

    # Ordina per reparto e profilo
    righe = sorted(agg.items(), key=lambda x: (x[0][0], x[0][1]))

    prev_reparto = None
    toggle = 0
    totali = {}

    for (reparto, profilo), d in righe:
        ti = d['ti']
        td = d['td']
        tot = d['tot']
        fabb = d['fabb']
        intensita = d['intensita']
        delta = tot - fabb

        if reparto != prev_reparto:
            toggle = 1 - toggle
            prev_reparto = reparto
        fill = FILL_A if toggle else FILL_B

        # Accumula totali per profilo
        if profilo not in totali:
            totali[profilo] = {'ti': 0, 'td': 0, 'tot': 0, 'fabb': 0}
        totali[profilo]['ti'] += ti
        totali[profilo]['td'] += td
        totali[profilo]['tot'] += tot
        totali[profilo]['fabb'] += fabb

        vals = [reparto, intensita, profilo, ti, td, tot, fabb, delta]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
            if ci >= 4:
                c.alignment = ALIGN_CENTER

        # Colore cella delta
        ws.cell(row=row, column=_N_COLS_DET).fill = _fill_delta(delta)
        row += 1

    return row, totali


# ============================================================
# TABELLA RIEPILOGO PER PROFILO (subtotale / totale aziendale)
# ============================================================

def _scrivi_riepilogo_profilo(ws, start_row, titolo, totali):
    """Scrive tabella riassuntiva per profilo con riga TOTALE."""
    if not totali:
        return start_row

    row = start_row + 1

    # Titolo sezione
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=_N_COLS_RIEP)
    ws.cell(row=row, column=1, value=titolo).font = FONT_SECTION
    row += 1

    # Intestazioni
    for ci, h in enumerate(_HEADERS_RIEPILOGO, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # Righe per profilo
    profili_ordine = sorted(totali.keys())
    toggle = 0
    for profilo in profili_ordine:
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1
        d = totali[profilo]
        delta = d['tot'] - d['fabb']

        vals = [profilo, d['ti'], d['td'], d['tot'], d['fabb'], delta]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
            if ci >= 2:
                c.alignment = ALIGN_CENTER

        ws.cell(row=row, column=_N_COLS_RIEP).fill = _fill_delta(delta)
        row += 1

    # Riga TOTALE complessiva
    tot_ti = sum(d['ti'] for d in totali.values())
    tot_td = sum(d['td'] for d in totali.values())
    tot_tot = sum(d['tot'] for d in totali.values())
    tot_fabb = sum(d['fabb'] for d in totali.values())
    tot_delta = tot_tot - tot_fabb

    vals = ['TOTALE', tot_ti, tot_td, tot_tot, tot_fabb, tot_delta]
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.fill = _FILL_SUBTOTAL
        c.font = _FONT_SUBTOTAL
        c.border = THIN_BORDER
        if ci >= 2:
            c.alignment = ALIGN_CENTER

    delta_cell = ws.cell(row=row, column=_N_COLS_RIEP)
    delta_cell.fill = _fill_delta(tot_delta)
    delta_cell.font = _FONT_SUBTOTAL
    row += 1

    return row


# ============================================================
# FUNZIONE PRINCIPALE
# ============================================================

def scrivi_foglio_riepilogo_fabbisogno_teorico(wb, grouped,
                                                livello_presidio):
    """Crea il foglio 'FABBISOGNO TEORICO' nel workbook del riepilogo
    aziendale.

    Per ogni presidio ospedaliero mostra, reparto per reparto:

    * personale in servizio (T.I., T.D., Totale)
    * fabbisogno teorico calcolato (posti letto × intensità)
    * delta (Totale − Fabbisogno Teorico)

    In calce, una tabella TOTALE AZIENDALE che aggrega tutti i presidi.
    """

    _AREE_SPECIALI = {'(Non assegnata)', 'LUNGHE ASSENZE',
                      'IN ATTESA DI ASSEGNAZIONE'}

    ws = wb.create_sheet(title='FABBISOGNO TEORICO')

    # Titolo principale
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=_N_COLS_DET)
    c = ws.cell(row=1, column=1,
                value="RIEPILOGO FABBISOGNO TEORICO - LIVELLO AZIENDALE")
    c.font = Font(bold=True, size=14, color='1F4E79')
    c.alignment = ALIGN_CENTER
    row = 3

    # Nota esplicativa
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=_N_COLS_DET)
    ws.cell(
        row=row, column=1,
        value=(
            "Fabbisogno basato su posti letto e intensità di cura. "
            "Inclusi solo reparti ospedalieri con indicatori validati "
            "(Dirigenti Medici, Infermieri, OSS). "
            "La metodologia non tiene in considerazione le attività ambulatoriali, "
            "le guardie interdivisionali, le limitazioni ex L. 104/92, "
            "le idoneità parziali o altri compiti trasversali: "
            "commisura esclusivamente i posti letto al personale necessario "
            "per gestirli. "
            "Per tale ragione il fabbisogno è sempre arrotondato per eccesso "
            "e per ciascun servizio è garantito un minimo di 2 unità "
            "per profilo al fine di assicurare la copertura di ferie "
            "e indisponibilità. "
            "Delta = Totale in servizio - Fabbisogno Teorico "
            "(valori negativi = carenza)."
        ),
    ).font = FONT_NOTE
    row += 2

    # ------------------------------------------------------------------
    # Filtra righe con fabbisogno numerico, escludendo Pronto Soccorso
    # (il PS ha pochi PL ma molto personale: il fabbisogno teorico
    # basato su posti letto non è significativo per questo servizio)
    # ------------------------------------------------------------------
    _PAT_PS = r'PRONTO SOCCORSO'
    df_num = grouped[
        grouped['Fabbisogno Teorico'].apply(_is_numeric)
        & ~grouped['_REPARTO'].str.contains(_PAT_PS, case=False, na=False)
    ].copy()

    # ------------------------------------------------------------------
    # Costruisco lista presidi ospedalieri
    # ------------------------------------------------------------------
    cities_data = []
    for citta, df_citta in df_num.groupby('_CITTA', sort=True):
        if citta in _AREE_SPECIALI:
            continue

        # Solo sedi ospedaliere (P.O.)
        df_po = df_citta[
            df_citta['_LUOGO'].str.contains('P.O.', na=False)
        ]
        if df_po.empty:
            continue

        presidio = None
        for pn in (livello_presidio or {}):
            if citta.upper() in pn.upper():
                presidio = pn
                break

        cities_data.append({
            'citta': citta,
            'df': df_po,
            'presidio': presidio,
        })

    # ------------------------------------------------------------------
    # Per ogni presidio: dettaglio + subtotale
    # ------------------------------------------------------------------
    grand_totali = {}

    for cd in cities_data:
        label = cd['presidio'] or cd['citta']

        # Divisore presidio
        row = _scrivi_divisore(ws, row, label)

        # Tabella dettaglio reparto × profilo
        row, totali_presidio = _scrivi_tabella_dettaglio(
            ws, row, cd['df'])

        # Subtotale per presidio
        row = _scrivi_riepilogo_profilo(
            ws, row, f"Subtotale - {label}", totali_presidio)
        row += 2  # spazio tra sezioni

        # Accumula grand totali
        for profilo, vals in totali_presidio.items():
            if profilo not in grand_totali:
                grand_totali[profilo] = {
                    'ti': 0, 'td': 0, 'tot': 0, 'fabb': 0}
            for k in ('ti', 'td', 'tot', 'fabb'):
                grand_totali[profilo][k] += vals[k]

    # ------------------------------------------------------------------
    # TOTALE AZIENDALE
    # ------------------------------------------------------------------
    if grand_totali:
        row = _scrivi_divisore(ws, row, "TOTALE AZIENDALE")
        row = _scrivi_riepilogo_profilo(
            ws, row - 1,
            "FABBISOGNO TEORICO - TOTALE AZIENDALE",
            grand_totali)

    # ------------------------------------------------------------------
    # Larghezza colonne automatica
    # ------------------------------------------------------------------
    auto_larghezza_colonne(ws)

    return ws
