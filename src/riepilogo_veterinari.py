"""
Foglio riepilogo VETERINARI per il file riepilogo_aziendale.xlsx.

Genera un foglio 'VETERINARI' che mostra lo stato attuale del personale
veterinario raggruppato per Area (A, B, C) e per distretto.

Le tre aree corrispondono ai reparti nel DB personale:
  Area A – Sanità animale
  Area B – Igiene prod., trasf., comm. alimenti orig. animale
  Area C – Igiene allevamenti e produzioni zootecniche
"""

import re as _re

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.stili_excel import (
    THIN_BORDER, FILL_A, FILL_B, FILL_HEADER,
    FONT_HEADER, ALIGN_CENTER,
    auto_larghezza_colonne,
)

# ============================================================
# COSTANTI DI STILE
# ============================================================

_FILL_DIVIDER = PatternFill(
    start_color='1F4E79', end_color='1F4E79', fill_type='solid')
_FONT_DIVIDER = Font(bold=True, size=14, color='FFFFFF')

_FILL_SUBTOTAL = PatternFill(
    start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
_FONT_SUBTOTAL = Font(bold=True, size=10)

_FILL_GRAND = PatternFill(
    start_color='4472C4', end_color='4472C4', fill_type='solid')
_FONT_GRAND = Font(bold=True, size=11, color='FFFFFF')

FONT_SECTION = Font(bold=True, size=12, color='1F4E79')

_ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
_ALIGN_LEFT = Alignment(horizontal='left', vertical='center')

# Mapping area → etichetta breve
_AREA_MAP = [
    ('AREA A', 'Area A – Sanità animale'),
    ('AREA B', 'Area B – Igiene alimenti orig. animale'),
    ('AREA C', 'Area C – Igiene allevamenti e prod. zootecniche'),
]

# Intestazioni tabella
_HEADERS = ['Sede', 'Area', 'T.I.', 'T.D.', 'Totale']
_N_COLS = len(_HEADERS)


# ============================================================
# UTILITÀ
# ============================================================

def _classifica_area(reparto):
    """Restituisce l'etichetta area da DESC_SC_SSD_SS / _REPARTO."""
    s = str(reparto).upper()
    for codice, etichetta in _AREA_MAP:
        if codice in s:
            return etichetta
    return 'Altra area'


def _cella(ws, row, col, value, font=None, fill=None, alignment=None,
           border=None, number_format=None):
    """Scrive una cella con stile."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


# ============================================================
# SCRITTURA BARRA DIVISORIO (nome distretto)
# ============================================================

def _scrivi_divisorio(ws, row, titolo, n_cols):
    """Barra colorata larga n_cols con il titolo del distretto."""
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _FILL_DIVIDER
        cell.border = THIN_BORDER
    cell_title = ws.cell(row=row, column=1, value=titolo)
    cell_title.font = _FONT_DIVIDER
    cell_title.alignment = _ALIGN_LEFT
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=n_cols
    )
    return row + 1


# ============================================================
# SCRITTURA HEADER TABELLA
# ============================================================

def _scrivi_header(ws, row):
    """Scrive la riga di intestazione della tabella."""
    for c, hdr in enumerate(_HEADERS, start=1):
        _cella(ws, row, c, hdr,
               font=FONT_HEADER, fill=FILL_HEADER,
               alignment=ALIGN_CENTER, border=THIN_BORDER)
    return row + 1


# ============================================================
# TABELLA DETTAGLIO PER DISTRETTO
# ============================================================

def _scrivi_tabella_distretto(ws, start_row, df_distretto):
    """
    Scrive le righe di dettaglio di un distretto:
    una riga per sede × area, poi un subtotale.

    Restituisce (next_row, dict_totali_per_area).
    """
    r = start_row
    totali_area = {}  # {etichetta_area: (ti, td, tot)}

    # Raggruppa per sede e area
    for _, grp_row in df_distretto.iterrows():
        sede = grp_row['_sede_breve']
        area = grp_row['_area']
        ti = int(grp_row['_ti'])
        td = int(grp_row['_td'])
        tot = ti + td

        # Accumula totali per area
        prev = totali_area.get(area, (0, 0, 0))
        totali_area[area] = (prev[0] + ti, prev[1] + td, prev[2] + tot)

        fill = FILL_A if (r - start_row) % 2 == 0 else FILL_B
        _cella(ws, r, 1, sede, fill=fill, border=THIN_BORDER,
               alignment=_ALIGN_LEFT)
        _cella(ws, r, 2, area, fill=fill, border=THIN_BORDER,
               alignment=_ALIGN_LEFT)
        _cella(ws, r, 3, ti, fill=fill, border=THIN_BORDER,
               alignment=_ALIGN_CENTER)
        _cella(ws, r, 4, td, fill=fill, border=THIN_BORDER,
               alignment=_ALIGN_CENTER)
        _cella(ws, r, 5, tot, fill=fill, border=THIN_BORDER,
               alignment=_ALIGN_CENTER)
        r += 1

    # Subtotale distretto
    sum_ti = sum(v[0] for v in totali_area.values())
    sum_td = sum(v[1] for v in totali_area.values())
    sum_tot = sum(v[2] for v in totali_area.values())

    _cella(ws, r, 1, 'Subtotale', font=_FONT_SUBTOTAL,
           fill=_FILL_SUBTOTAL, border=THIN_BORDER, alignment=_ALIGN_LEFT)
    _cella(ws, r, 2, '', fill=_FILL_SUBTOTAL, border=THIN_BORDER)
    _cella(ws, r, 3, sum_ti, font=_FONT_SUBTOTAL,
           fill=_FILL_SUBTOTAL, border=THIN_BORDER, alignment=_ALIGN_CENTER)
    _cella(ws, r, 4, sum_td, font=_FONT_SUBTOTAL,
           fill=_FILL_SUBTOTAL, border=THIN_BORDER, alignment=_ALIGN_CENTER)
    _cella(ws, r, 5, sum_tot, font=_FONT_SUBTOTAL,
           fill=_FILL_SUBTOTAL, border=THIN_BORDER, alignment=_ALIGN_CENTER)
    r += 1

    return r, totali_area


# ============================================================
# TABELLA RIEPILOGO PER AREA (AZIENDALE)
# ============================================================

def _scrivi_riepilogo_aree(ws, start_row, totali_globali):
    """
    Scrive la tabella riepilogativa aziendale per area.
    totali_globali: {etichetta_area: (ti, td, tot)}
    """
    r = start_row + 1  # riga vuota di separazione

    # Titolo
    _cella(ws, r, 1, 'TOTALE AZIENDALE PER AREA VETERINARIA',
           font=FONT_SECTION)
    r += 2

    # Header
    headers = ['Area', 'T.I.', 'T.D.', 'Totale']
    for c, hdr in enumerate(headers, start=1):
        _cella(ws, r, c, hdr,
               font=FONT_HEADER, fill=FILL_HEADER,
               alignment=ALIGN_CENTER, border=THIN_BORDER)
    r += 1

    grand_ti = grand_td = grand_tot = 0
    idx = 0
    for _, etichetta in _AREA_MAP:
        ti, td, tot = totali_globali.get(etichetta, (0, 0, 0))
        grand_ti += ti
        grand_td += td
        grand_tot += tot

        fill = FILL_A if idx % 2 == 0 else FILL_B
        _cella(ws, r, 1, etichetta, fill=fill, border=THIN_BORDER,
               alignment=_ALIGN_LEFT)
        _cella(ws, r, 2, ti, fill=fill, border=THIN_BORDER,
               alignment=_ALIGN_CENTER)
        _cella(ws, r, 3, td, fill=fill, border=THIN_BORDER,
               alignment=_ALIGN_CENTER)
        _cella(ws, r, 4, tot, fill=fill, border=THIN_BORDER,
               alignment=_ALIGN_CENTER)
        r += 1
        idx += 1

    # Riga Altra area (se presente)
    for area_key, vals in totali_globali.items():
        if area_key not in [e for _, e in _AREA_MAP]:
            ti, td, tot = vals
            grand_ti += ti
            grand_td += td
            grand_tot += tot
            fill = FILL_A if idx % 2 == 0 else FILL_B
            _cella(ws, r, 1, area_key, fill=fill, border=THIN_BORDER,
                   alignment=_ALIGN_LEFT)
            _cella(ws, r, 2, ti, fill=fill, border=THIN_BORDER,
                   alignment=_ALIGN_CENTER)
            _cella(ws, r, 3, td, fill=fill, border=THIN_BORDER,
                   alignment=_ALIGN_CENTER)
            _cella(ws, r, 4, tot, fill=fill, border=THIN_BORDER,
                   alignment=_ALIGN_CENTER)
            r += 1
            idx += 1

    # Grand total
    _cella(ws, r, 1, 'TOTALE', font=_FONT_GRAND,
           fill=_FILL_GRAND, border=THIN_BORDER, alignment=_ALIGN_LEFT)
    _cella(ws, r, 2, grand_ti, font=_FONT_GRAND,
           fill=_FILL_GRAND, border=THIN_BORDER, alignment=_ALIGN_CENTER)
    _cella(ws, r, 3, grand_td, font=_FONT_GRAND,
           fill=_FILL_GRAND, border=THIN_BORDER, alignment=_ALIGN_CENTER)
    _cella(ws, r, 4, grand_tot, font=_FONT_GRAND,
           fill=_FILL_GRAND, border=THIN_BORDER, alignment=_ALIGN_CENTER)
    r += 1

    return r


# ============================================================
# TABELLA RIEPILOGO PER DISTRETTO (CROSS-TAB)
# ============================================================

def _scrivi_riepilogo_distretti(ws, start_row, totali_per_distretto):
    """
    Scrive una cross-tab distretto × area.
    totali_per_distretto: {nome_distretto: {etichetta_area: (ti, td, tot)}}
    """
    r = start_row + 1

    _cella(ws, r, 1, 'RIEPILOGO PER DISTRETTO',
           font=FONT_SECTION)
    r += 2

    # Colonne: Distretto | Area A TI | TD | Tot | Area B TI | TD | Tot | ...
    # Semplifichiamo: Distretto | Area A | Area B | Area C | Totale
    area_labels = [e for _, e in _AREA_MAP]
    headers = ['Distretto'] + [f'Area {chr(65+i)}' for i in range(len(area_labels))] + ['Totale']
    for c, hdr in enumerate(headers, start=1):
        _cella(ws, r, c, hdr,
               font=FONT_HEADER, fill=FILL_HEADER,
               alignment=ALIGN_CENTER, border=THIN_BORDER)
    r += 1

    grand = [0] * (len(area_labels) + 1)
    idx = 0
    for distretto in sorted(totali_per_distretto.keys()):
        area_totals = totali_per_distretto[distretto]
        fill = FILL_A if idx % 2 == 0 else FILL_B

        _cella(ws, r, 1, distretto, fill=fill, border=THIN_BORDER,
               alignment=_ALIGN_LEFT)
        row_tot = 0
        for ai, etichetta in enumerate(area_labels):
            _, _, tot = area_totals.get(etichetta, (0, 0, 0))
            _cella(ws, r, 2 + ai, tot, fill=fill, border=THIN_BORDER,
                   alignment=_ALIGN_CENTER)
            grand[ai] += tot
            row_tot += tot
        _cella(ws, r, 2 + len(area_labels), row_tot, fill=fill,
               border=THIN_BORDER, alignment=_ALIGN_CENTER)
        grand[-1] += row_tot
        r += 1
        idx += 1

    # Grand total
    _cella(ws, r, 1, 'TOTALE', font=_FONT_GRAND,
           fill=_FILL_GRAND, border=THIN_BORDER, alignment=_ALIGN_LEFT)
    for ci, val in enumerate(grand):
        _cella(ws, r, 2 + ci, val, font=_FONT_GRAND,
               fill=_FILL_GRAND, border=THIN_BORDER, alignment=_ALIGN_CENTER)
    r += 1

    return r


# ============================================================
# FUNZIONE PRINCIPALE
# ============================================================

def scrivi_foglio_veterinari(wb, grouped):
    """
    Aggiunge un foglio 'VETERINARI' al workbook *wb* con il
    riepilogo del personale veterinario raggruppato per area
    e per distretto.

    *grouped* è il DataFrame finale (dopo rename) con le colonne
    standard (Sede, _REPARTO, Profilo Professionale, Quantità T.I.,
    Quantità T.D., _CITTA, _LUOGO, …).
    """
    # Filtra solo veterinari
    mask = grouped['Profilo Professionale'].str.contains(
        'VETERINAR', case=False, na=False
    )
    df_vet = grouped[mask].copy()

    if df_vet.empty:
        return

    # Classifica area
    df_vet['_area'] = df_vet['_REPARTO'].apply(_classifica_area)

    # Sede breve (es. "VIA ROMA" da "CAMPOBASSO - VIA ROMA")
    df_vet['_sede_breve'] = df_vet['Sede'].apply(
        lambda s: s.split(' - ', 1)[1].strip() if ' - ' in str(s) else str(s)
    )

    # T.I. e T.D.
    df_vet['_ti'] = df_vet['Quantità T.I.'].fillna(0).astype(int)
    df_vet['_td'] = df_vet['Quantità T.D.'].fillna(0).astype(int)

    # Distretto = _CITTA (CAMPOBASSO, ISERNIA, TERMOLI, AGNONE)
    df_vet['_distretto'] = df_vet['_CITTA'].fillna('(Non assegnato)')

    # Ordina
    df_vet = df_vet.sort_values(
        ['_distretto', '_sede_breve', '_area']
    ).reset_index(drop=True)

    # Crea foglio
    ws = wb.create_sheet(title='VETERINARI')

    r = 1
    _cella(ws, r, 1, 'PERSONALE VETERINARIO – STATO ATTUALE',
           font=Font(bold=True, size=16, color='1F4E79'))
    r += 2

    # Struttura: per ogni distretto → divisorio + header + dettaglio
    distretti = sorted(df_vet['_distretto'].unique())
    totali_globali = {}       # {area: (ti, td, tot)}
    totali_per_distretto = {} # {distretto: {area: (ti, td, tot)}}

    for distretto in distretti:
        df_dist = df_vet[df_vet['_distretto'] == distretto]

        r = _scrivi_divisorio(ws, r, distretto, _N_COLS)
        r = _scrivi_header(ws, r)

        r, totali_area = _scrivi_tabella_distretto(ws, r, df_dist)
        r += 1  # riga vuota tra distretti

        totali_per_distretto[distretto] = totali_area

        # Accumula globali
        for area, (ti, td, tot) in totali_area.items():
            prev = totali_globali.get(area, (0, 0, 0))
            totali_globali[area] = (
                prev[0] + ti, prev[1] + td, prev[2] + tot
            )

    # Riepilogo per area aziendale
    r = _scrivi_riepilogo_aree(ws, r, totali_globali)

    # Cross-tab distretto × area
    r = _scrivi_riepilogo_distretti(ws, r, totali_per_distretto)

    # Larghezze colonne
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 45
    for col_letter in ['C', 'D', 'E']:
        ws.column_dimensions[col_letter].width = 12
