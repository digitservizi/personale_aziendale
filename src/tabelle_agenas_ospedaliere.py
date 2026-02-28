"""
Tabelle AGENAS ospedaliere (range-based per livello di presidio).
Estratte da report_fabbisogno.py per modularizzazione.

Funzioni incluse:
  - _scrivi_tabella_agenas_materno_infantile
  - _scrivi_tabella_agenas_radiologia
  - _scrivi_tabella_agenas_emergenza_urgenza
  - _scrivi_tabella_agenas_terapia_intensiva
  - _scrivi_tabella_agenas_sale_operatorie
  - _scrivi_tabella_agenas_area_ti_bo
  - _scrivi_tabella_agenas_anatomia_patologica
  - _scrivi_tabella_agenas_laboratorio
  - _scrivi_tabella_agenas_tecnici_laboratorio
  - _scrivi_tabella_agenas_medicina_legale
"""

from openpyxl.styles import Font, Alignment, PatternFill

from src.stili_excel import (
    THIN_BORDER, FILL_A, FILL_B, FILL_HEADER,
    FONT_HEADER, ALIGN_CENTER,
)


# ============================================================
# TABELLA AGENAS MATERNO-INFANTILE IN CALCE AL FOGLIO
# ============================================================

def _scrivi_tabella_agenas_materno_infantile(
        ws, start_row, df_area, fabb_presidio,
        mapping_uo, mapping_profili, sede_completa):
    """Aggiunge una tabella riepilogativa AGENAS materno-infantile
    in calce al foglio di dettaglio di un presidio ospedaliero.

    Restituisce la riga successiva libera.
    """
    import re as _re

    if not fabb_presidio:
        return start_row

    # --- Stili locali ---
    FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
    FONT_NORMAL  = Font(size=10)

    row = start_row + 1
    N_COLS_MERGE = 6  # per merge titolo/sottotitolo

    # Titolo sezione
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    c = ws.cell(row=row, column=1,
                value="AREA MATERNO INFANTILE - Standard AGENAS")
    c.font = FONT_SECTION
    row += 1

    # Nota presidio
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value=f"Presidio: {sede_completa}"
            ).font = FONT_NORMAL
    row += 2

    # Intestazioni
    headers = ['Profilo', 'T.I.', 'T.D.', 'Totale', 'Range AGENAS',
               'Esito']
    N_COLS = len(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # Nomi leggibili dei profili AGENAS materno-infantile
    nomi_profili = {
        'DIRIGENTE_MEDICO_PEDIATRIA':  'Dir. Medici Pediatria/Neonat.',
        'DIRIGENTE_MEDICO_OSTETRICIA': 'Dir. Medici Ostetricia/Ginec.',
        'OSTETRICA':                   'Ostetriche',
        'INFERMIERE':                  'Infermieri',
        'OPERATORE_SOCIO_SANITARIO':   'Operatori Socio Sanitari',
    }

    # Calcola gli in servizio per ciascun profilo AGENAS
    in_servizio = {}
    in_servizio_td = {}
    for _, r in df_area.iterrows():
        ssd = str(r.get('_REPARTO', ''))
        profilo = str(r.get('Profilo Professionale', '')
                      ).strip().upper()
        quantita_ti = int(r.get('Quantità T.I.', 0))
        quantita_td = int(r.get('Quantità T.D.', 0))

        # Verifica se la UO è dell'area materno-infantile
        uo_match = False
        profilo_uo = None
        for m_uo in mapping_uo:
            if _re.search(m_uo['pattern'], ssd.upper()):
                uo_match = True
                profilo_uo = m_uo.get('profilo_agenas')
                break
        if not uo_match:
            continue

        # Mappa il profilo
        for m_prof in mapping_profili:
            if _re.search(m_prof['pattern'], profilo):
                pa = m_prof['profilo_agenas']
                # I dirigenti medici prendono il profilo dalla UO
                if pa in ('DIRIGENTE_MEDICO_PEDIATRIA',
                          'DIRIGENTE_MEDICO_OSTETRICIA'):
                    if profilo_uo:
                        pa = profilo_uo
                    else:
                        continue
                in_servizio[pa] = in_servizio.get(pa, 0) + quantita_ti
                in_servizio_td[pa] = in_servizio_td.get(pa, 0) + quantita_td
                break

    # Righe dati
    FILL_OK = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                          fill_type='solid')
    FILL_CARENZA = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                               fill_type='solid')
    toggle = 0
    for prof_key, rng in fabb_presidio.items():
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1
        servizio_ti = in_servizio.get(prof_key, 0)
        servizio_td = in_servizio_td.get(prof_key, 0)
        servizio = servizio_ti + servizio_td
        v_min = rng['min']
        v_max = rng['max']
        range_str = f'{v_min} - {v_max}'

        if servizio < v_min:
            esito = f'CARENZA (min {v_min - servizio})'
            fill_esito = FILL_CARENZA
        elif servizio > v_max:
            esito = f'ECCEDENZA (+{servizio - v_max})'
            fill_esito = FILL_OK
        else:
            esito = 'IN RANGE'
            fill_esito = FILL_OK

        vals = [
            nomi_profili.get(prof_key, prof_key),
            servizio_ti, servizio_td, servizio, range_str, esito,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        # Colora la cella Esito
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        row += 1

    # Note piè di pagina
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="(*) I medici comprendono guardie e apicalità. "
                  "Maggiorazione +15% per turnazione, guardie, "
                  "ferie, malattie e altre indisponibilità."
            ).font = Font(
                italic=True, size=9, color='555555')

    return row + 1


# ============================================================
# TABELLA AGENAS RADIOLOGIA IN CALCE AL FOGLIO
# ============================================================

def _scrivi_tabella_agenas_radiologia(
        ws, start_row, df_luogo, fabb_radio_presidio,
        mapping_uo_radio, mapping_profili_radio, sede_completa,
        livello_label, df_area=None):
    """Aggiunge una tabella riepilogativa AGENAS radiologia
    in calce al foglio di dettaglio di un presidio ospedaliero.

    Se df_area è fornito, contabilizza anche il personale
    radiologico presente in sedi territoriali della stessa area.

    Restituisce la riga successiva libera.
    """
    import re as _re

    if not fabb_radio_presidio:
        return start_row

    # --- Stili locali ---
    FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
    FONT_NORMAL  = Font(size=10)

    row = start_row + 1
    N_COLS_MERGE = 6  # per merge titolo/sottotitolo

    # Titolo sezione
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    c = ws.cell(row=row, column=1,
                value="AREA SERVIZI DI RADIOLOGIA - Standard AGENAS")
    c.font = FONT_SECTION
    row += 1

    # Nota presidio + livello
    livello_leggibile = {
        'OSPEDALE_DI_BASE':     'Ospedale di Base',
        'PRESIDIO_I_LIVELLO':   'Presidio di I livello',
        'PRESIDIO_II_LIVELLO':  'Presidio di II livello',
    }.get(livello_label, livello_label)

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value=f"Presidio: {sede_completa}  –  {livello_leggibile}"
            ).font = FONT_NORMAL
    row += 2

    # Intestazioni
    headers = ['Profilo', 'T.I.', 'T.D.', 'Totale', 'Range AGENAS',
               'Esito']
    N_COLS = len(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # Nomi leggibili dei profili AGENAS radiologia
    nomi_profili = {
        'DIRIGENTE_MEDICO_RADIOLOGIA': 'Dir. Medici Radiologia (*)',
        'TECNICO_RADIOLOGIA':          'Tecnici di Radiologia',
        'INFERMIERE':                  'Infermieri',
        'OPERATORE_SOCIO_SANITARIO':   'Operatori Socio Sanitari',
    }

    # Calcola gli in servizio per ciascun profilo AGENAS radiologia
    in_servizio = {}
    in_servizio_td = {}

    def _conta_radio(df_src, solo_territoriali=False):
        """Conta personale radiologico nel DataFrame fornito."""
        for _, r in df_src.iterrows():
            ssd = str(r.get('_REPARTO', ''))
            sede_r = str(r.get('Sede', ''))
            profilo = str(r.get('Profilo Professionale', '')
                          ).strip().upper()
            quantita_ti = int(r.get('Quantità T.I.', 0))
            quantita_td = int(r.get('Quantità T.D.', 0))

            # Se richiesto solo territoriali, salta i P.O.
            if solo_territoriali and 'P.O.' in sede_r:
                continue

            # Verifica se la UO è dell'area radiologica
            uo_radio = False
            for m_uo in mapping_uo_radio:
                if _re.search(m_uo['pattern'], ssd.upper()):
                    uo_radio = True
                    break

            # Per sedi territoriali: verifica anche CDC
            if not uo_radio and 'P.O.' not in sede_r:
                cdc_val = str(r.get('Centro di Costo', '')).upper()
                if _re.search(
                        r'CENTRO RADIOLOGICO|RADIOLOGIA', cdc_val):
                    uo_radio = True

            # Per sedi territoriali: TS RADIOLOGIA conta sempre
            if not uo_radio and 'P.O.' not in sede_r:
                for m_prof in mapping_profili_radio:
                    if (m_prof['profilo_agenas'] == 'TECNICO_RADIOLOGIA'
                            and _re.search(m_prof['pattern'], profilo)):
                        uo_radio = True
                        break

            if not uo_radio:
                continue

            # Mappa il profilo
            for m_prof in mapping_profili_radio:
                if _re.search(m_prof['pattern'], profilo):
                    pa = m_prof['profilo_agenas']
                    in_servizio[pa] = in_servizio.get(pa, 0) + quantita_ti
                    in_servizio_td[pa] = in_servizio_td.get(pa, 0) + quantita_td
                    break

    # Conta personale P.O. (da df_luogo)
    _conta_radio(df_luogo)

    # Conta anche personale da sedi territoriali dell'area
    if df_area is not None:
        _conta_radio(df_area, solo_territoriali=True)

    # Righe dati
    FILL_OK = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                          fill_type='solid')
    FILL_CARENZA = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                               fill_type='solid')
    toggle = 0
    for prof_key, rng in fabb_radio_presidio.items():
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1
        servizio_ti = in_servizio.get(prof_key, 0)
        servizio_td = in_servizio_td.get(prof_key, 0)
        servizio = servizio_ti + servizio_td
        v_min = rng['min']
        v_max = rng['max']
        range_str = f'{v_min} - {v_max}'

        if servizio < v_min:
            esito = f'CARENZA (min {v_min - servizio})'
            fill_esito = FILL_CARENZA
        elif servizio > v_max:
            esito = f'ECCEDENZA (+{servizio - v_max})'
            fill_esito = FILL_OK
        else:
            esito = 'IN RANGE'
            fill_esito = FILL_OK

        vals = [
            nomi_profili.get(prof_key, prof_key),
            servizio_ti, servizio_td, servizio, range_str, esito,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        # Colora la cella Esito
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        row += 1

    # Note piè di pagina
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="(*) comprese guardie e apicalità. "
                  "Maggiorazione +15% per turnazione, guardie, "
                  "ferie, malattie e altre indisponibilità."
            ).font = Font(
                italic=True, size=9, color='555555')

    if df_area is not None:
        row += 1
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS_MERGE)
        ws.cell(row=row, column=1,
                value="I dati includono il personale delle sedi "
                      "territoriali dell'area."
                ).font = Font(italic=True, size=9, color='555555')

    return row + 1


# ============================================================
# TABELLA AGENAS EMERGENZA-URGENZA IN CALCE AL FOGLIO
# ============================================================

def _scrivi_tabella_agenas_emergenza_urgenza(
        ws, start_row, df_area, fabb_emergenza_presidio,
        mapping_uo_emergenza, mapping_profili_emergenza,
        sede_completa, livello_label):
    """Aggiunge una tabella riepilogativa AGENAS per l'area
    dell'emergenza-urgenza (Tabella 20) nel foglio RIEPILOGO.

    Conta il personale T.I. nelle UO che matchano i pattern
    (Pronto Soccorso / MCAU) e confronta con i range AGENAS
    in base al livello del presidio (PS, DEA I, DEA II).

    Restituisce la riga successiva libera.
    """
    import re as _re

    if not fabb_emergenza_presidio:
        return start_row

    # --- Stili locali ---
    FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
    FONT_NORMAL  = Font(size=10)

    row = start_row + 1
    N_COLS_MERGE = 6

    # Titolo sezione
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="AREA EMERGENZA-URGENZA - Standard AGENAS"
            ).font = FONT_SECTION
    row += 1

    # Nota presidio + livello
    livello_leggibile = {
        'OSPEDALE_DI_BASE':    'Pronto Soccorso',
        'PRESIDIO_I_LIVELLO':  'DEA I',
        'PRESIDIO_II_LIVELLO': 'DEA II',
    }.get(livello_label, livello_label)

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value=f"Presidio: {sede_completa}  –  {livello_leggibile}"
            ).font = FONT_NORMAL
    row += 2

    # Intestazioni
    headers = ['Profilo', 'T.I.', 'T.D.', 'Totale', 'Range AGENAS', 'Esito']
    N_COLS = len(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # Nomi leggibili dei profili
    nomi_profili = {
        'DIRIGENTE_MEDICO':          'Dirigenti Medici',
        'INFERMIERE':                'Infermieri',
        'OPERATORE_SOCIO_SANITARIO': 'Operatori Socio Sanitari',
    }

    # --- Conta in servizio per ciascun profilo ---
    in_servizio = {}
    in_servizio_td = {}
    for _, r in df_area.iterrows():
        ssd = str(r.get('_REPARTO', ''))
        cdc = str(r.get('Centro di Costo', ''))
        profilo = str(r.get('Profilo Professionale', '')).strip().upper()
        quantita_ti = int(r.get('Quantità T.I.', 0))
        quantita_td = int(r.get('Quantità T.D.', 0))

        # Verifica se la UO è dell'area emergenza-urgenza
        uo_match = False
        for m_uo in mapping_uo_emergenza:
            if (_re.search(m_uo['pattern'], ssd, _re.IGNORECASE)
                    or _re.search(m_uo['pattern'], cdc, _re.IGNORECASE)):
                uo_match = True
                break
        if not uo_match:
            continue

        # Mappa il profilo
        for m_prof in mapping_profili_emergenza:
            if _re.search(m_prof['pattern'], profilo):
                pa = m_prof['profilo_agenas']
                in_servizio[pa] = in_servizio.get(pa, 0) + quantita_ti
                in_servizio_td[pa] = in_servizio_td.get(pa, 0) + quantita_td
                break

    # --- Righe dati ---
    FILL_OK = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                          fill_type='solid')
    FILL_CARENZA = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                               fill_type='solid')
    toggle = 0
    for prof_key, rng in fabb_emergenza_presidio.items():
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1
        servizio_ti = in_servizio.get(prof_key, 0)
        servizio_td = in_servizio_td.get(prof_key, 0)
        servizio = servizio_ti + servizio_td
        v_min = rng['min']
        v_max = rng['max']
        range_str = f'{v_min} - {v_max}'

        if servizio < v_min:
            esito = f'CARENZA (min {v_min - servizio})'
            fill_esito = FILL_CARENZA
        elif servizio > v_max:
            esito = f'ECCEDENZA (+{servizio - v_max})'
            fill_esito = FILL_OK
        else:
            esito = 'IN RANGE'
            fill_esito = FILL_OK

        vals = [
            nomi_profili.get(prof_key, prof_key),
            servizio_ti, servizio_td, servizio, range_str, esito,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        row += 1

    # Note piè di pagina
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="(*) Valori minimi in FTE riferiti ad apertura "
                  "proporzionata sulle 24 ore (DM 70/2015)."
            ).font = Font(italic=True, size=9, color='555555')

    return row + 1


# ============================================================
# TABELLA AGENAS TERAPIA INTENSIVA (§ 8.1.1) IN CALCE AL FOGLIO
# ============================================================

def _scrivi_tabella_agenas_terapia_intensiva(
        ws, start_row, df_area, fabb_ti_presidio,
        mapping_uo_ti, mapping_profili_ti,
        sede_completa):
    """Aggiunge una tabella riepilogativa AGENAS per l'area
    Terapia Intensiva (§ 8.1.1) nel foglio RIEPILOGO.

    Il fabbisogno è calcolato con la formula:
      FTE = (PL / rapporto_letti) × (ore_turno × 365) / ore_annue_eff

    Colonne: Profilo | Rapporto | PL | FTE Atteso | T.I. | T.D. |
             Totale | Esito

    Restituisce la riga successiva libera.
    """
    import re as _re

    if not fabb_ti_presidio:
        return start_row

    # --- Stili locali ---
    FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
    FONT_NORMAL  = Font(size=10)

    row = start_row + 1
    N_COLS_MERGE = 8

    # Titolo sezione
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="AREA TERAPIA INTENSIVA - Standard AGENAS (§ 8.1.1)"
            ).font = FONT_SECTION
    row += 1

    # Nota presidio
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value=f"Presidio: {sede_completa}"
            ).font = FONT_NORMAL
    row += 2

    # Intestazioni
    headers = ['Profilo', 'Rapporto', 'PL Intensivi',
               'FTE Atteso', 'T.I.', 'T.D.', 'Totale', 'Esito']
    N_COLS = len(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # Nomi leggibili dei profili
    nomi_profili = {
        'DIRIGENTE_MEDICO': 'Dirigenti Medici',
        'INFERMIERE':       'Infermieri',
    }

    # --- Conta in servizio per ciascun profilo ---
    # Filtra solo il personale il cui Centro di Costo contiene
    # "TERAPIA INTENSIVA" (esclude Blocco Operatorio e CDC condivisi
    # come "Anestesia e T.I. – Camere Operatorie").
    _RE_CDC_TI = _re.compile(r'TERAPIA INTENSIVA', _re.IGNORECASE)
    in_servizio = {}
    in_servizio_td = {}
    for _, r in df_area.iterrows():
        cdc = str(r.get('Centro di Costo', ''))
        profilo = str(r.get('Profilo Professionale', '')).strip().upper()
        quantita_ti = int(r.get('Quantità T.I.', 0))
        quantita_td = int(r.get('Quantità T.D.', 0))

        if not _RE_CDC_TI.search(cdc):
            continue

        # Mappa il profilo
        for m_prof in mapping_profili_ti:
            if _re.search(m_prof['pattern'], profilo):
                pa = m_prof['profilo_agenas']
                in_servizio[pa] = in_servizio.get(pa, 0) + quantita_ti
                in_servizio_td[pa] = in_servizio_td.get(pa, 0) + quantita_td
                break

    # --- Righe dati ---
    FILL_OK = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                          fill_type='solid')
    FILL_CARENZA = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                               fill_type='solid')
    toggle = 0
    for prof_key, dati in fabb_ti_presidio.items():
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1

        fte_atteso = dati['fte_atteso']
        rapporto = dati['rapporto']
        pl = dati['pl']

        servizio_ti = in_servizio.get(prof_key, 0)
        servizio_td = in_servizio_td.get(prof_key, 0)
        servizio = servizio_ti + servizio_td

        if servizio < fte_atteso:
            delta = fte_atteso - servizio
            esito = f'CARENZA ({delta})'
            fill_esito = FILL_CARENZA
        elif servizio > fte_atteso:
            delta = servizio - fte_atteso
            esito = f'ECCEDENZA (+{delta})'
            fill_esito = FILL_OK
        else:
            esito = 'CONFORME'
            fill_esito = FILL_OK

        rapporto_str = f'1 ogni {rapporto} letti'

        vals = [
            nomi_profili.get(prof_key, prof_key),
            rapporto_str, pl, fte_atteso,
            servizio_ti, servizio_td, servizio, esito,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        row += 1

    # Note piè di pagina
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="Formula: FTE = (PL / Rapporto) × (24 × 365) "
                  "/ Ore annue effettive × 1,15.  "
                  "Fonte: AGENAS § 8.1.1 – Area Intensiva.  "
                  "Maggiorazione +15% per turnazione, guardie, "
                  "ferie, malattie e altre indisponibilità."
            ).font = Font(italic=True, size=9, color='555555')

    return row + 1


# ============================================================
# TABELLA AGENAS SALE OPERATORIE IN CALCE AL FOGLIO
# ============================================================

def _scrivi_tabella_agenas_sale_operatorie(
        ws, start_row, df_luogo, fabb_so_presidio,
        mapping_uo_so, mapping_profili_so, sede_completa):
    """Aggiunge una tabella riepilogativa AGENAS sale operatorie (§ 8.1.2)
    in calce al foglio di dettaglio di un presidio ospedaliero.

    Restituisce la riga successiva libera.
    """
    import re as _re

    if not fabb_so_presidio:
        return start_row

    # --- Stili locali ---
    FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
    FONT_NORMAL  = Font(size=10)

    row = start_row + 1
    N_COLS_MERGE = 8  # colonne: Profilo|Staff/Sala|N.Sale|FTE|TI|TD|Tot|Esito

    # Titolo sezione
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    c = ws.cell(row=row, column=1,
                value="AREA SALE OPERATORIE - Standard AGENAS (§ 8.1.2)")
    c.font = FONT_SECTION
    row += 1

    # Recupero n_sale dal primo profilo (uguale per tutti)
    first_prof = next(iter(fabb_so_presidio.values()))
    n_sale = first_prof['n_sale']
    ore_cop = first_prof['ore_copertura']
    gg_anno = first_prof['giorni_anno']

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value=f"Presidio: {sede_completa}  –  "
                  f"N. sale operative: {n_sale}  –  "
                  f"Copertura: {ore_cop}h × {gg_anno} gg/anno"
            ).font = FONT_NORMAL
    row += 2

    # Intestazioni
    headers = ['Profilo', 'Staff/Sala', 'N. Sale', 'FTE Atteso',
               'T.I.', 'T.D.', 'Totale', 'Esito']
    N_COLS = len(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # Nomi leggibili
    nomi_profili = {
        'DIRIGENTE_MEDICO_ANESTESISTA': 'Medici Anestesisti (*)',
        'INFERMIERE':                   'Infermieri',
        'OPERATORE_SOCIO_SANITARIO':    'Operatori Socio Sanitari',
    }

    # Calcola gli in servizio – filtra per UO di Blocco Operatorio
    # e per UO di Anestesia/Rianimazione (anestesisti)
    in_servizio = {}
    in_servizio_td = {}
    for _, r in df_luogo.iterrows():
        cdc = str(r.get('Centro di Costo', ''))
        reparto = str(r.get('_REPARTO', ''))
        profilo = str(r.get('Profilo Professionale', '')).strip().upper()
        quantita_ti = int(r.get('Quantità T.I.', 0))
        quantita_td = int(r.get('Quantità T.D.', 0))

        # Verifica se la UO/CDC è dell'area sale operatorie
        uo_match = False
        for m_uo in mapping_uo_so:
            pat = m_uo['pattern']
            if (_re.search(pat, cdc, _re.IGNORECASE)
                    or _re.search(pat, reparto, _re.IGNORECASE)):
                uo_match = True
                break
        if not uo_match:
            continue

        # Escludi il personale il cui CDC è "Terapia Intensiva":
        # è già conteggiato nella tabella TI (§ 8.1.1) e non va
        # doppio-contato nella tabella Sale Operatorie.
        if _re.search(r'TERAPIA INTENSIVA', cdc, _re.IGNORECASE):
            continue

        for m_prof in mapping_profili_so:
            if _re.search(m_prof['pattern'], profilo):
                pa = m_prof['profilo_agenas']
                in_servizio[pa] = in_servizio.get(pa, 0) + quantita_ti
                in_servizio_td[pa] = in_servizio_td.get(pa, 0) + quantita_td
                break

    # Righe dati
    FILL_OK = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                          fill_type='solid')
    FILL_CARENZA = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                               fill_type='solid')
    toggle = 0
    for prof_key, dati in fabb_so_presidio.items():
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1

        fte_atteso = dati['fte_atteso']
        pps = dati['personale_per_sala']

        servizio_ti = in_servizio.get(prof_key, 0)
        servizio_td = in_servizio_td.get(prof_key, 0)
        servizio = servizio_ti + servizio_td

        if servizio < fte_atteso:
            delta = fte_atteso - servizio
            esito = f'CARENZA ({delta})'
            fill_esito = FILL_CARENZA
        elif servizio > fte_atteso:
            delta = servizio - fte_atteso
            esito = f'ECCEDENZA (+{delta})'
            fill_esito = FILL_OK
        else:
            esito = 'CONFORME'
            fill_esito = FILL_OK

        vals = [
            nomi_profili.get(prof_key, prof_key),
            pps, n_sale, fte_atteso,
            servizio_ti, servizio_td, servizio, esito,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        row += 1

    # Note piè di pagina
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="(*) Solo anestesisti; i medici specialisti (chirurghi, "
                  "ortopedici, ecc.) sono già conteggiati nelle tabelle "
                  "per livello di presidio."
            ).font = Font(italic=True, size=9, color='555555')
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="Formula: FTE = N.Sale × Staff/Sala × "
                  f"({ore_cop} × {gg_anno}) / Ore annue eff. × 1,15.  "
                  "Fonte: AGENAS § 8.1.2 – Sale Operatorie.  "
                  "Maggiorazione +15% per turnazione, guardie, "
                  "ferie, malattie e altre indisponibilità."
            ).font = Font(italic=True, size=9, color='555555')

    return row + 1


# ============================================================
# TABELLA COMBINATA AREA TI + BLOCCO OPERATORIO
# ============================================================

def _scrivi_tabella_agenas_area_ti_bo(
        ws, start_row, df_area, fabb_ti_presidio, fabb_so_presidio,
        mapping_uo_ti, sede_completa):
    """Tabella riepilogativa AREA ANESTESIA, TERAPIA INTENSIVA E
    BLOCCO OPERATORIO.

    Unisce i fabbisogni TI (§ 8.1.1) e BO (§ 8.1.2), raccoglie
    TUTTO il personale afferente alla SC/SSD di Anestesia e Terapia
    Intensiva e calcola le carenze/eccedenze sull'area combinata.

    Restituisce la riga successiva libera.
    """
    import re as _re

    fabb_ti = fabb_ti_presidio or {}
    fabb_so = fabb_so_presidio or {}

    if not fabb_ti and not fabb_so:
        return start_row

    # --- Stili locali ---
    FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
    FONT_NORMAL  = Font(size=10)

    row = start_row + 1
    N_COLS_MERGE = 8

    # Titolo sezione
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="AREA ANESTESIA, TERAPIA INTENSIVA E BLOCCO OPERATORIO"
            ).font = FONT_SECTION
    row += 1

    # PL intensivi e n. sale dal fabbisogno
    pl_ti = 0
    for d in fabb_ti.values():
        pl_ti = d.get('pl', 0)
        break
    n_sale = 0
    for d in fabb_so.values():
        n_sale = d.get('n_sale', 0)
        break

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value=f"Presidio: {sede_completa}  \u2013  "
                  f"PL Intensivi: {pl_ti}  \u2013  "
                  f"Sale Operative: {n_sale}"
            ).font = FONT_NORMAL
    row += 2

    # Intestazioni
    headers = ['Profilo', 'FTE TI', 'FTE BO', 'FTE Area',
               'T.I.', 'T.D.', 'Totale', 'Esito']
    N_COLS = len(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # Profili unificati (merge TI + BO)
    _PROFILI_AREA = [
        {
            'label': 'Dirigenti Medici',
            'key': 'DIRIGENTE_MEDICO',
            'ti_key': 'DIRIGENTE_MEDICO',
            'so_key': 'DIRIGENTE_MEDICO_ANESTESISTA',
        },
        {
            'label': 'Infermieri',
            'key': 'INFERMIERE',
            'ti_key': 'INFERMIERE',
            'so_key': 'INFERMIERE',
        },
        {
            'label': 'Operatori Socio Sanitari',
            'key': 'OPERATORE_SOCIO_SANITARIO',
            'ti_key': None,
            'so_key': 'OPERATORE_SOCIO_SANITARIO',
        },
    ]

    # Mapping profilo personale → profilo unificato dell'area
    _MAPPING_PROFILI_AREA = [
        {'pattern': r'DIRIGENTE MEDICO|DIR\.? MED',
         'profilo_agenas': 'DIRIGENTE_MEDICO'},
        {'pattern': r'INFERMIERE|INFERMIERA',
         'profilo_agenas': 'INFERMIERE'},
        {'pattern': r'OPERATORE SOCIO',
         'profilo_agenas': 'OPERATORE_SOCIO_SANITARIO'},
    ]

    # --- Conta in servizio: TUTTO il personale della UO ---
    in_servizio = {}
    in_servizio_td = {}
    for _, r in df_area.iterrows():
        ssd = str(r.get('_REPARTO', ''))
        profilo = str(r.get('Profilo Professionale', '')).strip().upper()
        quantita_ti = int(r.get('Quantità T.I.', 0))
        quantita_td = int(r.get('Quantità T.D.', 0))

        # Verifica se la UO è dell'area Anestesia e TI
        uo_match = False
        for m_uo in mapping_uo_ti:
            if _re.search(m_uo['pattern'], ssd, _re.IGNORECASE):
                uo_match = True
                break
        if not uo_match:
            continue

        # Mappa il profilo (mapping unificato)
        for m_prof in _MAPPING_PROFILI_AREA:
            if _re.search(m_prof['pattern'], profilo):
                pa = m_prof['profilo_agenas']
                in_servizio[pa] = in_servizio.get(pa, 0) + quantita_ti
                in_servizio_td[pa] = in_servizio_td.get(pa, 0) + quantita_td
                break

    # --- Righe dati ---
    FILL_OK = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                          fill_type='solid')
    FILL_CARENZA = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                               fill_type='solid')
    toggle = 0
    for prof_info in _PROFILI_AREA:
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1

        fte_ti = 0
        if prof_info['ti_key'] and prof_info['ti_key'] in fabb_ti:
            fte_ti = fabb_ti[prof_info['ti_key']]['fte_atteso']
        fte_so = 0
        if prof_info['so_key'] and prof_info['so_key'] in fabb_so:
            fte_so = fabb_so[prof_info['so_key']]['fte_atteso']
        fte_area = fte_ti + fte_so

        servizio_ti = in_servizio.get(prof_info['key'], 0)
        servizio_td = in_servizio_td.get(prof_info['key'], 0)
        servizio = servizio_ti + servizio_td

        if servizio < fte_area:
            delta = fte_area - servizio
            esito = f'CARENZA ({delta})'
            fill_esito = FILL_CARENZA
        elif servizio > fte_area:
            delta = servizio - fte_area
            esito = f'ECCEDENZA (+{delta})'
            fill_esito = FILL_OK
        else:
            esito = 'CONFORME'
            fill_esito = FILL_OK

        vals = [
            prof_info['label'],
            fte_ti, fte_so, fte_area,
            servizio_ti, servizio_td, servizio, esito,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        row += 1

    # Note piè di pagina
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="FTE TI = (PL / Rapporto) \u00d7 (24 \u00d7 365) "
                  "/ Ore_annue \u00d7 1,15  |  "
                  "FTE BO = N.Sale \u00d7 Staff/Sala \u00d7 (24 \u00d7 250) "
                  "/ Ore_annue \u00d7 1,15.  "
                  "Le carenze sono calcolate sull\u2019area complessiva: "
                  "tutto il personale afferente alla SC/SSD di Anestesia "
                  "e Terapia Intensiva (inclusi i CDC condivisi)."
            ).font = Font(italic=True, size=9, color='555555')

    return row + 1


# ============================================================
# TABELLA AGENAS ANATOMIA PATOLOGICA IN CALCE AL FOGLIO
# ============================================================

def _scrivi_tabella_agenas_anatomia_patologica(
        ws, start_row, df_luogo, fabb_anapato_presidio,
        mapping_uo_anapato, mapping_profili_anapato, sede_completa,
        livello_label):
    """Aggiunge una tabella riepilogativa AGENAS anatomia patologica
    in calce al foglio di dettaglio di un presidio ospedaliero.

    Restituisce la riga successiva libera.
    """
    import re as _re

    if not fabb_anapato_presidio:
        return start_row

    # --- Stili locali ---
    FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
    FONT_NORMAL  = Font(size=10)

    row = start_row + 1
    N_COLS_MERGE = 6

    # Titolo sezione
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    c = ws.cell(row=row, column=1,
                value="AREA ANATOMIA PATOLOGICA - Standard AGENAS")
    c.font = FONT_SECTION
    row += 1

    # Nota presidio + livello
    livello_leggibile = {
        'OSPEDALE_DI_BASE':     'Ospedale di Base',
        'PRESIDIO_I_LIVELLO':   'Presidio di I livello',
        'PRESIDIO_II_LIVELLO':  'Presidio di II livello',
    }.get(livello_label, livello_label)

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value=f"Presidio: {sede_completa}  \u2013  {livello_leggibile}"
            ).font = FONT_NORMAL
    row += 2

    # Intestazioni
    headers = ['Profilo', 'T.I.', 'T.D.', 'Totale', 'Range AGENAS',
               'Esito']
    N_COLS = len(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # Nomi leggibili dei profili AGENAS anatomia patologica
    nomi_profili = {
        'DIRIGENTE_SANITARIO':       'Dir. Sanitari (*)',
        'INFERMIERE':                'Infermieri',
        'OPERATORE_SOCIO_SANITARIO': 'Operatori Socio Sanitari',
    }

    # Calcola gli in servizio per ciascun profilo
    in_servizio = {}
    in_servizio_td = {}
    for _, r in df_luogo.iterrows():
        ssd = str(r.get('_REPARTO', ''))
        profilo = str(r.get('Profilo Professionale', '')).strip().upper()
        quantita_ti = int(r.get('Quantità T.I.', 0))
        quantita_td = int(r.get('Quantità T.D.', 0))

        uo_ok = False
        for m_uo in mapping_uo_anapato:
            if _re.search(m_uo['pattern'], ssd.upper()):
                uo_ok = True
                break
        if not uo_ok:
            continue

        for m_prof in mapping_profili_anapato:
            if _re.search(m_prof['pattern'], profilo):
                pa = m_prof['profilo_agenas']
                in_servizio[pa] = in_servizio.get(pa, 0) + quantita_ti
                in_servizio_td[pa] = in_servizio_td.get(pa, 0) + quantita_td
                break

    # Righe dati
    FILL_OK = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                          fill_type='solid')
    FILL_CARENZA = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                               fill_type='solid')
    toggle = 0
    for prof_key, rng in fabb_anapato_presidio.items():
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1
        servizio_ti = in_servizio.get(prof_key, 0)
        servizio_td = in_servizio_td.get(prof_key, 0)
        servizio = servizio_ti + servizio_td
        v_min = rng['min']
        v_max = rng['max']
        range_str = f'{v_min} - {v_max}'

        if servizio < v_min:
            esito = f'CARENZA (min {v_min - servizio})'
            fill_esito = FILL_CARENZA
        elif servizio > v_max:
            esito = f'ECCEDENZA (+{servizio - v_max})'
            fill_esito = FILL_OK
        else:
            esito = 'IN RANGE'
            fill_esito = FILL_OK

        vals = [
            nomi_profili.get(prof_key, prof_key),
            servizio_ti, servizio_td, servizio, range_str, esito,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        row += 1

    # Nota piè di pagina
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="(*) Dirigenti Sanitari Medici e Non Medici "
                  "(comprese guardie e apicalità). "
                  "Maggiorazione +15% per turnazione, guardie, "
                  "ferie, malattie e altre indisponibilità."
            ).font = Font(italic=True, size=9, color='555555')

    return row + 1


# ============================================================
# TABELLA AGENAS SERVIZI DI LABORATORIO IN CALCE AL FOGLIO
# ============================================================

def _scrivi_tabella_agenas_laboratorio(
        ws, start_row, df_luogo, fabb_lab_presidio,
        mapping_uo_lab, mapping_profili_lab, esclusioni_lab,
        sede_completa, livello_label):
    """Aggiunge una tabella riepilogativa AGENAS servizi di laboratorio
    in calce al foglio di dettaglio di un presidio ospedaliero.

    Restituisce la riga successiva libera.
    """
    import re as _re

    if not fabb_lab_presidio:
        return start_row

    # --- Stili locali ---
    FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
    FONT_NORMAL  = Font(size=10)

    row = start_row + 1
    N_COLS_MERGE = 6

    # Titolo sezione
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    c = ws.cell(row=row, column=1,
                value="AREA SERVIZI DI LABORATORIO - Standard AGENAS")
    c.font = FONT_SECTION
    row += 1

    # Nota presidio + livello
    livello_leggibile = {
        'OSPEDALE_DI_BASE':     'Ospedale di Base',
        'PRESIDIO_I_LIVELLO':   'Presidio di I livello',
        'PRESIDIO_II_LIVELLO':  'Presidio di II livello',
    }.get(livello_label, livello_label)

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value=f"Presidio: {sede_completa}  \u2013  {livello_leggibile}"
            ).font = FONT_NORMAL
    row += 2

    # Intestazioni
    headers = ['Profilo', 'T.I.', 'T.D.', 'Totale', 'Range AGENAS',
               'Esito']
    N_COLS = len(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # Nomi leggibili dei profili AGENAS laboratorio
    nomi_profili = {
        'DIRIGENTE_SANITARIO':       'Dir. Sanitari (*)',
        'INFERMIERE':                'Infermieri',
        'OPERATORE_SOCIO_SANITARIO': 'Operatori Socio Sanitari',
    }

    # Calcola gli in servizio per ciascun profilo
    in_servizio = {}
    in_servizio_td = {}
    for _, r in df_luogo.iterrows():
        ssd = str(r.get('_REPARTO', ''))
        profilo = str(r.get('Profilo Professionale', '')).strip().upper()
        quantita_ti = int(r.get('Quantità T.I.', 0))
        quantita_td = int(r.get('Quantità T.D.', 0))

        uo_ok = False
        for m_uo in mapping_uo_lab:
            if _re.search(m_uo['pattern'], ssd.upper()):
                uo_ok = True
                break
        if not uo_ok:
            continue

        # Escludi UO che hanno tabelle AGENAS dedicate
        esclusa = False
        for excl_pattern in esclusioni_lab:
            if _re.search(excl_pattern, ssd.upper()):
                esclusa = True
                break
        if esclusa:
            continue

        for m_prof in mapping_profili_lab:
            if _re.search(m_prof['pattern'], profilo):
                pa = m_prof['profilo_agenas']
                in_servizio[pa] = in_servizio.get(pa, 0) + quantita_ti
                in_servizio_td[pa] = in_servizio_td.get(pa, 0) + quantita_td
                break

    # Righe dati
    FILL_OK = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                          fill_type='solid')
    FILL_CARENZA = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                               fill_type='solid')
    toggle = 0
    for prof_key, rng in fabb_lab_presidio.items():
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1
        servizio_ti = in_servizio.get(prof_key, 0)
        servizio_td = in_servizio_td.get(prof_key, 0)
        servizio = servizio_ti + servizio_td
        v_min = rng['min']
        v_max = rng['max']
        range_str = f'{v_min} - {v_max}'

        if servizio < v_min:
            esito = f'CARENZA (min {v_min - servizio})'
            fill_esito = FILL_CARENZA
        elif servizio > v_max:
            esito = f'ECCEDENZA (+{servizio - v_max})'
            fill_esito = FILL_OK
        else:
            esito = 'IN RANGE'
            fill_esito = FILL_OK

        vals = [
            nomi_profili.get(prof_key, prof_key),
            servizio_ti, servizio_td, servizio, range_str, esito,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        row += 1

    # Nota piè di pagina
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="(*) Dirigenti Sanitari Medici e Non Medici "
                  "(comprese guardie e apicalità). "
                  "Maggiorazione +15% per turnazione, guardie, "
                  "ferie, malattie e altre indisponibilità."
            ).font = Font(italic=True, size=9, color='555555')

    return row + 1


# ============================================================
# TABELLA AGENAS TECNICI DI LABORATORIO IN CALCE AL FOGLIO
# ============================================================

def _scrivi_tabella_agenas_tecnici_laboratorio(
        ws, start_row, df_luogo, fabb_teclab_presidio,
        mapping_profili_teclab, sede_completa, livello_label):
    """Aggiunge una tabella riepilogativa AGENAS tecnici di laboratorio
    (Tab. 17) in calce al foglio di dettaglio di un presidio ospedaliero.

    Il conteggio è per ruolo su tutto il presidio, senza filtro UO.
    Restituisce la riga successiva libera.
    """
    import re as _re

    if not fabb_teclab_presidio:
        return start_row

    # --- Stili locali ---
    FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
    FONT_NORMAL  = Font(size=10)

    row = start_row + 1
    N_COLS_MERGE = 6

    # Titolo sezione
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    c = ws.cell(row=row, column=1,
                value="TECNICI DI LABORATORIO - Standard AGENAS")
    c.font = FONT_SECTION
    row += 1

    # Nota presidio + livello
    livello_leggibile = {
        'OSPEDALE_DI_BASE':     'Ospedale di Base',
        'PRESIDIO_I_LIVELLO':   'Presidio di I livello',
        'PRESIDIO_II_LIVELLO':  'Presidio di II livello',
    }.get(livello_label, livello_label)

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value=f"Presidio: {sede_completa}  \u2013  {livello_leggibile}"
            ).font = FONT_NORMAL
    row += 2

    # Intestazioni
    headers = ['Profilo', 'T.I.', 'T.D.', 'Totale', 'Range AGENAS',
               'Esito']
    N_COLS = len(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # Nomi leggibili
    nomi_profili = {
        'TECNICO_LABORATORIO': 'Tecnici di Laboratorio',
    }

    # Calcola gli in servizio per ciascun profilo (tutto il presidio)
    in_servizio = {}
    in_servizio_td = {}
    for _, r in df_luogo.iterrows():
        profilo = str(r.get('Profilo Professionale', '')).strip().upper()
        quantita_ti = int(r.get('Quantità T.I.', 0))
        quantita_td = int(r.get('Quantità T.D.', 0))

        for m_prof in mapping_profili_teclab:
            if _re.search(m_prof['pattern'], profilo):
                pa = m_prof['profilo_agenas']
                in_servizio[pa] = in_servizio.get(pa, 0) + quantita_ti
                in_servizio_td[pa] = in_servizio_td.get(pa, 0) + quantita_td
                break

    # Righe dati
    FILL_OK = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                          fill_type='solid')
    FILL_CARENZA = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                               fill_type='solid')
    toggle = 0
    for prof_key, rng in fabb_teclab_presidio.items():
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1
        servizio_ti = in_servizio.get(prof_key, 0)
        servizio_td = in_servizio_td.get(prof_key, 0)
        servizio = servizio_ti + servizio_td
        v_min = rng['min']
        v_max = rng['max']
        range_str = f'{v_min} - {v_max}'

        if servizio < v_min:
            esito = f'CARENZA (min {v_min - servizio})'
            fill_esito = FILL_CARENZA
        elif servizio > v_max:
            esito = f'ECCEDENZA (+{servizio - v_max})'
            fill_esito = FILL_OK
        else:
            esito = 'IN RANGE'
            fill_esito = FILL_OK

        vals = [
            nomi_profili.get(prof_key, prof_key),
            servizio_ti, servizio_td, servizio, range_str, esito,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        row += 1

    # Nota piè di pagina
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value="Il conteggio è per ruolo su tutto il presidio, "
                  "indipendentemente dall'unità operativa di assegnazione "
                  "(§7.1.9 AGENAS). "
                  "Maggiorazione +15% per turnazione, guardie, "
                  "ferie, malattie e altre indisponibilità."
            ).font = Font(italic=True, size=9, color='555555')

    return row + 1


# ============================================================
# TABELLA AGENAS MEDICINA LEGALE IN CALCE AL FOGLIO
# ============================================================

def _scrivi_tabella_agenas_medicina_legale(
        ws, start_row, df_luogo, fabb_medleg_presidio,
        mapping_uo_medleg, mapping_profili_medleg, sede_completa,
        livello_label):
    """Aggiunge una tabella riepilogativa AGENAS medicina legale
    (Tab. 18) in calce al foglio di dettaglio di un presidio ospedaliero.

    Restituisce la riga successiva libera.
    """
    import re as _re

    if not fabb_medleg_presidio:
        return start_row

    # --- Stili locali ---
    FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
    FONT_NORMAL  = Font(size=10)

    row = start_row + 1
    N_COLS_MERGE = 6

    # Titolo sezione
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    c = ws.cell(row=row, column=1,
                value="MEDICINA LEGALE - Standard AGENAS")
    c.font = FONT_SECTION
    row += 1

    # Nota presidio + livello
    livello_leggibile = {
        'OSPEDALE_DI_BASE':     'Ospedale di Base',
        'PRESIDIO_I_LIVELLO':   'Presidio di I livello',
        'PRESIDIO_II_LIVELLO':  'Presidio di II livello',
    }.get(livello_label, livello_label)

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS_MERGE)
    ws.cell(row=row, column=1,
            value=f"Presidio: {sede_completa}  \u2013  {livello_leggibile}"
            ).font = FONT_NORMAL
    row += 2

    # Intestazioni
    headers = ['Profilo', 'T.I.', 'T.D.', 'Totale', 'Range AGENAS',
               'Esito']
    N_COLS = len(headers)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # Nomi leggibili
    nomi_profili = {
        'DIRIGENTE_MEDICO':          'Dirigenti Medici',
        'INFERMIERE':                'Infermieri',
        'OPERATORE_SOCIO_SANITARIO': 'Operatori Socio Sanitari',
    }

    # Calcola gli in servizio per ciascun profilo
    in_servizio = {}
    in_servizio_td = {}
    for _, r in df_luogo.iterrows():
        ssd = str(r.get('_REPARTO', ''))
        profilo = str(r.get('Profilo Professionale', '')).strip().upper()
        quantita_ti = int(r.get('Quantità T.I.', 0))
        quantita_td = int(r.get('Quantità T.D.', 0))

        uo_ok = False
        for m_uo in mapping_uo_medleg:
            if _re.search(m_uo['pattern'], ssd.upper()):
                uo_ok = True
                break
        if not uo_ok:
            continue

        for m_prof in mapping_profili_medleg:
            if _re.search(m_prof['pattern'], profilo):
                pa = m_prof['profilo_agenas']
                in_servizio[pa] = in_servizio.get(pa, 0) + quantita_ti
                in_servizio_td[pa] = in_servizio_td.get(pa, 0) + quantita_td
                break

    # Righe dati
    FILL_OK = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                          fill_type='solid')
    FILL_CARENZA = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                               fill_type='solid')
    toggle = 0
    for prof_key, rng in fabb_medleg_presidio.items():
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1
        servizio_ti = in_servizio.get(prof_key, 0)
        servizio_td = in_servizio_td.get(prof_key, 0)
        servizio = servizio_ti + servizio_td
        v_min = rng['min']
        v_max = rng['max']
        range_str = f'{v_min} - {v_max}'

        if servizio < v_min:
            esito = f'CARENZA (min {v_min - servizio})'
            fill_esito = FILL_CARENZA
        elif servizio > v_max:
            esito = f'ECCEDENZA (+{servizio - v_max})'
            fill_esito = FILL_OK
        else:
            esito = 'IN RANGE'
            fill_esito = FILL_OK

        vals = [
            nomi_profili.get(prof_key, prof_key),
            servizio_ti, servizio_td, servizio, range_str, esito,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        row += 1

    # Nota piè di pagina
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=6)
    ws.cell(row=row, column=1,
            value="Maggiorazione +15% per turnazione, guardie, "
                  "ferie, malattie e altre indisponibilità."
            ).font = Font(italic=True, size=9, color='555555')

    return row + 1
