"""
Dettaglio nominativo del personale medico e non medico per atto aziendale.

Genera un file Excel con:
 - Un foglio per ogni disciplina (medici/veterinari) con l'elenco nominativo
 - Un foglio per ogni profilo professionale (altri) con l'elenco nominativo

Colonne per ogni dipendente:
  Matricola | Cognome | Nome | Data Nascita | Forma Contrattuale |
  Disciplina / Profilo | Sede Fisica | Centro di Costo | SSD | Data Assunzione
"""

import os
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.caricamento_dati import (
    carica_dataframe,
    normalizza_colonne_personale,
)
from src.caricamento_xml import (
    carica_medici_atto_aziendale,
    carica_profili_atto_aziendale,
)
from src.config import (
    FILE_MEDICI_ATTO_AZIENDALE,
    FILE_PROFILI_ATTO_AZIENDALE,
)
from src.stili_excel import auto_larghezza_colonne

# ─────────────────────────────────────────────────────────────
# COSTANTI STILE
# ─────────────────────────────────────────────────────────────
_FILL_HEADER  = PatternFill('solid', fgColor='1F4E79')  # blu scuro
_FILL_SUBHDR  = PatternFill('solid', fgColor='2E75B6')  # blu medio (titolo foglio)
_FILL_A       = PatternFill('solid', fgColor='DCE6F1')  # azzurro chiaro
_FILL_B       = PatternFill('solid', fgColor='FFFFFF')  # bianco
_FILL_TOTALE  = PatternFill('solid', fgColor='BDD7EE')  # azzurro totale

_FONT_HEADER  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
_FONT_TITOLO  = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
_FONT_BODY    = Font(name='Calibri', size=10)
_FONT_TOTALE  = Font(name='Calibri', bold=True, size=10)

_THIN = Side(style='thin', color='B8CCE4')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_C  = Alignment(horizontal='center', vertical='center', wrap_text=False)
_ALIGN_L  = Alignment(horizontal='left',   vertical='center', wrap_text=False)
_ALIGN_CW = Alignment(horizontal='center', vertical='center', wrap_text=True)


# ─────────────────────────────────────────────────────────────
# MAPPING FORMA CONTRATTUALE
# ─────────────────────────────────────────────────────────────
def _forma_contrattuale(desc_natura: str) -> str:
    """Traduce DESC_NATURA in etichetta leggibile."""
    n = str(desc_natura).strip().upper()
    if n.startswith('COMANDATO IN USCITA'):
        return 'T.I. – Comando in uscita'
    if n in ('PENITENZIARIO INDETERMINATO',):
        return 'T.I.'
    if n.startswith('TEMPO INDETERMINATO ASP'):
        return 'T.I.'
    if n == 'TEMPO INDETERMINATO':
        return 'T.I.'
    if '15_OCTIES' in n:
        return 'T.D. – Art. 15 Octies'
    if n == 'TEMPO DETERMINATO ART. 15 SEPTIES DLGS 502/92':
        return 'T.D. – Art. 15 Septies'
    if n == 'TEMPO DETERMINATO':
        return 'T.D.'
    if n == 'T.D. SPECIALIZZANDI':
        return 'Specializzando'
    if n == 'UNIVERSITARI H19':
        return 'Universitario'
    return str(desc_natura).strip()


# Ordine di priorità per la forma contrattuale nell'ordinamento righe
_ORDINE_NATURA = {
    'TEMPO INDETERMINATO': 1,
    'PENITENZIARIO INDETERMINATO': 2,
}
# TI – Comando in uscita → 3
# TI – ASP → 4
# TD puro → 5
# TD 15 Septies → 6
# TD 15 Octies → 7
# Specializzando → 8
# Universitario → 9


def _ordine_natura(desc_natura: str) -> int:
    """Restituisce un intero di ordinamento per DESC_NATURA (TI prima, poi TD, poi altro)."""
    n = str(desc_natura).strip().upper()
    if n == 'TEMPO INDETERMINATO':
        return 1
    if n == 'PENITENZIARIO INDETERMINATO':
        return 2
    if n.startswith('COMANDATO IN USCITA'):
        return 3
    if n.startswith('TEMPO INDETERMINATO ASP'):
        return 4
    if n == 'TEMPO DETERMINATO':
        return 5
    if n == 'TEMPO DETERMINATO ART. 15 SEPTIES DLGS 502/92':
        return 6
    if '15_OCTIES' in n:
        return 7
    if n == 'T.D. SPECIALIZZANDI':
        return 8
    if n == 'UNIVERSITARI H19':
        return 9
    return 99


# ─────────────────────────────────────────────────────────────
# SCRITTURA FOGLIO NOMINATIVO
# ─────────────────────────────────────────────────────────────
def _scrivi_foglio_nominativo(wb: Workbook, sheet_name: str,
                               titolo: str, df_pers: pd.DataFrame,
                               col_gruppo: str,
                               col_colore: str = 'DESC_SEDE_FISICA') -> None:
    """
    Scrive un foglio Excel con l'elenco nominativo dei dipendenti.

    Riga 1  → titolo (merge su tutta la larghezza)
    Riga 2  → intestazioni colonne
    Riga 3+ → dati, alternanza colori per gruppo (default: sede fisica)
    Ultima  → contatore totale dipendenti
    """
    ws = wb.create_sheet(title=sheet_name[:31])

    COLS = [
        'Matricola', 'Cognome', 'Nome', 'Data Nascita',
        'Forma Contrattuale',
        col_gruppo,
        'Sede Fisica', 'Centro di Costo', 'SSD', 'Data Assunzione',
        'Data Cessazione',
    ]
    n_cols = len(COLS)

    # ── Riga 1: Titolo ────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1, value=titolo)
    tc.font      = _FONT_TITOLO
    tc.fill      = _FILL_SUBHDR
    tc.alignment = _ALIGN_CW
    tc.border    = _BORDER
    ws.row_dimensions[1].height = 22

    # ── Riga 2: Intestazioni ──────────────────────────────────
    for ci, col in enumerate(COLS, 1):
        c = ws.cell(row=2, column=ci, value=col)
        c.font      = _FONT_HEADER
        c.fill      = _FILL_HEADER
        c.alignment = _ALIGN_C
        c.border    = _BORDER
    ws.row_dimensions[2].height = 18

    # ── Dati ──────────────────────────────────────────────────
    color_toggle = 0
    prev_grp     = None
    row          = 3

    for _, r in df_pers.iterrows():
        grp = r.get(col_colore, '')
        if grp != prev_grp:
            color_toggle = 1 - color_toggle
            prev_grp = grp
        fill = _FILL_A if color_toggle else _FILL_B

        # Formatta data nascita e data assunzione
        dt_nasc  = r.get('DT_NASCITA', '')
        dt_ass   = r.get('PRIMA_DATA_ASSUNZIONE', '')
        dt_cess  = r.get('DT_CESSAZIONE', '')
        try:
            dt_nasc = pd.to_datetime(dt_nasc).strftime('%d/%m/%Y')
        except Exception:
            dt_nasc = str(dt_nasc) if pd.notna(dt_nasc) else ''
        try:
            dt_ass = pd.to_datetime(dt_ass).strftime('%d/%m/%Y')
        except Exception:
            dt_ass = str(dt_ass) if pd.notna(dt_ass) else ''
        try:
            dt_cess = pd.to_datetime(dt_cess).strftime('%d/%m/%Y')
        except Exception:
            dt_cess = str(dt_cess) if pd.notna(dt_cess) else ''

        valori = [
            r.get('MATR.', ''),
            r.get('PF_COGNOME', ''),
            r.get('PF_NOME', ''),
            dt_nasc,
            _forma_contrattuale(r.get('DESC_NATURA', '')),
            r.get(col_gruppo, ''),
            r.get('DESC_SEDE_FISICA', ''),
            r.get('DESC_TIPO_CDC', ''),
            r.get('DESC_SC_SSD', ''),
            dt_ass,
            dt_cess,
        ]

        for ci, val in enumerate(valori, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font      = _FONT_BODY
            c.fill      = fill
            c.border    = _BORDER
            c.alignment = _ALIGN_C if ci in (1, 4, 5, 10, 11) else _ALIGN_L
        row += 1

    # ── Riga totale ───────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=n_cols)
    tc2 = ws.cell(row=row, column=1,
                  value=f'Totale dipendenti: {len(df_pers)}')
    tc2.font      = _FONT_TOTALE
    tc2.fill      = _FILL_TOTALE
    tc2.alignment = _ALIGN_C
    tc2.border    = _BORDER

    # ── Larghezze colonne ─────────────────────────────────────
    auto_larghezza_colonne(ws, COLS)
    # Aggiustamenti manuali per colonne note
    ws.column_dimensions['A'].width = 12   # Matricola
    ws.column_dimensions['B'].width = 20   # Cognome
    ws.column_dimensions['C'].width = 16   # Nome
    ws.column_dimensions['D'].width = 14   # Data nascita
    ws.column_dimensions['E'].width = 26   # Forma contrattuale
    ws.column_dimensions['F'].width = 30   # Disciplina / Profilo
    ws.column_dimensions['G'].width = 28   # Sede fisica
    ws.column_dimensions['H'].width = 42   # Centro di costo
    ws.column_dimensions['I'].width = 36   # SSD
    ws.column_dimensions['J'].width = 16   # Data assunzione
    ws.column_dimensions['K'].width = 16   # Data cessazione

    # Blocca le prime 2 righe (titolo + intestazioni)
    ws.freeze_panes = 'A3'


# ─────────────────────────────────────────────────────────────
# FOGLIO RIEPILOGO INIZIALE
# ─────────────────────────────────────────────────────────────
def _scrivi_foglio_riepilogo(wb: Workbook, titolo_wb: str,
                              righe: list[dict], anno_analisi: int) -> None:
    """
    Scrive un foglio 'RIEPILOGO' in testa al workbook con il conteggio
    per disciplina/profilo, dotazione atto e delta.
    """
    ws = wb.create_sheet(title='RIEPILOGO', index=0)

    COLS = ['Disciplina / Profilo', 'Dotazione Atto',
            'T.I.', 'T.D.', 'Univ./Spec.', 'Totale', 'Delta']
    n_cols = len(COLS)

    # Titolo
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1, value=titolo_wb)
    tc.font      = _FONT_TITOLO
    tc.fill      = _FILL_SUBHDR
    tc.alignment = _ALIGN_CW
    tc.border    = _BORDER
    ws.row_dimensions[1].height = 22

    # Intestazioni
    for ci, col in enumerate(COLS, 1):
        c = ws.cell(row=2, column=ci, value=col)
        c.font      = _FONT_HEADER
        c.fill      = _FILL_HEADER
        c.alignment = _ALIGN_C
        c.border    = _BORDER
    ws.row_dimensions[2].height = 18

    # Dati
    row = 3
    for idx, r in enumerate(righe):
        fill = _FILL_A if idx % 2 == 0 else _FILL_B
        delta = r.get('delta', '')
        valori = [
            r['nome'],
            r.get('dotazione', ''),
            r.get('n_ti', 0),
            r.get('n_td', 0),
            r.get('n_us', 0),
            r.get('n_tot', 0),
            delta,
        ]
        for ci, val in enumerate(valori, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font      = _FONT_BODY
            c.fill      = fill
            c.border    = _BORDER
            c.alignment = _ALIGN_C if ci > 1 else _ALIGN_L

            # Colora delta rosso se negativo
            if ci == 7 and isinstance(val, (int, float)) and val < 0:
                c.font = Font(name='Calibri', bold=True,
                              color='C00000', size=10)
        row += 1

    # Riga totali
    tot_dot = sum(r.get('dotazione', 0) for r in righe
                  if isinstance(r.get('dotazione'), int))
    tot_ti  = sum(r.get('n_ti', 0) for r in righe)
    tot_td  = sum(r.get('n_td', 0) for r in righe)
    tot_us  = sum(r.get('n_us', 0) for r in righe)
    tot_tot = sum(r.get('n_tot', 0) for r in righe)
    tot_del = tot_tot - tot_dot

    tot_vals = ['TOTALE', tot_dot, tot_ti, tot_td, tot_us, tot_tot, tot_del]
    for ci, val in enumerate(tot_vals, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font      = _FONT_TOTALE
        c.fill      = _FILL_TOTALE
        c.border    = _BORDER
        c.alignment = _ALIGN_C if ci > 1 else _ALIGN_L
        if ci == 7 and isinstance(val, (int, float)) and val < 0:
            c.font = Font(name='Calibri', bold=True, color='C00000', size=10)

    # Larghezze
    ws.column_dimensions['A'].width = 40
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 16
    ws.freeze_panes = 'A3'


# ─────────────────────────────────────────────────────────────
# GENERA DETTAGLIO MEDICI
# ─────────────────────────────────────────────────────────────
def genera_dettaglio_medici(personale_file: str,
                             mapper_atto_aziendale: str,
                             anno_analisi: int,
                             output_dir: str,
                             pensionamenti_file: str = None) -> str:
    """
    Genera 'dettaglio_medici_{anno}.xlsx':
    - Foglio 'RIEPILOGO' (conteggi per disciplina)
    - Un foglio per ogni disciplina dell'atto con l'elenco nominativo
    - Un foglio 'FUORI ATTO' per medici in discipline non mappate
    """
    personale_df = carica_dataframe(personale_file)
    personale_df = normalizza_colonne_personale(personale_df)

    # DT_CESSAZIONE è già nel file personale; assicura la colonna esista
    if 'DT_CESSAZIONE' not in personale_df.columns:
        personale_df['DT_CESSAZIONE'] = ''

    # ── Escludi dipendenti già cessati ────────────────────────
    _oggi = date.today()
    _dt_cess = pd.to_datetime(personale_df['DT_CESSAZIONE'], errors='coerce')
    _mask_cessati = _dt_cess.notna() & (_dt_cess.dt.date <= _oggi)
    _n_cessati = int(_mask_cessati.sum())
    if _n_cessati:
        print(f"  [Dettaglio medici] Esclusi {_n_cessati} dipendenti già cessati alla data {_oggi}")
        personale_df = personale_df[~_mask_cessati].copy()

    # Solo dirigenti medici e veterinari
    medici = personale_df[
        personale_df['PROFILO_RAGGRUPPATO'].str.upper().isin(
            ['DIRIGENTE MEDICO', 'DIRIGENTE VETERINARIO']
        )
    ].copy()
    medici['DISC_UPPER'] = (
        medici['DESC_DISCIPLINE'].str.upper().str.strip()
        .fillna('DIRIGENTE MEDICO')
    )

    natura_upper = medici['DESC_NATURA'].str.upper()
    mask_td  = (
        natura_upper.isin([
            'TEMPO DETERMINATO',
            'TEMPO DETERMINATO ART. 15 SEPTIES DLGS 502/92',
        ]) | natura_upper.str.contains('15_OCTIES', na=False)
    )
    mask_us  = natura_upper.isin(['UNIVERSITARI H19', 'T.D. SPECIALIZZANDI'])
    mask_ti  = ~mask_td & ~mask_us

    mapper = carica_medici_atto_aziendale(mapper_atto_aziendale)

    wb = Workbook()
    wb.remove(wb.active)

    righe_riep  = []
    disc_mappate = set()

    for disc in mapper:
        nome_atto   = disc['nome_atto']
        dotazione   = disc['dotazione']
        voci_db     = disc['discipline_db']

        df_disc = medici[medici['DISC_UPPER'].isin(voci_db)].copy()
        # Aggiunge colonna leggibile usata nel foglio nominativo
        df_disc['Disciplina'] = nome_atto

        n_ti  = int(mask_ti[df_disc.index].sum())
        n_td  = int(mask_td[df_disc.index].sum())
        n_us  = int(mask_us[df_disc.index].sum())
        n_tot = len(df_disc)

        righe_riep.append({
            'nome':      nome_atto,
            'dotazione': dotazione,
            'n_ti':      n_ti,
            'n_td':      n_td,
            'n_us':      n_us,
            'n_tot':     n_tot,
            'delta':     n_tot - dotazione,
        })
        disc_mappate.update(voci_db)

        # Ordina: sede → centro di costo → forma contrattuale → cognome → nome
        df_disc['_ORD_NATURA'] = df_disc['DESC_NATURA'].apply(_ordine_natura)
        df_disc = df_disc.sort_values(
            ['DESC_SEDE_FISICA', 'DESC_TIPO_CDC', '_ORD_NATURA',
             'PF_COGNOME', 'PF_NOME']
        ).drop(columns=['_ORD_NATURA'])

        titolo = (
            f'{nome_atto} – {anno_analisi} '
            f'(Dotaz. atto: {dotazione} | In servizio: {n_tot})'
        )
        # Nome foglio: max 31 caratteri, safe
        sheet_name = nome_atto[:31]
        _scrivi_foglio_nominativo(
            wb, sheet_name, titolo, df_disc, 'Disciplina'
        )

    # Medici fuori atto
    df_fuori = medici[~medici['DISC_UPPER'].isin(disc_mappate)].copy()
    if not df_fuori.empty:
        df_fuori['Disciplina'] = df_fuori['DESC_DISCIPLINE'].fillna(
            'N.D.'
        ).str.title()
        df_fuori['_ORD_NATURA'] = df_fuori['DESC_NATURA'].apply(_ordine_natura)
        df_fuori = df_fuori.sort_values(
            ['Disciplina', 'DESC_SEDE_FISICA', 'DESC_TIPO_CDC',
             '_ORD_NATURA', 'PF_COGNOME', 'PF_NOME']
        ).drop(columns=['_ORD_NATURA'])

        n_ti_f  = int(mask_ti[df_fuori.index].sum())
        n_td_f  = int(mask_td[df_fuori.index].sum())
        n_us_f  = int(mask_us[df_fuori.index].sum())
        righe_riep.append({
            'nome':      '⚠ FUORI ATTO',
            'dotazione': '–',
            'n_ti':      n_ti_f,
            'n_td':      n_td_f,
            'n_us':      n_us_f,
            'n_tot':     len(df_fuori),
            'delta':     '',
        })
        _scrivi_foglio_nominativo(
            wb, 'FUORI ATTO', 'Medici – Discipline fuori Atto Aziendale',
            df_fuori, 'Disciplina'
        )

    # Riepilogo in testa
    _scrivi_foglio_riepilogo(
        wb,
        f'Fabbisogno Atto Aziendale – Medici e Veterinari ({anno_analisi})',
        righe_riep,
        anno_analisi,
    )

    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir,
                               f'dettaglio_medici_{anno_analisi}.xlsx')
    wb.save(output_file)
    print(f"  Dettaglio medici salvato in: {output_file}")
    return output_file


# ─────────────────────────────────────────────────────────────
# GENERA DETTAGLIO ALTRI PROFILI
# ─────────────────────────────────────────────────────────────
def genera_dettaglio_altri(personale_file: str,
                            profili_atto_xml: str,
                            anno_analisi: int,
                            output_dir: str,
                            pensionamenti_file: str = None) -> str:
    """
    Genera 'dettaglio_altri_{anno}.xlsx':
    - Foglio 'RIEPILOGO' (conteggi per profilo)
    - Un foglio per ogni profilo professionale dell'atto con l'elenco
    - Un foglio 'FUORI ATTO' per personale non mappato
    """
    _PROFILI_MEDICI = {'DIRIGENTE MEDICO', 'DIRIGENTE VETERINARIO'}

    personale_df = carica_dataframe(personale_file)
    personale_df = normalizza_colonne_personale(personale_df)

    # DT_CESSAZIONE è già nel file personale; assicura la colonna esista
    if 'DT_CESSAZIONE' not in personale_df.columns:
        personale_df['DT_CESSAZIONE'] = ''

    # ── Escludi dipendenti già cessati ────────────────────────
    _oggi = date.today()
    _dt_cess = pd.to_datetime(personale_df['DT_CESSAZIONE'], errors='coerce')
    _mask_cessati = _dt_cess.notna() & (_dt_cess.dt.date <= _oggi)
    _n_cessati = int(_mask_cessati.sum())
    if _n_cessati:
        print(f"  [Dettaglio altri] Esclusi {_n_cessati} dipendenti già cessati alla data {_oggi}")
        personale_df = personale_df[~_mask_cessati].copy()

    # Escludi medici/veterinari
    personale_df = personale_df[
        ~personale_df['PROFILO_RAGGRUPPATO'].str.upper().isin(_PROFILI_MEDICI)
    ].copy()

    natura_upper = personale_df['DESC_NATURA'].str.upper()
    mask_td = (
        natura_upper.isin([
            'TEMPO DETERMINATO',
            'TEMPO DETERMINATO ART. 15 SEPTIES DLGS 502/92',
        ]) | natura_upper.str.contains('15_OCTIES', na=False)
    )
    mask_us  = natura_upper.isin(['UNIVERSITARI H19', 'T.D. SPECIALIZZANDI'])
    mask_ti  = ~mask_td & ~mask_us

    mapper = carica_profili_atto_aziendale(profili_atto_xml)
    # Escludi profili medici (già nel file dettaglio_medici)
    mapper = [p for p in mapper
              if p['nome_atto'].upper() not in _PROFILI_MEDICI]

    wb = Workbook()
    wb.remove(wb.active)

    righe_riep    = []
    profili_mappati = set()

    for profilo in mapper:
        nome_atto  = profilo['nome_atto']
        dotazione  = profilo['dotazione']

        # Il match usa PROFILO_RAGGRUPPATO (già calcolato da normalizza_colonne_personale)
        # che corrisponde direttamente al nome_atto del profilo
        df_prof = personale_df[
            personale_df['PROFILO_RAGGRUPPATO'].str.upper() == nome_atto.upper()
        ].copy()
        df_prof['Profilo'] = nome_atto

        n_ti  = int(mask_ti[df_prof.index].sum())
        n_td  = int(mask_td[df_prof.index].sum())
        n_us  = int(mask_us[df_prof.index].sum())
        n_tot = len(df_prof)

        righe_riep.append({
            'nome':      nome_atto,
            'dotazione': dotazione,
            'n_ti':      n_ti,
            'n_td':      n_td,
            'n_us':      n_us,
            'n_tot':     n_tot,
            'delta':     n_tot - dotazione,
        })
        profili_mappati.add(nome_atto.upper())

        # Ordina: sede → centro di costo → forma contrattuale → cognome → nome
        df_prof['_ORD_NATURA'] = df_prof['DESC_NATURA'].apply(_ordine_natura)
        df_prof = df_prof.sort_values(
            ['DESC_SEDE_FISICA', 'DESC_TIPO_CDC', '_ORD_NATURA',
             'PF_COGNOME', 'PF_NOME']
        ).drop(columns=['_ORD_NATURA'])
        titolo  = (
            f'{nome_atto} – {anno_analisi} '
            f'(Dotaz. atto: {dotazione} | In servizio: {n_tot})'
        )
        sheet_name = nome_atto[:31]
        _scrivi_foglio_nominativo(
            wb, sheet_name, titolo, df_prof, 'Profilo'
        )

    # Personale fuori atto
    df_fuori = personale_df[
        ~personale_df['PROFILO_RAGGRUPPATO'].str.upper().isin(profili_mappati)
    ].copy()
    if not df_fuori.empty:
        df_fuori['Profilo'] = df_fuori['PROFILO_RAGGRUPPATO'].fillna(
            'N.D.'
        ).str.title()
        df_fuori['_ORD_NATURA'] = df_fuori['DESC_NATURA'].apply(_ordine_natura)
        df_fuori = df_fuori.sort_values(
            ['Profilo', 'DESC_SEDE_FISICA', 'DESC_TIPO_CDC',
             '_ORD_NATURA', 'PF_COGNOME', 'PF_NOME']
        ).drop(columns=['_ORD_NATURA'])

        n_ti_f  = int(mask_ti[df_fuori.index].sum())
        n_td_f  = int(mask_td[df_fuori.index].sum())
        n_us_f  = int(mask_us[df_fuori.index].sum())
        righe_riep.append({
            'nome':      '⚠ FUORI ATTO',
            'dotazione': '–',
            'n_ti':      n_ti_f,
            'n_td':      n_td_f,
            'n_us':      n_us_f,
            'n_tot':     len(df_fuori),
            'delta':     '',
        })
        _scrivi_foglio_nominativo(
            wb, 'FUORI ATTO', 'Altri profili – Fuori Atto Aziendale',
            df_fuori, 'Profilo'
        )

    # Riepilogo in testa
    _scrivi_foglio_riepilogo(
        wb,
        f'Fabbisogno Atto Aziendale – Altri Profili ({anno_analisi})',
        righe_riep,
        anno_analisi,
    )

    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir,
                               f'dettaglio_altri_{anno_analisi}.xlsx')
    wb.save(output_file)
    print(f"  Dettaglio altri profili salvato in: {output_file}")
    return output_file
