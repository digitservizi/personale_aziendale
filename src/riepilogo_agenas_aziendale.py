"""
Foglio riepilogo AGENAS aziendale per il file riepilogo_aziendale.xlsx.

Genera un foglio 'FABBISOGNO AGENAS' che consolida le tabelle AGENAS
di tutte le aree (Campobasso, Isernia, Termoli) e aggiunge una tabella
TOTALE AZIENDALE per ciascun indicatore.

Le eccedenze/carenze TOTALI sono calcolate come somma delle
eccedenze/carenze dei singoli presidi, NON come valutazione globale
contro il range sommato.
"""

import re as _re

from openpyxl.styles import Font, PatternFill

from src.stili_excel import (
    THIN_BORDER, FILL_A, FILL_B, FILL_HEADER,
    FONT_HEADER, ALIGN_CENTER,
    auto_larghezza_colonne,
)
from src.tabelle_agenas import (
    _scrivi_tabella_agenas_materno_infantile,
    _scrivi_tabella_agenas_radiologia,
    _scrivi_tabella_agenas_emergenza_urgenza,
    _scrivi_tabella_agenas_terapia_intensiva,
    _scrivi_tabella_agenas_sale_operatorie,
    _scrivi_tabella_agenas_area_ti_bo,
    _scrivi_tabella_agenas_anatomia_patologica,
    _scrivi_tabella_agenas_laboratorio,
    _scrivi_tabella_agenas_tecnici_laboratorio,
    _scrivi_tabella_agenas_medicina_legale,
    _scrivi_tabella_agenas_trasfusionale,
    _scrivi_tabella_fabbisogno_uoc_trasfusionale,
    _scrivi_tabella_agenas_territoriale,
)



# ============================================================
# STILI COMUNI
# ============================================================

_FILL_DIVIDER = PatternFill(
    start_color='1F4E79', end_color='1F4E79', fill_type='solid')
_FONT_DIVIDER = Font(bold=True, size=14, color='FFFFFF')

FILL_OK = PatternFill(
    start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FILL_CARENZA = PatternFill(
    start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
FILL_REGIME = PatternFill(
    start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')

FONT_SECTION = Font(bold=True, size=12, color='1F4E79')
FONT_NOTE = Font(italic=True, size=9, color='555555')

FILL_ALTRE_AREE = PatternFill(
    start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')   # giallino


# ============================================================
# HELPER: CONTEGGIO PERSONALE FUORI AREA (ALTRE AREE/MANSIONI)
# ============================================================

def _conta_fuori_area(personale_all_df, discipline_db, tot_in_area):
    """Calcola per ogni profilo_agenas quante persone nel DB hanno la
    disciplina corrispondente ma NON risultano conteggiate in area.

    Parametri
    ---------
    personale_all_df : DataFrame  – dati completi personale (tutte le sedi)
    discipline_db : list[dict]    – lista di {pattern, profilo_agenas,
                                    uo_in_area (opz.)} dal tag
                                    <discipline_db> dell'XML
    tot_in_area : dict            – {profilo_agenas: n_in_area}
                                    totale conteggiato nei presidi

    Ritorna
    -------
    (dict, list)
        dict  {profilo_agenas: n_fuori_area}
        list  nomi UO (DESC_SC_SSD_SS) unici delle persone fuori area
    """
    if not discipline_db or personale_all_df is None:
        return {}, []

    import pandas as _pd

    result = {}
    # Maschera cumulativa "fuori area" per raccogliere le UO
    fuori_mask_global = _pd.Series(False, index=personale_all_df.index)

    # Recupera il pattern uo_in_area (uguale per tutti gli entry)
    uo_in_area_pat = None
    for entry in discipline_db:
        if entry.get('uo_in_area'):
            uo_in_area_pat = entry['uo_in_area']
            break

    sede_col = personale_all_df['DESC_SEDE_FISICA'].astype(str).str.upper()
    uo_col = personale_all_df['DESC_SC_SSD_SS'].astype(str).str.upper()

    # Maschera: persona è "in area" se è in un P.O. OPPURE la sua UO
    # matcha il pattern uo_in_area dell'area AGENAS
    in_po_mask = sede_col.str.contains('P\\.O\\.', na=False)
    if uo_in_area_pat:
        in_uo_area_mask = uo_col.str.contains(
            uo_in_area_pat, na=False, regex=True)
    else:
        in_uo_area_mask = _pd.Series(False, index=personale_all_df.index)
    in_area_mask = in_po_mask | in_uo_area_mask

    for entry in discipline_db:
        pat = entry.get('pattern', '')
        prof = entry.get('profilo_agenas', '')
        if not pat or not prof:
            continue

        disc_col = personale_all_df['DESC_DISCIPLINE'].astype(str).str.upper()
        match_mask = disc_col.str.contains(pat, na=False, regex=True)

        # Fuori area: ha la disciplina ma non è "in area"
        fuori_mask = match_mask & ~in_area_mask
        n_fuori = int(fuori_mask.sum())
        result[prof] = result.get(prof, 0) + n_fuori
        fuori_mask_global = fuori_mask_global | fuori_mask

    # Identifica UO dei fuori area
    uo_fuori = (
        personale_all_df.loc[fuori_mask_global, 'DESC_SC_SSD_SS']
        .dropna().astype(str).str.strip()
    )
    # Aggiungi la sede tra parentesi per disambiguare
    sede_fuori = (
        personale_all_df.loc[fuori_mask_global, 'DESC_SEDE_FISICA']
        .fillna('').astype(str).str.strip()
    )
    uo_con_sede = []
    for u, s in zip(uo_fuori, sede_fuori):
        if not u or u == 'nan':
            continue
        # Estrai sigla breve della sede (es. "IS" da "ISERNIA - ...")
        sigla = s.split(' -')[0].strip() if s else ''
        label = f"{u} ({sigla})" if sigla else u
        uo_con_sede.append(label)
    uo_uniche = sorted(set(uo_con_sede))

    return result, uo_uniche


# ============================================================
# HELPER: TABELLA DISTRIBUZIONE ANESTESISTI PER CDC
# ============================================================

def _scrivi_distribuzione_anestesisti(ws, start_row, personale_all_df):
    """Scrive una tabellina con la distribuzione degli anestesisti per CDC.

    Mostra in quali centri di costo sono distribuiti i dirigenti medici
    con disciplina Anestesia, suddivisi per presidio.

    Ritorna la riga successiva libera.
    """
    if personale_all_df is None:
        return start_row

    import pandas as _pd

    disc_col = personale_all_df['DESC_DISCIPLINE'].astype(str).str.upper()
    prof_col = (personale_all_df['DESC_PROFILO_PROFESSIONALE']
                .astype(str).str.upper())
    mask = (disc_col.str.contains('ANESTESIA', na=False, regex=True)
            & prof_col.str.contains('DIRIGENTE MEDICO', na=False))
    anest = personale_all_df[mask].copy()

    if anest.empty:
        return start_row

    anest['DESC_TIPO_CDC'] = anest['DESC_TIPO_CDC'].fillna('(non assegnato)')
    anest['DESC_SEDE_FISICA'] = anest['DESC_SEDE_FISICA'].fillna('(vuoto)')

    # Calcolo sigla sede (CB, IS, TE, ...)
    anest['_SIGLA_SEDE'] = (
        anest['DESC_SEDE_FISICA'].astype(str)
        .str.split(' -').str[0].str.strip()
        .str[:2]   # prime 2 lettere es. CA→CB approx
    )
    # Meglio usare la prima parola abbreviata
    def _sigla(sede):
        s = str(sede).upper().strip()
        if 'CAMPOBASSO' in s:
            return 'CB'
        if 'ISERNIA' in s:
            return 'IS'
        if 'TERMOLI' in s or 'LARINO' in s:
            return 'TE'
        return s.split(' ')[0][:4]
    anest['_SIGLA_SEDE'] = anest['DESC_SEDE_FISICA'].apply(_sigla)

    # Semplifica nome CDC (rimuove prefisso "OSP. xxx -")
    # e accorpa i costi comuni con la terapia intensiva
    def _cdc_breve(cdc):
        s = str(cdc)
        # Rimuove "OSP. CARDARELLI - ", "OSP. S. TIMOTEO - ", ecc.
        parts = s.split(' - ', 1)
        if len(parts) > 1 and parts[0].startswith('OSP.'):
            s = parts[1].strip()
        # Accorpa "COSTI COMUNI" → "TERAPIA INTENSIVA - DEGENZE ORD."
        if 'COSTI COMUNI' in s.upper():
            s = 'TERAPIA INTENSIVA - DEGENZE ORD.'
        return s
    anest['_CDC_BREVE'] = anest['DESC_TIPO_CDC'].apply(_cdc_breve)

    # Raggruppa per CDC breve
    grp = (anest.groupby('_CDC_BREVE')
           .agg(
               N=('_CDC_BREVE', 'size'),
               sedi=('_SIGLA_SEDE', lambda x: x.value_counts()
                     .to_dict()),
           )
           .sort_values('N', ascending=False)
           .reset_index())

    row = start_row
    # Titolo
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=5)
    ws.cell(row=row, column=1,
            value="Distribuzione Dirigenti Medici Anestesisti "
                  f"per Centro di Costo (tot. {len(anest)})"
            ).font = Font(bold=True, size=10, color='1F4E79')
    row += 1

    # Header
    headers = ['Centro di Costo', 'N.', 'Distribuzione per sede']
    fills_h = [FILL_HEADER, FILL_HEADER, FILL_HEADER]
    for ci, h in enumerate(headers):
        c = ws.cell(row=row, column=ci + 1, value=h)
        c.font = FONT_HEADER
        c.fill = fills_h[ci]
        c.border = THIN_BORDER
        c.alignment = ALIGN_CENTER
    row += 1

    # Righe dati
    for i, r in grp.iterrows():
        fill = FILL_A if (i % 2 == 0) else FILL_B
        sedi_str = ", ".join(
            f"{k}={v}" for k, v in sorted(r['sedi'].items()))
        vals = [r['_CDC_BREVE'], r['N'], sedi_str]
        for ci, v in enumerate(vals):
            c = ws.cell(row=row, column=ci + 1, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        row += 1

    # Nota
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=5)
    # Conta quanti sono in P.O.
    n_po = int(anest['DESC_SEDE_FISICA'].astype(str).str.upper()
               .str.contains('P\\.O\\.', na=False).sum())
    n_fuori = len(anest) - n_po
    ws.cell(row=row, column=1,
            value=f"Di cui {n_po} assegnati ai presidi ospedalieri "
                  f"e {n_fuori} ad altre strutture/mansioni "
                  f"(cfr. colonna \"Altre aree\")."
            ).font = FONT_NOTE

    return row + 2


def _scrivi_distribuzione_radiologi(ws, start_row, personale_all_df):
    """Scrive una tabellina con la distribuzione dei radiologi per CDC.

    Mostra in quali centri di costo sono distribuiti i dirigenti medici
    con disciplina Radiodiagnostica / Medicina Nucleare, suddivisi per
    presidio.

    Ritorna la riga successiva libera.
    """
    if personale_all_df is None:
        return start_row

    import pandas as _pd

    disc_col = personale_all_df['DESC_DISCIPLINE'].astype(str).str.upper()
    prof_col = (personale_all_df['DESC_PROFILO_PROFESSIONALE']
                .astype(str).str.upper())
    mask = (disc_col.str.contains(
                'RADIODIAGNOSTICA|RADIOTERAPIA|NEURORADIOLOGIA'
                '|MEDICINA NUCLEARE', na=False, regex=True)
            & prof_col.str.contains('DIRIGENTE MEDICO', na=False))
    radio = personale_all_df[mask].copy()

    if radio.empty:
        return start_row

    radio['DESC_TIPO_CDC'] = radio['DESC_TIPO_CDC'].fillna('(non assegnato)')
    radio['DESC_SEDE_FISICA'] = radio['DESC_SEDE_FISICA'].fillna('(vuoto)')

    # Sigla sede
    def _sigla(sede):
        s = str(sede).upper().strip()
        if 'CAMPOBASSO' in s:
            return 'CB'
        if 'ISERNIA' in s:
            return 'IS'
        if 'TERMOLI' in s or 'LARINO' in s:
            return 'TE'
        return s.split(' ')[0][:4]
    radio['_SIGLA_SEDE'] = radio['DESC_SEDE_FISICA'].apply(_sigla)

    # Semplifica nome CDC
    def _cdc_breve(cdc):
        s = str(cdc)
        parts = s.split(' - ', 1)
        if len(parts) > 1 and parts[0].startswith('OSP.'):
            s = parts[1].strip()
        return s
    radio['_CDC_BREVE'] = radio['DESC_TIPO_CDC'].apply(_cdc_breve)

    # Raggruppa per CDC breve
    grp = (radio.groupby('_CDC_BREVE')
           .agg(
               N=('_CDC_BREVE', 'size'),
               sedi=('_SIGLA_SEDE', lambda x: x.value_counts()
                     .to_dict()),
           )
           .sort_values('N', ascending=False)
           .reset_index())

    row = start_row
    # Titolo
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=5)
    ws.cell(row=row, column=1,
            value="Distribuzione Dirigenti Medici Radiologia "
                  f"per Centro di Costo (tot. {len(radio)})"
            ).font = Font(bold=True, size=10, color='1F4E79')
    row += 1

    # Header
    headers = ['Centro di Costo', 'N.', 'Distribuzione per sede']
    for ci, h in enumerate(headers):
        c = ws.cell(row=row, column=ci + 1, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.border = THIN_BORDER
        c.alignment = ALIGN_CENTER
    row += 1

    # Righe dati
    for i, r in grp.iterrows():
        fill = FILL_A if (i % 2 == 0) else FILL_B
        sedi_str = ", ".join(
            f"{k}={v}" for k, v in sorted(r['sedi'].items()))
        vals = [r['_CDC_BREVE'], r['N'], sedi_str]
        for ci, v in enumerate(vals):
            c = ws.cell(row=row, column=ci + 1, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        row += 1

    # Nota
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=5)
    n_po = int(radio['DESC_SEDE_FISICA'].astype(str).str.upper()
               .str.contains('P\\.O\\.', na=False).sum())
    n_fuori = len(radio) - n_po
    ws.cell(row=row, column=1,
            value=f"Di cui {n_po} assegnati ai presidi ospedalieri "
                  f"e {n_fuori} ad altre strutture/mansioni "
                  f"(cfr. colonna \"Altre aree\")."
            ).font = FONT_NOTE

    return row + 2


# ============================================================
# HELPER: DIVISORE VISUALE TRA SEZIONI
# ============================================================

def _scrivi_divisore_area(ws, row, titolo_area):
    """Barra colorata larga 8 colonne come separatore tra sezioni."""
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value=titolo_area)
    c.font = _FONT_DIVIDER
    c.fill = _FILL_DIVIDER
    c.alignment = ALIGN_CENTER
    c.border = THIN_BORDER
    ws.row_dimensions[row].height = 30
    return row + 2


# ============================================================
# FUNZIONI DI CONTEGGIO PER-PRESIDIO
# ============================================================

def _conta_generico(df, mapping_uo, mapping_profili,
                    esclusioni=None, match_cdc=True):
    """Conteggio generico: match UO su _REPARTO (e opzionalmente CDC),
    poi match profilo.

    Usato da: emergenza, anatomia patologica, laboratorio,
    medicina legale.

    Returns: (in_servizio_ti, in_servizio_td) dicts {profilo_agenas: n}.
    """
    in_s = {}
    in_td = {}
    for _, r in df.iterrows():
        ssd = str(r.get('_REPARTO', '')).upper()
        cdc = str(r.get('Centro di Costo', '')).upper()
        profilo = str(r.get('Profilo Professionale', '')).strip().upper()
        qti = int(r.get('Quantità T.I.', 0))
        qtd = int(r.get('Quantità T.D.', 0))

        uo_ok = False
        for m in mapping_uo:
            if _re.search(m['pattern'], ssd, _re.IGNORECASE):
                uo_ok = True
                break
            if match_cdc and _re.search(m['pattern'], cdc, _re.IGNORECASE):
                uo_ok = True
                break
        if not uo_ok:
            continue

        if esclusioni:
            skip = False
            for ep in esclusioni:
                if _re.search(ep, ssd):
                    skip = True
                    break
            if skip:
                continue

        for mp in mapping_profili:
            if _re.search(mp['pattern'], profilo):
                k = mp['profilo_agenas']
                in_s[k] = in_s.get(k, 0) + qti
                in_td[k] = in_td.get(k, 0) + qtd
                break
    return in_s, in_td


def _conta_materno_infantile(df, mapping_uo, mapping_profili):
    """Conteggio materno-infantile con logica dirigente → UO."""
    in_s = {}
    in_td = {}
    for _, r in df.iterrows():
        ssd = str(r.get('_REPARTO', '')).upper()
        profilo = str(r.get('Profilo Professionale', '')).strip().upper()
        qti = int(r.get('Quantità T.I.', 0))
        qtd = int(r.get('Quantità T.D.', 0))

        uo_ok = False
        profilo_uo = None
        for m in mapping_uo:
            if _re.search(m['pattern'], ssd):
                uo_ok = True
                profilo_uo = m.get('profilo_agenas')
                break
        if not uo_ok:
            continue

        for mp in mapping_profili:
            if _re.search(mp['pattern'], profilo):
                pa = mp['profilo_agenas']
                if pa in ('DIRIGENTE_MEDICO_PEDIATRIA',
                          'DIRIGENTE_MEDICO_OSTETRICIA'):
                    pa = profilo_uo if profilo_uo else None
                    if not pa:
                        continue
                in_s[pa] = in_s.get(pa, 0) + qti
                in_td[pa] = in_td.get(pa, 0) + qtd
                break
    return in_s, in_td


def _conta_radiologia(df, mapping_uo, mapping_profili):
    """Conteggio radiologia con logica territoriale extra."""
    in_s = {}
    in_td = {}
    for _, r in df.iterrows():
        ssd = str(r.get('_REPARTO', '')).upper()
        sede_r = str(r.get('Sede', ''))
        cdc = str(r.get('Centro di Costo', '')).upper()
        profilo = str(r.get('Profilo Professionale', '')).strip().upper()
        qti = int(r.get('Quantità T.I.', 0))
        qtd = int(r.get('Quantità T.D.', 0))

        uo_ok = False
        for m in mapping_uo:
            if _re.search(m['pattern'], ssd):
                uo_ok = True
                break

        # Territoriali: verifica CDC
        if not uo_ok and 'P.O.' not in sede_r:
            if _re.search(r'CENTRO RADIOLOGICO|RADIOLOGIA', cdc):
                uo_ok = True

        # Territoriali: TS RADIOLOGIA conta sempre
        if not uo_ok and 'P.O.' not in sede_r:
            for mp in mapping_profili:
                if (mp['profilo_agenas'] == 'TECNICO_RADIOLOGIA'
                        and _re.search(mp['pattern'], profilo)):
                    uo_ok = True
                    break

        if not uo_ok:
            continue

        for mp in mapping_profili:
            if _re.search(mp['pattern'], profilo):
                k = mp['profilo_agenas']
                in_s[k] = in_s.get(k, 0) + qti
                in_td[k] = in_td.get(k, 0) + qtd
                break
    return in_s, in_td


def _conta_tecnici_lab(df, mapping_profili):
    """Conteggio tecnici laboratorio (no filtro UO, tutto il presidio)."""
    in_s = {}
    in_td = {}
    for _, r in df.iterrows():
        profilo = str(r.get('Profilo Professionale', '')).strip().upper()
        qti = int(r.get('Quantità T.I.', 0))
        qtd = int(r.get('Quantità T.D.', 0))
        for mp in mapping_profili:
            if _re.search(mp['pattern'], profilo):
                k = mp['profilo_agenas']
                in_s[k] = in_s.get(k, 0) + qti
                in_td[k] = in_td.get(k, 0) + qtd
                break
    return in_s, in_td


def _conta_terapia_intensiva(df, mapping_profili):
    """Conteggio TI: solo righe il cui CDC contiene 'TERAPIA INTENSIVA'."""
    _PAT = _re.compile(r'TERAPIA INTENSIVA', _re.IGNORECASE)
    in_s = {}
    in_td = {}
    for _, r in df.iterrows():
        cdc = str(r.get('Centro di Costo', ''))
        if not _PAT.search(cdc):
            continue
        profilo = str(r.get('Profilo Professionale', '')).strip().upper()
        qti = int(r.get('Quantità T.I.', 0))
        qtd = int(r.get('Quantità T.D.', 0))
        for mp in mapping_profili:
            if _re.search(mp['pattern'], profilo):
                k = mp['profilo_agenas']
                in_s[k] = in_s.get(k, 0) + qti
                in_td[k] = in_td.get(k, 0) + qtd
                break
    return in_s, in_td


def _conta_sale_operatorie(df, mapping_uo, mapping_profili):
    """Conteggio SO: match UO su CDC/reparto, escludi CDC 'TERAPIA INTENSIVA'."""
    in_s = {}
    in_td = {}
    for _, r in df.iterrows():
        cdc = str(r.get('Centro di Costo', ''))
        rep = str(r.get('_REPARTO', ''))
        profilo = str(r.get('Profilo Professionale', '')).strip().upper()
        qti = int(r.get('Quantità T.I.', 0))
        qtd = int(r.get('Quantità T.D.', 0))

        uo_ok = False
        for m in mapping_uo:
            if (_re.search(m['pattern'], cdc, _re.IGNORECASE)
                    or _re.search(m['pattern'], rep, _re.IGNORECASE)):
                uo_ok = True
                break
        if not uo_ok:
            continue
        if _re.search(r'TERAPIA INTENSIVA', cdc, _re.IGNORECASE):
            continue

        for mp in mapping_profili:
            if _re.search(mp['pattern'], profilo):
                k = mp['profilo_agenas']
                in_s[k] = in_s.get(k, 0) + qti
                in_td[k] = in_td.get(k, 0) + qtd
                break
    return in_s, in_td


def _conta_ti_bo(df, mapping_uo_ti):
    """Conteggio combinato TI+BO: tutto il personale nella UO
    Anestesia/TI con mapping profilo unificato."""
    _MAP = [
        {'pattern': r'DIRIGENTE MEDICO|DIR\.? MED',
         'key': 'DIRIGENTE_MEDICO'},
        {'pattern': r'INFERMIERE|INFERMIERA',
         'key': 'INFERMIERE'},
        {'pattern': r'OPERATORE SOCIO',
         'key': 'OPERATORE_SOCIO_SANITARIO'},
    ]
    in_s = {}
    in_td = {}
    for _, r in df.iterrows():
        ssd = str(r.get('_REPARTO', ''))
        profilo = str(r.get('Profilo Professionale', '')).strip().upper()
        qti = int(r.get('Quantità T.I.', 0))
        qtd = int(r.get('Quantità T.D.', 0))

        uo_ok = False
        for m in mapping_uo_ti:
            if _re.search(m['pattern'], ssd, _re.IGNORECASE):
                uo_ok = True
                break
        if not uo_ok:
            continue

        for mp in _MAP:
            if _re.search(mp['pattern'], profilo):
                k = mp['key']
                in_s[k] = in_s.get(k, 0) + qti
                in_td[k] = in_td.get(k, 0) + qtd
                break
    return in_s, in_td


def _conta_territoriale(df, uo_patterns, profili_attesi):
    """Conteggio territoriale: match UO/CDC + qualifiche.

    Returns: lista di (ti, td) allineata con profili_attesi.
    """
    result = []
    for prof in profili_attesi:
        ti = 0
        td = 0
        for _, r in df.iterrows():
            rep = str(r.get('_REPARTO', ''))
            cdc = str(r.get('Centro di Costo', ''))
            qual = str(r.get('Profilo Professionale', '')).strip()
            qti = int(r.get('Quantità T.I.', 0))
            qtd = int(r.get('Quantità T.D.', 0))

            if uo_patterns:
                uo_ok = False
                for pat in uo_patterns:
                    if (_re.search(pat, rep, _re.IGNORECASE)
                            or _re.search(pat, cdc, _re.IGNORECASE)):
                        uo_ok = True
                        break
                if not uo_ok:
                    continue
            # se nessun uo_patterns, conta tutto

            for q in prof['qualifiche']:
                if _re.search(q, qual, _re.IGNORECASE):
                    ti += qti
                    td += qtd
                    break
        result.append((ti, td))
    return result


# ============================================================
# RENDERER: TOTALE AZIENDALE RANGE-BASED
# ============================================================

def _scrivi_totale_range(ws, start_row, titolo, nomi_profili,
                         per_presidio_data, fuori_area=None,
                         uo_fuori_area=None):
    """Scrive tabella TOTALE AZIENDALE per aree range-based.

    per_presidio_data: lista di dict con chiavi
        'fabb': {profilo_key: {'min': n, 'max': n}},
        'srv_ti': {profilo_key: n},
        'srv_td': {profilo_key: n},
    fuori_area: dict {profilo_key: n} oppure None
    uo_fuori_area: list[str] – nomi UO dove sono assegnati i fuori area

    Delta per presidio: 0 se in range, eccedenza se > max, carenza se < min.
    Esito totale = somma carenze e somma eccedenze separatamente.
    """
    if not per_presidio_data:
        return start_row

    has_fuori = fuori_area and any(v > 0 for v in fuori_area.values())

    row = start_row + 1
    N_COLS = 7 if has_fuori else 6

    # Titolo
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1,
            value=f"{titolo} - TOTALE AZIENDALE").font = FONT_SECTION
    row += 1

    # Nota
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1,
            value="Somma dei fabbisogni per singolo presidio"
            ).font = Font(size=10)
    row += 2

    # Intestazioni
    headers = ['Profilo', 'T.I.', 'T.D.', 'Totale',
               'Range AGENAS', 'Esito']
    if has_fuori:
        headers.insert(4, 'Altre aree')
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # Profili in ordine (dal primo presidio)
    profili_ordine = list(per_presidio_data[0]['fabb'].keys())

    toggle = 0
    for prof_key in profili_ordine:
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1

        tot_ti = 0
        tot_td = 0
        tot_range_min = 0
        tot_range_max = 0

        for pd in per_presidio_data:
            fabb = pd['fabb'].get(prof_key, {'min': 0, 'max': 0})
            srv_ti = pd['srv_ti'].get(prof_key, 0)
            srv_td = pd['srv_td'].get(prof_key, 0)
            v_min, v_max = fabb['min'], fabb['max']

            tot_ti += srv_ti
            tot_td += srv_td
            tot_range_min += v_min
            tot_range_max += v_max

        tot = tot_ti + tot_td
        range_str = f'{tot_range_min} - {tot_range_max}'

        # Esito: valutazione globale contro range sommato
        if tot < tot_range_min:
            esito = f'CARENZA (min {tot_range_min - tot})'
            fill_esito = FILL_CARENZA
        elif tot > tot_range_max:
            esito = f'ECCEDENZA (+{tot - tot_range_max})'
            fill_esito = FILL_OK
        else:
            esito = 'IN RANGE'
            fill_esito = FILL_OK

        if has_fuori:
            n_fuori = fuori_area.get(prof_key, 0)
            vals = [nomi_profili.get(prof_key, prof_key),
                    tot_ti, tot_td, tot,
                    n_fuori if n_fuori else '',
                    range_str, esito]
        else:
            vals = [nomi_profili.get(prof_key, prof_key),
                    tot_ti, tot_td, tot, range_str, esito]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        if has_fuori and fuori_area.get(prof_key, 0) > 0:
            ws.cell(row=row, column=5).fill = FILL_ALTRE_AREE
        row += 1

    # Nota piè di pagina
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    nota_testo = (
        "(*) Maggiorazione +15% per turnazione, guardie, "
        "ferie, malattie e altre indisponibilità. "
        "Le eccedenze/carenze sono calcolate come somma "
        "delle eccedenze/carenze dei singoli presidi."
    )
    ws.cell(row=row, column=1, value=nota_testo).font = FONT_NOTE
    if has_fuori:
        row += 1
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS)
        ws.cell(row=row, column=1,
                value="\"Altre aree\": personale con disciplina "
                      "corrispondente ma assegnato a UO al di fuori "
                      "dell'area specifica (possibili limitazioni "
                      "o incarichi diversi)."
                ).font = FONT_NOTE
        if uo_fuori_area:
            row += 1
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=N_COLS)
            ws.cell(row=row, column=1,
                    value="UO di assegnazione: "
                          + ", ".join(uo_fuori_area)
                    ).font = FONT_NOTE

    return row + 1


# ============================================================
# RENDERER: TOTALE AZIENDALE FTE-BASED (TI, SO)
# ============================================================

def _scrivi_totale_fte(ws, start_row, titolo, nomi_profili,
                       per_presidio_data, fuori_area=None,
                       uo_fuori_area=None):
    """Scrive tabella TOTALE AZIENDALE per aree FTE-based.

    per_presidio_data: lista di dict con chiavi
        'fabb': {profilo_key: {'fte_atteso': n, ...}},
        'srv_ti': {profilo_key: n},
        'srv_td': {profilo_key: n},
    fuori_area: dict {profilo_key: n} oppure None
    uo_fuori_area: list[str] – nomi UO dove sono assegnati i fuori area
    """
    if not per_presidio_data:
        return start_row

    has_fuori = fuori_area and any(v > 0 for v in fuori_area.values())

    row = start_row + 1
    N_COLS = 7 if has_fuori else 6

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1,
            value=f"{titolo} - TOTALE AZIENDALE").font = FONT_SECTION
    row += 1

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1,
            value="Somma FTE e conteggi per singolo presidio"
            ).font = Font(size=10)
    row += 2

    headers = ['Profilo', 'FTE Atteso', 'T.I.', 'T.D.', 'Totale', 'Esito']
    if has_fuori:
        headers.insert(5, 'Altre aree')
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    profili_ordine = list(per_presidio_data[0]['fabb'].keys())

    toggle = 0
    for prof_key in profili_ordine:
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1

        tot_fte = 0
        tot_ti = 0
        tot_td = 0

        for pd in per_presidio_data:
            dati = pd['fabb'].get(prof_key, {'fte_atteso': 0})
            fte = dati['fte_atteso']
            srv_ti = pd['srv_ti'].get(prof_key, 0)
            srv_td = pd['srv_td'].get(prof_key, 0)

            tot_fte += fte
            tot_ti += srv_ti
            tot_td += srv_td

        tot = tot_ti + tot_td

        if tot < tot_fte:
            esito = f'CARENZA ({tot_fte - tot})'
            fill_esito = FILL_CARENZA
        elif tot > tot_fte:
            esito = f'ECCEDENZA (+{tot - tot_fte})'
            fill_esito = FILL_OK
        else:
            esito = 'CONFORME'
            fill_esito = FILL_OK

        if has_fuori:
            n_fuori = fuori_area.get(prof_key, 0)
            vals = [nomi_profili.get(prof_key, prof_key),
                    tot_fte, tot_ti, tot_td, tot,
                    n_fuori if n_fuori else '',
                    esito]
        else:
            vals = [nomi_profili.get(prof_key, prof_key),
                    tot_fte, tot_ti, tot_td, tot, esito]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        if has_fuori and fuori_area.get(prof_key, 0) > 0:
            ws.cell(row=row, column=6).fill = FILL_ALTRE_AREE
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1,
            value="Le eccedenze/carenze sono calcolate come somma "
                  "delle eccedenze/carenze dei singoli presidi."
            ).font = FONT_NOTE
    if has_fuori:
        row += 1
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS)
        ws.cell(row=row, column=1,
                value="\"Altre aree\": personale con disciplina "
                      "corrispondente ma assegnato a UO al di fuori "
                      "dell'area specifica (possibili limitazioni "
                      "o incarichi diversi)."
                ).font = FONT_NOTE
        if uo_fuori_area:
            row += 1
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=N_COLS)
            ws.cell(row=row, column=1,
                    value="UO di assegnazione: "
                          + ", ".join(sorted(uo_fuori_area))
                    ).font = FONT_NOTE

    return row + 1


# ============================================================
# RENDERER: TOTALE AZIENDALE COMBINATA TI + BO
# ============================================================

def _scrivi_totale_fte_combo(ws, start_row, per_presidio_data,
                             fabb_ti_per_presidio,
                             fabb_so_per_presidio,
                             fuori_area=None,
                             uo_fuori_area=None):
    """Scrive tabella TOTALE AZIENDALE per area combinata TI+BO.

    per_presidio_data: lista di dict con chiavi
        'presidio': nome_presidio,
        'srv_ti': {profilo_key: n},
        'srv_td': {profilo_key: n},
    fuori_area: dict {profilo_key: n} oppure None
                (la chiave è la 'key' del profilo combo, es.
                DIRIGENTE_MEDICO)
    """
    if not per_presidio_data:
        return start_row

    has_fuori = fuori_area and any(v > 0 for v in fuori_area.values())

    _PROFILI = [
        {'label': 'Dirigenti Medici', 'key': 'DIRIGENTE_MEDICO',
         'ti_key': 'DIRIGENTE_MEDICO',
         'so_key': 'DIRIGENTE_MEDICO_ANESTESISTA'},
        {'label': 'Infermieri', 'key': 'INFERMIERE',
         'ti_key': 'INFERMIERE', 'so_key': 'INFERMIERE'},
        {'label': 'Operatori Socio Sanitari',
         'key': 'OPERATORE_SOCIO_SANITARIO',
         'ti_key': None, 'so_key': 'OPERATORE_SOCIO_SANITARIO'},
    ]

    row = start_row + 1
    N_COLS = 9 if has_fuori else 8

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1,
            value="AREA ANESTESIA, T.I. E BLOCCO OPERATORIO "
                  "- TOTALE AZIENDALE").font = FONT_SECTION
    row += 1

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1,
            value="Somma FTE e conteggi per singolo presidio"
            ).font = Font(size=10)
    row += 2

    headers = ['Profilo', 'FTE TI', 'FTE BO', 'FTE Area',
               'T.I.', 'T.D.', 'Totale', 'Esito']
    if has_fuori:
        headers.insert(7, 'Altre aree')
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    toggle = 0
    for prof_info in _PROFILI:
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1

        tot_fte_ti = 0
        tot_fte_so = 0
        tot_srv_ti = 0
        tot_srv_td = 0

        for pd in per_presidio_data:
            presidio = pd['presidio']
            fabb_ti = fabb_ti_per_presidio.get(presidio, {})
            fabb_so = fabb_so_per_presidio.get(presidio, {})

            fte_ti = 0
            if (prof_info['ti_key']
                    and prof_info['ti_key'] in fabb_ti):
                fte_ti = fabb_ti[prof_info['ti_key']]['fte_atteso']
            fte_so = 0
            if (prof_info['so_key']
                    and prof_info['so_key'] in fabb_so):
                fte_so = fabb_so[prof_info['so_key']]['fte_atteso']

            srv_ti = pd['srv_ti'].get(prof_info['key'], 0)
            srv_td = pd['srv_td'].get(prof_info['key'], 0)

            tot_fte_ti += fte_ti
            tot_fte_so += fte_so
            tot_srv_ti += srv_ti
            tot_srv_td += srv_td

        tot = tot_srv_ti + tot_srv_td
        tot_fte_area = tot_fte_ti + tot_fte_so

        if tot < tot_fte_area:
            esito = f'CARENZA ({tot_fte_area - tot})'
            fill_esito = FILL_CARENZA
        elif tot > tot_fte_area:
            esito = f'ECCEDENZA (+{tot - tot_fte_area})'
            fill_esito = FILL_OK
        else:
            esito = 'CONFORME'
            fill_esito = FILL_OK

        if has_fuori:
            n_fuori = fuori_area.get(prof_info['key'], 0)
            vals = [prof_info['label'],
                    tot_fte_ti, tot_fte_so, tot_fte_area,
                    tot_srv_ti, tot_srv_td, tot,
                    n_fuori if n_fuori else '',
                    esito]
        else:
            vals = [prof_info['label'],
                    tot_fte_ti, tot_fte_so, tot_fte_area,
                    tot_srv_ti, tot_srv_td, tot, esito]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        if has_fuori and fuori_area.get(prof_info['key'], 0) > 0:
            ws.cell(row=row, column=8).fill = FILL_ALTRE_AREE
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1,
            value="Le eccedenze/carenze sono calcolate come somma "
                  "delle eccedenze/carenze dei singoli presidi."
            ).font = FONT_NOTE
    if has_fuori:
        row += 1
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=N_COLS)
        ws.cell(row=row, column=1,
                value="\"Altre aree\": personale con disciplina "
                      "corrispondente ma assegnato a UO al di fuori "
                      "dell'area specifica (possibili limitazioni "
                      "o incarichi diversi)."
                ).font = FONT_NOTE
        if uo_fuori_area:
            row += 1
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=N_COLS)
            ws.cell(row=row, column=1,
                    value="UO di assegnazione: "
                          + ", ".join(sorted(uo_fuori_area))
                    ).font = FONT_NOTE

    return row + 1


# ============================================================
# RENDERER: TOTALE AZIENDALE TERRITORIALE
# ============================================================

def _scrivi_totale_territoriale(ws, start_row, indicatori,
                                per_area_data):
    """Scrive tabella TOTALE AZIENDALE per aree territoriali.

    per_area_data: lista di dict con chiavi
        'fabb': {profili: [...], ...},
        'servizio': [(ti, td), ...],  # allineato con profili
    """
    if not per_area_data:
        return start_row

    titolo = indicatori.get('titolo', 'Standard AGENAS')
    first_fabb = per_area_data[0]['fabb']
    profili = first_fabb['profili']
    has_regime = any(
        p.get('atteso_regime') is not None for p in profili)

    row = start_row + 1

    if has_regime:
        headers = ['Profilo', 'T.I.', 'T.D.', 'Totale',
                   'Atteso min.', 'Atteso a regime', 'Esito']
    else:
        headers = ['Profilo', 'T.I.', 'T.D.', 'Totale',
                   'Standard min.', 'Esito']
    N_COLS = len(headers)

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1,
            value=f"{titolo} - TOTALE AZIENDALE").font = FONT_SECTION
    row += 1

    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1,
            value="Somma dei fabbisogni per singola area"
            ).font = Font(size=10)
    row += 2

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    row += 1

    # Somme per i totali complessivi
    grand_ti = 0
    grand_td = 0
    grand_amin = 0
    grand_areg = 0

    toggle = 0
    for pi, prof_template in enumerate(profili):
        fill = FILL_A if toggle % 2 == 0 else FILL_B
        toggle += 1

        tot_ti = 0
        tot_td = 0
        tot_amin = 0
        tot_areg = 0
        tot_carenza = 0  # somma carenze per-area

        for pad in per_area_data:
            prof_info = pad['fabb']['profili'][pi]
            srv_ti, srv_td = pad['servizio'][pi]
            srv = srv_ti + srv_td
            amin = int(round(prof_info['atteso_min']))
            areg = (int(round(prof_info['atteso_regime']))
                    if prof_info.get('atteso_regime') is not None
                    else None)

            tot_ti += srv_ti
            tot_td += srv_td
            tot_amin += amin
            if areg is not None:
                tot_areg += areg

            # Carenza per-area
            if srv < amin:
                tot_carenza += (amin - srv)

        tot = tot_ti + tot_td
        grand_ti += tot_ti
        grand_td += tot_td
        grand_amin += tot_amin
        grand_areg += tot_areg

        # Esito
        if has_regime:
            if tot_carenza > 0:
                esito = f'CARENZA ({tot_carenza})'
                fill_esito = FILL_CARENZA
            elif tot >= tot_areg:
                esito = 'A REGIME'
                fill_esito = FILL_REGIME
            elif tot >= tot_amin:
                esito = 'IN RANGE'
                fill_esito = FILL_OK
            else:
                esito = f'CARENZA ({tot_amin - tot})'
                fill_esito = FILL_CARENZA
            vals = [prof_template['nome'], tot_ti, tot_td, tot,
                    tot_amin, tot_areg, esito]
        else:
            if tot_carenza > 0:
                esito = f'CARENZA ({tot_carenza})'
                fill_esito = FILL_CARENZA
            else:
                esito = 'OK'
                fill_esito = FILL_OK
            vals = [prof_template['nome'], tot_ti, tot_td, tot,
                    tot_amin, esito]

        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.border = THIN_BORDER
        ws.cell(row=row, column=N_COLS).fill = fill_esito
        row += 1

    # Riga TOTALE
    grand_tot = grand_ti + grand_td
    if has_regime:
        if grand_tot < grand_amin:
            esito_tot = f'CARENZA ({grand_amin - grand_tot})'
            fill_tot = FILL_CARENZA
        elif grand_tot >= grand_areg:
            esito_tot = 'A REGIME'
            fill_tot = FILL_REGIME
        else:
            esito_tot = 'IN RANGE'
            fill_tot = FILL_OK
        vals_tot = ['TOTALE', grand_ti, grand_td, grand_tot,
                    grand_amin, grand_areg, esito_tot]
    else:
        if grand_tot >= grand_amin:
            esito_tot = 'OK'
            fill_tot = FILL_OK
        else:
            esito_tot = f'CARENZA ({grand_amin - grand_tot})'
            fill_tot = FILL_CARENZA
        vals_tot = ['TOTALE', grand_ti, grand_td, grand_tot,
                    grand_amin, esito_tot]

    for ci, v in enumerate(vals_tot, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = Font(bold=True, size=10)
        c.fill = FILL_HEADER
        c.border = THIN_BORDER
    ws.cell(row=row, column=N_COLS).fill = fill_tot
    row += 1

    # Note
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=N_COLS)
    ws.cell(row=row, column=1,
            value="Le carenze sono calcolate come somma "
                  "delle carenze delle singole aree. "
                  "Maggiorazione +15% per turnazione, guardie, "
                  "ferie, malattie e altre indisponibilità."
            ).font = FONT_NOTE

    return row + 1


# ============================================================
# RIEPILOGO TRASFUSIONALE AZIENDALE (custom)
# ============================================================

def _scrivi_riepilogo_trasfusionale_aziendale(
        ws, start_row, grouped,
        fabb_trasf_per_presidio,
        mapping_uo_trasf, mapping_profili_trasf,
        fuori_area=None, uo_fuori_area=None):
    """Tabella riepilogativa aziendale medicina trasfusionale.

    Per la trasfusionale c'è tipicamente un solo presidio, quindi
    il conteggio globale coincide con la somma per presidio.
    """
    if not fabb_trasf_per_presidio:
        return start_row

    nomi_profili = {
        'DIRIGENTE_SANITARIO':       'Dir. Sanitari (*)',
        'INFERMIERE':                'Infermieri',
        'OPERATORE_SOCIO_SANITARIO': 'Operatori Socio Sanitari',
    }

    # Conta per singolo presidio e accumula
    per_pres = []
    for presidio, fabb in fabb_trasf_per_presidio.items():
        # Filtra df per il presidio
        # (la trasfusionale di solito ha 1 presidio, ma gestiamo N)
        srv_ti, srv_td = _conta_generico(
            grouped, mapping_uo_trasf, mapping_profili_trasf,
            match_cdc=False)
        per_pres.append({
            'fabb': fabb,
            'srv_ti': srv_ti,
            'srv_td': srv_td,
        })

    if not per_pres:
        return start_row

    return _scrivi_totale_range(
        ws, start_row,
        "AREA MEDICINA TRASFUSIONALE",
        nomi_profili,
        per_pres,
        fuori_area=fuori_area,
        uo_fuori_area=uo_fuori_area,
    )


# ============================================================
# FUNZIONE PRINCIPALE
# ============================================================

# --- Nomi leggibili per profilo (usati nei TOTALI) ---

_NOMI_MATERNO = {
    'DIRIGENTE_MEDICO_PEDIATRIA':  'Dir. Medici Pediatria/Neonat.',
    'DIRIGENTE_MEDICO_OSTETRICIA': 'Dir. Medici Ostetricia/Ginec.',
    'OSTETRICA':                   'Ostetriche',
    'INFERMIERE':                  'Infermieri',
    'OPERATORE_SOCIO_SANITARIO':   'Operatori Socio Sanitari',
}
_NOMI_RADIOLOGIA = {
    'DIRIGENTE_MEDICO_RADIOLOGIA': 'Dir. Medici Radiologia (*)',
    'TECNICO_RADIOLOGIA':          'Tecnici di Radiologia',
    'INFERMIERE':                  'Infermieri',
    'OPERATORE_SOCIO_SANITARIO':   'Operatori Socio Sanitari',
}
_NOMI_EMERGENZA = {
    'DIRIGENTE_MEDICO':          'Dirigenti Medici',
    'INFERMIERE':                'Infermieri',
    'OPERATORE_SOCIO_SANITARIO': 'Operatori Socio Sanitari',
}
_NOMI_ANAPATO = {
    'DIRIGENTE_SANITARIO':       'Dir. Sanitari (*)',
    'INFERMIERE':                'Infermieri',
    'OPERATORE_SOCIO_SANITARIO': 'Operatori Socio Sanitari',
}
_NOMI_LABORATORIO = {
    'DIRIGENTE_SANITARIO':       'Dir. Sanitari (*)',
    'INFERMIERE':                'Infermieri',
    'OPERATORE_SOCIO_SANITARIO': 'Operatori Socio Sanitari',
}
_NOMI_TECLAB = {
    'TECNICO_LABORATORIO': 'Tecnici di Laboratorio',
}
_NOMI_MEDLEG = {
    'DIRIGENTE_MEDICO':          'Dirigenti Medici',
    'INFERMIERE':                'Infermieri',
    'OPERATORE_SOCIO_SANITARIO': 'Operatori Socio Sanitari',
}
_NOMI_TI = {
    'DIRIGENTE_MEDICO': 'Dirigenti Medici',
    'INFERMIERE':       'Infermieri',
}
_NOMI_SO = {
    'DIRIGENTE_MEDICO_ANESTESISTA': 'Medici Anestesisti (*)',
    'INFERMIERE':                   'Infermieri',
    'OPERATORE_SOCIO_SANITARIO':    'Operatori Socio Sanitari',
}


def scrivi_foglio_riepilogo_agenas(
        wb, grouped, livello_presidio,
        # --- Ospedaliere ---
        fabb_agenas_per_presidio,
        mapping_uo_agenas, mapping_profili_agenas,
        fabb_radio_per_presidio,
        mapping_uo_radio, mapping_profili_radio,
        fabb_anapato_per_presidio,
        mapping_uo_anapato, mapping_profili_anapato,
        fabb_lab_per_presidio,
        mapping_uo_lab, mapping_profili_lab, esclusioni_lab,
        fabb_teclab_per_presidio, mapping_profili_teclab,
        fabb_medleg_per_presidio,
        mapping_uo_medleg, mapping_profili_medleg,
        fabb_trasf_per_presidio,
        mapping_uo_trasf, mapping_profili_trasf,
        fabb_trasf_speciale,
        fabb_emergenza_per_presidio,
        mapping_uo_emergenza, mapping_profili_emergenza,
        fabb_ti_per_presidio,
        mapping_uo_ti, mapping_profili_ti,
        fabb_so_per_presidio,
        mapping_uo_so, mapping_profili_so,
        # --- Territoriali ---
        indicatori_salute_mentale, fabb_salute_mentale,
        indicatori_dipendenze, fabb_dipendenze,
        indicatori_npia, fabb_npia,
        indicatori_carcere, fabb_carcere,
        # --- Dati completi per colonna "Altre aree" ---
        personale_all_df=None,
        discipline_db_map=None,
        parti_per_presidio=None,
):
    """Crea il foglio 'FABBISOGNO AGENAS' nel workbook del riepilogo
    aziendale.

    Per ogni area AGENAS:
      1. Tabelle per singola area/città (stesse dei file per-città)
      2. Tabella TOTALE AZIENDALE che somma carenze/eccedenze
         dei singoli presidi (non valutazione globale su range sommato)
    """

    _AREE_SPECIALI = {'(Non assegnata)', 'LUNGHE ASSENZE',
                      'IN ATTESA DI ASSEGNAZIONE'}

    ws = wb.create_sheet(title='FABBISOGNO AGENAS')

    # Titolo principale
    FONT_MAIN = Font(bold=True, size=14, color='1F4E79')
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=8)
    c = ws.cell(row=1, column=1,
                value="RIEPILOGO FABBISOGNI AGENAS - LIVELLO AZIENDALE")
    c.font = FONT_MAIN
    c.alignment = ALIGN_CENTER
    row = 3

    # ------------------------------------------------------------------
    # Costruisco lista città con presidio associato
    # ------------------------------------------------------------------
    cities_data = []
    for citta, df_citta in grouped.groupby('_CITTA', sort=True):
        if citta in _AREE_SPECIALI:
            continue
        presidio = None
        for pn in (livello_presidio or {}):
            if citta.upper() in pn.upper():
                presidio = pn
                break
        livello_label = (
            livello_presidio.get(presidio, '')
            if livello_presidio and presidio else ''
        )
        cities_data.append({
            'citta': citta,
            'df': df_citta,
            'presidio': presidio,
            'livello': livello_label,
        })

    # ------------------------------------------------------------------
    # Helper per calcolo "Altre aree" (personale fuori area)
    # ------------------------------------------------------------------
    _disc_map = discipline_db_map or {}

    def _calc_fuori(area_key, per_pres):
        """Calcola fuori_area per l'area indicata.

        Somma il tot TI+TD in area dai per_pres e poi confronta
        col totale nel DB tramite _conta_fuori_area.

        Ritorna (fuori_area_dict, uo_list) oppure (None, []).
        """
        disc_list = _disc_map.get(area_key, [])
        if not disc_list or personale_all_df is None:
            return None, []
        # Somma in-area per profilo
        tot_in_area = {}
        for pd in per_pres:
            for prof_key in pd['srv_ti']:
                tot_in_area[prof_key] = (
                    tot_in_area.get(prof_key, 0)
                    + pd['srv_ti'].get(prof_key, 0)
                    + pd['srv_td'].get(prof_key, 0)
                )
        return _conta_fuori_area(personale_all_df, disc_list,
                                 tot_in_area)

    # ==================================================================
    # AREA MATERNO INFANTILE
    # ==================================================================
    if fabb_agenas_per_presidio:
        row = _scrivi_divisore_area(ws, row,
                                    "AREA MATERNO INFANTILE")
        per_pres = []
        for cd in cities_data:
            if (cd['presidio']
                    and fabb_agenas_per_presidio.get(cd['presidio'])):
                row = _scrivi_tabella_agenas_materno_infantile(
                    ws, row, cd['df'],
                    fabb_agenas_per_presidio[cd['presidio']],
                    mapping_uo_agenas, mapping_profili_agenas,
                    cd['presidio'],
                )
                srv_ti, srv_td = _conta_materno_infantile(
                    cd['df'], mapping_uo_agenas, mapping_profili_agenas)
                per_pres.append({
                    'fabb': fabb_agenas_per_presidio[cd['presidio']],
                    'srv_ti': srv_ti,
                    'srv_td': srv_td,
                })
        if per_pres:
            _fa_mat, _uo_fa_mat = _calc_fuori(
                'materno_infantile', per_pres)
            row = _scrivi_totale_range(
                ws, row + 1, "AREA MATERNO INFANTILE",
                _NOMI_MATERNO, per_pres,
                fuori_area=_fa_mat,
                uo_fuori_area=_uo_fa_mat)
            # Nota con numero di parti rilevati per presidio
            if parti_per_presidio:
                _parti_items = []
                for cd in cities_data:
                    if cd['presidio']:
                        n = parti_per_presidio.get(
                            cd['presidio'], 0)
                        if n:
                            _parti_items.append(
                                f"{cd['citta']} {n}")
                if _parti_items:
                    _tot_parti = sum(
                        parti_per_presidio.get(
                            cd['presidio'], 0)
                        for cd in cities_data
                        if cd['presidio'])
                    row += 1
                    _ncols_m = 7 if (_fa_mat and any(
                        v > 0 for v in _fa_mat.values()
                    )) else 6
                    ws.merge_cells(
                        start_row=row, start_column=1,
                        end_row=row, end_column=_ncols_m)
                    ws.cell(
                        row=row, column=1,
                        value="Parti rilevati: "
                              + ", ".join(_parti_items)
                              + f"  (Totale: {_tot_parti})"
                    ).font = FONT_NOTE

    # ==================================================================
    # AREA SERVIZI DI RADIOLOGIA
    # ==================================================================
    if fabb_radio_per_presidio:
        row = _scrivi_divisore_area(ws, row + 1,
                                    "AREA SERVIZI DI RADIOLOGIA")
        per_pres = []
        for cd in cities_data:
            if (cd['presidio']
                    and fabb_radio_per_presidio.get(cd['presidio'])):
                row = _scrivi_tabella_agenas_radiologia(
                    ws, row, cd['df'],
                    fabb_radio_per_presidio[cd['presidio']],
                    mapping_uo_radio, mapping_profili_radio,
                    cd['presidio'], cd['livello'],
                    df_area=cd['df'],
                )
                srv_ti, srv_td = _conta_radiologia(
                    cd['df'], mapping_uo_radio, mapping_profili_radio)
                per_pres.append({
                    'fabb': fabb_radio_per_presidio[cd['presidio']],
                    'srv_ti': srv_ti,
                    'srv_td': srv_td,
                })
        if per_pres:
            _fa_radio, _uo_fa_radio = _calc_fuori(
                'radiologia', per_pres)
            row = _scrivi_totale_range(
                ws, row + 1, "AREA SERVIZI DI RADIOLOGIA",
                _NOMI_RADIOLOGIA, per_pres,
                fuori_area=_fa_radio,
                uo_fuori_area=_uo_fa_radio)
            # Nota esplicativa composizione Dir. Medici Radiologia
            _ncols_r = 7 if (_fa_radio and any(
                v > 0 for v in _fa_radio.values())) else 6
            row += 1
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=_ncols_r)
            ws.cell(row=row, column=1,
                    value="N.B. I Dir. Medici Radiologia comprendono "
                          "i medici con disciplina Radiodiagnostica e "
                          "Medicina Nucleare afferenti all'area dei "
                          "servizi di Radiologia (Tab. 13 AGENAS)."
                    ).font = FONT_NOTE
            # Tabella distribuzione radiologi per CDC
            row = _scrivi_distribuzione_radiologi(
                ws, row + 2, personale_all_df)

    # ==================================================================
    # AREA ANATOMIA PATOLOGICA
    # ==================================================================
    if fabb_anapato_per_presidio:
        row = _scrivi_divisore_area(ws, row + 1,
                                    "AREA ANATOMIA PATOLOGICA")
        per_pres = []
        for cd in cities_data:
            if (cd['presidio']
                    and fabb_anapato_per_presidio.get(cd['presidio'])):
                row = _scrivi_tabella_agenas_anatomia_patologica(
                    ws, row, cd['df'],
                    fabb_anapato_per_presidio[cd['presidio']],
                    mapping_uo_anapato, mapping_profili_anapato,
                    cd['presidio'], cd['livello'],
                )
                srv_ti, srv_td = _conta_generico(
                    cd['df'], mapping_uo_anapato,
                    mapping_profili_anapato, match_cdc=False)
                per_pres.append({
                    'fabb': fabb_anapato_per_presidio[cd['presidio']],
                    'srv_ti': srv_ti,
                    'srv_td': srv_td,
                })
        if per_pres:
            _fa_anap, _uo_fa_anap = _calc_fuori(
                'anatomia_pat', per_pres)
            row = _scrivi_totale_range(
                ws, row + 1, "AREA ANATOMIA PATOLOGICA",
                _NOMI_ANAPATO, per_pres,
                fuori_area=_fa_anap,
                uo_fuori_area=_uo_fa_anap)

    # ==================================================================
    # AREA SERVIZI DI LABORATORIO
    # ==================================================================
    if fabb_lab_per_presidio:
        row = _scrivi_divisore_area(ws, row + 1,
                                    "AREA SERVIZI DI LABORATORIO")
        per_pres = []
        for cd in cities_data:
            if (cd['presidio']
                    and fabb_lab_per_presidio.get(cd['presidio'])):
                row = _scrivi_tabella_agenas_laboratorio(
                    ws, row, cd['df'],
                    fabb_lab_per_presidio[cd['presidio']],
                    mapping_uo_lab, mapping_profili_lab,
                    esclusioni_lab,
                    cd['presidio'], cd['livello'],
                )
                srv_ti, srv_td = _conta_generico(
                    cd['df'], mapping_uo_lab, mapping_profili_lab,
                    esclusioni=esclusioni_lab, match_cdc=False)
                per_pres.append({
                    'fabb': fabb_lab_per_presidio[cd['presidio']],
                    'srv_ti': srv_ti,
                    'srv_td': srv_td,
                })
        if per_pres:
            _fa_lab, _uo_fa_lab = _calc_fuori(
                'laboratorio', per_pres)
            row = _scrivi_totale_range(
                ws, row + 1, "AREA SERVIZI DI LABORATORIO",
                _NOMI_LABORATORIO, per_pres,
                fuori_area=_fa_lab,
                uo_fuori_area=_uo_fa_lab)

    # ==================================================================
    # TECNICI DI LABORATORIO
    # ==================================================================
    if fabb_teclab_per_presidio:
        row = _scrivi_divisore_area(ws, row + 1,
                                    "TECNICI DI LABORATORIO")
        per_pres = []
        for cd in cities_data:
            if (cd['presidio']
                    and fabb_teclab_per_presidio.get(cd['presidio'])):
                row = _scrivi_tabella_agenas_tecnici_laboratorio(
                    ws, row, cd['df'],
                    fabb_teclab_per_presidio[cd['presidio']],
                    mapping_profili_teclab,
                    cd['presidio'], cd['livello'],
                )
                srv_ti, srv_td = _conta_tecnici_lab(
                    cd['df'], mapping_profili_teclab)
                per_pres.append({
                    'fabb': fabb_teclab_per_presidio[cd['presidio']],
                    'srv_ti': srv_ti,
                    'srv_td': srv_td,
                })
        if per_pres:
            row = _scrivi_totale_range(
                ws, row + 1, "TECNICI DI LABORATORIO",
                _NOMI_TECLAB, per_pres)

    # ==================================================================
    # MEDICINA LEGALE
    # ==================================================================
    if fabb_medleg_per_presidio:
        row = _scrivi_divisore_area(ws, row + 1,
                                    "MEDICINA LEGALE")
        per_pres = []
        for cd in cities_data:
            if (cd['presidio']
                    and fabb_medleg_per_presidio.get(cd['presidio'])):
                row = _scrivi_tabella_agenas_medicina_legale(
                    ws, row, cd['df'],
                    fabb_medleg_per_presidio[cd['presidio']],
                    mapping_uo_medleg, mapping_profili_medleg,
                    cd['presidio'], cd['livello'],
                )
                srv_ti, srv_td = _conta_generico(
                    cd['df'], mapping_uo_medleg,
                    mapping_profili_medleg, match_cdc=False)
                per_pres.append({
                    'fabb': fabb_medleg_per_presidio[cd['presidio']],
                    'srv_ti': srv_ti,
                    'srv_td': srv_td,
                })
        if per_pres:
            _fa_medl, _uo_fa_medl = _calc_fuori(
                'medicina_legale', per_pres)
            row = _scrivi_totale_range(
                ws, row + 1, "MEDICINA LEGALE",
                _NOMI_MEDLEG, per_pres,
                fuori_area=_fa_medl,
                uo_fuori_area=_uo_fa_medl)

    # ==================================================================
    # AREA MEDICINA TRASFUSIONALE
    # ==================================================================
    if fabb_trasf_per_presidio:
        row = _scrivi_divisore_area(ws, row + 1,
                                    "AREA MEDICINA TRASFUSIONALE")
        for cd in cities_data:
            if (cd['presidio']
                    and fabb_trasf_per_presidio.get(cd['presidio'])):
                row = _scrivi_tabella_agenas_trasfusionale(
                    ws, row, cd['df'],
                    fabb_trasf_per_presidio[cd['presidio']],
                    mapping_uo_trasf, mapping_profili_trasf,
                    cd['presidio'], cd['livello'],
                    df_completo=grouped,
                )
        # Totale aziendale (tabella semplice)
        # Calcolo fuori_area per trasfusionale
        _fa_trasf = None
        _uo_fa_trasf = []
        _disc_trasf = _disc_map.get('trasfusionale', [])
        if _disc_trasf and personale_all_df is not None:
            _srv_ti_t, _srv_td_t = _conta_generico(
                grouped, mapping_uo_trasf, mapping_profili_trasf,
                match_cdc=False)
            _tot_in_t = {k: _srv_ti_t.get(k, 0) + _srv_td_t.get(k, 0)
                         for k in set(list(_srv_ti_t) + list(_srv_td_t))}
            _fa_trasf, _uo_fa_trasf = _conta_fuori_area(
                personale_all_df, _disc_trasf, _tot_in_t)
        row = _scrivi_riepilogo_trasfusionale_aziendale(
            ws, row + 1, grouped,
            fabb_trasf_per_presidio,
            mapping_uo_trasf, mapping_profili_trasf,
            fuori_area=_fa_trasf,
            uo_fuori_area=_uo_fa_trasf,
        )

        # UOC Trasfusionale – Fabbisogno Primaria (se configurato)
        if fabb_trasf_speciale and fabb_trasf_speciale.get('sedi'):
            sede_princ = fabb_trasf_speciale['sedi'][0]['nome']
            presidio_ref = None
            for cd in cities_data:
                if cd['presidio'] and (
                        cd['citta'].upper() in sede_princ.upper()
                        or sede_princ.upper()
                        in cd['presidio'].upper()):
                    presidio_ref = cd['presidio']
                    break
            row = _scrivi_tabella_fabbisogno_uoc_trasfusionale(
                ws, row + 1, grouped,
                fabb_trasf_speciale,
                presidio_ref or "TOTALE AZIENDALE",
            )

    # ==================================================================
    # AREA EMERGENZA-URGENZA
    # ==================================================================
    if fabb_emergenza_per_presidio:
        row = _scrivi_divisore_area(ws, row + 1,
                                    "AREA EMERGENZA-URGENZA")
        per_pres = []
        for cd in cities_data:
            if (cd['presidio']
                    and fabb_emergenza_per_presidio.get(
                        cd['presidio'])):
                row = _scrivi_tabella_agenas_emergenza_urgenza(
                    ws, row, cd['df'],
                    fabb_emergenza_per_presidio[cd['presidio']],
                    mapping_uo_emergenza, mapping_profili_emergenza,
                    cd['presidio'], cd['livello'],
                )
                srv_ti, srv_td = _conta_generico(
                    cd['df'], mapping_uo_emergenza,
                    mapping_profili_emergenza, match_cdc=True)
                per_pres.append({
                    'fabb': fabb_emergenza_per_presidio[
                        cd['presidio']],
                    'srv_ti': srv_ti,
                    'srv_td': srv_td,
                })
        if per_pres:
            _fa_emer, _uo_fa_emer = _calc_fuori(
                'emergenza_urgenza', per_pres)
            row = _scrivi_totale_range(
                ws, row + 1, "AREA EMERGENZA-URGENZA",
                _NOMI_EMERGENZA, per_pres,
                fuori_area=_fa_emer,
                uo_fuori_area=_uo_fa_emer)

    # ==================================================================
    # AREA TERAPIA INTENSIVA (§ 8.1.1)
    # ==================================================================
    if fabb_ti_per_presidio:
        row = _scrivi_divisore_area(
            ws, row + 1,
            "AREA TERAPIA INTENSIVA (§ 8.1.1)")
        per_pres = []
        for cd in cities_data:
            if (cd['presidio']
                    and fabb_ti_per_presidio.get(cd['presidio'])):
                row = _scrivi_tabella_agenas_terapia_intensiva(
                    ws, row, cd['df'],
                    fabb_ti_per_presidio[cd['presidio']],
                    mapping_uo_ti, mapping_profili_ti,
                    cd['presidio'],
                )
                srv_ti, srv_td = _conta_terapia_intensiva(
                    cd['df'], mapping_profili_ti)
                per_pres.append({
                    'fabb': fabb_ti_per_presidio[cd['presidio']],
                    'srv_ti': srv_ti,
                    'srv_td': srv_td,
                })
        if per_pres:
            _fa_ti, _uo_fa_ti = _calc_fuori(
                'terapia_intensiva', per_pres)
            row = _scrivi_totale_fte(
                ws, row + 1,
                "AREA TERAPIA INTENSIVA (§ 8.1.1)",
                _NOMI_TI, per_pres,
                fuori_area=_fa_ti,
                uo_fuori_area=_uo_fa_ti)

    # ==================================================================
    # AREA SALE OPERATORIE (§ 8.1.2)
    # ==================================================================
    if fabb_so_per_presidio:
        row = _scrivi_divisore_area(
            ws, row + 1,
            "AREA SALE OPERATORIE (§ 8.1.2)")
        per_pres = []
        for cd in cities_data:
            if (cd['presidio']
                    and fabb_so_per_presidio.get(cd['presidio'])):
                row = _scrivi_tabella_agenas_sale_operatorie(
                    ws, row, cd['df'],
                    fabb_so_per_presidio[cd['presidio']],
                    mapping_uo_so, mapping_profili_so,
                    cd['presidio'],
                )
                srv_ti, srv_td = _conta_sale_operatorie(
                    cd['df'], mapping_uo_so, mapping_profili_so)
                per_pres.append({
                    'fabb': fabb_so_per_presidio[cd['presidio']],
                    'srv_ti': srv_ti,
                    'srv_td': srv_td,
                })
        if per_pres:
            _fa_so, _uo_fa_so = _calc_fuori(
                'sale_operatorie', per_pres)
            row = _scrivi_totale_fte(
                ws, row + 1,
                "AREA SALE OPERATORIE (§ 8.1.2)",
                _NOMI_SO, per_pres,
                fuori_area=_fa_so,
                uo_fuori_area=_uo_fa_so)

    # ==================================================================
    # AREA ANESTESIA, T.I. E BLOCCO OPERATORIO (combinata)
    # ==================================================================
    if fabb_ti_per_presidio or fabb_so_per_presidio:
        row = _scrivi_divisore_area(
            ws, row + 1,
            "AREA ANESTESIA, T.I. E BLOCCO OPERATORIO")
        # Tabellina distribuzione anestesisti per CDC
        row = _scrivi_distribuzione_anestesisti(
            ws, row, personale_all_df)
        per_pres_combo = []
        for cd in cities_data:
            if cd['presidio'] and (
                    fabb_ti_per_presidio.get(cd['presidio'])
                    or fabb_so_per_presidio.get(cd['presidio'])):
                row = _scrivi_tabella_agenas_area_ti_bo(
                    ws, row, cd['df'],
                    fabb_ti_per_presidio.get(cd['presidio'], {}),
                    fabb_so_per_presidio.get(cd['presidio'], {}),
                    mapping_uo_ti, cd['presidio'],
                )
                srv_ti, srv_td = _conta_ti_bo(
                    cd['df'], mapping_uo_ti)
                per_pres_combo.append({
                    'presidio': cd['presidio'],
                    'srv_ti': srv_ti,
                    'srv_td': srv_td,
                })
        if per_pres_combo:
            # fuori_area combo: usa discipline_db di TI (ANESTESIA →
            # DIRIGENTE_MEDICO) che corrisponde alla chiave combo
            _fa_combo, _uo_fa_combo = _calc_fuori(
                'terapia_intensiva', per_pres_combo)
            row = _scrivi_totale_fte_combo(
                ws, row + 1, per_pres_combo,
                fabb_ti_per_presidio or {},
                fabb_so_per_presidio or {},
                fuori_area=_fa_combo,
                uo_fuori_area=_uo_fa_combo,
            )

    # ==================================================================
    # SALUTE MENTALE ADULTI (territoriale)
    # ==================================================================
    if fabb_salute_mentale:
        row = _scrivi_divisore_area(ws, row + 1,
                                    "SALUTE MENTALE ADULTI")
        uo_pats_sm = indicatori_salute_mentale.get(
            'unita_operative', [])
        per_area = []
        for cd in cities_data:
            citta_upper = cd['citta'].upper()
            if fabb_salute_mentale.get(citta_upper):
                row = _scrivi_tabella_agenas_territoriale(
                    ws, row, cd['df'],
                    indicatori_salute_mentale,
                    fabb_salute_mentale[citta_upper],
                    citta_upper,
                )
                fabb_area = fabb_salute_mentale[citta_upper]
                profili = fabb_area['profili']
                srv_list = _conta_territoriale(
                    cd['df'], uo_pats_sm, profili)
                per_area.append({
                    'fabb': fabb_area,
                    'servizio': srv_list,
                })
        if per_area:
            row = _scrivi_totale_territoriale(
                ws, row + 1, indicatori_salute_mentale, per_area)

    # ==================================================================
    # DIPENDENZE PATOLOGICHE (SerD)
    # ==================================================================
    if fabb_dipendenze:
        row = _scrivi_divisore_area(ws, row + 1,
                                    "DIPENDENZE PATOLOGICHE (SerD)")
        uo_pats_dip = indicatori_dipendenze.get(
            'unita_operative', [])
        per_area = []
        for cd in cities_data:
            citta_upper = cd['citta'].upper()
            if fabb_dipendenze.get(citta_upper):
                row = _scrivi_tabella_agenas_territoriale(
                    ws, row, cd['df'],
                    indicatori_dipendenze,
                    fabb_dipendenze[citta_upper],
                    citta_upper,
                )
                fabb_area = fabb_dipendenze[citta_upper]
                profili = fabb_area['profili']
                srv_list = _conta_territoriale(
                    cd['df'], uo_pats_dip, profili)
                per_area.append({
                    'fabb': fabb_area,
                    'servizio': srv_list,
                })
        if per_area:
            row = _scrivi_totale_territoriale(
                ws, row + 1, indicatori_dipendenze, per_area)

    # ==================================================================
    # NEUROPSICHIATRIA INFANZIA E ADOLESCENZA (NPIA)
    # ==================================================================
    if fabb_npia:
        row = _scrivi_divisore_area(
            ws, row + 1,
            "NEUROPSICHIATRIA INFANZIA E ADOLESCENZA")
        uo_pats_npia = indicatori_npia.get('unita_operative', [])
        per_area = []
        for cd in cities_data:
            citta_upper = cd['citta'].upper()
            if fabb_npia.get(citta_upper):
                row = _scrivi_tabella_agenas_territoriale(
                    ws, row, cd['df'],
                    indicatori_npia,
                    fabb_npia[citta_upper],
                    citta_upper,
                )
                fabb_area = fabb_npia[citta_upper]
                profili = fabb_area['profili']
                srv_list = _conta_territoriale(
                    cd['df'], uo_pats_npia, profili)
                per_area.append({
                    'fabb': fabb_area,
                    'servizio': srv_list,
                })
        if per_area:
            row = _scrivi_totale_territoriale(
                ws, row + 1, indicatori_npia, per_area)

    # ==================================================================
    # SALUTE IN CARCERE
    # ==================================================================
    if fabb_carcere:
        row = _scrivi_divisore_area(ws, row + 1,
                                    "SALUTE IN CARCERE")
        uo_pats_carc = indicatori_carcere.get(
            'unita_operative', [])
        per_area = []
        for cd in cities_data:
            citta_upper = cd['citta'].upper()
            if fabb_carcere.get(citta_upper):
                row = _scrivi_tabella_agenas_territoriale(
                    ws, row, cd['df'],
                    indicatori_carcere,
                    fabb_carcere[citta_upper],
                    citta_upper,
                )
                fabb_area = fabb_carcere[citta_upper]
                profili = fabb_area['profili']
                srv_list = _conta_territoriale(
                    cd['df'], uo_pats_carc, profili)
                per_area.append({
                    'fabb': fabb_area,
                    'servizio': srv_list,
                })
        if per_area:
            row = _scrivi_totale_territoriale(
                ws, row + 1, indicatori_carcere, per_area)

    # ------------------------------------------------------------------
    # Larghezza colonne automatica
    # ------------------------------------------------------------------
    auto_larghezza_colonne(ws)

    return ws
